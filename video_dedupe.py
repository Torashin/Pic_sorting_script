# video_dedupe.py
#
# Deduplication pipeline with selectable refine modes:
#   - "anchors": start/end anchors (subset-aware).
#   - "legacy":  sparse sampler + aligner with FAISS shortlist.
#   - "both":    FAISS shortlist + anchors/legacy union.
#   - "audio":   audio fingerprint offset-voting.
#   - "timeline": full-timeline pHash offset-voting.
#
# Cache layout:
#   - resources/cache/video_cache.json (legacy/anchors/audio/timeline entries).
#   - reports/ for Excel outputs.
#
# Tuning:
#   - See the "Knobs ? tweak here" section below for parameter details and defaults.
#
# Requirements:
#   - FFmpeg on PATH.
#   - OpenCV, numpy, pandas, faiss-cpu, imagehash, pillow, openpyxl, PySimpleGUI.

from __future__ import annotations

import os
import json
import glob
import atexit
import time
import datetime
from pathlib import Path
from typing import List, Tuple, Dict, Callable
import concurrent.futures
import math
import re
import subprocess
import platform
import sys
import multiprocessing as mp
import tempfile
from collections import Counter

import numpy as np
import pandas as pd
import cv2
from PIL import Image
import imagehash
import faiss
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers
from openpyxl import load_workbook
from pandas import ExcelWriter
import PySimpleGUI as sg

# Audio helpers  estimate head/tail silence to trim dead air (fast).
from audio_processing import (
    detect_head_tail_silence_ffmpeg,
    merge_close_regions,
    compute_audio_fingerprints,
    ffprobe_duration_seconds,
)

try:
    from camcorder_timestamp_ocr import (
        _winrt_ocr_text as _osd_ocr_text,
        _timestamp_roi_tophat as _osd_roi_tophat,
        _roi_likely_has_timestamp_text as _osd_roi_has_timestamp_like_text,
        _extract_month as _osd_extract_month,
        _extract_year as _osd_extract_year,
        _extract_day as _osd_extract_day,
        _extract_time as _osd_extract_time,
    )
except Exception:
    _osd_ocr_text = None
    _osd_roi_tophat = None
    _osd_roi_has_timestamp_like_text = None
    _osd_extract_month = None
    _osd_extract_year = None
    _osd_extract_day = None
    _osd_extract_time = None


# 
# Paths & layout
# 

RESOURCES_DIR = os.path.join("resources")
CACHE_DIR     = os.path.join(RESOURCES_DIR, "cache")
REPORTS_DIR   = os.path.join("reports")

os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

# Unified cache file: legacy avg/seq AND anchor hashes live here.
VIDEO_HASH_STORE_PATH   = os.path.join(CACHE_DIR, "video_cache.json")
DEFAULT_REPORT_PATH     = os.path.join(REPORTS_DIR, "video_duplicates.xlsx")


# 
# Knobs ? tweak here
#
# How to tune (high level):
# - More recall: loosen thresholds (higher Hamming), lower min votes/overlap, increase sampling density.
# - Fewer false positives: tighten thresholds, raise min votes/overlap, reduce sampling density.
# - HDDs: keep worker counts low and avoid small step sizes.
#
# Legacy sparse sampler & FAISS coarse gate
MAX_SAMPLES          = 40    # max sparse frames per video (legacy); raise for recall, slower.
MIN_SAMPLES          = 5     # min sparse frames per video (legacy); raise to stabilize short clips.
FAISS_THRESHOLD      = 12    # L2 distance on 64-bit avg hash; raise for noisy analog.
ALIGN_THRESHOLD      = 20.0  # mean Hamming over 60 s; lower = stricter.
ALIGN_OFFSET_LIMIT_S = 60.0  # max allowed time offset (s); raise if trims drift.
TOP_K                = 5     # FAISS neighbours per video; increase if matches are missed.
EXTS                 = {'.mp4', '.mov', '.avi', '.m4v', '.mpg', '.mkv'}  # file extensions to scan.
SAVE_EVERY           = 5     # save cache every N new legacy entries.
CACHE_AUTOSAVE_SECONDS = 30  # flush dirty cache periodically during long runs.
MAX_WORKERS          = 8     # workers for metadata/legacy; keep low for HDDs.

# Anchor mode (subset-aware)
ANCHOR_WINDOW_S       = 120.0   # seconds sampled at start and end; raise for long clips.
ANCHOR_STEP_S         = 1.0     # seconds between frames; lower = more accurate, slower.
ANCHOR_HAMMING_THRESH = 18      # per-frame pHash Hamming; higher tolerates noise.
ANCHOR_MIN_FRACTION   = 0.25    # min fraction of frames that must match; lower = more permissive.
ANCHOR_MAX_MAD_S      = 5.0     # max allowed offset scatter (s); higher = looser alignment.
DURATION_RATIO_MIN    = 0.20    # skip if min/max duration < this; lower allows bigger mismatches.
SHORT_CLIP_DURATION_RELAX_MAX_S = 180.0  # for very short clips, relax duration-ratio gating instead of hard-blocking.
SHORT_CLIP_DURATION_RATIO_MIN   = 0.03   # lets ~2 min clips still compare against long transfers when evidence is strong.

# FFmpeg settings (anchors/timeline)
FFMPEG_THREADS        = 1       # 0 = auto; use 1 for HDDs.
ANCHOR_WORKERS        = 1       # concurrent FFmpeg readers for anchors.

# Audio fingerprinting (offset voting)
AUDIO_TARGET_SR       = 8000    # decode sample rate (Hz).
AUDIO_WIN_S           = 1.0     # window size (s).
AUDIO_HOP_S           = 1.0     # hop size (s); lower = denser hashes.
AUDIO_N_BANDS         = 16      # spectral bands; raise for more detail.
AUDIO_MIN_FREQ        = 200.0   # low cutoff (Hz).
AUDIO_MAX_FREQ        = 3400.0  # high cutoff (Hz).
AUDIO_RMS_THRESH_DB   = -65.0   # keep quiet sections, but still drop near-silence floor.
AUDIO_HAMMING_THRESH  = 13      # loose pass threshold for offset voting.
AUDIO_STRICT_HAMMING_THRESH = 10  # strict pass threshold for contiguous verification.
AUDIO_BIN_S           = 1.0     # offset histogram bin width; larger = more forgiving.
AUDIO_MIN_VOTES       = 8       # minimum loose votes in best bin.
AUDIO_MIN_VOTE_FRACTION = 0.018 # dynamic min votes relative to shorter sequence length.
AUDIO_MIN_OVERLAP_S   = 45.0    # minimum loose contiguous overlap (s).
AUDIO_STRICT_MIN_OVERLAP_S = 45.0  # strict contiguous overlap gate (s).
AUDIO_PEAK_RATIO_MIN  = 1.06    # best offset-bin votes must dominate second-best by this ratio.
AUDIO_PEAK_MARGIN     = 1       # and exceed second-best by this many votes.
AUDIO_LSH_CHUNKS      = 7       # balance collisions and recall.
AUDIO_BRUTE_MAX       = 2_000_000  # avoid expensive O(N*M) on long clips.
AUDIO_PAIR_WORKERS    = 0       # 0 = auto; audio pair-eval is CPU-only, so use a few workers.
AUDIO_PAIR_WORKERS_AUTO_MAX = 6  # cap auto worker count to avoid memory blowups on large pair sets.
AUDIO_MIN_BITCOUNT    = 3       # drop low-information hashes.
AUDIO_MAX_BITCOUNT    = 28      # drop low-information hashes.
AUDIO_MIN_FILTERED_FRAMES = 15  # skip clips with too few usable fingerprints.
AUDIO_MIN_UNIQUE_RATIO = 0.05   # skip clips dominated by repetitive hashes.
AUDIO_DURATION_RATIO_MIN = 0.10 # skip pair if min/max duration ratio is below this.
AUDIO_MIN_HASHSET_INTERSECT_RATIO = 0.08  # coarse gate on shared unique hashes (faster + fewer random pairs).
# Playback-speed drift tolerance (consumer camcorder vs pro deck transfers).
AUDIO_SPEED_RATIO_MIN = 0.881
AUDIO_SPEED_RATIO_MAX = 1.041
AUDIO_SPEED_STEPS     = 13
AUDIO_MAX_CANDIDATES_PER_FRAME = 2000  # downsample overly-colliding hashes instead of skipping.
# Rescue true long overlaps when offset peak is ambiguous (repetitive audio can split votes).
AUDIO_LONG_OVERLAP_OVERRIDE_S = 380.0      # seconds of strict contiguous overlap to bypass peak-dominance gate.
AUDIO_LONG_OVERLAP_VOTE_MULT  = 1.0        # strict votes must be at least this x dynamic min-votes.
AUDIO_MUTUAL_OVERLAP_RATIO_MIN = 0.55      # require overlap support on both A/B timelines, not one-sided only.

# Timeline video fingerprints (offset voting)
TIMELINE_STEP_S         = 1.0   # sampling interval (s); lower = more accurate, slower.
TIMELINE_HAMMING_THRESH = 16    # loose per-frame threshold for offset voting.
TIMELINE_STRICT_HAMMING_THRESH = 13  # strict threshold for contiguous verification.
TIMELINE_BIN_S          = 1.0   # offset histogram bin width.
TIMELINE_MIN_VOTES      = 50    # minimum loose votes in best bin.
TIMELINE_MIN_VOTE_FRACTION = 0.03  # dynamic min votes relative to shorter sequence length.
TIMELINE_MIN_OVERLAP_S  = 120.0 # required loose contiguous overlap length (s); capped per pair for short clips.
TIMELINE_STRICT_MIN_OVERLAP_S = 95.0  # strict contiguous overlap gate (also capped per pair for short clips).
TIMELINE_PEAK_RATIO_MIN = 1.15  # best bin votes must dominate second-best by this ratio.
TIMELINE_PEAK_MARGIN    = 1     # and by at least this many votes.
TIMELINE_LSH_CHUNKS     = 8    # more chunks = fewer collisions, potentially fewer candidates.
TIMELINE_BRUTE_MAX      = 15_000_000     # brute-force if pair size <= this.
TIMELINE_MAX_CANDIDATES_PER_FRAME = 1500  # downsample huge LSH candidate sets instead of skipping.
# Speed-ratio search (consumer/pro playback drift). Use explicit toggle.
TIMELINE_ENABLE_SPEED_SWEEP = False  # if False, timeline uses fixed ratio only.
TIMELINE_SPEED_RATIO_FIXED = 1.0
TIMELINE_SPEED_RATIO_MIN = 0.997
TIMELINE_SPEED_RATIO_MAX = 1.003
TIMELINE_SPEED_STEPS     = 7
TIMELINE_LONG_OVERLAP_OVERRIDE_S = 139.7  # accept very long strict overlap even if peak is split.
TIMELINE_LONG_OVERLAP_VOTE_MULT  = 1.0
TIMELINE_MIN_BITCOUNT   = 6     # filter low-information hashes (black/static).
TIMELINE_MAX_BITCOUNT   = 58     # filter overexposed/flat hashes.
TIMELINE_MIN_FILTERED_FRAMES = 30  # skip clips with too few usable timeline hashes.
TIMELINE_MIN_UNIQUE_RATIO = 0.05   # skip clips dominated by repetitive timeline hashes.
TIMELINE_DURATION_RATIO_MIN = 0.10 # min/max duration sanity gate for timeline candidates.
TIMELINE_PAIR_WORKERS   = 0     # 0 = auto (cpu_count-1). Use 1 to disable multiprocessing.
TIMELINE_POSTFIT_ENABLE = True  # refine accepted timeline pairs with drift fit + segment extraction.
TIMELINE_POSTFIT_MAX_MATCHES = 250_000  # cap detailed match points per pair to bound CPU/memory.
TIMELINE_POSTFIT_SEED_GAP_S = 60.0  # small fixed merge before gaussian smoothing creates the seed mask.
TIMELINE_POSTFIT_GAUSS_SIGMA_S = 60.0  # Broad smoothing to bridge long noisy valleys between true segments.
TIMELINE_POSTFIT_GAUSS_SEED_BLEND = 0.0  # 0=only evidence, 1=only seed prior.
TIMELINE_POSTFIT_GAUSS_SCORE_MIN = 0.02  # Lower threshold for continuity; raise (~0.03-0.05) to be stricter.
TIMELINE_POSTFIT_GAUSS_RAW_CLAMP_MIN = 0.04  # Clamp final segment ends to bins with at least this raw evidence.
TIMELINE_POSTFIT_LOCAL_SEARCH_TOL_S = 18.0  # Allow local time wobble around the fitted line when building dense support.
TIMELINE_POSTFIT_ALT_MAX_LINES = 8  # Probe a few strong secondary offset lines for piecewise overlaps.
TIMELINE_POSTFIT_ALT_OFFSET_SEP_S = 5.0  # Ignore alternative lines that are effectively the same offset regime.
TIMELINE_POSTFIT_ALT_MIN_STRICT_S = 90.0  # Secondary line must contribute meaningful strict overlap.
TIMELINE_POSTFIT_ALT_MIN_ADDED_S = 45.0  # Secondary line must add this much new A-side coverage.
TIMELINE_POSTFIT_ALT_MAX_B_OVERLAP_FRAC = 0.20  # Reject secondary lines that mostly re-explain the same B region.

# Contiguous-run tolerance for overlap estimation.
# Higher values are more tolerant to dropped/noisy frames but may raise false links if too high.
AUDIO_RUN_GAP_MULT = 3.0
TIMELINE_RUN_GAP_MULT = 4.21

# Consolidated timeline-coverage estimate knobs (reporting only; does not affect match detection).
# For each file_b, coverage intervals are merged if the gap is <= gap_tol_s, where:
#   gap_tol_s = clamp(TIMELINE_COVERAGE_BRIDGE_FRAC_OF_B * duration_b_s,
#                     TIMELINE_COVERAGE_BRIDGE_MIN_S,
#                     TIMELINE_COVERAGE_BRIDGE_MAX_S)
# This lets us treat short "holes" between nearby matched sections as continuous coverage.
TIMELINE_COVERAGE_BRIDGE_MIN_S = 90.0   # floor on bridge tolerance, even for short file_b clips.
TIMELINE_COVERAGE_BRIDGE_FRAC_OF_B = 0.20  # main scaler: allow bridging up to 20% of file_b duration.
TIMELINE_COVERAGE_BRIDGE_MAX_S = 900.0  # cap on bridge tolerance for very long file_b clips.
TIMELINE_HASH_EDGE_TRIM_MIN_S = 20.0  # infer trim from timeline-hash support only if edge gap is at least this long.
# Coverage labeling/flags:
# Treat as "fully covered" if either:
# - coverage percent is at least this threshold, OR
# - unmatched tail/head leftovers are at most this many seconds.
TIMELINE_COVERAGE_FULL_MIN_PCT = 97
TIMELINE_COVERAGE_FULL_MAX_MISSING_S = 10.0
TIMELINE_COVERAGE_MOSTLY_MIN_PCT = 94.0
# Create remux-planning rows only for unmatched A-side unique segments at or above
# this duration. Shorter leftovers are usually not worth preserving as standalone clips.
REMUX_PLAN_MIN_UNIQUE_SEGMENT_S = 30.0
REMUX_PLAN_SHORT_MIN_UNIQUE_SEGMENT_S = 5.0
# Exclude long internal dead sections (typically black + near-silent tape gaps)
# from coverage denominators. Derived from long gaps in cached audio fingerprints.
TIMELINE_INTERNAL_DEAD_ENABLE = True
TIMELINE_INTERNAL_DEAD_AUDIO_GAP_MIN_S = 5.0
TIMELINE_INTERNAL_DEAD_EDGE_GUARD_S = 30.0
TIMELINE_INTERNAL_DEAD_VISUAL_SAMPLE_COUNT = 3  # start / middle / end sparse frame checks per candidate gap.
TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_W = 320
TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_H = 240
TIMELINE_INTERNAL_DEAD_BLACK_MEAN_MAX = 18.0  # near-black frame brightness threshold.
TIMELINE_INTERNAL_DEAD_BLACK_STD_MAX = 18.0  # near-black frame variation threshold.
TIMELINE_INTERNAL_DEAD_NOISE_SAT_MAX = 35.0  # tape snow is usually near-monochrome.
TIMELINE_INTERNAL_DEAD_NOISE_STD_MIN = 20.0  # tape snow still has visible luminance variation.
TIMELINE_INTERNAL_DEAD_NOISE_NEIGHBOR_CORR_MAX = 0.35  # random noise has weak pixel-to-pixel correlation.
TIMELINE_INTERNAL_DEAD_STRIPED_SAT_MAX = 12.0  # monochrome striped tape static has very low saturation.
TIMELINE_INTERNAL_DEAD_STRIPED_STD_MIN = 25.0  # but still carries substantial luminance variation.
TIMELINE_INTERNAL_DEAD_STRIPED_CORR_HIGH_MIN = 0.55  # one axis keeps noticeable line continuity.
TIMELINE_INTERNAL_DEAD_STRIPED_CORR_LOW_MAX = 0.35  # the orthogonal axis remains weakly correlated.
TIMELINE_INTERNAL_DEAD_STRIPED_ANISOTROPY_MIN = 0.40  # require a strong directional imbalance to avoid live footage.
TIMELINE_INTERNAL_DEAD_CACHE_VERSION = 2
# Targeted reconciliation for anchor-full<->full vs timeline-partial conflicts.
# We do not blindly trust anchors. When timeline placement looks too rigid, we
# optionally recover extra A-side coverage from cached timeline hashes, without
# expanding B-side coverage semantics.
TIMELINE_GAP_VERIFY_ENABLE = True
TIMELINE_GAP_VERIFY_MIN_GAP_S = 10.0      # only probe substantial uncovered gaps.
TIMELINE_GAP_VERIFY_PROBES_PER_GAP = 3    # evenly spaced probes per uncovered gap.
TIMELINE_GAP_VERIFY_SEARCH_RADIUS_S = 4.0 # search near predicted B time to absorb seek/drift jitter.
TIMELINE_GAP_VERIFY_SEARCH_STEP_S = 1.0   # B-side search step within the local window.
TIMELINE_GAP_VERIFY_SCORE_MIN = 0.84      # frame-descriptor similarity needed to treat a probe as matched.
TIMELINE_GAP_VERIFY_MIN_PASS_FRACTION = 0.67  # fraction of probes in a gap that must pass.
TIMELINE_GAP_VERIFY_MAX_ROWS = 4          # cap slow pair-level verification during consolidation.
REMUX_BORDERLINE_SEARCH_STEP_S = 1.0
REMUX_BORDERLINE_REFINE_RADIUS_S = 4.0
REMUX_BORDERLINE_REFINE_STEP_S = 0.5
REMUX_BORDERLINE_PROBE_COUNT = 3
REMUX_BORDERLINE_SCORE_MIN = 0.50
REMUX_BORDERLINE_SCORE_MARGIN = 0.05
REMUX_BORDERLINE_MAX_SEGMENT_S = 30.0
# Approximate self-repeat detection for file_a unique-content preservation.
# Later repeated playback in file_a is treated as redundant when estimating
# whether unique content from A is preserved somewhere in B.
TIMELINE_A_SELF_REPEAT_ENABLE = True
TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S = 20.0
TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S = 30.0
TIMELINE_A_SELF_REPEAT_MIN_BIN_VOTES = 18
TIMELINE_A_SELF_REPEAT_BIN_NEIGHBOR_BINS = 1
# Self-repeat matching can be looser than cross-file matching.
TIMELINE_A_SELF_REPEAT_HAMMING_THRESH = 20
# Trigger only when most of A is already covered and leftover edge gap is suspicious.
TIMELINE_A_SELF_REPEAT_TRIGGER_MIN_COVERAGE_PCT = 70.0
# Downsample very long timeline sequences for self-repeat detection to bound runtime.
TIMELINE_A_SELF_REPEAT_MAX_SEQ_POINTS = 1800
TIMELINE_A_SELF_REPEAT_CACHE_VERSION = 1
# If both A and B have unresolved edge gaps (start/end) on a strong match, credit
# the shared edge gaps as unresolved common content (timeline under-detection),
# not A-unique loss.
TIMELINE_A_COMMON_EDGE_RECOVERY_ENABLE = True
TIMELINE_A_COMMON_EDGE_RECOVERY_MIN_CONF = 90.0
TIMELINE_A_COMMON_EDGE_RECOVERY_FULL_IF_RESIDUAL_LE_S = 90.0

# Camcorder on-screen timestamp OCR scan (optional).
# This is independent of dedupe matching and can be run standalone or with other scans.
TIMESTAMP_SCAN_STEP_S = 9.0     # sample every N seconds across trimmed video.
TIMESTAMP_SCAN_FRAME_W = 768      # decode width before ROI extraction.
TIMESTAMP_SCAN_FRAME_H = 576      # decode height before ROI extraction.
TIMESTAMP_SCAN_WORKERS = 0        # CPU OCR workers for timestamp pipeline (I/O stage is always 1 sequential reader).
TIMESTAMP_SCAN_WORKERS_AUTO_MAX = 16  # cap auto CPU workers to avoid excessive memory pressure.
TIMESTAMP_OCR_BATCH_FRAMES = 0     # 0 = auto from inflight budget; higher = less overhead, more memory.
TIMESTAMP_OCR_MAX_INFLIGHT_BATCHES = 0  # 0 = auto (2x workers) to bound queued-frame memory.
TIMESTAMP_PROGRESS_HEARTBEAT_S = 20.0  # concise status heartbeat interval for timestamp pipeline.
TIMESTAMP_PROGRESS_FILE_TICK = 5      # print additional read progress every N files per stage.
TIMESTAMP_YEAR_MIN = 1980         # clamp parsed years to plausible camcorder range.
TIMESTAMP_YEAR_MAX = 2035
TIMESTAMP_MIN_REPEAT_FOR_ROBUST = 3  # prefer month-year values seen at least this many times.
TIMESTAMP_MONTH_ONLY_INFER_NEARBY_S = 1800.0  # infer year for month-only OCR hits from nearby full-date hits.
TIMESTAMP_EDGE_SINGLETON_MAX_MONTH_JUMP = 3  # allow a single edge month to extend the range by up to this many months.
TIMESTAMP_SUMMARY_CLUSTER_MAX_GAP_MONTHS = 2  # split sparse OCR month-years into clusters if separated by larger gaps.
TIMESTAMP_SUMMARY_MIN_CLUSTER_SHARE = 0.35  # keep a distant month-year cluster only if it has a meaningful share of support.
TIMESTAMP_SUMMARY_MIN_DISTINCT_MONTHS_FOR_ALT_CLUSTER = 2  # keep distant clusters with multiple corroborating months.
TIMESTAMP_SUMMARY_MIN_EVENT_SPAN_S_FOR_ALT_CLUSTER = 120.0  # keep distant single-month clusters only if they recur over a meaningful span.
TIMESTAMP_SUMMARY_VERSION = 4  # bump when timestamp candidate summarization logic changes without requiring rescans.
TIMESTAMP_CACHE_VERSION = 9       # bump when OCR / ROI / timestamp parsing logic changes.
TIMESTAMP_BURST_ENABLE = True     # after coarse pass, densify locally around text hits.
TIMESTAMP_BURST_WINDOW_S = 3.0    # scan this much before/after each coarse hit.
TIMESTAMP_BURST_STEP_S = 1.0      # local dense sampling interval in seconds.
TIMESTAMP_BURST_MAX_WINDOWS = 120 # cap local windows per file to bound runtime.
TIMESTAMP_BURST_MERGE_GAP_S = 12.0  # merge nearby burst windows to avoid lots of tiny FFmpeg seeks.
TIMESTAMP_BURST_DRAIN_EVERY_WINDOWS = 4  # let several burst windows queue before forcing a drain.

TIMESTAMP_SUMMARY_FIELDS = (
    "osd_month_year_start",
    "osd_month_year_end",
    "osd_month_year_first_in_video",
    "osd_month_year_last_in_video",
    "osd_month_year_span_months",
    "osd_out_of_order_in_video",
    "osd_unique_month_year_count",
    "osd_top_month_year_counts",
)

def _hex_to_vec(hex_str: str) -> np.ndarray:
    """Convert a 64-bit hex string into a 64-dim 0/1 float32 vector (legacy FAISS embedding)."""
    ba   = bytes.fromhex(hex_str)
    bits = np.unpackbits(np.frombuffer(ba, dtype=np.uint8))
    return bits.astype("float32")


def _calculate_target_samples(duration: float) -> int:
    """Legacy: choose sparse sample count via log curve; clamp to [MIN_SAMPLES, MAX_SAMPLES]."""
    if duration <= 0:
        return MIN_SAMPLES
    samples = int(math.log(duration + 1, 10) * 12)
    return max(MIN_SAMPLES, min(samples, MAX_SAMPLES))


def _average_hex(pairs: List[Tuple[str, float]]) -> str:
    """Legacy: average many 64-bit pHashes (hex) into one 64-bit hex by bit-wise majority."""
    hexes = [h for (h, _) in pairs]
    if not hexes:
        return "0" * 16
    bits = np.stack([_hex_to_vec(h) for h in hexes])
    avg_bits = (bits.mean(axis=0) >= 0.5).astype(np.uint8)
    return np.packbits(avg_bits).tobytes().hex()


def _hamming_hex(a: str, b: str) -> int:
    """Legacy: Hamming distance between two 64-bit pHashes stored as hex strings."""
    return bin(int(a, 16) ^ int(b, 16)).count("1")


def _format_mmss(seconds: float | None) -> str | None:
    """Pretty mm:ss for subset offsets; returns None if value is None/NaN/inf."""
    if seconds is None or not math.isfinite(seconds):
        return None
    s = int(round(seconds))
    m, s = divmod(abs(s), 60)
    sign = "-" if seconds < 0 else ""
    return f"{sign}{m:d}:{s:02d}"


def _format_duration_auto(seconds: float | None) -> str:
    """
    Human-readable duration for logs:
      - < 60 s: seconds
      - < 60 min: minutes
      - otherwise: hours
    """
    if seconds is None:
        return "n/a"
    try:
        s = float(seconds)
    except Exception:
        return "n/a"
    if not math.isfinite(s):
        return "n/a"
    s = max(0.0, s)
    if s < 60.0:
        return f"{s:.1f}s"
    if s < 3600.0:
        return f"{(s / 60.0):.1f}m"
    return f"{(s / 3600.0):.1f}h"


def _safe_text_cell(value: object) -> str:
    """Normalize missing-ish values to empty text for Excel-facing string cells."""
    if value is None:
        return ""
    if isinstance(value, float) and np.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "null", "nat"}:
        return ""
    return str(value)


def _month_year_str(year: int | None, month: int | None) -> str:
    if year is None or month is None:
        return ""
    y = int(year)
    m = int(month)
    if y <= 0 or m < 1 or m > 12:
        return ""
    return f"{y:04d}-{m:02d}"


def _month_year_index(year: int, month: int) -> int:
    return int(year) * 12 + int(month)


def _spawn_pool_supported_in_this_launch() -> tuple[bool, str]:
    """
    Windows multiprocessing "spawn" requires a real script path for __main__.
    Interactive launches (stdin / -c) expose virtual paths (<stdin>/<string>)
    and child bootstrap fails. Return (supported, descriptor).
    """
    main_mod = sys.modules.get("__main__")
    main_file = getattr(main_mod, "__file__", "") if main_mod else ""
    descriptor = str(main_file or "")
    if not descriptor:
        return False, "<no-main-file>"
    if "<stdin>" in descriptor or "<string>" in descriptor:
        return False, descriptor
    if descriptor.startswith("<") and descriptor.endswith(">"):
        return False, descriptor
    return True, descriptor


def _resolve_timeline_speed_search(
    enable_sweep: bool,
    ratio_min: float,
    ratio_max: float,
    steps: int,
    *,
    fixed_ratio: float = TIMELINE_SPEED_RATIO_FIXED,
) -> tuple[float, float, int]:
    """
    Resolve timeline speed search settings into effective values.
    Returns (min_ratio, max_ratio, steps). If sweep is disabled or invalid,
    returns fixed-ratio search (fixed, fixed, 1).
    """
    rmin = float(min(ratio_min, ratio_max))
    rmax = float(max(ratio_min, ratio_max))
    n = int(max(1, steps))
    if bool(enable_sweep) and n > 1 and (rmax - rmin) > 1e-6:
        return rmin, rmax, n
    rfix = float(fixed_ratio)
    return rfix, rfix, 1


def _hamming_int(a: int, b: int) -> int:
    return (a ^ b).bit_count()


def _bytes_to_int64(h: bytes) -> int:
    return int.from_bytes(h, byteorder="big", signed=False)


def _filter_hash_seq(
    seq: list[tuple[float, int]],
    *,
    min_bitcount: int | None = None,
    max_bitcount: int | None = None,
    collapse_runs: bool = True,
) -> list[tuple[float, int]]:
    """
    Remove low-information hashes and optionally collapse consecutive duplicate hashes.
    This reduces repeated false votes from steady hum/silence-like segments.
    """
    out: list[tuple[float, int]] = []
    last_h: int | None = None
    for t, h in seq:
        bc = h.bit_count()
        if min_bitcount is not None and bc < min_bitcount:
            continue
        if max_bitcount is not None and bc > max_bitcount:
            continue
        if collapse_runs and last_h == h:
            continue
        out.append((t, h))
        last_h = h
    return out


def _longest_time_run(times: list[float], hop_s: float, *, gap_mult: float = 1.5) -> float:
    if not times:
        return 0.0
    # Unique-and-sort to avoid double-counting duplicate t_a hits.
    seq = sorted(set(times))
    run_start = seq[0]
    run_last = seq[0]
    longest = 0.0
    max_gap = max(float(hop_s), float(hop_s) * max(1.0, float(gap_mult)))
    for t in seq[1:]:
        if t - run_last <= max_gap:
            run_last = t
        else:
            longest = max(longest, run_last - run_start)
            run_start = t
            run_last = t
    longest = max(longest, run_last - run_start)
    return longest + hop_s


def _merge_intervals_simple(intervals: list[tuple[float, float]], *, gap_tolerance_s: float = 0.0) -> list[tuple[float, float]]:
    """Merge [start,end] intervals with optional gap tolerance."""
    if not intervals:
        return []
    segs = sorted((float(a), float(b)) for a, b in intervals if b > a)
    if not segs:
        return []
    out: list[tuple[float, float]] = []
    cur_s, cur_e = segs[0]
    gap = max(0.0, float(gap_tolerance_s))
    for s, e in segs[1:]:
        if s <= (cur_e + gap):
            if e > cur_e:
                cur_e = e
        else:
            out.append((cur_s, cur_e))
            cur_s, cur_e = s, e
    out.append((cur_s, cur_e))
    return out


def _build_audio_lsh(seq: list[tuple[float, int]], total_bits: int, chunks: int) -> tuple[list[dict[int, list[int]]], int]:
    chunk_size = math.ceil(total_bits / chunks)
    mask = (1 << chunk_size) - 1
    tables: list[dict[int, list[int]]] = [dict() for _ in range(chunks)]
    for idx, (_t, h) in enumerate(seq):
        for c in range(chunks):
            val = (h >> (c * chunk_size)) & mask
            bucket = tables[c].setdefault(val, [])
            bucket.append(idx)
    return tables, chunk_size


def _collect_lsh_candidates(
    *,
    hash_value: int,
    tables: list[dict[int, list[int]]],
    chunk_size: int,
    chunks: int,
    max_candidates_per_frame: int | None,
) -> list[int]:
    """
    Collect candidate indices from LSH buckets with bounded memory.
    For very hot buckets, sample deterministically instead of materializing giant unions.
    """
    mask = (1 << chunk_size) - 1
    bucket_lists: list[list[int]] = []
    total_bucket_items = 0
    for c in range(chunks):
        val = (hash_value >> (c * chunk_size)) & mask
        bucket = tables[c].get(val)
        if not bucket:
            continue
        bucket_lists.append(bucket)
        total_bucket_items += len(bucket)

    if not bucket_lists:
        return []

    max_cands = None if max_candidates_per_frame is None else max(1, int(max_candidates_per_frame))
    if max_cands is None:
        candidates: set[int] = set()
        for bucket in bucket_lists:
            candidates.update(bucket)
        return list(candidates)

    # Small enough: build exact union, then downsample if needed.
    if total_bucket_items <= max(20_000, max_cands * 20):
        candidates_exact: set[int] = set()
        for bucket in bucket_lists:
            candidates_exact.update(bucket)
        if len(candidates_exact) <= max_cands:
            return list(candidates_exact)
        cand_list = sorted(candidates_exact)
        step = max(1, len(cand_list) // max_cands)
        return cand_list[::step][:max_cands]

    # Hot-bucket path: deterministic sampling without giant unions.
    picked: set[int] = set()
    per_bucket_target = max(1, max_cands // len(bucket_lists))
    for bucket in bucket_lists:
        blen = len(bucket)
        step = max(1, blen // per_bucket_target)
        start = (blen // 3) % step
        for idx in range(start, blen, step):
            picked.add(bucket[idx])
            if len(picked) >= max_cands:
                break
        if len(picked) >= max_cands:
            break

    if len(picked) < max_cands:
        # Top-up with simple edge picks.
        for bucket in bucket_lists:
            if len(picked) >= max_cands:
                break
            if bucket:
                picked.add(bucket[0])
                if len(picked) >= max_cands:
                    break
                picked.add(bucket[-1])

    if len(picked) > max_cands:
        return sorted(picked)[:max_cands]
    return list(picked)


def _audio_match_offset(
    seq_a: list[tuple[float, int]],
    seq_b: list[tuple[float, int]],
    *,
    total_bits: int,
    hamming_thresh: int,
    bin_s: float,
    min_votes: int,
    min_overlap_s: float,
    hop_s: float,
    lsh_chunks: int,
    brute_limit: int | None = None,
    strict_hamming_thresh: int | None = None,
    strict_min_overlap_s: float | None = None,
    min_vote_fraction: float = 0.0,
    peak_ratio_min: float = 1.0,
    peak_margin: int = 0,
    min_bitcount: int | None = None,
    max_bitcount: int | None = None,
    speed_ratio_min: float = 1.0,
    speed_ratio_max: float = 1.0,
    speed_steps: int = 1,
    max_candidates_per_frame: int | None = None,
    long_overlap_override_s: float | None = None,
    long_overlap_vote_mult: float = 1.0,
    run_gap_mult: float = 1.5,
    mutual_overlap_ratio_min: float = 0.0,
) -> tuple[bool, float | None, int, float]:
    def _seq_span_s(seq_local: list[tuple[float, int]], hop_local: float) -> float:
        if not seq_local:
            return 0.0
        try:
            first_t = float(seq_local[0][0])
            last_t = float(seq_local[-1][0])
        except Exception:
            return 0.0
        return max(float(hop_local), (last_t - first_t) + float(hop_local))

    def _adaptive_overlap_req(requested_s: float, shorter_span_s: float, frac: float) -> float:
        req = max(0.0, float(requested_s))
        shorter = max(0.0, float(shorter_span_s))
        if shorter <= 0.0:
            return req
        return min(req, max(float(hop_s), shorter * max(0.0, float(frac))))

    # Backward-compatible path (used by timeline mode): legacy acceptance gates,
    # with optional candidate downsampling to bound hot buckets.
    if (
        strict_hamming_thresh is None
        and strict_min_overlap_s is None
        and min_vote_fraction <= 0.0
        and peak_ratio_min <= 1.0
        and peak_margin <= 0
        and min_bitcount is None
        and max_bitcount is None
    ):
        if not seq_a or not seq_b:
            return False, None, 0, 0.0

        hist: dict[int, int] = {}
        matches_by_bin: dict[int, list[tuple[float, float]]] = {}

        use_brute = False
        if brute_limit is not None:
            if len(seq_a) * len(seq_b) <= brute_limit:
                use_brute = True

        if use_brute:
            for t_a, h_a in seq_a:
                for t_b, h_b in seq_b:
                    if _hamming_int(h_a, h_b) <= hamming_thresh:
                        offset = t_b - t_a
                        bin_key = int(round(offset / bin_s))
                        hist[bin_key] = hist.get(bin_key, 0) + 1
                        matches_by_bin.setdefault(bin_key, []).append((t_a, t_b))
        else:
            tables, chunk_size = _build_audio_lsh(seq_b, total_bits, lsh_chunks)
            for t_a, h_a in seq_a:
                cand_iter = _collect_lsh_candidates(
                    hash_value=h_a,
                    tables=tables,
                    chunk_size=chunk_size,
                    chunks=lsh_chunks,
                    max_candidates_per_frame=max_candidates_per_frame,
                )
                for idx in cand_iter:
                    t_b, h_b = seq_b[idx]
                    if _hamming_int(h_a, h_b) <= hamming_thresh:
                        offset = t_b - t_a
                        bin_key = int(round(offset / bin_s))
                        hist[bin_key] = hist.get(bin_key, 0) + 1
                        matches_by_bin.setdefault(bin_key, []).append((t_a, t_b))

        if not hist:
            return False, None, 0, 0.0

        shorter_span_s = min(_seq_span_s(seq_a, hop_s), _seq_span_s(seq_b, hop_s))
        eff_min_overlap_s = _adaptive_overlap_req(min_overlap_s, shorter_span_s, frac=0.85)
        best_bin = max(hist, key=hist.get)
        best_votes = hist[best_bin]
        offset_s = best_bin * bin_s
        if best_votes < min_votes:
            return False, offset_s, best_votes, 0.0

        matches = matches_by_bin.get(best_bin, [])
        if not matches:
            return False, offset_s, best_votes, 0.0

        overlap_s = _longest_time_run(
            [t_a for (t_a, _t_b) in matches],
            hop_s=hop_s,
            gap_mult=run_gap_mult,
        )

        if overlap_s < eff_min_overlap_s:
            return False, offset_s, best_votes, overlap_s

        return True, offset_s, best_votes, overlap_s

    # Stricter path for audio mode.
    if min_bitcount is None and max_bitcount is None:
        seq_a_f = seq_a
        seq_b_f = seq_b
    else:
        seq_a_f = _filter_hash_seq(seq_a, min_bitcount=min_bitcount, max_bitcount=max_bitcount, collapse_runs=False)
        seq_b_f = _filter_hash_seq(seq_b, min_bitcount=min_bitcount, max_bitcount=max_bitcount, collapse_runs=False)
    if not seq_a_f or not seq_b_f:
        return False, None, 0, 0.0

    hist: dict[int, int] = {}
    # Store (t_a, t_b, loose_hamming) to support strict second-pass verification.
    matches_by_bin: dict[int, list[tuple[float, float, int]]] = {}

    use_brute = False
    if brute_limit is not None:
        if len(seq_a_f) * len(seq_b_f) <= brute_limit:
            use_brute = True

    if use_brute:
        for t_a, h_a in seq_a_f:
            for t_b, h_b in seq_b_f:
                hd = _hamming_int(h_a, h_b)
                if hd <= hamming_thresh:
                    offset = t_b - t_a
                    bin_key = int(round(offset / bin_s))
                    hist[bin_key] = hist.get(bin_key, 0) + 1
                    row = (t_a, t_b, hd)
                    matches_by_bin.setdefault(bin_key, []).append(row)
    else:
        tables, chunk_size = _build_audio_lsh(seq_b_f, total_bits, lsh_chunks)
        for t_a, h_a in seq_a_f:
            cand_iter = _collect_lsh_candidates(
                hash_value=h_a,
                tables=tables,
                chunk_size=chunk_size,
                chunks=lsh_chunks,
                max_candidates_per_frame=max_candidates_per_frame,
            )
            for idx in cand_iter:
                t_b, h_b = seq_b_f[idx]
                hd = _hamming_int(h_a, h_b)
                if hd <= hamming_thresh:
                    offset = t_b - t_a
                    bin_key = int(round(offset / bin_s))
                    hist[bin_key] = hist.get(bin_key, 0) + 1
                    row = (t_a, t_b, hd)
                    matches_by_bin.setdefault(bin_key, []).append(row)

    if not hist:
        return False, None, 0, 0.0

    dyn_min_votes = max(
        int(min_votes),
        int(math.ceil(float(min_vote_fraction) * min(len(seq_a_f), len(seq_b_f))))
    )
    strict_thresh = hamming_thresh if strict_hamming_thresh is None else min(hamming_thresh, strict_hamming_thresh)
    shorter_span_s = min(_seq_span_s(seq_a_f, hop_s), _seq_span_s(seq_b_f, hop_s))
    eff_min_overlap_s = _adaptive_overlap_req(min_overlap_s, shorter_span_s, frac=0.85)
    req_strict_overlap_base = min_overlap_s if strict_min_overlap_s is None else strict_min_overlap_s
    req_strict_overlap = _adaptive_overlap_req(req_strict_overlap_base, shorter_span_s, frac=0.75)
    req_strict_votes = max(6, int(math.ceil(dyn_min_votes * 0.5)))

    def _evaluate_alignment(
        *,
        bins: dict[int, list[tuple[float, float, int]]],
        slope: float,
        best_bin_local: int,
        best_votes_local: int,
        second_votes_local: int,
        min_votes_req: int,
        ratio_req: float,
        margin_req: int,
        overlap_req: float,
        strict_overlap_req: float,
        strict_votes_req: int,
        long_overlap_req: float | None,
        long_votes_mult: float,
        mutual_overlap_ratio_req: float,
    ) -> tuple[bool, float, int, float]:
        offset_local = best_bin_local * bin_s
        if best_votes_local < min_votes_req:
            return False, offset_local, best_votes_local, 0.0

        near: list[tuple[float, float, int]] = []
        for k in (best_bin_local - 1, best_bin_local, best_bin_local + 1):
            near.extend(bins.get(k, []))
        if not near:
            return False, offset_local, best_votes_local, 0.0
        near = [m for m in near if abs((m[1] - (slope * m[0])) - offset_local) <= (1.5 * bin_s)]
        if not near:
            return False, offset_local, best_votes_local, 0.0

        loose_overlap_s = _longest_time_run(
            [t_a for (t_a, _t_b, _hd) in near],
            hop_s=hop_s,
            gap_mult=run_gap_mult,
        )
        if loose_overlap_s < overlap_req:
            return False, offset_local, best_votes_local, loose_overlap_s

        strict = [m for m in near if m[2] <= strict_thresh]
        strict_overlap_a_s = _longest_time_run(
            [t_a for (t_a, _t_b, _hd) in strict],
            hop_s=hop_s,
            gap_mult=run_gap_mult,
        )
        strict_overlap_b_s = _longest_time_run(
            [t_b for (_t_a, t_b, _hd) in strict],
            hop_s=hop_s,
            gap_mult=run_gap_mult,
        )
        slope_abs = max(1e-6, abs(float(slope)))
        strict_overlap_b_norm_s = strict_overlap_b_s / slope_abs
        strict_overlap_s = min(strict_overlap_a_s, strict_overlap_b_norm_s)
        strict_overlap_max_s = max(strict_overlap_a_s, strict_overlap_b_norm_s)
        strict_overlap_balance = strict_overlap_s / max(1e-6, strict_overlap_max_s)

        # Use unique timestamps on both sides to avoid inflated votes from many-to-one collisions.
        strict_unique_a = len({t_a for (t_a, _t_b, _hd) in strict})
        strict_unique_b = len({t_b for (_t_a, t_b, _hd) in strict})
        strict_votes_local = min(strict_unique_a, strict_unique_b)

        if strict_votes_local < strict_votes_req:
            return False, offset_local, strict_votes_local, strict_overlap_s
        if strict_overlap_s < strict_overlap_req:
            return False, offset_local, strict_votes_local, strict_overlap_s
        if strict_overlap_balance < max(0.0, float(mutual_overlap_ratio_req)):
            return False, offset_local, strict_votes_local, strict_overlap_s
        if long_overlap_req is not None and long_overlap_req > 0:
            long_votes_req = max(strict_votes_req, int(math.ceil(min_votes_req * max(0.0, float(long_votes_mult)))))
            if strict_overlap_s >= float(long_overlap_req) and strict_votes_local >= long_votes_req:
                return True, offset_local, strict_votes_local, strict_overlap_s
        if second_votes_local > 0:
            if best_votes_local < int(math.ceil(second_votes_local * max(1.0, ratio_req))):
                return False, offset_local, strict_votes_local, strict_overlap_s
            if (best_votes_local - second_votes_local) < int(max(0, margin_req)):
                return False, offset_local, strict_votes_local, strict_overlap_s
        return True, offset_local, strict_votes_local, strict_overlap_s

    def _rebin_matches_for_slope(
        slope_local: float,
    ) -> tuple[dict[int, int], dict[int, list[tuple[float, float, int]]]]:
        hist_local: dict[int, int] = {}
        bins_local: dict[int, list[tuple[float, float, int]]] = {}
        for rows in matches_by_bin.values():
            for t_a, t_b, hd in rows:
                delta = t_b - (float(slope_local) * t_a)
                k = int(round(delta / bin_s))
                hist_local[k] = hist_local.get(k, 0) + 1
                bins_local.setdefault(k, []).append((t_a, t_b, hd))
        return hist_local, bins_local

    # First try fixed-offset alignment.
    best_bin = max(hist, key=hist.get)
    best_votes = int(hist[best_bin])
    second_votes = max((v for k, v in hist.items() if k != best_bin), default=0)
    ok, offset_s, votes_s, overlap_s = _evaluate_alignment(
        bins=matches_by_bin,
        slope=1.0,
        best_bin_local=best_bin,
        best_votes_local=best_votes,
        second_votes_local=second_votes,
        min_votes_req=dyn_min_votes,
        ratio_req=peak_ratio_min,
        margin_req=peak_margin,
        overlap_req=eff_min_overlap_s,
        strict_overlap_req=req_strict_overlap,
        strict_votes_req=req_strict_votes,
        long_overlap_req=long_overlap_override_s,
        long_votes_mult=long_overlap_vote_mult,
        mutual_overlap_ratio_req=mutual_overlap_ratio_min,
    )
    if ok:
        return True, offset_s, votes_s, overlap_s

    fail_offset, fail_votes, fail_overlap = offset_s, votes_s, overlap_s

    # Rescue fallback: if base alignment has strong evidence but fails strict
    # dominance gates, retry with slightly relaxed gates. This now works even
    # when speed sweep is disabled (speed_steps=1), so fixed-ratio runs do not
    # silently lose otherwise clear matches.
    n_steps = int(max(1, speed_steps))
    s_min = float(min(speed_ratio_min, speed_ratio_max))
    s_max = float(max(speed_ratio_min, speed_ratio_max))
    can_drift = (
        n_steps > 1
        and (s_max - s_min) > 1e-6
        and sum(len(v) for v in matches_by_bin.values()) >= max(10, int(math.ceil(dyn_min_votes * 0.6)))
    )
    strong_near_miss = (
        fail_overlap >= max(45.0, float(eff_min_overlap_s) * 0.75, float(req_strict_overlap) * 0.75)
        and fail_votes >= max(req_strict_votes * 2, int(math.ceil(dyn_min_votes * 1.5)))
    )
    if strong_near_miss:
        relaxed_min_votes = max(8, int(math.ceil(dyn_min_votes * 0.8)))
        relaxed_ratio = max(1.05, float(peak_ratio_min) - 0.15)
        relaxed_margin = max(1, int(peak_margin) - 3)
        relaxed_overlap = max(float(hop_s), float(eff_min_overlap_s) * 0.75)
        relaxed_strict_overlap = max(float(hop_s), float(req_strict_overlap) * 0.75)
        relaxed_strict_votes = max(5, int(math.ceil(req_strict_votes * 0.75)))
        # Timeline mode uses mutual_overlap_ratio_min=0 by design; keep it disabled there.
        if float(mutual_overlap_ratio_min) > 0.0:
            relaxed_mutual_overlap_ratio = max(0.35, float(mutual_overlap_ratio_min) * 0.8)
        else:
            relaxed_mutual_overlap_ratio = 0.0

        best_candidate: tuple[tuple[float, int], float, int, float] | None = None
        slope_candidates: list[float] = [1.0]
        if can_drift:
            slope_candidates.extend(float(s) for s in np.linspace(s_min, s_max, n_steps))
        # Stable dedupe ordering for reproducibility.
        slope_candidates = sorted(set(slope_candidates))

        for slope in slope_candidates:
            hist_d, bins_d_all = _rebin_matches_for_slope(float(slope))
            if not hist_d:
                continue

            best_bin_d = max(hist_d, key=hist_d.get)
            best_votes_d = int(hist_d[best_bin_d])
            second_votes_d = max((v for k, v in hist_d.items() if k != best_bin_d), default=0)
            keep_bins = {best_bin_d - 1, best_bin_d, best_bin_d + 1}
            bins_d = {k: v for k, v in bins_d_all.items() if k in keep_bins}
            ok_d, off_d, votes_d, overlap_d = _evaluate_alignment(
                bins=bins_d,
                slope=float(slope),
                best_bin_local=best_bin_d,
                best_votes_local=best_votes_d,
                second_votes_local=second_votes_d,
                min_votes_req=relaxed_min_votes,
                ratio_req=relaxed_ratio,
                margin_req=relaxed_margin,
                overlap_req=relaxed_overlap,
                strict_overlap_req=relaxed_strict_overlap,
                strict_votes_req=relaxed_strict_votes,
                long_overlap_req=long_overlap_override_s,
                long_votes_mult=long_overlap_vote_mult,
                mutual_overlap_ratio_req=relaxed_mutual_overlap_ratio,
            )
            if not ok_d:
                continue

            score = (float(overlap_d), int(votes_d))
            if best_candidate is None or score > best_candidate[0]:
                best_candidate = (score, off_d, votes_d, overlap_d)

        if best_candidate is not None:
            _score, off_d, votes_d, overlap_d = best_candidate
            return True, off_d, votes_d, overlap_d

    return False, fail_offset, fail_votes, fail_overlap


def _effective_duration_ratio_min(shorter_duration_s: float, base_ratio_min: float) -> float:
    """
    Relax duration-ratio gating for genuinely short clips.

    Otherwise a 1-3 minute excerpt can be rejected before matching even starts
    when compared against a much longer full-tape transfer.
    """
    shorter = max(0.0, float(shorter_duration_s))
    base = max(0.0, float(base_ratio_min))
    if shorter <= float(SHORT_CLIP_DURATION_RELAX_MAX_S):
        return min(base, float(SHORT_CLIP_DURATION_RATIO_MIN))
    return base


def _segments_to_text(segments: list[tuple[float, float]], *, max_items: int = 64) -> str:
    """Compact segment serializer for Excel cells."""
    if not segments:
        return ""
    parts: list[str] = []
    for i, (s0, s1) in enumerate(segments):
        if i >= max_items:
            parts.append("...")
            break
        parts.append(f"{float(s0):.1f}-{float(s1):.1f}")
    return "; ".join(parts)


def _gaussian_smooth_1d(values: np.ndarray, sigma_bins: float) -> np.ndarray:
    """Centered Gaussian smoother with sigma in bins."""
    if values.size == 0:
        return values.copy()
    sigma = float(sigma_bins)
    if sigma <= 0.5:
        return values.astype(np.float64, copy=True)
    radius = int(max(1, math.ceil(4.0 * sigma)))
    x = np.arange(-radius, radius + 1, dtype=np.float64)
    kernel = np.exp(-0.5 * (x / sigma) ** 2)
    ksum = float(np.sum(kernel))
    if ksum <= 0.0:
        return values.astype(np.float64, copy=True)
    kernel /= ksum
    out = np.convolve(values.astype(np.float64), kernel, mode="same")
    # np.convolve(mode="same") returns length max(len(values), len(kernel)).
    # When kernel is wider than the signal (large sigma on short clips), center-crop
    # back to signal length so downstream masks always align.
    n = int(values.size)
    if out.size == n:
        return out
    if out.size > n:
        start = int((out.size - n) // 2)
        return out[start:start + n]
    # Defensive (should not happen): pad edges to expected length.
    pad = n - int(out.size)
    left = pad // 2
    right = pad - left
    return np.pad(out, (left, right), mode="edge")

def _timeline_postfit_details(
    seq_a: list[tuple[float, int]],
    seq_b: list[tuple[float, int]],
    *,
    total_bits: int,
    loose_hamming_thresh: int,
    strict_hamming_thresh: int,
    bin_s: float,
    step_s: float,
    lsh_chunks: int,
    brute_limit: int | None,
    speed_ratio_min: float,
    speed_ratio_max: float,
    speed_steps: int,
    max_candidates_per_frame: int | None,
    run_gap_mult: float,
    max_matches: int,
    seed_gap_s: float,
    return_debug: bool = False,
) -> dict:
    """
    Post-fit accepted timeline pairs with a robust drift-aware line:
      t_b ~= offset + drift_ratio * t_a
    and derive segment-level overlap statistics using gaussian-only bridging.
    """
    out = {
        "offset_s": None,
        "drift_ratio": 1.0,
        "votes_raw": 0,
        "votes_strict": 0,
        "overlap_raw_s": 0.0,
        "overlap_strict_unbridged_s": 0.0,
        "overlap_strict_s": 0.0,
        "segment_count_raw": 0,
        "segment_count_strict_unbridged": 0,
        "segment_count_strict": 0,
        "segments_a_raw": "",
        "segments_b_raw": "",
        "segments_a_strict": "",
        "segments_b_strict": "",
    }
    if return_debug:
        out["debug"] = {}
    if not seq_a or not seq_b:
        return out

    matches: list[tuple[float, float, int]] = []
    use_brute = bool(brute_limit is not None and (len(seq_a) * len(seq_b) <= int(brute_limit)))
    if use_brute:
        for t_a, h_a in seq_a:
            for t_b, h_b in seq_b:
                hd = _hamming_int(h_a, h_b)
                if hd <= loose_hamming_thresh:
                    matches.append((float(t_a), float(t_b), int(hd)))
    else:
        tables, chunk_size = _build_audio_lsh(seq_b, total_bits, lsh_chunks)
        for t_a, h_a in seq_a:
            cand_iter = _collect_lsh_candidates(
                hash_value=h_a,
                tables=tables,
                chunk_size=chunk_size,
                chunks=lsh_chunks,
                max_candidates_per_frame=max_candidates_per_frame,
            )
            for idx in cand_iter:
                t_b, h_b = seq_b[idx]
                hd = _hamming_int(h_a, h_b)
                if hd <= loose_hamming_thresh:
                    matches.append((float(t_a), float(t_b), int(hd)))

    if not matches:
        return out

    max_keep = max(10_000, int(max_matches))
    if len(matches) > max_keep:
        # Keep representative points while preferring lower hamming distances.
        matches.sort(key=lambda r: (r[2], r[0], r[1]))
        step = max(1, len(matches) // max_keep)
        matches = matches[::step][:max_keep]

    s_min = float(min(speed_ratio_min, speed_ratio_max))
    s_max = float(max(speed_ratio_min, speed_ratio_max))
    n_steps = int(max(1, speed_steps))
    if n_steps <= 1 or abs(s_max - s_min) <= 1e-6:
        slope_grid = [1.0]
    else:
        slope_grid = sorted({1.0, *[float(v) for v in np.linspace(s_min, s_max, n_steps)]})

    max_t = max((float(t_a) for t_a, _h in seq_a), default=0.0)
    min_t_b = min((float(t_b) for t_b, _h_b in seq_b), default=0.0)
    max_t_b = max((float(t_b) for t_b, _h_b in seq_b), default=0.0) + float(step_s)
    n_bins = max(1, int(math.ceil(max_t / max(1e-6, float(step_s)))) + 2)
    b_times = np.array([float(t_b) for t_b, _h_b in seq_b], dtype=np.float64)
    b_hashes = [int(h_b) for _t_b, h_b in seq_b]
    align_tol_s = max(
        float(step_s) * 2.0,
        float(bin_s) * 1.5,
        float(TIMELINE_POSTFIT_LOCAL_SEARCH_TOL_S),
    )
    search_radius_bins = max(1, int(math.ceil(align_tol_s / max(1e-6, float(step_s)))))
    dense_max_hd = max(int(loose_hamming_thresh) + 4, int(strict_hamming_thresh) + 4)

    def _build_line_candidate(
        *,
        slope_local: float,
        bin_key_local: int,
        votes_local: int,
    ) -> dict | None:
        offset_guess = float(bin_key_local * bin_s)
        near_local = [
            m for m in matches
            if abs((m[1] - (float(slope_local) * m[0])) - offset_guess) <= (1.5 * bin_s)
        ]
        if not near_local:
            return None
        offset_refined = float(np.median([t_b - (float(slope_local) * t_a) for t_a, t_b, _hd in near_local]))
        near_local = [
            m for m in near_local
            if abs((m[1] - (float(slope_local) * m[0])) - offset_refined) <= (1.5 * bin_s)
        ]
        if not near_local:
            return None
        return {
            "score": (len(near_local), int(votes_local), -abs(float(slope_local) - 1.0)),
            "slope": float(slope_local),
            "offset_s": float(offset_refined),
            "near_matches": near_local,
            "votes_bin": int(votes_local),
        }

    line_candidates: list[dict] = []
    for slope in slope_grid:
        hist: dict[int, int] = {}
        for t_a, t_b, _hd in matches:
            k = int(round((t_b - (float(slope) * t_a)) / bin_s))
            hist[k] = hist.get(k, 0) + 1
        if not hist:
            continue
        top_bins = sorted(hist.items(), key=lambda kv: (-kv[1], abs(kv[0])))[: max(1, int(TIMELINE_POSTFIT_ALT_MAX_LINES))]
        for bin_key, votes_local in top_bins:
            cand = _build_line_candidate(
                slope_local=float(slope),
                bin_key_local=int(bin_key),
                votes_local=int(votes_local),
            )
            if cand is not None:
                line_candidates.append(cand)

    if not line_candidates:
        return out

    best_line = max(line_candidates, key=lambda row: row["score"])
    slope = float(best_line["slope"])
    offset_s = float(best_line["offset_s"])
    out["offset_s"] = float(offset_s)
    out["drift_ratio"] = float(slope)

    def _merge_ranges(ranges: list[tuple[float, float]]) -> list[tuple[float, float]]:
        if not ranges:
            return []
        segs = sorted((float(a), float(b)) for a, b in ranges if b > a)
        if not segs:
            return []
        out_ranges: list[tuple[float, float]] = []
        cur_s, cur_e = segs[0]
        for s, e in segs[1:]:
            if s <= cur_e:
                if e > cur_e:
                    cur_e = e
            else:
                out_ranges.append((cur_s, cur_e))
                cur_s, cur_e = s, e
        out_ranges.append((cur_s, cur_e))
        return out_ranges

    def _segments_from_matches(
        rows: list[tuple[float, float, int]],
        *,
        slope_local: float,
        offset_local: float,
    ) -> tuple[list[tuple[float, float]], list[tuple[float, float]]]:
        if not rows:
            return [], []
        # For duplicate t_a values, keep the point closest to fitted line.
        best_by_ta: dict[float, tuple[float, float]] = {}
        for t_a, t_b, _hd in rows:
            k = round(float(t_a), 3)
            resid = abs((t_b - (float(slope_local) * t_a)) - float(offset_local))
            prev = best_by_ta.get(k)
            if (prev is None) or (resid < prev[0]):
                best_by_ta[k] = (resid, float(t_b))
        pts = sorted((float(k), float(v[1])) for k, v in best_by_ta.items())
        if not pts:
            return [], []

        seg_a: list[tuple[float, float]] = []
        seg_b: list[tuple[float, float]] = []
        gap_a = max(float(step_s), float(step_s) * max(1.0, float(run_gap_mult)))
        slope_abs = max(1e-6, abs(float(slope_local)))
        gap_resid = max(float(step_s), float(step_s) * slope_abs * max(1.0, float(run_gap_mult)))

        start_a, start_b = pts[0]
        prev_a, prev_b = pts[0]
        for t_a, t_b in pts[1:]:
            da = float(t_a - prev_a)
            expected_db = float(slope * da)
            db = float(t_b - prev_b)
            resid_db = abs(db - expected_db)
            contiguous = (da <= gap_a) and (resid_db <= gap_resid)
            if contiguous:
                prev_a, prev_b = t_a, t_b
                continue
            a0, a1 = float(start_a), float(prev_a + step_s)
            b0, b1 = float(start_b), float(prev_b + (step_s * slope_abs))
            if a1 > a0:
                seg_a.append((a0, a1))
            if b1 > b0:
                seg_b.append((b0, b1))
            start_a, start_b = t_a, t_b
            prev_a, prev_b = t_a, t_b

        a0, a1 = float(start_a), float(prev_a + step_s)
        b0, b1 = float(start_b), float(prev_b + (step_s * slope_abs))
        if a1 > a0:
            seg_a.append((a0, a1))
        if b1 > b0:
            seg_b.append((b0, b1))
        return _merge_ranges(seg_a), _merge_ranges(seg_b)

    def _project_a_segments_to_b(
        seg_a: list[tuple[float, float]],
        *,
        slope_local: float,
        offset_local: float,
    ) -> list[tuple[float, float]]:
        slope_f = float(slope_local)
        out_b: list[tuple[float, float]] = []
        for a0, a1 in seg_a:
            b0 = (slope_f * float(a0)) + float(offset_local)
            b1 = (slope_f * float(a1)) + float(offset_local)
            if b1 < b0:
                b0, b1 = b1, b0
            b0 = max(float(min_t_b), min(float(max_t_b), float(b0)))
            b1 = max(float(min_t_b), min(float(max_t_b), float(b1)))
            if b1 > b0:
                out_b.append((b0, b1))
        return _merge_intervals_simple(out_b, gap_tolerance_s=0.0)

    def _segments_to_mask(segments: list[tuple[float, float]]) -> np.ndarray:
        mask = np.zeros(n_bins, dtype=bool)
        for s, e in segments:
            i0 = max(0, int(math.floor(float(s) / max(1e-6, float(step_s)))))
            i1 = min(n_bins, int(math.ceil(float(e) / max(1e-6, float(step_s)))))
            if i1 > i0:
                mask[i0:i1] = True
        return mask

    def _mask_to_segments(mask: np.ndarray, *, require_seed_mask: np.ndarray | None = None) -> list[tuple[float, float]]:
        if mask.size == 0:
            return []
        segs: list[tuple[float, float]] = []
        i = 0
        n = int(mask.size)
        while i < n:
            if not bool(mask[i]):
                i += 1
                continue
            j = i + 1
            while j < n and bool(mask[j]):
                j += 1
            if require_seed_mask is not None and not bool(np.any(require_seed_mask[i:j])):
                i = j
                continue
            s = float(i) * float(step_s)
            e = float(j) * float(step_s)
            if e > s:
                segs.append((s, e))
            i = j
        return _merge_intervals_simple(segs, gap_tolerance_s=0.0)

    def _clamp_segments_to_raw_support(
        segments: list[tuple[float, float]],
        *,
        raw_support_mask: np.ndarray,
    ) -> list[tuple[float, float]]:
        if not segments:
            return []
        clamped: list[tuple[float, float]] = []
        for s, e in segments:
            i0 = max(0, int(math.floor(float(s) / max(1e-6, float(step_s)))))
            i1 = min(n_bins - 1, int(math.ceil(float(e) / max(1e-6, float(step_s)))))
            if i1 <= i0:
                continue
            local = raw_support_mask[i0:i1 + 1]
            if local.size == 0:
                continue
            nz = np.flatnonzero(local)
            if nz.size == 0:
                continue
            new_i0 = i0 + int(nz[0])
            new_i1 = i0 + int(nz[-1]) + 1
            new_s = max(float(s), float(new_i0) * float(step_s))
            new_e = min(float(e), float(new_i1) * float(step_s))
            if new_e > new_s:
                clamped.append((new_s, new_e))
        return _merge_intervals_simple(clamped, gap_tolerance_s=0.0)

    def _evaluate_line(
        line: dict,
        *,
        include_debug: bool = False,
    ) -> dict:
        slope_local = float(line["slope"])
        offset_local = float(line["offset_s"])
        near_matches_local = list(line["near_matches"])
        strict_matches_local = [m for m in near_matches_local if m[2] <= strict_hamming_thresh]
        seg_a_raw_local, _seg_b_raw_local = _segments_from_matches(
            near_matches_local,
            slope_local=slope_local,
            offset_local=offset_local,
        )
        seg_a_strict_unbridged_local, _seg_b_strict_unbridged_local = _segments_from_matches(
            strict_matches_local,
            slope_local=slope_local,
            offset_local=offset_local,
        )
        seg_a_strict_seed_local = _merge_intervals_simple(
            seg_a_strict_unbridged_local,
            gap_tolerance_s=float(max(0.0, seed_gap_s)),
        )
        seg_a_strict_final_local = list(seg_a_strict_seed_local)

        sparse_loose_signal = np.zeros(n_bins, dtype=np.float64)
        dense_loose_signal = np.zeros(n_bins, dtype=np.float64)
        strict_signal = np.zeros(n_bins, dtype=np.float64)
        dense_best_hd = np.full(n_bins, np.nan, dtype=np.float64)
        dense_best_dt = np.full(n_bins, np.nan, dtype=np.float64)
        loose_denom = max(1.0, float(loose_hamming_thresh))

        for t_a, _t_b, hd in near_matches_local:
            idx = int(round(float(t_a) / max(1e-6, float(step_s))))
            if idx < 0 or idx >= n_bins:
                continue
            score = max(0.0, 1.0 - (float(hd) / loose_denom))
            if score > sparse_loose_signal[idx]:
                sparse_loose_signal[idx] = score
        for t_a, _t_b, _hd in strict_matches_local:
            idx = int(round(float(t_a) / max(1e-6, float(step_s))))
            if 0 <= idx < n_bins:
                strict_signal[idx] = 1.0

        if b_times.size > 0 and dense_max_hd > 0:
            for t_a, h_a in seq_a:
                idx = int(round(float(t_a) / max(1e-6, float(step_s))))
                if idx < 0 or idx >= n_bins:
                    continue
                pred_b = float(offset_local) + (float(slope_local) * float(t_a))
                center = int(np.searchsorted(b_times, pred_b, side="left"))
                lo = max(0, center - search_radius_bins - 1)
                hi = min(int(b_times.size), center + search_radius_bins + 2)
                best_hd = None
                best_dt = None
                for j in range(lo, hi):
                    dt = abs(float(b_times[j]) - pred_b)
                    if dt > align_tol_s:
                        continue
                    hd = _hamming_int(int(h_a), b_hashes[j])
                    if (best_hd is None) or (hd < best_hd) or (hd == best_hd and (best_dt is None or dt < best_dt)):
                        best_hd = int(hd)
                        best_dt = float(dt)
                if best_hd is None or best_hd > dense_max_hd:
                    continue
                dense_best_hd[idx] = float(best_hd)
                dense_best_dt[idx] = float(best_dt or 0.0)
                dense_score = max(0.0, 1.0 - (float(best_hd) / float(dense_max_hd)))
                if dense_score > dense_loose_signal[idx]:
                    dense_loose_signal[idx] = dense_score

        can_try_gaussian = (len(seg_a_strict_unbridged_local) >= 2) and (len(near_matches_local) > 0)
        raw_score = np.maximum(np.maximum(sparse_loose_signal, dense_loose_signal), strict_signal)
        seed_mask = _segments_to_mask(seg_a_strict_seed_local)
        gauss_signal = np.zeros(n_bins, dtype=np.float64)
        gauss_mask = np.zeros(n_bins, dtype=bool)
        if can_try_gaussian:
            sigma_bins = float(TIMELINE_POSTFIT_GAUSS_SIGMA_S) / max(1e-6, float(step_s))
            seed_blend = max(0.0, min(1.0, float(TIMELINE_POSTFIT_GAUSS_SEED_BLEND)))
            gauss_raw = _gaussian_smooth_1d(raw_score, sigma_bins)
            gauss_seed = _gaussian_smooth_1d(seed_mask.astype(np.float64), sigma_bins)
            gauss_signal = ((1.0 - seed_blend) * gauss_raw) + (seed_blend * gauss_seed)
            gauss_mask = gauss_signal >= float(TIMELINE_POSTFIT_GAUSS_SCORE_MIN)
            gauss_mask = np.logical_or(gauss_mask, seed_mask)
            seg_a_strict_gaussian = _mask_to_segments(gauss_mask, require_seed_mask=seed_mask)
            gauss_support = (raw_score >= float(TIMELINE_POSTFIT_GAUSS_RAW_CLAMP_MIN)) | (strict_signal > 0.0)
            seg_a_strict_gaussian = _clamp_segments_to_raw_support(
                seg_a_strict_gaussian,
                raw_support_mask=gauss_support,
            )
            if seg_a_strict_gaussian:
                seg_a_strict_final_local = seg_a_strict_gaussian

        seg_b_strict_final_local = _project_a_segments_to_b(
            seg_a_strict_final_local,
            slope_local=slope_local,
            offset_local=offset_local,
        )
        result = {
            "line": line,
            "slope": float(slope_local),
            "offset_s": float(offset_local),
            "near_matches": near_matches_local,
            "strict_matches": strict_matches_local,
            "votes_raw": int(min(len({m[0] for m in near_matches_local}), len({m[1] for m in near_matches_local}))),
            "votes_strict": int(min(len({m[0] for m in strict_matches_local}), len({m[1] for m in strict_matches_local}))),
            "seg_a_raw": seg_a_raw_local,
            "seg_a_strict_unbridged": seg_a_strict_unbridged_local,
            "seg_a_strict_final": seg_a_strict_final_local,
            "seg_b_strict_final": seg_b_strict_final_local,
            "overlap_raw_s": float(sum((b - a) for a, b in seg_a_raw_local)),
            "overlap_strict_unbridged_s": float(sum((b - a) for a, b in seg_a_strict_unbridged_local)),
            "overlap_strict_s": float(sum((b - a) for a, b in seg_a_strict_final_local)),
        }
        if include_debug:
            result["debug"] = {
                "time_a_s": (np.arange(n_bins, dtype=np.float64) * float(step_s)).round(3).tolist(),
                "predicted_b_s": (
                    (np.arange(n_bins, dtype=np.float64) * float(step_s) * float(slope_local)) + float(offset_local)
                ).round(3).tolist(),
                "sparse_loose_score": sparse_loose_signal.round(6).tolist(),
                "dense_loose_score": dense_loose_signal.round(6).tolist(),
                "combined_raw_score": raw_score.round(6).tolist(),
                "strict_seed_signal": strict_signal.round(6).tolist(),
                "seed_mask": seed_mask.astype(int).tolist(),
                "gauss_signal": gauss_signal.round(6).tolist(),
                "gauss_mask": gauss_mask.astype(int).tolist(),
                "final_mask": _segments_to_mask(seg_a_strict_final_local).astype(int).tolist(),
                "dense_best_hd": np.where(np.isfinite(dense_best_hd), dense_best_hd, -1).astype(int).tolist(),
                "dense_best_dt_s": np.where(np.isfinite(dense_best_dt), dense_best_dt, -1.0).round(3).tolist(),
            }
        return result

    primary_eval = _evaluate_line(best_line, include_debug=return_debug)
    accepted_evals: list[dict] = [primary_eval]
    accepted_a_strict = _merge_intervals_simple(primary_eval["seg_a_strict_final"], gap_tolerance_s=0.0)
    accepted_b_strict = _merge_intervals_simple(primary_eval["seg_b_strict_final"], gap_tolerance_s=0.0)

    alt_summaries: list[dict[str, object]] = []
    alt_lines = sorted(line_candidates, key=lambda row: row["score"], reverse=True)
    for cand in alt_lines:
        slope_c = float(cand["slope"])
        offset_c = float(cand["offset_s"])
        if any(
            abs(offset_c - float(prev["offset_s"])) <= float(TIMELINE_POSTFIT_ALT_OFFSET_SEP_S)
            and abs(slope_c - float(prev["slope"])) <= 1e-6
            for prev in accepted_evals
        ):
            continue
        alt_eval = _evaluate_line(cand, include_debug=False)
        if alt_eval["overlap_strict_s"] < float(TIMELINE_POSTFIT_ALT_MIN_STRICT_S):
            alt_summaries.append({
                "offset_s": round(offset_c, 3),
                "slope": round(slope_c, 6),
                "status": "rejected_small_overlap",
                "overlap_strict_s": round(float(alt_eval["overlap_strict_s"]), 1),
            })
            continue
        uncovered_a = _interval_complement(
            accepted_a_strict,
            start_s=0.0,
            end_s=float(max_t + float(step_s)),
        )
        added_a = _intersect_intervals_simple(alt_eval["seg_a_strict_final"], uncovered_a)
        added_a = _merge_intervals_simple(added_a, gap_tolerance_s=0.0)
        added_a_s = float(sum((e - s) for s, e in added_a))
        if added_a_s < float(TIMELINE_POSTFIT_ALT_MIN_ADDED_S):
            alt_summaries.append({
                "offset_s": round(offset_c, 3),
                "slope": round(slope_c, 6),
                "status": "rejected_small_gain",
                "added_a_s": round(added_a_s, 1),
            })
            continue
        added_b = _project_a_segments_to_b(
            added_a,
            slope_local=slope_c,
            offset_local=offset_c,
        )
        added_b_s = float(sum((e - s) for s, e in added_b))
        overlap_b = _intersect_intervals_simple(added_b, accepted_b_strict)
        overlap_b_s = float(sum((e - s) for s, e in overlap_b))
        if added_b_s <= 0.0 or overlap_b_s > (added_b_s * float(TIMELINE_POSTFIT_ALT_MAX_B_OVERLAP_FRAC)):
            alt_summaries.append({
                "offset_s": round(offset_c, 3),
                "slope": round(slope_c, 6),
                "status": "rejected_b_overlap",
                "added_a_s": round(added_a_s, 1),
                "added_b_s": round(added_b_s, 1),
                "overlap_b_s": round(overlap_b_s, 1),
            })
            continue
        accepted_evals.append(alt_eval)
        accepted_a_strict = _merge_intervals_simple(accepted_a_strict + added_a, gap_tolerance_s=0.0)
        accepted_b_strict = _merge_intervals_simple(accepted_b_strict + added_b, gap_tolerance_s=0.0)
        alt_summaries.append({
            "offset_s": round(offset_c, 3),
            "slope": round(slope_c, 6),
            "status": "accepted",
            "added_a_s": round(added_a_s, 1),
            "added_b_s": round(added_b_s, 1),
        })

    all_raw = _merge_intervals_simple(
        [seg for row in accepted_evals for seg in row["seg_a_raw"]],
        gap_tolerance_s=0.0,
    )
    all_unbridged = _merge_intervals_simple(
        [seg for row in accepted_evals for seg in row["seg_a_strict_unbridged"]],
        gap_tolerance_s=0.0,
    )
    all_final = _merge_intervals_simple(accepted_a_strict, gap_tolerance_s=0.0)
    all_b_final = _merge_intervals_simple(accepted_b_strict, gap_tolerance_s=0.0)
    raw_a_times = {t for row in accepted_evals for (t, _tb, _hd) in row["near_matches"]}
    raw_b_times = {tb for row in accepted_evals for (_ta, tb, _hd) in row["near_matches"]}
    strict_a_times = {t for row in accepted_evals for (t, _tb, _hd) in row["strict_matches"]}
    strict_b_times = {tb for row in accepted_evals for (_ta, tb, _hd) in row["strict_matches"]}

    out["votes_raw"] = int(min(len(raw_a_times), len(raw_b_times)))
    out["votes_strict"] = int(min(len(strict_a_times), len(strict_b_times)))
    out["segment_count_raw"] = int(len(all_raw))
    out["segment_count_strict_unbridged"] = int(len(all_unbridged))
    out["segment_count_strict"] = int(len(all_final))
    out["overlap_raw_s"] = float(sum((b - a) for a, b in all_raw))
    out["overlap_strict_unbridged_s"] = float(sum((b - a) for a, b in all_unbridged))
    out["overlap_strict_s"] = float(sum((b - a) for a, b in all_final))
    out["segments_a_raw"] = _segments_to_text(all_raw)
    out["segments_b_raw"] = _segments_to_text(
        _merge_intervals_simple(
            [
                seg
                for row in accepted_evals
                for seg in _project_a_segments_to_b(
                    row["seg_a_raw"],
                    slope_local=float(row["slope"]),
                    offset_local=float(row["offset_s"]),
                )
            ],
            gap_tolerance_s=0.0,
        )
    )
    out["segments_a_strict"] = _segments_to_text(all_final)
    out["segments_b_strict"] = _segments_to_text(all_b_final)
    if return_debug:
        out["debug"] = dict(primary_eval.get("debug", {}))
        out["debug"]["accepted_lines"] = [
            {
                "offset_s": round(float(row["offset_s"]), 3),
                "slope": round(float(row["slope"]), 6),
                "overlap_strict_s": round(float(row["overlap_strict_s"]), 1),
                "segments_a_strict": _segments_to_text(row["seg_a_strict_final"]),
                "segments_b_strict": _segments_to_text(row["seg_b_strict_final"]),
            }
            for row in accepted_evals
        ]
        out["debug"]["alt_line_decisions"] = alt_summaries
    return out


def export_timeline_pair_diagnostics(
    pairs: list[tuple[str, str]],
    *,
    output_dir: str | None = None,
    step_s: float = TIMELINE_STEP_S,
    speed_ratio_min: float = TIMELINE_SPEED_RATIO_FIXED,
    speed_ratio_max: float = TIMELINE_SPEED_RATIO_FIXED,
    speed_steps: int = 1,
) -> pd.DataFrame:
    """
    Export per-bin timeline evidence for selected pairs.
    This is intended for focused debugging of under-covered matches without
    re-fingerprinting the whole dataset.
    """
    if output_dir is None:
        output_dir = os.path.join(REPORTS_DIR, "timeline_debug")
    os.makedirs(output_dir, exist_ok=True)

    vhs = VideoHashStore()
    summary_rows: list[dict[str, object]] = []
    for idx, (file_a, file_b) in enumerate(pairs, 1):
        timeline_a_raw, _trim_a = vhs.get_timeline(file_a, step_s=step_s)
        timeline_b_raw, _trim_b = vhs.get_timeline(file_b, step_s=step_s)
        seq_a = _filter_hash_seq(
            [(t, _bytes_to_int64(h)) for t, h in timeline_a_raw],
            min_bitcount=TIMELINE_MIN_BITCOUNT,
            max_bitcount=TIMELINE_MAX_BITCOUNT,
            collapse_runs=False,
        )
        seq_b = _filter_hash_seq(
            [(t, _bytes_to_int64(h)) for t, h in timeline_b_raw],
            min_bitcount=TIMELINE_MIN_BITCOUNT,
            max_bitcount=TIMELINE_MAX_BITCOUNT,
            collapse_runs=False,
        )
        details = _timeline_postfit_details(
            seq_a,
            seq_b,
            total_bits=64,
            loose_hamming_thresh=TIMELINE_HAMMING_THRESH,
            strict_hamming_thresh=TIMELINE_STRICT_HAMMING_THRESH,
            bin_s=TIMELINE_BIN_S,
            step_s=step_s,
            lsh_chunks=TIMELINE_LSH_CHUNKS,
            brute_limit=TIMELINE_BRUTE_MAX,
            speed_ratio_min=speed_ratio_min,
            speed_ratio_max=speed_ratio_max,
            speed_steps=speed_steps,
            max_candidates_per_frame=TIMELINE_MAX_CANDIDATES_PER_FRAME,
            run_gap_mult=TIMELINE_RUN_GAP_MULT,
            max_matches=TIMELINE_POSTFIT_MAX_MATCHES,
            seed_gap_s=TIMELINE_POSTFIT_SEED_GAP_S,
            return_debug=True,
        )
        debug = details.get("debug", {}) or {}
        debug_df = pd.DataFrame({
            "time_a_s": debug.get("time_a_s", []),
            "predicted_b_s": debug.get("predicted_b_s", []),
            "sparse_loose_score": debug.get("sparse_loose_score", []),
            "dense_loose_score": debug.get("dense_loose_score", []),
            "combined_raw_score": debug.get("combined_raw_score", []),
            "strict_seed_signal": debug.get("strict_seed_signal", []),
            "seed_mask": debug.get("seed_mask", []),
            "gauss_signal": debug.get("gauss_signal", []),
            "gauss_mask": debug.get("gauss_mask", []),
            "final_mask": debug.get("final_mask", []),
            "dense_best_hd": debug.get("dense_best_hd", []),
            "dense_best_dt_s": debug.get("dense_best_dt_s", []),
        })
        safe_name = (
            f"{idx:02d}__"
            + "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in Path(file_a).stem)[:60]
            + "__"
            + "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in Path(file_b).stem)[:60]
            + ".csv"
        )
        debug_path = os.path.join(output_dir, safe_name)
        debug_df.to_csv(debug_path, index=False)
        summary_rows.append(
            {
                "file_a": file_a,
                "file_b": file_b,
                "seq_a": len(seq_a),
                "seq_b": len(seq_b),
                "offset_s": round(float(details.get("offset_s") or 0.0), 3),
                "drift_ratio": round(float(details.get("drift_ratio") or 1.0), 6),
                "overlap_raw_s": round(float(details.get("overlap_raw_s") or 0.0), 1),
                "overlap_strict_unbridged_s": round(float(details.get("overlap_strict_unbridged_s") or 0.0), 1),
                "overlap_strict_s": round(float(details.get("overlap_strict_s") or 0.0), 1),
                "segment_count_raw": int(details.get("segment_count_raw") or 0),
                "segment_count_strict_unbridged": int(details.get("segment_count_strict_unbridged") or 0),
                "segment_count_strict": int(details.get("segment_count_strict") or 0),
                "segments_a_raw": str(details.get("segments_a_raw") or ""),
                "segments_b_raw": str(details.get("segments_b_raw") or ""),
                "segments_a_strict": str(details.get("segments_a_strict") or ""),
                "segments_b_strict": str(details.get("segments_b_strict") or ""),
                "debug_csv": debug_path,
            }
        )

    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df.to_excel(os.path.join(output_dir, "timeline_pair_diagnostics.xlsx"), index=False)
    return summary_df


def _parse_segments_text(text: str) -> list[tuple[float, float]]:
    if not isinstance(text, str) or not text.strip():
        return []
    out: list[tuple[float, float]] = []
    for part in text.split(";"):
        part = part.strip()
        if not part or "-" not in part:
            continue
        left, right = part.split("-", 1)
        try:
            s = float(left.strip())
            e = float(right.strip())
        except Exception:
            continue
        if e > s:
            out.append((s, e))
    return _merge_intervals_simple(out, gap_tolerance_s=0.0)


def _interval_complement(
    intervals: list[tuple[float, float]],
    *,
    start_s: float,
    end_s: float,
) -> list[tuple[float, float]]:
    lo = float(start_s)
    hi = float(end_s)
    if hi <= lo:
        return []
    segs = _merge_intervals_simple(
        [(max(lo, float(s)), min(hi, float(e))) for s, e in intervals if float(e) > float(s)],
        gap_tolerance_s=0.0,
    )
    out: list[tuple[float, float]] = []
    cursor = lo
    for s, e in segs:
        if s > cursor:
            out.append((cursor, s))
        cursor = max(cursor, e)
    if cursor < hi:
        out.append((cursor, hi))
    return [(s, e) for s, e in out if e > s]


def _intersect_intervals_simple(
    left: list[tuple[float, float]],
    right: list[tuple[float, float]],
) -> list[tuple[float, float]]:
    """Return intersection of two interval lists."""
    if not left or not right:
        return []
    a = _merge_intervals_simple(left, gap_tolerance_s=0.0)
    b = _merge_intervals_simple(right, gap_tolerance_s=0.0)
    out: list[tuple[float, float]] = []
    i = 0
    j = 0
    while i < len(a) and j < len(b):
        a0, a1 = a[i]
        b0, b1 = b[j]
        x0 = max(float(a0), float(b0))
        x1 = min(float(a1), float(b1))
        if x1 > x0:
            out.append((x0, x1))
        if a1 <= b1:
            i += 1
        else:
            j += 1
    return out


def _detect_self_repeat_segments_from_timeline_hashes(
    seq: list[tuple[float, int]],
    *,
    total_bits: int = 64,
    hamming_thresh: int = TIMELINE_HAMMING_THRESH,
    bin_s: float = TIMELINE_BIN_S,
    step_s: float = TIMELINE_STEP_S,
    lsh_chunks: int = TIMELINE_LSH_CHUNKS,
    brute_limit: int | None = TIMELINE_BRUTE_MAX,
    max_candidates_per_frame: int | None = TIMELINE_MAX_CANDIDATES_PER_FRAME,
    run_gap_mult: float = TIMELINE_RUN_GAP_MULT,
    min_segment_s: float = TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S,
    min_bin_votes: int = TIMELINE_A_SELF_REPEAT_MIN_BIN_VOTES,
    neighbor_bins: int = TIMELINE_A_SELF_REPEAT_BIN_NEIGHBOR_BINS,
    min_offset_s: float = TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S,
    max_seq_points: int = TIMELINE_A_SELF_REPEAT_MAX_SEQ_POINTS,
) -> dict:
    """
    Find repeated timeline regions within a single file.

    We only keep positive non-trivial offsets (later playback matching earlier
    playback). This gives an approximate "redundant later playback" mask and a
    way to project coverage from a later repeated section back onto the earlier
    canonical content interval.
    """
    out = {
        "repeat_pairs": [],
        "canonical_segments": [],
        "repeated_segments": [],
        "bin_count": 0,
        "notes": "",
    }
    if not seq:
        return out
    if int(max_seq_points) > 0 and len(seq) > int(max_seq_points):
        stride = max(1, int(math.ceil(len(seq) / float(max_seq_points))))
        seq = seq[::stride]

    matches: list[tuple[float, float, int]] = []
    use_brute = bool(brute_limit is not None and (len(seq) * len(seq) <= int(brute_limit)))
    if use_brute:
        for i, (t_a, h_a) in enumerate(seq):
            for j in range(i + 1, len(seq)):
                t_b, h_b = seq[j]
                if (float(t_b) - float(t_a)) < float(min_offset_s):
                    continue
                hd = _hamming_int(h_a, h_b)
                if hd <= hamming_thresh:
                    matches.append((float(t_a), float(t_b), int(hd)))
    else:
        tables, chunk_size = _build_audio_lsh(seq, total_bits, lsh_chunks)
        for i, (t_a, h_a) in enumerate(seq):
            cand_iter = _collect_lsh_candidates(
                hash_value=h_a,
                tables=tables,
                chunk_size=chunk_size,
                chunks=lsh_chunks,
                max_candidates_per_frame=max_candidates_per_frame,
            )
            for j in cand_iter:
                if j <= i:
                    continue
                t_b, h_b = seq[j]
                if (float(t_b) - float(t_a)) < float(min_offset_s):
                    continue
                hd = _hamming_int(h_a, h_b)
                if hd <= hamming_thresh:
                    matches.append((float(t_a), float(t_b), int(hd)))
    if not matches:
        return out

    offset_hist: dict[int, int] = {}
    for t_a, t_b, _hd in matches:
        k = int(round((float(t_b) - float(t_a)) / max(1e-6, float(bin_s))))
        if k <= 0:
            continue
        offset_hist[k] = offset_hist.get(k, 0) + 1
    if not offset_hist:
        return out

    candidate_bins = [k for k, v in sorted(offset_hist.items(), key=lambda kv: (-kv[1], kv[0])) if v >= int(min_bin_votes)]
    if not candidate_bins:
        return out

    def _segments_from_points(points: list[float]) -> list[tuple[float, float]]:
        pts = sorted(set(float(p) for p in points))
        if not pts:
            return []
        gap_a = max(float(step_s), float(step_s) * max(1.0, float(run_gap_mult)))
        segs: list[tuple[float, float]] = []
        start_a = pts[0]
        prev_a = pts[0]
        for t_a in pts[1:]:
            if (float(t_a) - float(prev_a)) <= gap_a:
                prev_a = float(t_a)
                continue
            segs.append((float(start_a), float(prev_a + float(step_s))))
            start_a = float(t_a)
            prev_a = float(t_a)
        segs.append((float(start_a), float(prev_a + float(step_s))))
        segs = _merge_intervals_simple(segs, gap_tolerance_s=0.0)
        return [(s, e) for s, e in segs if (float(e) - float(s)) >= float(min_segment_s)]

    repeat_pairs: list[dict[str, object]] = []
    keep_delta = int(max(0, neighbor_bins))
    all_canonical: list[tuple[float, float]] = []
    all_repeated: list[tuple[float, float]] = []
    accepted_bins: list[int] = []
    for k in candidate_bins:
        rows = [
            m for m in matches
            if abs(int(round((float(m[1]) - float(m[0])) / max(1e-6, float(bin_s)))) - int(k)) <= keep_delta
        ]
        if len(rows) < int(min_bin_votes):
            continue
        offset_s = float(k) * float(bin_s)
        earlier = _segments_from_points([t_a for t_a, _t_b, _hd in rows])
        later = _segments_from_points([t_b for _t_a, t_b, _hd in rows])
        if not earlier or not later:
            continue
        repeat_pairs.append({
            "offset_s": offset_s,
            "earlier_segments": earlier,
            "later_segments": later,
        })
        accepted_bins.append(int(k))
        all_canonical.extend(earlier)
        all_repeated.extend(later)

    out["repeat_pairs"] = repeat_pairs
    out["canonical_segments"] = _merge_intervals_simple(all_canonical, gap_tolerance_s=0.0)
    out["repeated_segments"] = _merge_intervals_simple(all_repeated, gap_tolerance_s=0.0)
    out["bin_count"] = len(accepted_bins)
    if accepted_bins:
        out["notes"] = f"self-repeat bins={len(accepted_bins)}"
    return out


def _detect_edge_repeat_segments_by_similarity(
    seq: list[tuple[float, int]],
    *,
    edge_intervals: list[tuple[float, float]],
    total_bits: int = 64,
    hamming_thresh: int = TIMELINE_A_SELF_REPEAT_HAMMING_THRESH,
    step_s: float = TIMELINE_STEP_S,
    lsh_chunks: int = TIMELINE_LSH_CHUNKS,
    max_candidates_per_frame: int | None = TIMELINE_MAX_CANDIDATES_PER_FRAME,
    run_gap_mult: float = TIMELINE_RUN_GAP_MULT,
    min_segment_s: float = TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S,
    min_offset_s: float = TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S,
    min_match_fraction: float = 0.90,
) -> list[tuple[float, float]]:
    """
    Fallback for repeated tail/head snippets without a single stable offset.
    Marks edge-region points as repeated when they strongly resemble earlier content.
    """
    if not seq or not edge_intervals:
        return []
    seq_sorted = sorted((float(t), int(h)) for t, h in seq)
    all_repeat: list[tuple[float, float]] = []
    run_gap = max(float(step_s), float(step_s) * max(1.0, float(run_gap_mult)))
    min_points = max(3, int(math.ceil(float(min_segment_s) / max(1e-6, float(step_s)))))

    for edge_s, edge_e in edge_intervals:
        edge_s = float(edge_s)
        edge_e = float(edge_e)
        if edge_e <= edge_s:
            continue
        earlier = [(t, h) for t, h in seq_sorted if t < (edge_s - float(min_offset_s))]
        later = [(t, h) for t, h in seq_sorted if edge_s <= t <= edge_e]
        if len(earlier) < min_points or len(later) < min_points:
            continue

        tables, chunk_size = _build_audio_lsh(earlier, total_bits, lsh_chunks)
        matched_times: list[float] = []
        for t_later, h_later in later:
            hit = False
            for j in _collect_lsh_candidates(
                hash_value=h_later,
                tables=tables,
                chunk_size=chunk_size,
                chunks=lsh_chunks,
                max_candidates_per_frame=max_candidates_per_frame,
            ):
                if _hamming_int(h_later, earlier[j][1]) <= int(hamming_thresh):
                    hit = True
                    break
            if hit:
                matched_times.append(float(t_later))
        if not matched_times:
            continue
        if (len(matched_times) / max(1, len(later))) < float(min_match_fraction):
            continue

        start_t = matched_times[0]
        prev_t = matched_times[0]
        segs: list[tuple[float, float]] = []
        for t in matched_times[1:]:
            if (float(t) - float(prev_t)) <= run_gap:
                prev_t = float(t)
                continue
            segs.append((float(start_t), float(prev_t + float(step_s))))
            start_t = float(t)
            prev_t = float(t)
        segs.append((float(start_t), float(prev_t + float(step_s))))
        segs = _merge_intervals_simple(segs, gap_tolerance_s=0.0)
        segs = [(s, e) for s, e in segs if (float(e) - float(s)) >= float(min_segment_s)]
        all_repeat.extend(segs)

    return _merge_intervals_simple(all_repeat, gap_tolerance_s=0.0)


def _timeline_probe_descriptor(frame: np.ndarray) -> tuple[np.ndarray, np.ndarray] | None:
    if frame is None or getattr(frame, "size", 0) <= 0:
        return None
    if len(getattr(frame, "shape", ())) == 2:
        gray = frame
    else:
        try:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        except Exception:
            return None
    h, w = gray.shape[:2]
    if h < 40 or w < 40:
        return None

    # Ignore borders and the typical camcorder timestamp region at bottom-right.
    x0 = int(round(w * 0.04))
    x1 = int(round(w * 0.96))
    y0 = int(round(h * 0.04))
    y1 = int(round(h * 0.92))
    roi = gray[y0:y1, x0:x1].copy()
    if roi.size <= 0:
        return None
    rh, rw = roi.shape[:2]
    ts_y0 = int(round(rh * 0.78))
    ts_x0 = int(round(rw * 0.60))
    fill_val = int(np.median(roi)) if roi.size else 0
    roi[ts_y0:rh, ts_x0:rw] = fill_val

    roi = cv2.GaussianBlur(roi, (5, 5), 0)
    small = cv2.resize(roi, (24, 18), interpolation=cv2.INTER_AREA).astype(np.float32)
    small = cv2.equalizeHist(np.clip(small, 0, 255).astype(np.uint8)).astype(np.float32)
    small -= float(np.mean(small))
    n_small = float(np.linalg.norm(small))
    if n_small <= 1e-6:
        return None
    small /= n_small

    gx = cv2.Sobel(roi, cv2.CV_32F, 1, 0, ksize=3)
    gy = cv2.Sobel(roi, cv2.CV_32F, 0, 1, ksize=3)
    grad = cv2.magnitude(gx, gy)
    grad = cv2.resize(grad, (24, 18), interpolation=cv2.INTER_AREA).astype(np.float32)
    grad -= float(np.mean(grad))
    n_grad = float(np.linalg.norm(grad))
    if n_grad <= 1e-6:
        return small.reshape(-1), np.zeros_like(small.reshape(-1))
    grad /= n_grad
    return small.reshape(-1), grad.reshape(-1)


def _timeline_probe_similarity(
    desc_a: tuple[np.ndarray, np.ndarray] | None,
    desc_b: tuple[np.ndarray, np.ndarray] | None,
) -> float:
    if desc_a is None or desc_b is None:
        return -1.0
    app = float(np.dot(desc_a[0], desc_b[0]))
    grad = float(np.dot(desc_a[1], desc_b[1]))
    return (0.65 * app) + (0.35 * grad)


def _build_timeline_descriptor_grid(
    path: str,
    *,
    start_s: float = 0.0,
    duration_s: float,
    step_s: float = 1.0,
    w: int = 96,
    h: int = 72,
) -> dict[float, tuple[np.ndarray, np.ndarray]]:
    out: dict[float, tuple[np.ndarray, np.ndarray]] = {}
    if (not path) or (not os.path.exists(path)) or float(duration_s) <= 0.0:
        return out
    fps = 1.0 / max(0.1, float(step_s))
    for t_s, frame in _stream_gray_frames_ffmpeg(
        path,
        float(start_s),
        float(duration_s),
        fps=float(fps),
        w=int(w),
        h=int(h),
    ):
        desc = _timeline_probe_descriptor(frame)
        if desc is not None:
            out[round(float(t_s), 1)] = desc
    return out


def _sample_segment_probe_descriptors(
    path: str,
    seg_start_s: float,
    seg_end_s: float,
    *,
    probe_count: int = 3,
    w: int = 96,
    h: int = 72,
) -> list[tuple[float, tuple[np.ndarray, np.ndarray]]]:
    seg_len_s = max(0.0, float(seg_end_s) - float(seg_start_s))
    if seg_len_s <= 0.0 or (not path) or (not os.path.exists(path)):
        return []
    frames = list(
        _stream_gray_frames_ffmpeg(
            path,
            float(seg_start_s),
            float(seg_len_s),
            fps=1.0,
            w=int(w),
            h=int(h),
        )
    )
    if not frames:
        return []
    probe_offsets = np.linspace(0.0, float(seg_len_s), max(1, int(probe_count)) + 2, dtype=np.float64)[1:-1]
    out: list[tuple[float, tuple[np.ndarray, np.ndarray]]] = []
    for probe_off in probe_offsets.tolist():
        best_frame = min(frames, key=lambda item: abs((float(item[0]) - float(seg_start_s)) - float(probe_off)))
        desc = _timeline_probe_descriptor(best_frame[1])
        if desc is not None:
            out.append((float(probe_off), desc))
    return out


def _score_segment_against_descriptor_grid(
    probe_descs: list[tuple[float, tuple[np.ndarray, np.ndarray]]],
    b_grid: dict[float, tuple[np.ndarray, np.ndarray]],
    candidate_start_s: float,
) -> tuple[float, int]:
    if not probe_descs or not b_grid:
        return -1.0, 0
    scores: list[float] = []
    for probe_off, desc_a in probe_descs:
        t_b = round(float(candidate_start_s) + float(probe_off), 1)
        desc_b = b_grid.get(t_b)
        if desc_b is None:
            continue
        scores.append(_timeline_probe_similarity(desc_a, desc_b))
    if not scores:
        return -1.0, 0
    return float(sum(scores) / len(scores)), int(len(scores))


def _verify_timeline_gap_via_frame_probes(
    *,
    file_a: str,
    file_b: str,
    gaps_a_abs: list[tuple[float, float]],
    offset_s: float,
    drift_ratio: float,
    playable_a: tuple[float, float],
    playable_b: tuple[float, float],
    probes_per_gap: int = TIMELINE_GAP_VERIFY_PROBES_PER_GAP,
    search_radius_s: float = TIMELINE_GAP_VERIFY_SEARCH_RADIUS_S,
    search_step_s: float = TIMELINE_GAP_VERIFY_SEARCH_STEP_S,
    score_min: float = TIMELINE_GAP_VERIFY_SCORE_MIN,
    min_pass_fraction: float = TIMELINE_GAP_VERIFY_MIN_PASS_FRACTION,
) -> dict:
    out = {
        "status": "not_run",
        "verified_gap_intervals_a": [],
        "verified_gap_intervals_b": [],
        "gaps_considered": 0,
        "gap_passes": 0,
        "probe_passes": 0,
        "probe_total": 0,
        "notes": "",
    }
    if not gaps_a_abs:
        out["status"] = "no_gap"
        return out
    if not os.path.exists(file_a) or not os.path.exists(file_b):
        out["status"] = "missing_file"
        return out

    cap_a = cv2.VideoCapture(file_a)
    cap_b = cv2.VideoCapture(file_b)
    try:
        if not cap_a.isOpened() or not cap_b.isOpened():
            out["status"] = "open_failed"
            return out

        def _read_desc(cap: cv2.VideoCapture, t_s: float) -> tuple[np.ndarray, np.ndarray] | None:
            cap.set(cv2.CAP_PROP_POS_MSEC, max(0.0, float(t_s)) * 1000.0)
            ok, frame = cap.read()
            if not ok or frame is None:
                return None
            return _timeline_probe_descriptor(frame)

        b_lo, b_hi = playable_b
        a_lo, a_hi = playable_a
        verified_a: list[tuple[float, float]] = []
        verified_b: list[tuple[float, float]] = []
        gap_notes: list[str] = []
        for gap_s, gap_e in gaps_a_abs:
            gap_s = max(a_lo, float(gap_s))
            gap_e = min(a_hi, float(gap_e))
            if gap_e <= gap_s:
                continue
            out["gaps_considered"] += 1
            gap_len = gap_e - gap_s
            probe_n = max(1, int(probes_per_gap))
            probe_times = np.linspace(gap_s, gap_e, probe_n + 2, dtype=np.float64)[1:-1]
            probe_passes = 0
            probe_total = 0
            for t_a in probe_times.tolist():
                desc_a = _read_desc(cap_a, float(t_a))
                if desc_a is None:
                    continue
                pred_b = float(offset_s) + (float(drift_ratio) * float(t_a))
                best_score = -1.0
                best_t_b = None
                t_b = pred_b - float(search_radius_s)
                while t_b <= (pred_b + float(search_radius_s) + 1e-6):
                    if b_lo <= t_b <= b_hi:
                        desc_b = _read_desc(cap_b, float(t_b))
                        score = _timeline_probe_similarity(desc_a, desc_b)
                        if score > best_score:
                            best_score = score
                            best_t_b = float(t_b)
                    t_b += float(search_step_s)
                probe_total += 1
                if best_score >= float(score_min):
                    probe_passes += 1
            out["probe_passes"] += int(probe_passes)
            out["probe_total"] += int(probe_total)
            pass_needed = max(1, int(math.ceil(float(min_pass_fraction) * max(1, probe_total))))
            if probe_total > 0 and probe_passes >= pass_needed:
                out["gap_passes"] += 1
                verified_a.append((gap_s, gap_e))
                b0 = (float(drift_ratio) * float(gap_s)) + float(offset_s)
                b1 = (float(drift_ratio) * float(gap_e)) + float(offset_s)
                if b1 < b0:
                    b0, b1 = b1, b0
                b0 = max(b_lo, b0)
                b1 = min(b_hi, b1)
                if b1 > b0:
                    verified_b.append((b0, b1))
                gap_notes.append(f"verified gap {gap_len:.0f}s")
            else:
                gap_notes.append(f"unverified gap {gap_len:.0f}s")

        out["verified_gap_intervals_a"] = _merge_intervals_simple(verified_a, gap_tolerance_s=0.0)
        out["verified_gap_intervals_b"] = _merge_intervals_simple(verified_b, gap_tolerance_s=0.0)
        if out["gaps_considered"] <= 0:
            out["status"] = "no_gap"
        elif out["gap_passes"] == out["gaps_considered"]:
            out["status"] = "verified"
        elif out["gap_passes"] > 0:
            out["status"] = "partial"
        else:
            out["status"] = "failed"
        out["notes"] = "; ".join(gap_notes)
        return out
    finally:
        try:
            cap_a.release()
        except Exception:
            pass
        try:
            cap_b.release()
        except Exception:
            pass


_PAIR_ENTRIES: list[dict] | None = None
_PAIR_PARAMS: dict | None = None


def _init_pair_worker(entries: list[dict], params: dict) -> None:
    global _PAIR_ENTRIES, _PAIR_PARAMS
    _PAIR_ENTRIES = entries
    _PAIR_PARAMS = params


def _eval_pair_worker(pair: tuple[int, int]) -> tuple[int, int, bool, float | None, int, float]:
    i, j = pair
    seq_a = _PAIR_ENTRIES[i]["seq"]
    seq_b = _PAIR_ENTRIES[j]["seq"]
    try:
        ok, offset_s, votes, overlap_s = _audio_match_offset(seq_a, seq_b, **_PAIR_PARAMS)
        return i, j, ok, offset_s, votes, overlap_s
    except MemoryError:
        # Keep long runs alive: skip this pair instead of crashing the full job.
        # votes = -1 is a sentinel used only for logging skipped memory-limited pairs.
        return i, j, False, None, -1, 0.0


def _order_pair(i: int, j: int, folder_ids: list[int], paths: list[str], prefer_folder: int | None) -> tuple[int, int]:
    """
    Return (a_idx, b_idx) in a consistent order.
    - If prefer_folder is set, try to put that folder on the A side when present.
    - Otherwise order by (folder_id, path) to make it stable across runs.
    """
    fi, fj = folder_ids[i], folder_ids[j]
    if prefer_folder is not None:
        if fi == prefer_folder and fj != prefer_folder:
            return i, j
        if fj == prefer_folder and fi != prefer_folder:
            return j, i
    # fallback: stable ordering by (folder, path)
    return (i, j) if (fi, paths[i]) <= (fj, paths[j]) else (j, i)


# 
# Legacy per-video sampler (kept for "legacy" / "both")
# 

def _sample_hashes_with_times(path: str) -> List[Tuple[str, float]]:
    """
    Legacy sampler: sparsely sample frames across the whole video and pHash them.
    Returns: [(hash_hex, timestamp_seconds), ...] up to ~MAX_SAMPLES elements.
    """
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        raise RuntimeError(f"Cannot open video: {path}")

    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) or 1
    duration = frame_count / fps

    k = _calculate_target_samples(duration)
    step = max(1, int(frame_count / k))
    out: List[Tuple[str, float]] = []

    for frame_idx in range(0, frame_count, step):
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ok, frame = cap.read()
        if not ok:
            break
        timestamp = frame_idx / fps
        img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        out.append((str(imagehash.phash(img)), timestamp))
        if len(out) >= k:
            break

    cap.release()
    return out


# 
# Legacy pair aligner
# 

def _aligned_distance_and_time_limited(
    seq_a: List[Tuple[str, float]],
    seq_b: List[Tuple[str, float]],
    max_shift_samples: int,
    offset_limit_s: float
) -> Tuple[float, float]:
    """
    Try small temporal shifts (max_shift_samples) and compute the mean Hamming
    of overlapping sparse samples, but only if their timestamps are within offset_limit_s.
    Returns: (best_mean_hamming, average_time_shift_seconds)
    """
    if not seq_a or not seq_b:
        return 64.0, 0.0

    best_dist = 64.0
    best_time_shift = 0.0
    L1, L2 = len(seq_a), len(seq_b)
    min_shift = max(-(L2 - 1), -max_shift_samples)
    max_shift = min(L1 - 1, max_shift_samples)

    for shift in range(min_shift, max_shift + 1):
        dists, ts_list = [], []
        for i, (ha, ta) in enumerate(seq_a):
            j = i - shift
            if 0 <= j < L2:
                hb, tb = seq_b[j]
                ts = tb - ta
                if abs(ts) <= offset_limit_s:
                    dists.append(_hamming_hex(ha, hb))
                    ts_list.append(ts)
        if dists:
            mean_d = sum(dists) / len(dists)
            if mean_d < best_dist:
                best_dist = mean_d
                best_time_shift = sum(ts_list) / len(ts_list)

    return best_dist, best_time_shift


# 
# Anchor-mode helpers (audio-guided trim + fast frame streaming)
# 

def estimate_trim_bounds(
    path: str,
    *,
    head_scan_s: float = 90.0,     # scan first N seconds for silence
    tail_scan_s: float = 1800.0,   # scan last N seconds for silence (30 min)
    noise_thresh_db: float = -55.0,
    min_silence_s: float = 1.5,
    merge_min_duration_s: float = 1.5,
    merge_max_gap_s: float = 0.8,
    guard_head_max_s: float = 60.0,
    guard_tail_max_s: float = 1800.0,
) -> tuple[float, float]:
    """
    Use audio to find head/tail silence and return (trim_head_s, trim_tail_s).
    If audio analysis fails (e.g. no audio stream), fall back to (0.0, 0.0).
    Times are rounded to 1 dp.
    """
    # Try audio-guided trim first
    try:
        res = detect_head_tail_silence_ffmpeg(
            path,
            noise_thresh_db=noise_thresh_db,
            min_silence_s=min_silence_s,
            head_scan_s=head_scan_s,
            tail_scan_s=tail_scan_s,
            downmix_mono=True,
        )
        duration = float(res.get("duration_s", 0.0))
        head_regions = merge_close_regions(
            res.get("head_silences", []),
            min_duration_s=merge_min_duration_s,
            max_gap_s=merge_max_gap_s,
        )
        tail_regions = merge_close_regions(
            res.get("tail_silences", []),
            min_duration_s=merge_min_duration_s,
            max_gap_s=merge_max_gap_s,
        )
    except Exception as e:
        # No audio stream or silencedetect failure  fall back gracefully
        # Get duration via OpenCV so tail logic can still clamp properly.
        try:
            cap = cv2.VideoCapture(path)
            fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
            cnt = cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0.0
            duration = float(cnt / max(fps, 1e-6))
            cap.release()
        except Exception:
            duration = 0.0
        head_regions, tail_regions = [], []
        try:
            fname = Path(path).name
        except Exception:
            fname = str(path)
        print(f"[audio] silencedetect unavailable for {fname} - assuming no trim. Reason: {e}", flush=True)

    # Compute trims from (possibly empty) regions
    trim_head = 0.0
    if head_regions:
        s0, e0 = head_regions[0]
        if s0 <= 0.3:  # only if silence starts ~at 0 s
            trim_head = min(e0, guard_head_max_s)

    trim_tail = 0.0
    if tail_regions and duration > 0.0:
        s_last, e_last = tail_regions[-1]
        if (duration - e_last) <= 1.0:  # ends ~at EOF
            trim_tail = min(duration - s_last, guard_tail_max_s)

    trim_head = round(max(0.0, min(trim_head, guard_head_max_s)), 1)
    trim_tail = round(max(0.0, min(trim_tail, guard_tail_max_s)), 1)
    return trim_head, trim_tail


def _stream_gray_frames_ffmpeg(
    path: str,
    start_s: float,
    duration_s: float,
    *,
    fps: float = 1.0,
    w: int = 32,
    h: int = 32,
    ffmpeg_path: str = "ffmpeg",
):
    """
    Stream grayscale frames at a fixed fps over [start_s, start_s + duration_s].
    FFmpeg does the decode + scale to 3232 gray. We silence logs and send stderr to DEVNULL
    to prevent pipe blocking on noisy sources.
    """
    start_s = max(0.0, float(start_s))
    duration_s = max(0.0, float(duration_s))
    if duration_s <= 0.0:
        return

    vf = f"fps={fps},scale={w}:{h}:flags=bicubic,format=gray"

    cmd = [
        ffmpeg_path,
        "-hide_banner", "-nostdin",
        "-loglevel", "quiet", "-nostats",  # hard-silence logs
        "-ss", f"{start_s:.3f}",
        "-t",  f"{duration_s:.3f}",
        "-i",  path,
        "-vf", vf,
        "-an", "-sn",
        "-f", "rawvideo", "-vcodec", "rawvideo",
        "-threads", str(FFMPEG_THREADS),
        "pipe:1",
    ]

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,   # <- avoid deadlock on stderr
        stdin=subprocess.DEVNULL,    # <- be explicit
        bufsize=0,
    )
    frame_size = w * h
    idx = 0
    try:
        while True:
            buf = proc.stdout.read(frame_size)
            if not buf or len(buf) < frame_size:
                break
            frame = np.frombuffer(buf, dtype=np.uint8).reshape(h, w)
            t = start_s + idx / fps
            yield (round(t, 1), frame)
            idx += 1
    finally:
        try:
            if proc.stdout:
                proc.stdout.close()
        except Exception:
            pass
        try:
            proc.wait(timeout=5)
        except Exception:
            proc.kill()


def _stream_bgr_frames_ffmpeg(
    path: str,
    start_s: float,
    duration_s: float,
    *,
    fps: float,
    w: int = TIMESTAMP_SCAN_FRAME_W,
    h: int = TIMESTAMP_SCAN_FRAME_H,
    ffmpeg_path: str = "ffmpeg",
):
    """
    Stream BGR frames at fixed fps over [start_s, start_s + duration_s].
    Used by timestamp OCR scan so we can run ROI extraction on reasonably sized frames.
    """
    start_s = max(0.0, float(start_s))
    duration_s = max(0.0, float(duration_s))
    fps = max(1e-6, float(fps))
    if duration_s <= 0.0:
        return

    vf = f"fps={fps:.8f},scale={int(w)}:{int(h)}:flags=bicubic,format=bgr24"
    cmd = [
        ffmpeg_path,
        "-hide_banner", "-nostdin",
        "-loglevel", "quiet", "-nostats",
        "-ss", f"{start_s:.3f}",
        "-t", f"{duration_s:.3f}",
        "-i", path,
        "-vf", vf,
        "-an", "-sn",
        "-pix_fmt", "bgr24",
        "-f", "rawvideo",
        "-vcodec", "rawvideo",
        "-threads", str(FFMPEG_THREADS),
        "pipe:1",
    ]

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        stdin=subprocess.DEVNULL,
        bufsize=0,
    )
    frame_size = int(w) * int(h) * 3
    idx = 0
    try:
        while True:
            if proc.stdout is None:
                break
            buf = bytearray()
            while len(buf) < frame_size:
                chunk = proc.stdout.read(frame_size - len(buf))
                if not chunk:
                    break
                buf.extend(chunk)
            if len(buf) < frame_size:
                break
            frame = np.frombuffer(buf, dtype=np.uint8).reshape(int(h), int(w), 3)
            t = start_s + (idx / fps)
            yield (round(float(t), 1), frame)
            idx += 1
    finally:
        try:
            if proc.stdout:
                proc.stdout.close()
        except Exception:
            pass
        try:
            proc.wait(timeout=5)
        except Exception:
            proc.kill()


def _internal_dead_crop_roi(frame: np.ndarray) -> np.ndarray:
    h, w = frame.shape[:2]
    if h <= 0 or w <= 0:
        return frame
    mx = max(0, int(round(w * 0.08)))
    my = max(0, int(round(h * 0.08)))
    x0 = min(mx, max(0, w - 1))
    x1 = max(x0 + 1, w - mx)
    y0 = min(my, max(0, h - 1))
    y1 = max(y0 + 1, h - my)
    return frame[y0:y1, x0:x1]


def _neighbor_correlation(gray: np.ndarray, axis: int) -> float:
    if gray.size <= 4:
        return 1.0
    if axis == 0:
        a = gray[:-1, :].astype(np.float32).ravel()
        b = gray[1:, :].astype(np.float32).ravel()
    else:
        a = gray[:, :-1].astype(np.float32).ravel()
        b = gray[:, 1:].astype(np.float32).ravel()
    if a.size <= 4 or b.size <= 4:
        return 1.0
    a = a - float(a.mean())
    b = b - float(b.mean())
    da = float(np.linalg.norm(a))
    db = float(np.linalg.norm(b))
    if da <= 1e-6 or db <= 1e-6:
        return 1.0
    return float(np.dot(a, b) / (da * db))


def _frame_is_internal_dead_like(frame: np.ndarray) -> bool:
    roi = _internal_dead_crop_roi(frame)
    if roi.size == 0:
        return False
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    mean_v = float(gray.mean())
    std_v = float(gray.std())
    if (
        mean_v <= float(TIMELINE_INTERNAL_DEAD_BLACK_MEAN_MAX)
        and std_v <= float(TIMELINE_INTERNAL_DEAD_BLACK_STD_MAX)
    ):
        return True

    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    sat_mean = float(hsv[:, :, 1].mean())
    if sat_mean > float(TIMELINE_INTERNAL_DEAD_NOISE_SAT_MAX):
        return False
    if std_v < float(TIMELINE_INTERNAL_DEAD_NOISE_STD_MIN):
        return False

    small = cv2.resize(gray, (128, 96), interpolation=cv2.INTER_AREA)
    corr_h = abs(_neighbor_correlation(small, axis=1))
    corr_v = abs(_neighbor_correlation(small, axis=0))
    if (
        corr_h <= float(TIMELINE_INTERNAL_DEAD_NOISE_NEIGHBOR_CORR_MAX)
        and corr_v <= float(TIMELINE_INTERNAL_DEAD_NOISE_NEIGHBOR_CORR_MAX)
    ):
        return True

    corr_hi = max(corr_h, corr_v)
    corr_lo = min(corr_h, corr_v)
    if (
        sat_mean <= float(TIMELINE_INTERNAL_DEAD_STRIPED_SAT_MAX)
        and std_v >= float(TIMELINE_INTERNAL_DEAD_STRIPED_STD_MIN)
        and corr_hi >= float(TIMELINE_INTERNAL_DEAD_STRIPED_CORR_HIGH_MIN)
        and corr_lo <= float(TIMELINE_INTERNAL_DEAD_STRIPED_CORR_LOW_MAX)
        and (corr_hi - corr_lo) >= float(TIMELINE_INTERNAL_DEAD_STRIPED_ANISOTROPY_MIN)
    ):
        return True
    return False


def _candidate_internal_dead_regions_from_audio_cache(audio_entry: dict) -> list[tuple[float, float]]:
    if not TIMELINE_INTERNAL_DEAD_ENABLE:
        return []
    if not isinstance(audio_entry, dict):
        return []
    rows = audio_entry.get("data")
    if not isinstance(rows, list) or len(rows) < 2:
        return []
    params = audio_entry.get("params")
    hop_s = 1.0
    if isinstance(params, dict):
        try:
            hop_s = max(0.1, float(params.get("hop_s", 1.0) or 1.0))
        except Exception:
            hop_s = 1.0

    times: list[float] = []
    for item in rows:
        if not isinstance(item, (list, tuple)) or len(item) < 1:
            continue
        try:
            t_s = float(item[0])
        except Exception:
            continue
        if np.isfinite(t_s):
            times.append(t_s)
    if len(times) < 2:
        return []
    times = sorted(set(times))
    min_gap_s = max(float(TIMELINE_INTERNAL_DEAD_AUDIO_GAP_MIN_S), 2.0 * hop_s)
    dead: list[tuple[float, float]] = []
    for i in range(1, len(times)):
        prev_t = float(times[i - 1])
        cur_t = float(times[i])
        gap_s = cur_t - prev_t
        if gap_s < min_gap_s:
            continue
        s = prev_t + hop_s
        e = cur_t
        if (e - s) >= min_gap_s:
            dead.append((s, e))
    return _merge_intervals_simple(dead, gap_tolerance_s=0.0)


def _confirm_internal_dead_regions_with_streaming(
    path: str,
    gaps: list[tuple[float, float]],
) -> list[tuple[float, float]]:
    if not gaps or (not os.path.exists(path)):
        return []
    sample_count = max(2, int(TIMELINE_INTERNAL_DEAD_VISUAL_SAMPLE_COUNT))
    kept: list[tuple[float, float]] = []
    for s, e in gaps:
        gap_len = max(0.0, float(e) - float(s))
        if gap_len < float(TIMELINE_INTERNAL_DEAD_AUDIO_GAP_MIN_S):
            continue
        window_start = float(s) + gap_len * 0.15
        window_end = float(s) + gap_len * 0.85
        window_duration = max(0.0, window_end - window_start)
        if window_duration <= 0.0:
            continue
        fps = max(1e-6, float(sample_count - 1) / window_duration)
        dead_votes = 0
        total_reads = 0
        for idx, (_t_s, frame) in enumerate(
            _stream_bgr_frames_ffmpeg(
                path,
                window_start,
                window_duration + 1e-3,
                fps=fps,
                w=TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_W,
                h=TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_H,
            )
        ):
            total_reads += 1
            if _frame_is_internal_dead_like(frame):
                dead_votes += 1
            if idx + 1 >= sample_count:
                break
        if total_reads <= 0:
            continue
        min_votes = max(total_reads - 1, int(math.ceil(total_reads * 0.75)))
        if dead_votes >= min_votes:
            kept.append((float(s), float(e)))
    return _merge_intervals_simple(kept, gap_tolerance_s=0.0)


def _load_playable_basis_maps(
    paths: list[str],
) -> tuple[
    dict[str, tuple[float, float]],
    dict[str, tuple[float, float]],
    dict[str, list[tuple[float, float]]],
    int,
    int,
    int,
]:
    """
    Reuse cached timeline trim/dead-gap analysis for a set of files.

    Returns trim bounds, active hash bounds, confirmed internal dead regions,
    trim cache hits, trim estimates performed, and internal-dead cache hits.
    """
    trim_bounds: dict[str, tuple[float, float]] = {}
    active_bounds: dict[str, tuple[float, float]] = {}
    internal_dead_regions: dict[str, list[tuple[float, float]]] = {}
    path_by_norm: dict[str, str] = {}
    for p in paths:
        if not isinstance(p, str) or (not p.strip()):
            continue
        path_by_norm.setdefault(_norm_path(p), p)
    target = set(path_by_norm.keys())
    if not target:
        return trim_bounds, active_bounds, internal_dead_regions, 0, 0, 0

    cache_hits = 0
    estimated_hits = 0
    internal_dead_hits = 0
    vhs_local = VideoHashStore()
    cache_data = vhs_local._data if isinstance(vhs_local._data, dict) else {}
    if isinstance(cache_data, dict):
        for raw_path, entry in cache_data.items():
            if not isinstance(raw_path, str) or not isinstance(entry, dict):
                continue
            norm_path = _norm_path(raw_path)
            if norm_path not in target:
                continue
            tl = entry.get("timeline")
            if not isinstance(tl, dict):
                continue
            trim = tl.get("trim")
            if not isinstance(trim, (list, tuple)) or len(trim) < 2:
                continue
            try:
                trim_head = max(0.0, float(trim[0]))
                trim_tail = max(0.0, float(trim[1]))
            except Exception:
                continue
            try:
                if os.path.exists(raw_path) and ("mtime" in tl):
                    cache_mtime = float(tl.get("mtime", 0.0) or 0.0)
                    file_mtime = float(os.path.getmtime(raw_path))
                    if abs(cache_mtime - file_mtime) > 0.5:
                        continue
            except Exception:
                pass
            trim_bounds[norm_path] = (trim_head, trim_tail)
            dead_regions, dead_cache_hit = vhs_local.get_confirmed_internal_dead_regions(
                raw_path,
                audio_entry=entry.get("audio"),
            )
            if dead_regions:
                internal_dead_regions[norm_path] = dead_regions
            if dead_cache_hit:
                internal_dead_hits += 1
            data_rows = tl.get("data")
            if isinstance(data_rows, list) and data_rows:
                first_t = None
                last_t = None
                for item in data_rows:
                    if not isinstance(item, (list, tuple)) or len(item) < 1:
                        continue
                    try:
                        t_s = float(item[0])
                    except Exception:
                        continue
                    if (first_t is None) or (t_s < first_t):
                        first_t = t_s
                    if (last_t is None) or (t_s > last_t):
                        last_t = t_s
                if (first_t is not None) and (last_t is not None) and (last_t > first_t):
                    active_bounds[norm_path] = (max(0.0, float(first_t)), max(0.0, float(last_t)))
            cache_hits += 1
    if vhs_local._dirty:
        vhs_local.save_if_dirty()

    for norm_path, raw_path in path_by_norm.items():
        if norm_path in trim_bounds:
            continue
        if (not isinstance(raw_path, str)) or (not os.path.exists(raw_path)):
            continue
        try:
            trim_head, trim_tail = estimate_trim_bounds(raw_path)
            trim_head = max(0.0, float(trim_head))
            trim_tail = max(0.0, float(trim_tail))
        except Exception:
            trim_head, trim_tail = 0.0, 0.0
        trim_bounds[norm_path] = (trim_head, trim_tail)
        estimated_hits += 1

    return trim_bounds, active_bounds, internal_dead_regions, cache_hits, estimated_hits, internal_dead_hits


def _resolve_playable_window_from_maps(
    path: str,
    dur_full_s: float,
    trim_bounds_map: dict[str, tuple[float, float]],
    hash_active_bounds_map: dict[str, tuple[float, float]],
    internal_dead_regions_map: dict[str, list[tuple[float, float]]],
) -> tuple[float, float, float, float, float, list[tuple[float, float]]]:
    norm_path = _norm_path(path)
    trim_head_s, trim_tail_s = trim_bounds_map.get(norm_path, (0.0, 0.0))
    hash_first_s, hash_last_s = hash_active_bounds_map.get(norm_path, (None, None))
    dur_full_s = max(0.0, float(dur_full_s))
    if dur_full_s > 0.0:
        if (trim_head_s <= 0.0) and (hash_first_s is not None):
            inferred_head = max(0.0, min(float(dur_full_s), float(hash_first_s)))
            if inferred_head >= float(TIMELINE_HASH_EDGE_TRIM_MIN_S):
                trim_head_s = inferred_head
        if (trim_tail_s <= 0.0) and (hash_last_s is not None):
            inferred_tail = max(0.0, float(dur_full_s) - float(hash_last_s))
            if inferred_tail >= float(TIMELINE_HASH_EDGE_TRIM_MIN_S):
                trim_tail_s = inferred_tail
    play_start = max(0.0, min(dur_full_s, float(trim_head_s)))
    play_end = max(play_start, dur_full_s - max(0.0, min(float(trim_tail_s), max(0.0, dur_full_s - play_start))))
    dead_regions = internal_dead_regions_map.get(norm_path, [])
    dead_regions_clamped: list[tuple[float, float]] = []
    if dead_regions and (play_end > play_start):
        edge_touch_tol = max(1.0, min(5.0, float(TIMELINE_STEP_S)))
        min_dead = max(1.0, float(TIMELINE_INTERNAL_DEAD_AUDIO_GAP_MIN_S))
        for s, e in dead_regions:
            s2 = max(play_start, float(s))
            e2 = min(play_end, float(e))
            if (e2 - s2) < min_dead:
                continue
            if (s2 - play_start) <= edge_touch_tol:
                continue
            if (play_end - e2) <= edge_touch_tol:
                continue
            dead_regions_clamped.append((s2, e2))
        dead_regions_clamped = _merge_intervals_simple(dead_regions_clamped, gap_tolerance_s=0.0)
    dead_total = float(sum((e - s) for s, e in dead_regions_clamped))
    play_dur = max(0.0, (play_end - play_start) - dead_total)
    return (
        float(trim_head_s),
        float(trim_tail_s),
        float(play_start),
        float(play_end),
        float(play_dur),
        dead_regions_clamped,
    )


def _get_filtered_timeline_seq_cached(
    vhs: "VideoHashStore",
    seq_cache: dict[str, list[tuple[float, int]]],
    path: str,
) -> list[tuple[float, int]]:
    cached = seq_cache.get(path)
    if cached is not None:
        return cached
    timeline_raw, _trim = vhs.get_timeline(path, step_s=TIMELINE_STEP_S)
    seq = _filter_hash_seq(
        [(t, _bytes_to_int64(h)) for t, h in timeline_raw],
        min_bitcount=TIMELINE_MIN_BITCOUNT,
        max_bitcount=TIMELINE_MAX_BITCOUNT,
        collapse_runs=False,
    )
    seq_cache[path] = seq
    return seq


def _get_cached_anchors(
    vhs: "VideoHashStore",
    anchor_cache: dict[str, dict[str, list[tuple[float, bytes]]]],
    path: str,
) -> dict[str, list[tuple[float, bytes]]]:
    cached = anchor_cache.get(path)
    if cached is not None:
        return cached
    anchors, _trim = vhs.get_anchors(path, window_s=ANCHOR_WINDOW_S, step_s=ANCHOR_STEP_S)
    anchor_cache[path] = anchors
    return anchors


def _serialize_self_repeat_info(repeat_info: dict) -> dict:
    def _round_seg_list(items: list[tuple[float, float]]) -> list[list[float]]:
        out: list[list[float]] = []
        for s, e in items or []:
            try:
                s_f = float(s)
                e_f = float(e)
            except Exception:
                continue
            if e_f > s_f:
                out.append([round(s_f, 1), round(e_f, 1)])
        return out

    pairs_out: list[dict[str, object]] = []
    for pair in (repeat_info or {}).get("repeat_pairs", []) or []:
        if not isinstance(pair, dict):
            continue
        try:
            offset_s = float(pair.get("offset_s", 0.0) or 0.0)
        except Exception:
            offset_s = 0.0
        pairs_out.append(
            {
                "offset_s": round(offset_s, 1),
                "earlier_segments": _round_seg_list(pair.get("earlier_segments", []) or []),
                "later_segments": _round_seg_list(pair.get("later_segments", []) or []),
            }
        )
    return {
        "repeat_pairs": pairs_out,
        "canonical_segments": _round_seg_list((repeat_info or {}).get("canonical_segments", []) or []),
        "repeated_segments": _round_seg_list((repeat_info or {}).get("repeated_segments", []) or []),
        "bin_count": int((repeat_info or {}).get("bin_count", 0) or 0),
        "notes": str((repeat_info or {}).get("notes", "") or ""),
    }


def _deserialize_self_repeat_info(payload: dict | None) -> dict:
    def _seg_list(items: object) -> list[tuple[float, float]]:
        out: list[tuple[float, float]] = []
        if not isinstance(items, list):
            return out
        for item in items:
            if not isinstance(item, (list, tuple)) or len(item) < 2:
                continue
            try:
                s = float(item[0])
                e = float(item[1])
            except Exception:
                continue
            if e > s:
                out.append((s, e))
        return out

    payload = payload if isinstance(payload, dict) else {}
    pairs_out: list[dict[str, object]] = []
    for pair in payload.get("repeat_pairs", []) or []:
        if not isinstance(pair, dict):
            continue
        try:
            offset_s = float(pair.get("offset_s", 0.0) or 0.0)
        except Exception:
            offset_s = 0.0
        pairs_out.append(
            {
                "offset_s": offset_s,
                "earlier_segments": _seg_list(pair.get("earlier_segments", [])),
                "later_segments": _seg_list(pair.get("later_segments", [])),
            }
        )
    return {
        "repeat_pairs": pairs_out,
        "canonical_segments": _seg_list(payload.get("canonical_segments", [])),
        "repeated_segments": _seg_list(payload.get("repeated_segments", [])),
        "bin_count": int(payload.get("bin_count", 0) or 0),
        "notes": str(payload.get("notes", "") or ""),
    }


def _phash64_from_gray(gray: np.ndarray) -> bytes:
    # Input is already 3232 gray from FFmpeg
    g32 = gray.astype(np.float32)
    dct = cv2.dct(g32)
    low = dct[:8, :8]
    med = float(np.median(low))
    bits = (low > med).astype(np.uint8).flatten()
    out = bytearray()
    byte = 0
    for i, b in enumerate(bits):
        byte = (byte << 1) | int(b)
        if (i % 8) == 7:
            out.append(byte & 0xFF)
            byte = 0
    return bytes(out)


def sample_anchor_hashes(
    path: str,
    *,
    window_s: float = ANCHOR_WINDOW_S,
    step_s: float = ANCHOR_STEP_S,   # 1.0  1 Hz. For speed use 2.0; for accuracy 0.5.
    trim: tuple[float, float] | None = None,
) -> dict[str, list[tuple[float, bytes]]]:
    """
    Sample 64-bit DCT pHash at fixed rate in the first/last window_s seconds, after trimming.
    Frames are streamed via FFmpeg (sequential reads)  much faster than random seeks.

    Returns: {"start": [(t_s, 8-byte_hash), ...], "end": [...]}
    """
    head_trim, tail_trim = (0.0, 0.0) if trim is None else trim
    fps = 1.0 / max(step_s, 1e-6)

    # Need total duration to position the tail window.
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        raise RuntimeError(f"Could not open video: {path}")
    vfps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    dur = float(cap.get(cv2.CAP_PROP_FRAME_COUNT) / max(vfps, 1e-6))
    cap.release()

    start_t0 = head_trim
    start_dur = max(0.0, window_s)

    end_t1 = dur - tail_trim
    end_t0 = max(head_trim, end_t1 - window_s)
    end_dur = max(0.0, end_t1 - end_t0)

    start_samples: list[tuple[float, bytes]] = []
    for t, gray in _stream_gray_frames_ffmpeg(path, start_t0, start_dur, fps=fps):
        start_samples.append((t, _phash64_from_gray(gray)))

    end_samples: list[tuple[float, bytes]] = []
    for t, gray in _stream_gray_frames_ffmpeg(path, end_t0, end_dur, fps=fps):
        end_samples.append((t, _phash64_from_gray(gray)))

    return {"start": start_samples, "end": end_samples}


def sample_timeline_hashes(
    path: str,
    *,
    step_s: float = TIMELINE_STEP_S,
    trim: tuple[float, float] | None = None,
    min_bitcount: int = TIMELINE_MIN_BITCOUNT,
    max_bitcount: int = TIMELINE_MAX_BITCOUNT,
) -> list[tuple[float, bytes]]:
    """
    Sample hashes across the full clip (after head/tail trim).
    Returns [(t, hash_bytes), ...] where t is absolute seconds.
    """
    if trim is None:
        trim = estimate_trim_bounds(path)
    head_trim, tail_trim = trim
    fps = 1.0 / max(step_s, 1e-6)

    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        raise RuntimeError(f"Cannot open video: {path}")
    fps_src = cap.get(cv2.CAP_PROP_FPS) or 25.0
    frame_count = cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0.0
    duration = float(frame_count / max(fps_src, 1e-6))
    cap.release()

    start_s = max(0.0, float(head_trim))
    end_s = max(start_s, duration - float(tail_trim))
    duration_s = end_s - start_s
    if duration_s <= 0.0:
        return []

    samples: list[tuple[float, bytes]] = []
    for t, gray in _stream_gray_frames_ffmpeg(path, start_s, duration_s, fps=fps):
        h = _phash64_from_gray(gray)
        bits = _bytes_to_int64(h).bit_count()
        if bits < min_bitcount or bits > max_bitcount:
            continue
        samples.append((t, h))
    return samples


def _summarize_osd_candidates(candidates: list[dict]) -> dict:
    """
    Summarize parsed OSD date candidates from one video.
    Resolve one month-year per inferred timestamp event, then summarize those
    event-level dates. This is more robust than counting individual OCR frames:
    each on-screen timestamp appearance should represent one date even if some
    frames inside that burst are noisy or incomplete.
    """
    out = {
        "osd_month_year_start": "",
        "osd_month_year_end": "",
        "osd_month_year_first_in_video": "",
        "osd_month_year_last_in_video": "",
        "osd_month_year_span_months": 0,
        "osd_out_of_order_in_video": False,
        "osd_unique_month_year_count": 0,
        "osd_top_month_year_counts": "",
    }
    if not candidates:
        return out

    parsed_hits: list[dict] = []
    for c in candidates:
        try:
            m = int(c.get("month"))
            t = float(c.get("t_s", 0.0))
        except Exception:
            continue
        if m < 1 or m > 12:
            continue
        y_raw = c.get("year")
        year = None
        if y_raw is not None and y_raw != "":
            try:
                year = int(y_raw)
            except Exception:
                year = None
        parsed_hits.append({"t_s": t, "month": m, "year": year})

    if not parsed_hits:
        return out

    parsed_hits.sort(key=lambda row: float(row["t_s"]))
    global_full_rows = [
        (int(row["year"]), int(row["month"]), float(row["t_s"]))
        for row in parsed_hits
        if row.get("year") is not None
    ]
    global_year_counts = Counter(y for y, _m, _t in global_full_rows)
    dominant_year = None
    dominant_year_count = 0
    single_year = False
    if global_year_counts:
        dominant_year, dominant_year_count = global_year_counts.most_common(1)[0]
        dominant_year = int(dominant_year)
        dominant_year_count = int(dominant_year_count)
        single_year = len(global_year_counts) == 1

    event_gap_s = max(2.0, float(TIMESTAMP_BURST_STEP_S) * 1.5)
    events: list[list[dict]] = []
    for hit in parsed_hits:
        if not events:
            events.append([hit])
            continue
        prev_t = float(events[-1][-1]["t_s"])
        if float(hit["t_s"]) - prev_t <= event_gap_s:
            events[-1].append(hit)
        else:
            events.append([hit])

    resolved_events: list[dict] = []
    for event in events:
        event_full_rows = [
            (int(row["year"]), int(row["month"]), float(row["t_s"]))
            for row in event
            if row.get("year") is not None
        ]
        event_year_counts = Counter(y for y, _m, _t in event_full_rows)
        event_dominant_year = None
        if event_year_counts:
            event_dominant_year = int(event_year_counts.most_common(1)[0][0])

        resolved_rows: list[tuple[int, int, float, bool]] = []
        for row in event:
            month = int(row["month"])
            t_s = float(row["t_s"])
            year = row.get("year")
            inferred = False
            if year is None:
                inferred_year = None
                if event_dominant_year is not None:
                    inferred_year = int(event_dominant_year)
                elif global_full_rows:
                    nearest_row = min(global_full_rows, key=lambda r: abs(float(r[2]) - t_s))
                    if abs(float(nearest_row[2]) - t_s) <= float(TIMESTAMP_MONTH_ONLY_INFER_NEARBY_S):
                        inferred_year = int(nearest_row[0])
                    elif single_year or dominant_year_count >= max(3, int(round(0.8 * len(global_full_rows)))):
                        inferred_year = int(dominant_year)
                elif single_year and dominant_year is not None:
                    inferred_year = int(dominant_year)
                if inferred_year is None:
                    continue
                year = inferred_year
                inferred = True
            resolved_rows.append((int(year), month, t_s, inferred))

        if not resolved_rows:
            continue

        ym_counts = Counter((y, m) for y, m, _t, _inferred in resolved_rows)
        full_counts = Counter((y, m) for y, m, _t, inferred in resolved_rows if not inferred)
        best_key = max(
            ym_counts.keys(),
            key=lambda k: (
                int(ym_counts.get(k, 0)),
                int(full_counts.get(k, 0)),
                -_month_year_index(k[0], k[1]),
            ),
        )
        resolved_events.append(
            {
                "key": best_key,
                "t_first": min(float(t) for _y, _m, t, _inf in resolved_rows),
                "t_last": max(float(t) for _y, _m, t, _inf in resolved_rows),
                "votes": int(ym_counts.get(best_key, 0)),
            }
        )

    if not resolved_events:
        return out

    counts = Counter(ev["key"] for ev in resolved_events)
    use_keys = sorted(counts.keys(), key=lambda k: _month_year_index(k[0], k[1]))
    cluster_gap = max(0, int(TIMESTAMP_SUMMARY_CLUSTER_MAX_GAP_MONTHS))
    ym_clusters: list[list[tuple[int, int]]] = []
    for key in use_keys:
        if not ym_clusters:
            ym_clusters.append([key])
            continue
        prev = ym_clusters[-1][-1]
        gap = _month_year_index(key[0], key[1]) - _month_year_index(prev[0], prev[1])
        if gap <= max(1, cluster_gap):
            ym_clusters[-1].append(key)
        else:
            ym_clusters.append([key])

    def _cluster_rows(cluster: list[tuple[int, int]]) -> list[dict]:
        cluster_keys = set(cluster)
        return [ev for ev in resolved_events if ev["key"] in cluster_keys]

    def _cluster_score(cluster: list[tuple[int, int]]) -> tuple[int, int, int, int]:
        first_idx = _month_year_index(cluster[0][0], cluster[0][1])
        last_idx = _month_year_index(cluster[-1][0], cluster[-1][1])
        total_votes = int(sum(counts.get(k, 0) for k in cluster))
        span = int(last_idx - first_idx)
        return (total_votes, span, len(cluster), -first_idx)

    cluster_infos = []
    for cluster in ym_clusters if ym_clusters else [use_keys]:
        rows = _cluster_rows(cluster)
        if not rows:
            continue
        total_votes = int(sum(counts.get(k, 0) for k in cluster))
        first_t = min(float(row["t_first"]) for row in rows)
        last_t = max(float(row["t_last"]) for row in rows)
        cluster_infos.append(
            {
                "cluster": list(cluster),
                "score": _cluster_score(cluster),
                "total_votes": total_votes,
                "distinct_months": int(len(cluster)),
                "event_span_s": max(0.0, last_t - first_t),
            }
        )
    if not cluster_infos:
        return out
    dominant_info = max(cluster_infos, key=lambda info: info["score"])
    dominant_votes = max(1, int(dominant_info["total_votes"]))
    retained_clusters: list[list[tuple[int, int]]] = []
    for info in cluster_infos:
        cluster = list(info["cluster"])
        if info is dominant_info:
            retained_clusters.append(cluster)
            continue
        if int(info["distinct_months"]) >= int(TIMESTAMP_SUMMARY_MIN_DISTINCT_MONTHS_FOR_ALT_CLUSTER):
            retained_clusters.append(cluster)
            continue
        if (
            float(info["event_span_s"]) >= float(TIMESTAMP_SUMMARY_MIN_EVENT_SPAN_S_FOR_ALT_CLUSTER)
            and (float(info["total_votes"]) / float(dominant_votes)) >= float(TIMESTAMP_SUMMARY_MIN_CLUSTER_SHARE)
        ):
            retained_clusters.append(cluster)
            continue
    retained_keys = {k for cluster in retained_clusters for k in cluster}
    if not retained_keys:
        retained_keys = set(dominant_info["cluster"])

    start_y, start_m = min(retained_keys, key=lambda k: _month_year_index(k[0], k[1]))
    end_y, end_m = max(retained_keys, key=lambda k: _month_year_index(k[0], k[1]))
    cluster_events = [ev for ev in resolved_events if ev["key"] in retained_keys]
    first_y, first_m = min(cluster_events, key=lambda ev: float(ev["t_first"]))["key"]
    last_y, last_m = max(cluster_events, key=lambda ev: float(ev["t_last"]))["key"]

    span_months = max(0, _month_year_index(end_y, end_m) - _month_year_index(start_y, start_m))
    out["osd_month_year_start"] = _month_year_str(start_y, start_m)
    out["osd_month_year_end"] = _month_year_str(end_y, end_m)
    out["osd_month_year_first_in_video"] = _month_year_str(first_y, first_m)
    out["osd_month_year_last_in_video"] = _month_year_str(last_y, last_m)
    out["osd_month_year_span_months"] = int(span_months)
    out["osd_out_of_order_in_video"] = (
        out["osd_month_year_first_in_video"] != out["osd_month_year_start"]
        or out["osd_month_year_last_in_video"] != out["osd_month_year_end"]
    )
    out["osd_unique_month_year_count"] = int(len(counts))
    top_parts = []
    for (y, m), n in counts.most_common(5):
        top_parts.append(f"{_month_year_str(y, m)}x{int(n)}")
    out["osd_top_month_year_counts"] = "; ".join(top_parts)
    return out


def _build_timestamp_burst_windows(
    hit_times: list[float],
    *,
    scan_start_s: float,
    scan_end_s: float,
    coarse_step_s: float,
) -> list[tuple[float, float]]:
    """
    Build local dense-scan windows around coarse OCR hits.
    Nearby windows are merged aggressively so burst mode behaves like a small
    number of streaming reads instead of many tiny FFmpeg seeks.
    """
    if not hit_times:
        return []
    burst_window_s = max(0.5, float(TIMESTAMP_BURST_WINDOW_S), float(coarse_step_s))
    burst_step_s = max(0.25, float(TIMESTAMP_BURST_STEP_S))
    merge_gap_s = max(float(TIMESTAMP_BURST_MERGE_GAP_S), burst_step_s)
    raw_windows: list[tuple[float, float]] = []
    for t_hit in sorted(set(round(float(t), 1) for t in hit_times)):
        lo = max(float(scan_start_s), float(t_hit) - burst_window_s)
        hi = min(float(scan_end_s), float(t_hit) + burst_window_s)
        if hi > lo:
            raw_windows.append((lo, hi))
    if not raw_windows:
        return []
    windows = _merge_intervals_simple(raw_windows, gap_tolerance_s=merge_gap_s)
    max_windows = max(1, int(TIMESTAMP_BURST_MAX_WINDOWS))
    if len(windows) > max_windows:
        windows = sorted(windows, key=lambda r: (r[1] - r[0]), reverse=True)[:max_windows]
        windows = sorted(windows, key=lambda r: r[0])
    return windows


def _compute_video_timestamp_scan_data(
    path: str,
    *,
    step_s: float = TIMESTAMP_SCAN_STEP_S,
    frame_w: int = TIMESTAMP_SCAN_FRAME_W,
    frame_h: int = TIMESTAMP_SCAN_FRAME_H,
) -> dict:
    """
    Scan one video for camcorder OSD date overlays and summarize month-year range.
    """
    base = {
        "duration_s": 0.0,
        "trim_head_s": 0.0,
        "trim_tail_s": 0.0,
        "scan_start_s": 0.0,
        "scan_end_s": 0.0,
        "scan_step_s": float(step_s),
        "sampled_frames": 0,
        "ocr_text_hits": 0,
        "osd_candidate_hits": 0,
        "status": "no_data",
        "candidates": [],
    }
    if (_osd_ocr_text is None) or (_osd_roi_tophat is None):
        base["status"] = "ocr_unavailable"
        return base

    try:
        duration_s = float(ffprobe_duration_seconds(path))
    except Exception:
        duration_s = 0.0
    base["duration_s"] = round(max(0.0, duration_s), 1)
    if duration_s <= 0.0:
        base["status"] = "duration_unknown"
        return base

    try:
        trim_head_s, trim_tail_s = estimate_trim_bounds(path)
    except Exception:
        trim_head_s, trim_tail_s = 0.0, 0.0
    trim_head_s = max(0.0, float(trim_head_s))
    trim_tail_s = max(0.0, float(trim_tail_s))
    scan_start_s = trim_head_s
    scan_end_s = max(scan_start_s, duration_s - trim_tail_s)
    span_s = max(0.0, scan_end_s - scan_start_s)
    base["trim_head_s"] = round(trim_head_s, 1)
    base["trim_tail_s"] = round(trim_tail_s, 1)
    base["scan_start_s"] = round(scan_start_s, 1)
    base["scan_end_s"] = round(scan_end_s, 1)
    if span_s <= 0.0:
        base["status"] = "empty_after_trim"
        return base

    step_eff = max(1.0, float(step_s))
    burst_enabled = bool(TIMESTAMP_BURST_ENABLE)
    base["scan_step_s"] = round(step_eff, 1)
    coarse_fps = 1.0 / max(1e-6, step_eff)
    base["burst_enabled"] = burst_enabled
    base["burst_windows_used"] = 0
    base["burst_sampled_frames"] = 0

    sampled_frames = 0
    ocr_text_hits = 0
    candidates: list[dict] = []
    coarse_text_hit_times: list[float] = []
    sampled_keys: set[int] = set()

    with tempfile.TemporaryDirectory() as td:
        roi_path = os.path.join(td, "osd_roi.png")

        def _process_frame(t_s: float, frame_bgr: np.ndarray, *, collect_hit_times: list[float] | None = None) -> None:
            nonlocal sampled_frames, ocr_text_hits
            key = int(round(float(t_s) * 10.0))
            if key in sampled_keys:
                return
            sampled_keys.add(key)
            sampled_frames += 1

            try:
                roi_img = _osd_roi_tophat(frame_bgr)
                cv2.imwrite(roi_path, roi_img)
                text = str(_osd_ocr_text(roi_path) or "").strip()
            except Exception:
                roi_img = None
                text = ""
            if text:
                ocr_text_hits += 1
            burst_trigger = bool(_osd_roi_has_timestamp_like_text(roi_img)) if _osd_roi_has_timestamp_like_text is not None and roi_img is not None else False
            if collect_hit_times is not None and (text or burst_trigger):
                collect_hit_times.append(float(t_s))

            try:
                month = int(_osd_extract_month(text)) if _osd_extract_month is not None else None
            except Exception:
                month = None
            try:
                year = int(_osd_extract_year(text)) if _osd_extract_year is not None else None
            except Exception:
                year = None
            if year is not None and not (int(TIMESTAMP_YEAR_MIN) <= int(year) <= int(TIMESTAMP_YEAR_MAX)):
                year = None
            if month is None or not (1 <= month <= 12):
                return

            try:
                day_val = _osd_extract_day(text) if _osd_extract_day is not None else None
            except Exception:
                day_val = None
            try:
                time_val = _osd_extract_time(text) if _osd_extract_time is not None else None
            except Exception:
                time_val = None
            candidates.append({
                "t_s": round(float(t_s), 1),
                "year": int(year) if year is not None else None,
                "month": int(month),
                "day": int(day_val) if day_val is not None else None,
                "time_text": str(time_val or ""),
            })

        # Coarse pass.
        for t_s, frame_bgr in _stream_bgr_frames_ffmpeg(
            path,
            scan_start_s,
            span_s,
            fps=coarse_fps,
            w=frame_w,
            h=frame_h,
        ):
            _process_frame(t_s, frame_bgr, collect_hit_times=coarse_text_hit_times)

        # Local burst pass around coarse OCR text hits.
        if burst_enabled and coarse_text_hit_times:
            burst_step_s = max(0.25, float(TIMESTAMP_BURST_STEP_S))
            burst_fps = 1.0 / max(1e-6, burst_step_s)
            windows = _build_timestamp_burst_windows(
                coarse_text_hit_times,
                scan_start_s=scan_start_s,
                scan_end_s=scan_end_s,
                coarse_step_s=step_eff,
            )

            burst_used = 0
            burst_frames = 0
            for lo, hi in windows:
                if (hi - lo) <= 0.0:
                    continue
                before = sampled_frames
                for t_s, frame_bgr in _stream_bgr_frames_ffmpeg(
                    path,
                    lo,
                    hi - lo,
                    fps=burst_fps,
                    w=frame_w,
                    h=frame_h,
                ):
                    _process_frame(t_s, frame_bgr, collect_hit_times=None)
                if sampled_frames > before:
                    burst_used += 1
                    burst_frames += (sampled_frames - before)
            base["burst_windows_used"] = int(burst_used)
            base["burst_sampled_frames"] = int(burst_frames)

    base["sampled_frames"] = int(sampled_frames)
    base["ocr_text_hits"] = int(ocr_text_hits)
    base["osd_candidate_hits"] = int(len(candidates))
    base["candidates"] = candidates
    base.update(_summarize_osd_candidates(candidates))
    base["summary_version"] = int(TIMESTAMP_SUMMARY_VERSION)

    if sampled_frames <= 0:
        base["status"] = "no_samples"
    elif len(candidates) > 0:
        base["status"] = "ok"
    elif ocr_text_hits > 0:
        base["status"] = "ocr_text_no_date"
    else:
        base["status"] = "no_ocr_text"
    return base


def _timestamp_scan_row_from_data(path: str, fid: int, data: dict) -> dict:
    return {
        "file_path": path,
        "file_name": Path(path).name,
        "folder_id": int(fid),
        "duration_s": round(float(data.get("duration_s", 0.0) or 0.0), 1),
        "trim_head_s": round(float(data.get("trim_head_s", 0.0) or 0.0), 1),
        "trim_tail_s": round(float(data.get("trim_tail_s", 0.0) or 0.0), 1),
        "scan_start_s": round(float(data.get("scan_start_s", 0.0) or 0.0), 1),
        "scan_end_s": round(float(data.get("scan_end_s", 0.0) or 0.0), 1),
        "scan_step_s": round(float(data.get("scan_step_s", 0.0) or 0.0), 1),
        "sampled_frames": int(data.get("sampled_frames", 0) or 0),
        "ocr_text_hits": int(data.get("ocr_text_hits", 0) or 0),
        "osd_candidate_hits": int(data.get("osd_candidate_hits", 0) or 0),
        "osd_month_year_start": _safe_text_cell(data.get("osd_month_year_start", "")),
        "osd_month_year_end": _safe_text_cell(data.get("osd_month_year_end", "")),
        "osd_month_year_first_in_video": _safe_text_cell(data.get("osd_month_year_first_in_video", "")),
        "osd_month_year_last_in_video": _safe_text_cell(data.get("osd_month_year_last_in_video", "")),
        "osd_month_year_span_months": int(data.get("osd_month_year_span_months", 0) or 0),
        "osd_out_of_order_in_video": bool(data.get("osd_out_of_order_in_video", False)),
        "osd_unique_month_year_count": int(data.get("osd_unique_month_year_count", 0) or 0),
        "osd_top_month_year_counts": _safe_text_cell(data.get("osd_top_month_year_counts", "")),
        "status": _safe_text_cell(data.get("status", "")),
    }


def _refresh_timestamp_report_from_cache(
    report_path: str,
    *,
    paths: list[str] | None = None,
) -> int:
    """
    Refresh stale month/year summary fields in dedupe_timestamps.xlsx from cached OCR
    candidates, without rescanning video frames.

    This is intentionally summary-only: it preserves the original OCR candidate data and
    simply reapplies the current summarization logic when cache/report summaries are old.
    """
    if not report_path or not os.path.exists(report_path):
        return 0

    wanted_norms = {_norm_path(p) for p in (paths or []) if isinstance(p, str) and p.strip()}

    try:
        wb = load_workbook(report_path)
    except Exception as e:
        print(f"[timestamps] warning: failed opening timestamp report {report_path}: {e}", flush=True)
        return 0

    ws = wb["Duplicates"] if "Duplicates" in wb.sheetnames else wb.active
    headers = {
        _normalize_col_name(cell.value): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None
    }
    c_path = headers.get("file_path") or headers.get("file") or headers.get("path")
    if not c_path:
        return 0

    field_to_col = {
        "osd_month_year_start": headers.get("osd_month_year_start"),
        "osd_month_year_end": headers.get("osd_month_year_end"),
        "osd_month_year_first_in_video": headers.get("osd_month_year_first_in_video"),
        "osd_month_year_last_in_video": headers.get("osd_month_year_last_in_video"),
        "osd_month_year_span_months": headers.get("osd_month_year_span_months"),
        "osd_out_of_order_in_video": headers.get("osd_out_of_order_in_video"),
        "osd_unique_month_year_count": headers.get("osd_unique_month_year_count"),
        "osd_top_month_year_counts": headers.get("osd_top_month_year_counts"),
        "osd_candidate_hits": headers.get("osd_candidate_hits") or headers.get("ym_hits"),
        "status": headers.get("status"),
    }

    vhs = VideoHashStore()
    changed_rows = 0

    for row_idx in range(2, ws.max_row + 1):
        path_cell = ws.cell(row=row_idx, column=c_path).value
        if not isinstance(path_cell, str) or not path_cell.strip():
            continue
        path_str = path_cell.strip()
        if wanted_norms and _norm_path(path_str) not in wanted_norms:
            continue

        entry = vhs._data.get(path_str) or {}
        cached = entry.get("osd_dates") if isinstance(entry, dict) else None
        if not isinstance(cached, dict):
            continue
        data_existing = dict(cached.get("data", {}) or {})
        candidates_existing = data_existing.get("candidates", []) or []
        if not candidates_existing:
            continue

        refreshed = _summarize_osd_candidates(candidates_existing)
        refreshed["summary_version"] = int(TIMESTAMP_SUMMARY_VERSION)

        row_changed = False
        for field, col_idx in field_to_col.items():
            if not col_idx:
                continue
            if field == "osd_candidate_hits":
                new_value = int(refreshed.get("osd_candidate_hits", len(candidates_existing)) or 0)
            elif field == "status":
                new_value = _safe_text_cell(data_existing.get("status", "") or refreshed.get("status", ""))
            elif field in {"osd_month_year_span_months", "osd_unique_month_year_count"}:
                new_value = int(refreshed.get(field, 0) or 0)
            elif field == "osd_out_of_order_in_video":
                new_value = bool(refreshed.get(field, False))
            else:
                new_value = _safe_text_cell(refreshed.get(field, ""))

            current_value = ws.cell(row=row_idx, column=col_idx).value
            if current_value != new_value:
                ws.cell(row=row_idx, column=col_idx).value = new_value
                row_changed = True

        if _timestamp_summary_differs(data_existing, refreshed):
            data_existing.update(refreshed)
            cached["data"] = data_existing
            entry["osd_dates"] = cached
            vhs._data[path_str] = entry
            vhs._dirty = True
            row_changed = True

        if row_changed:
            changed_rows += 1

    if vhs._dirty:
        try:
            vhs.save_if_dirty()
        except Exception as e:
            print(f"[timestamps] warning: failed saving refreshed timestamp cache: {e}", flush=True)

    if changed_rows:
        try:
            if os.path.exists(report_path):
                _backup_reports([report_path], label="refresh")
            wb.save(report_path)
        except Exception as e:
            print(f"[timestamps] warning: failed saving refreshed timestamp report {report_path}: {e}", flush=True)
            return 0
        print(
            f"[timestamps] refreshed report summaries from cache: {changed_rows} row(s)",
            flush=True,
        )

    return changed_rows


def _timestamp_summary_differs(existing: dict, refreshed: dict) -> bool:
    for field in TIMESTAMP_SUMMARY_FIELDS:
        if existing.get(field) != refreshed.get(field):
            return True
    return int(existing.get("summary_version", 0) or 0) != int(TIMESTAMP_SUMMARY_VERSION)


def _remux_segment_summary(
    seg_status: str,
    seg_len_s: float,
    uncertain_overlap_s: float,
) -> str:
    def _fmt(seconds: float) -> str:
        total = max(0, int(round(float(seconds))))
        h = total // 3600
        m = (total % 3600) // 60
        s = total % 60
        return f"{h:d}:{m:02d}:{s:02d}"

    seg_len_s = max(0.0, float(seg_len_s))
    uncertain_overlap_s = max(0.0, min(seg_len_s, float(uncertain_overlap_s)))
    if seg_len_s <= 0.0:
        return "snippet: empty interval"
    if seg_status == "definite_unique":
        return f"snippet: definite unique; duration {_fmt(seg_len_s)}"
    if seg_status == "borderline_maybe_match":
        return (
            f"snippet: borderline maybe match; duration {_fmt(seg_len_s)}; "
            f"borderline across entire snippet"
        )
    if seg_status == "uncertain_unresolved":
        return (
            f"snippet: unresolved/uncertain; duration {_fmt(seg_len_s)}; "
            f"no confirmed B-side location"
        )
    uncertain_pct = 100.0 * (uncertain_overlap_s / seg_len_s) if seg_len_s > 0.0 else 0.0
    return (
        f"snippet: mixed unique and borderline; duration {_fmt(seg_len_s)}; "
        f"borderline in ~{uncertain_pct:.1f}% of snippet"
    )


def _init_timestamp_scan_state(
    path: str,
    *,
    step_s: float,
    cached_trim: tuple[float, float] | None = None,
    cached_duration_s: float | None = None,
) -> dict:
    """
    Build per-file timestamp scan state used by the two-stage pipeline.
    Stage 1 (I/O): sequential frame reads.
    Stage 2 (CPU): OCR + date parsing on queued frame batches.
    """
    data = {
        "duration_s": 0.0,
        "trim_head_s": 0.0,
        "trim_tail_s": 0.0,
        "scan_start_s": 0.0,
        "scan_end_s": 0.0,
        "scan_step_s": float(step_s),
        "sampled_frames": 0,
        "ocr_text_hits": 0,
        "osd_candidate_hits": 0,
        "status": "no_data",
        "candidates": [],
        "burst_enabled": bool(TIMESTAMP_BURST_ENABLE),
        "burst_windows_used": 0,
        "burst_sampled_frames": 0,
    }

    state = {
        "path": path,
        "data": data,
        "coarse_fps": 0.0,
        "burst_fps": 1.0 / max(1e-6, float(TIMESTAMP_BURST_STEP_S)),
        "coarse_target_frames": 0,
        "estimated_total_frames": 0,
        "scan_start_s": 0.0,
        "scan_end_s": 0.0,
        "scan_span_s": 0.0,
        "queued_keys": set(),
        "processed_keys": set(),
        "candidates": [],
        "ocr_text_hits": 0,
        "coarse_text_hit_times": [],
        "coarse_submitted_batches": 0,
        "coarse_done_batches": 0,
        "coarse_submit_finished": False,
        "coarse_committed": False,
        "burst_submitted_batches": 0,
        "burst_done_batches": 0,
        "burst_submit_finished": False,
        "burst_committed": False,
        "burst_windows": [],
        "burst_windows_used": 0,
        "burst_sampled_frames": 0,
        "ready_for_io": False,
    }

    if (_osd_ocr_text is None) or (_osd_roi_tophat is None):
        data["status"] = "ocr_unavailable"
        return state

    duration_s = 0.0
    if cached_duration_s is not None:
        try:
            duration_s = float(cached_duration_s)
        except Exception:
            duration_s = 0.0
    if duration_s <= 0.0:
        try:
            duration_s = float(ffprobe_duration_seconds(path))
        except Exception:
            duration_s = 0.0
    data["duration_s"] = round(max(0.0, duration_s), 1)
    if duration_s <= 0.0:
        data["status"] = "duration_unknown"
        return state

    if cached_trim is not None:
        try:
            trim_head_s = max(0.0, float(cached_trim[0]))
            trim_tail_s = max(0.0, float(cached_trim[1]))
        except Exception:
            trim_head_s, trim_tail_s = 0.0, 0.0
    else:
        # Timestamp scan should not spend minutes per file on fresh silencedetect.
        # Reuse cached trim bounds when available; otherwise scan the full runtime.
        trim_head_s, trim_tail_s = 0.0, 0.0
    trim_head_s = max(0.0, float(trim_head_s))
    trim_tail_s = max(0.0, float(trim_tail_s))
    scan_start_s = trim_head_s
    scan_end_s = max(scan_start_s, duration_s - trim_tail_s)
    scan_span_s = max(0.0, scan_end_s - scan_start_s)
    data["trim_head_s"] = round(trim_head_s, 1)
    data["trim_tail_s"] = round(trim_tail_s, 1)
    data["scan_start_s"] = round(scan_start_s, 1)
    data["scan_end_s"] = round(scan_end_s, 1)
    if scan_span_s <= 0.0:
        data["status"] = "empty_after_trim"
        return state

    step_eff = max(1.0, float(step_s))
    coarse_target_frames = max(1, int(math.floor(scan_span_s / step_eff)) + 1)
    data["scan_step_s"] = round(step_eff, 1)
    coarse_fps = 1.0 / max(1e-6, step_eff)
    burst_window_s = max(0.5, float(TIMESTAMP_BURST_WINDOW_S))
    burst_window_s = max(burst_window_s, step_eff)
    burst_step_s = max(0.25, float(TIMESTAMP_BURST_STEP_S))
    burst_frames_per_window = max(1, int(math.floor((2.0 * burst_window_s) / burst_step_s)) + 1)
    estimated_total_frames = int(coarse_target_frames)
    if bool(TIMESTAMP_BURST_ENABLE):
        estimated_total_frames += int(max(1, int(TIMESTAMP_BURST_MAX_WINDOWS)) * burst_frames_per_window)

    state.update(
        {
            "coarse_fps": float(coarse_fps),
            "coarse_target_frames": int(coarse_target_frames),
            "estimated_total_frames": int(estimated_total_frames),
            "scan_start_s": float(scan_start_s),
            "scan_end_s": float(scan_end_s),
            "scan_span_s": float(scan_span_s),
            "ready_for_io": True,
        }
    )
    data["status"] = "pending"
    return state


def _timestamp_ocr_batch_worker(
    task: tuple[str, str, list[tuple[float, np.ndarray]]]
) -> tuple[str, str, list[dict], dict]:
    """
    OCR/parse a batch of timestamp frames (CPU stage of pipeline).
    Returns per-frame rows with minimal parsed info; aggregation happens in caller thread.
    """
    path, stage, frames = task
    out_rows: list[dict] = []
    if not frames:
        return path, stage, out_rows, {"elapsed_s": 0.0, "frame_count": 0, "text_hits": 0, "candidate_hits": 0}

    batch_t0 = time.time()
    text_hits = 0
    candidate_hits = 0
    with tempfile.TemporaryDirectory() as td:
        for i, (t_s, frame_bgr) in enumerate(frames):
            roi_path = os.path.join(td, f"roi_{i:04d}.png")
            try:
                roi_img = _osd_roi_tophat(frame_bgr)
                cv2.imwrite(roi_path, roi_img)
                text = str(_osd_ocr_text(roi_path) or "").strip()
            except Exception:
                roi_img = None
                text = ""

            burst_trigger = bool(_osd_roi_has_timestamp_like_text(roi_img)) if _osd_roi_has_timestamp_like_text is not None and roi_img is not None else False
            text_hit = bool(text)
            if text_hit:
                text_hits += 1
            try:
                month = int(_osd_extract_month(text)) if _osd_extract_month is not None else None
            except Exception:
                month = None
            try:
                year = int(_osd_extract_year(text)) if _osd_extract_year is not None else None
            except Exception:
                year = None
            if year is not None and not (int(TIMESTAMP_YEAR_MIN) <= int(year) <= int(TIMESTAMP_YEAR_MAX)):
                year = None

            candidate = None
            if month is not None and 1 <= month <= 12:
                try:
                    day_val = _osd_extract_day(text) if _osd_extract_day is not None else None
                except Exception:
                    day_val = None
                try:
                    time_val = _osd_extract_time(text) if _osd_extract_time is not None else None
                except Exception:
                    time_val = None
                candidate = {
                    "t_s": round(float(t_s), 1),
                    "year": int(year) if year is not None else None,
                    "month": int(month),
                    "day": int(day_val) if day_val is not None else None,
                    "time_text": str(time_val or ""),
                }
                candidate_hits += 1

            out_rows.append(
                {
                    "key": int(round(float(t_s) * 10.0)),
                    "t_s": round(float(t_s), 1),
                    "text_hit": bool(text_hit),
                    "burst_trigger": bool(burst_trigger),
                    "candidate": candidate,
                }
            )
    return path, stage, out_rows, {
        "elapsed_s": max(0.0, time.time() - batch_t0),
        "frame_count": int(len(frames)),
        "text_hits": int(text_hits),
        "candidate_hits": int(candidate_hits),
    }


def _timestamp_scan_worker(task: tuple[str, int, float, int, int]) -> tuple[str, int, dict]:
    path, fid, step_s, frame_w, frame_h = task
    try:
        data = _compute_video_timestamp_scan_data(
            path,
            step_s=float(step_s),
            frame_w=int(frame_w),
            frame_h=int(frame_h),
        )
    except Exception as e:
        data = {
            "status": f"scan_error:{e.__class__.__name__}",
            "duration_s": 0.0,
            "trim_head_s": 0.0,
            "trim_tail_s": 0.0,
            "scan_start_s": 0.0,
            "scan_end_s": 0.0,
            "scan_step_s": float(step_s),
            "sampled_frames": 0,
            "ocr_text_hits": 0,
            "osd_candidate_hits": 0,
            "osd_month_year_start": "",
            "osd_month_year_end": "",
            "osd_month_year_first_in_video": "",
            "osd_month_year_last_in_video": "",
            "osd_month_year_span_months": 0,
            "osd_out_of_order_in_video": False,
            "osd_unique_month_year_count": 0,
            "osd_top_month_year_counts": "",
            "candidates": [],
        }
    return path, int(fid), data


def _hamming64(a: bytes, b: bytes) -> int:
    """Hamming distance between two 8-byte pHashes (64 bits)."""
    if len(a) != 8 or len(b) != 8:
        raise ValueError("Expected 8-byte hashes")
    return (int.from_bytes(a, "big") ^ int.from_bytes(b, "big")).bit_count()


def match_anchor_sets(
    a: list[tuple[float, bytes]],
    b: list[tuple[float, bytes]],
    *,
    hamming_thresh: int = ANCHOR_HAMMING_THRESH,
) -> tuple[float, float, float]:
    """
    For each element of the smaller set, find nearest neighbour in the larger set by Hamming.
    Keep pairs with Hamming  hamming_thresh and collect time offsets.
    Returns: (match_fraction_wrt_smaller, median_offset_s, MAD_offset_s)
    """
    if not a or not b:
        return 0.0, 0.0, float("inf")

    queries, targets = (a, b) if len(a) <= len(b) else (b, a)
    t_hashes = [h for _, h in targets]

    offsets: list[float] = []
    num_good = 0
    for t_s, qh in queries:
        best = 65
        best_idx = -1
        for idx, th in enumerate(t_hashes):
            d = _hamming64(qh, th)
            if d < best:
                best = d
                best_idx = idx
                if best == 0:
                    break
        if best_idx >= 0 and best <= hamming_thresh:
            num_good += 1
            ref_s = targets[best_idx][0]
            offsets.append(t_s - ref_s)

    if num_good == 0:
        return 0.0, 0.0, float("inf")

    offs = np.array(offsets, dtype=np.float64)
    med = float(np.median(offs))
    mad = float(np.median(np.abs(offs - med)))
    return num_good / float(len(queries)), med, mad


def compare_anchors(
    anchorsA: dict[str, list[tuple[float, bytes]]],
    anchorsB: dict[str, list[tuple[float, bytes]]],
    *,
    hamming_thresh: int = ANCHOR_HAMMING_THRESH,
) -> dict:
    """Compare startstart and endend anchor sets and return fractions + robust offsets."""
    f1, o1, m1 = match_anchor_sets(anchorsA.get("start", []), anchorsB.get("start", []), hamming_thresh=hamming_thresh)
    f2, o2, m2 = match_anchor_sets(anchorsA.get("end",   []), anchorsB.get("end",   []), hamming_thresh=hamming_thresh)
    return {
        "start_fraction": f1, "start_offset_s": o1, "start_mad_s": m1,
        "end_fraction":   f2, "end_offset_s":   o2, "end_mad_s":   m2,
    }


def _median_offset_ordered(
    a_list: list[tuple[float, bytes]],
    b_list: list[tuple[float, bytes]],
    *,
    hamming_thresh: int = ANCHOR_HAMMING_THRESH,
) -> tuple[float, float, float]:
    """
    Compute offsets in a fixed order: for each (t_a, h_a) in A, find nearest in B.
    Returns (median_offset_seconds = t_a - t_b, MAD_seconds, match_fraction_wrt_A).
    """
    if not a_list or not b_list:
        return 0.0, float("inf"), 0.0
    offsets: list[float] = []
    good = 0
    b_hashes = [h for _, h in b_list]
    for t_a, h_a in a_list:
        best = 65
        best_idx = -1
        for idx, h_b in enumerate(b_hashes):
            d = _hamming64(h_a, h_b)
            if d < best:
                best = d
                best_idx = idx
                if best == 0:
                    break
        if best_idx >= 0 and best <= hamming_thresh:
            good += 1
            t_b = b_list[best_idx][0]
            offsets.append(t_a - t_b)
    if good == 0:
        return 0.0, float("inf"), 0.0
    offs = np.array(offsets, dtype=np.float64)
    med = float(np.median(offs))
    mad = float(np.median(np.abs(offs - med)))
    frac = good / float(len(a_list))
    return med, mad, frac


def decide_subset_match(
    stats: dict,
    *,
    min_fraction: float = ANCHOR_MIN_FRACTION,
    max_mad_s: float = ANCHOR_MAX_MAD_S,
) -> tuple[bool, str]:
    """
    Decide the relation label based on match strength at start and end.
    Labels: {"fullfull", "first_partfull", "second_partfull", "ambiguous"}.
    """
    sf, sm = stats["start_fraction"], stats["start_mad_s"]
    ef, em = stats["end_fraction"],   stats["end_mad_s"]

    start_strong = (sf >= min_fraction) and (sm <= max_mad_s)
    end_strong   = (ef >= min_fraction) and (em <= max_mad_s)

    if start_strong and end_strong:
        return True, "fullfull"
    if start_strong and not end_strong:
        return True, "first_partfull"
    if end_strong and not start_strong:
        return True, "second_partfull"
    return False, "ambiguous"


def refine_prefix_suffix(
    pathA: str,
    pathB: str,
    *,
    window_s: float = ANCHOR_WINDOW_S,
    step_s: float = ANCHOR_STEP_S,
    hamming_thresh: int = ANCHOR_HAMMING_THRESH,
    min_fraction: float = ANCHOR_MIN_FRACTION,
    max_mad_s: float = ANCHOR_MAX_MAD_S,
    precomputed: tuple[dict, dict] | None = None,
) -> dict:
    """
    Run the anchor refine for A and B. If precomputed=(ancA, ancB) is provided, it uses that.
    Returns a dict with: ok, relation, start_* / end_* stats, subset_start_in_full_s.
    """
    if precomputed is not None:
        ancA, ancB = precomputed
    else:
        trA = estimate_trim_bounds(pathA)
        trB = estimate_trim_bounds(pathB)
        ancA = sample_anchor_hashes(pathA, window_s=window_s, step_s=step_s, trim=trA)
        ancB = sample_anchor_hashes(pathB, window_s=window_s, step_s=step_s, trim=trB)

    stats = compare_anchors(ancA, ancB, hamming_thresh=hamming_thresh)
    ok, relation = decide_subset_match(stats, min_fraction=min_fraction, max_mad_s=max_mad_s)

    # Directional roles by duration (5% tolerance = ambiguous)
    subset_start_in_full_s = None
    relation_dir = relation
    full_clip = None
    subset_clip = None

    if ok and relation in ("first_partfull", "second_partfull"):
        # decide roles by duration
        # (use external durations in caller to set full/subset columns consistently)
        pass  # roles are assigned in the main loop where durations are known

    return {
        "ok": ok,
        "relation": relation,
        "start_fraction": stats["start_fraction"],
        "start_offset_s": stats["start_offset_s"],
        "start_mad_s": stats["start_mad_s"],
        "end_fraction": stats["end_fraction"],
        "end_offset_s": stats["end_offset_s"],
        "end_mad_s": stats["end_mad_s"],
        "subset_start_in_full_s": subset_start_in_full_s,  # caller fills when roles resolved
        "relation_dir": relation_dir,                      # caller fills when roles resolved
        "full_clip": full_clip,                            # caller fills
        "subset_clip": subset_clip,                        # caller fills
    }


# 
# JSON encode/decode helpers for anchors (bytes  hex strings)
# 

def _hash_bytes_to_hex_list(seq: list[tuple[float, bytes]]) -> list[tuple[float, str]]:
    """JSON cant store bytes; store each 8-byte pHash as hex. Timestamps already rounded to 0.1 s."""
    return [(float(t), h.hex()) for t, h in seq]


def _hash_hex_to_bytes_list(seq: list[tuple[float, str]]) -> list[tuple[float, bytes]]:
    """Decode JSON-stored [(t, 'hex'), ] to [(t, bytes), ] for runtime matching."""
    return [(float(t), bytes.fromhex(hx)) for t, hx in seq]


# 
# Persistent cache: legacy hashes + NEW anchors in one JSON
# 

class VideoHashStore:
    """
    Caches per-file:
       legacy 'avg' (64-bit hex) and 'seq' (sparse [(hex, t), ...])
       anchors: {'trim':(head, tail), 'data':{'start':[(t,hex)], 'end':[...]} , 'params', 'mtime'}
    Key: absolute file path.
    """
    _inst = None

    def __new__(cls, path=VIDEO_HASH_STORE_PATH):
        if cls._inst is None:
            cls._inst = super().__new__(cls)
            cls._inst._init(path)
        return cls._inst

    def _init(self, path: str):
        self.path = path
        self._data: Dict[str, Dict] = {}
        self._dirty = False
        self._new_count = 0

        # Load new cache, or migrate old root-level cache if present.
        if os.path.exists(path):
            try:
                with open(path) as f:
                    self._data = json.load(f)
                print(f"[HashStore] Loaded {len(self._data)} cached item(s).", flush=True)
            except json.JSONDecodeError as e:
                recovered = None
                raw = None
                try:
                    with open(path) as f:
                        raw = f.read()
                    cut = raw.rfind('\n  "', 0, e.pos)
                    if cut != -1:
                        truncated = raw[:cut].rstrip()
                        if truncated.endswith(','):
                            truncated = truncated[:-1].rstrip()
                        recovered_text = truncated + '\n}'
                        recovered = json.loads(recovered_text)
                        print(f"[HashStore] Cache file partially recovered; truncated after last complete entry.", flush=True)
                except Exception:
                    recovered = None

                base, ext = os.path.splitext(path)
                backup_path = f"{base}-corrupt{ext}"
                if os.path.exists(backup_path):
                    i = 2
                    while os.path.exists(f"{base}-corrupt_{i}{ext}"):
                        i += 1
                    backup_path = f"{base}-corrupt_{i}{ext}"
                try:
                    os.replace(path, backup_path)
                    print(f"[HashStore] Cache file corrupt. Backed up to {backup_path}.", flush=True)
                except OSError:
                    print(f"[HashStore] Cache file corrupt and could not be backed up: {path}", flush=True)

                if recovered is None:
                    self._data = {}
                else:
                    self._data = recovered
                    try:
                        tmp_path = f"{path}.tmp"
                        with open(tmp_path, "w") as f:
                            json.dump(recovered, f, indent=2)
                        os.replace(tmp_path, path)
                        print(f"[HashStore] Wrote recovered cache to {path}.", flush=True)
                    except OSError:
                        print(f"[HashStore] Could not write recovered cache to {path}.", flush=True)
                    print(f"[HashStore] Recovered {len(self._data)} cached item(s) from corrupted file.", flush=True)
        atexit.register(self.save_if_dirty)

    #  Legacy getters (avg + sparse seq) 
    def get(self, filepath: str) -> Tuple[str, List[Tuple[str, float]]]:
        """Fetch or compute legacy data for one file. Returns (avg_hex, sparse_seq[(hex, t), ])."""
        global PROCESSED_VIDEOS, TOTAL_VIDEOS
        mtime = os.path.getmtime(filepath)
        entry = self._data.get(filepath)

        if not entry or entry.get("mtime") != mtime or "avg" not in entry or "seq" not in entry:
            seq_pairs = _sample_hashes_with_times(filepath)
            avg_hex   = _average_hex(seq_pairs)
            self._data[filepath] = {"mtime": mtime, "avg": avg_hex, "seq": seq_pairs}
            self._dirty = True
            self._new_count += 1
            print(f"[HashStore] NEW {Path(filepath).name} ({len(seq_pairs)} samples)", flush=True)
            if self._new_count >= SAVE_EVERY:
                self.save_if_dirty()
                self._new_count = 0
                pct = (PROCESSED_VIDEOS / TOTAL_VIDEOS * 100) if TOTAL_VIDEOS else 0
                print(f"[find_video_duplicates] Saved {PROCESSED_VIDEOS}/{TOTAL_VIDEOS} videos to JSON ({pct:.1f}%)")
            PROCESSED_VIDEOS += 1
            return avg_hex, seq_pairs

        PROCESSED_VIDEOS += 1
        print(f"[HashStore] CACHE {Path(filepath).name}", flush=True)
        return entry["avg"], entry["seq"]

    #  Anchors 
    def get_anchors(
        self,
        path: str,
        *,
        window_s: float = ANCHOR_WINDOW_S,
        step_s: float = ANCHOR_STEP_S,
        force: bool = False,
    ) -> tuple[dict, tuple[float, float]]:
        """
        Fetch (or compute & cache) start/end anchor hashes for a file.
        Returns: (anchors_dict, trim_tuple)
        """
        mtime = os.path.getmtime(path)
        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        params = {"version": 1, "window_s": float(window_s), "step_s": float(step_s)}
        anc = entry.get("anchors")
        valid = False
        if anc and not force:
            try:
                valid = (
                    isinstance(anc, dict)
                    and anc.get("params") == params
                    and float(anc.get("mtime", 0.0)) == float(mtime)
                    and "data" in anc and "trim" in anc
                )
            except Exception:
                valid = False

        if not valid:
            trim = estimate_trim_bounds(path)
            data = sample_anchor_hashes(path, window_s=window_s, step_s=step_s, trim=trim)
            anc = {
                "mtime": mtime,
                "params": params,
                "trim": (round(float(trim[0]), 1), round(float(trim[1]), 1)),
                "data": {
                    "start": _hash_bytes_to_hex_list(data.get("start", [])),
                    "end":   _hash_bytes_to_hex_list(data.get("end",   [])),
                },
            }
            entry["anchors"] = anc
            self._data[path] = entry
            self._dirty = True

        runtime_data = {
            "start": _hash_hex_to_bytes_list(anc["data"].get("start", [])),
            "end":   _hash_hex_to_bytes_list(anc["data"].get("end",   [])),
        }
        trim_tuple = (float(anc["trim"][0]), float(anc["trim"][1]))
        return runtime_data, trim_tuple

    def get_timeline(
        self,
        path: str,
        *,
        step_s: float = TIMELINE_STEP_S,
        force: bool = False,
    ) -> tuple[list[tuple[float, bytes]], tuple[float, float]]:
        mtime = os.path.getmtime(path)
        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        params = {"version": 1, "step_s": float(step_s)}
        tl = entry.get("timeline")
        valid = False
        if tl and not force:
            try:
                valid = (
                    isinstance(tl, dict)
                    and tl.get("params") == params
                    and float(tl.get("mtime", 0.0)) == float(mtime)
                    and "data" in tl and "trim" in tl
                )
            except Exception:
                valid = False

        if not valid:
            trim = estimate_trim_bounds(path)
            data = sample_timeline_hashes(path, step_s=step_s, trim=trim)
            tl = {
                "mtime": mtime,
                "params": params,
                "trim": (round(float(trim[0]), 1), round(float(trim[1]), 1)),
                "data": _hash_bytes_to_hex_list(data),
            }
            entry["timeline"] = tl
            self._data[path] = entry
            self._dirty = True

        runtime_seq = _hash_hex_to_bytes_list(tl.get("data", []))
        trim_tuple = (float(tl["trim"][0]), float(tl["trim"][1]))
        return runtime_seq, trim_tuple

    def get_cached_trim_bounds(
        self,
        path: str,
    ) -> tuple[tuple[float, float] | None, float | None, str | None]:
        """
        Fetch cached trim bounds without triggering recomputation.
        Preference order:
        1. timeline trim
        2. anchors trim
        3. prior timestamp-scan trim
        Returns: (trim_tuple|None, duration_s|None, source|None)
        """
        entry = self._data.get(path)
        if not isinstance(entry, dict):
            return None, None, None

        try:
            file_mtime = float(os.path.getmtime(path))
        except Exception:
            file_mtime = None

        def _mtime_ok(block: dict | None) -> bool:
            if not isinstance(block, dict):
                return False
            if file_mtime is None:
                return True
            try:
                return abs(float(block.get("mtime", 0.0) or 0.0) - file_mtime) <= 0.5
            except Exception:
                return False

        for key, source in (("timeline", "timeline"), ("anchors", "anchors")):
            block = entry.get(key)
            if not _mtime_ok(block):
                continue
            trim = block.get("trim")
            if not isinstance(trim, (list, tuple)) or len(trim) < 2:
                continue
            try:
                trim_head = max(0.0, float(trim[0]))
                trim_tail = max(0.0, float(trim[1]))
            except Exception:
                continue
            return (trim_head, trim_tail), None, source

        block = entry.get("osd_dates")
        if _mtime_ok(block):
            data = block.get("data")
            if isinstance(data, dict):
                try:
                    trim_head = max(0.0, float(data.get("trim_head_s", 0.0) or 0.0))
                    trim_tail = max(0.0, float(data.get("trim_tail_s", 0.0) or 0.0))
                    duration_s = float(data.get("duration_s", 0.0) or 0.0)
                    return (trim_head, trim_tail), (duration_s if duration_s > 0.0 else None), "timestamp_cache"
                except Exception:
                    pass
        return None, None, None

    def inspect_internal_dead_cache(
        self,
        path: str,
        *,
        force: bool = False,
    ) -> dict:
        try:
            mtime = float(os.path.getmtime(path))
        except Exception:
            mtime = 0.0
        params = {
            "version": int(TIMELINE_INTERNAL_DEAD_CACHE_VERSION),
            "min_gap_s": float(TIMELINE_INTERNAL_DEAD_AUDIO_GAP_MIN_S),
            "sample_count": int(TIMELINE_INTERNAL_DEAD_VISUAL_SAMPLE_COUNT),
            "frame_w": int(TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_W),
            "frame_h": int(TIMELINE_INTERNAL_DEAD_VISUAL_FRAME_H),
            "black_mean_max": float(TIMELINE_INTERNAL_DEAD_BLACK_MEAN_MAX),
            "black_std_max": float(TIMELINE_INTERNAL_DEAD_BLACK_STD_MAX),
            "noise_sat_max": float(TIMELINE_INTERNAL_DEAD_NOISE_SAT_MAX),
            "noise_std_min": float(TIMELINE_INTERNAL_DEAD_NOISE_STD_MIN),
            "noise_neighbor_corr_max": float(TIMELINE_INTERNAL_DEAD_NOISE_NEIGHBOR_CORR_MAX),
        }
        if force:
            return {
                "cache_hit": False,
                "cache_miss_reason": "forced_rebuild",
                "params": params,
                "mtime": mtime,
            }
        entry = self._data.get(path)
        if not isinstance(entry, dict):
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_entry",
                "params": params,
                "mtime": mtime,
            }
        cached = entry.get("internal_dead")
        if cached is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_internal_dead_cache",
                "params": params,
                "mtime": mtime,
            }
        try:
            if not isinstance(cached, dict):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "invalid_internal_dead_cache",
                    "params": params,
                    "mtime": mtime,
                }
            if cached.get("params") != params:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "params_changed",
                    "params": params,
                    "mtime": mtime,
                }
            if float(cached.get("mtime", 0.0)) != float(mtime):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "mtime_changed",
                    "params": params,
                    "mtime": mtime,
                }
            if "data" not in cached:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_internal_dead_data",
                    "params": params,
                    "mtime": mtime,
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_internal_dead_cache",
                "params": params,
                "mtime": mtime,
            }
        return {
            "cache_hit": True,
            "cache_miss_reason": None,
            "params": params,
            "mtime": mtime,
        }

    def get_confirmed_internal_dead_regions(
        self,
        path: str,
        *,
        audio_entry: dict | None = None,
        force: bool = False,
    ) -> tuple[list[tuple[float, float]], bool]:
        status = self.inspect_internal_dead_cache(path, force=force)
        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        cached = entry.get("internal_dead")
        if status["cache_hit"] and isinstance(cached, dict):
            data = cached.get("data", []) or []
            out = []
            for item in data:
                if not isinstance(item, (list, tuple)) or len(item) < 2:
                    continue
                try:
                    s = float(item[0])
                    e = float(item[1])
                except Exception:
                    continue
                if e > s:
                    out.append((s, e))
            return out, True

        if audio_entry is None and isinstance(entry, dict):
            audio_entry = entry.get("audio")
        candidate_gaps = _candidate_internal_dead_regions_from_audio_cache(audio_entry)
        confirmed = _confirm_internal_dead_regions_with_streaming(path, candidate_gaps)
        entry["internal_dead"] = {
            "mtime": float(status["mtime"]),
            "params": status["params"],
            "data": [[round(float(s), 1), round(float(e), 1)] for s, e in confirmed],
        }
        self._data[path] = entry
        self._dirty = True
        return confirmed, False

    def inspect_a_self_repeat_cache(
        self,
        path: str,
        *,
        force: bool = False,
    ) -> dict:
        mtime = os.path.getmtime(path)
        params = {
            "version": int(TIMELINE_A_SELF_REPEAT_CACHE_VERSION),
            "hamming_thresh": int(TIMELINE_A_SELF_REPEAT_HAMMING_THRESH),
            "min_segment_s": float(TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S),
            "min_bin_votes": int(TIMELINE_A_SELF_REPEAT_MIN_BIN_VOTES),
            "neighbor_bins": int(TIMELINE_A_SELF_REPEAT_BIN_NEIGHBOR_BINS),
            "min_offset_s": float(TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S),
            "max_seq_points": int(TIMELINE_A_SELF_REPEAT_MAX_SEQ_POINTS),
            "step_s": float(TIMELINE_STEP_S),
            "bin_s": float(TIMELINE_BIN_S),
            "lsh_chunks": int(TIMELINE_LSH_CHUNKS),
            "brute_limit": None if TIMELINE_BRUTE_MAX is None else int(TIMELINE_BRUTE_MAX),
            "max_candidates_per_frame": (
                None if TIMELINE_MAX_CANDIDATES_PER_FRAME is None else int(TIMELINE_MAX_CANDIDATES_PER_FRAME)
            ),
            "run_gap_mult": float(TIMELINE_RUN_GAP_MULT),
        }
        if force:
            return {"cache_hit": False, "cache_miss_reason": "force", "params": params, "mtime": mtime}
        entry = self._data.get(path)
        if not isinstance(entry, dict):
            return {"cache_hit": False, "cache_miss_reason": "missing_entry", "params": params, "mtime": mtime}
        cached = entry.get("a_self_repeat")
        if cached is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_a_self_repeat_cache",
                "params": params,
                "mtime": mtime,
            }
        try:
            if not isinstance(cached, dict):
                raise TypeError
            if cached.get("params") != params:
                return {"cache_hit": False, "cache_miss_reason": "params_changed", "params": params, "mtime": mtime}
            if float(cached.get("mtime", 0.0)) != float(mtime):
                return {"cache_hit": False, "cache_miss_reason": "mtime_changed", "params": params, "mtime": mtime}
            if "data" not in cached:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_a_self_repeat_data",
                    "params": params,
                    "mtime": mtime,
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_a_self_repeat_cache",
                "params": params,
                "mtime": mtime,
            }
        return {"cache_hit": True, "cache_miss_reason": None, "params": params, "mtime": mtime}

    def get_a_self_repeat_analysis(
        self,
        path: str,
        *,
        seq: list[tuple[float, int]] | None = None,
        force: bool = False,
    ) -> tuple[dict, bool]:
        status = self.inspect_a_self_repeat_cache(path, force=force)
        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        cached = entry.get("a_self_repeat")
        if status["cache_hit"] and isinstance(cached, dict):
            return _deserialize_self_repeat_info(cached.get("data")), True

        if seq is None:
            timeline_raw, _trim = self.get_timeline(path, step_s=TIMELINE_STEP_S)
            seq = _filter_hash_seq(
                [(t, _bytes_to_int64(h)) for t, h in timeline_raw],
                min_bitcount=TIMELINE_MIN_BITCOUNT,
                max_bitcount=TIMELINE_MAX_BITCOUNT,
                collapse_runs=False,
            )
        repeat_info = _detect_self_repeat_segments_from_timeline_hashes(
            seq,
            total_bits=64,
            hamming_thresh=TIMELINE_A_SELF_REPEAT_HAMMING_THRESH,
            bin_s=TIMELINE_BIN_S,
            step_s=TIMELINE_STEP_S,
            lsh_chunks=TIMELINE_LSH_CHUNKS,
            brute_limit=TIMELINE_BRUTE_MAX,
            max_candidates_per_frame=TIMELINE_MAX_CANDIDATES_PER_FRAME,
            run_gap_mult=TIMELINE_RUN_GAP_MULT,
            min_segment_s=TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S,
            min_bin_votes=TIMELINE_A_SELF_REPEAT_MIN_BIN_VOTES,
            neighbor_bins=TIMELINE_A_SELF_REPEAT_BIN_NEIGHBOR_BINS,
            min_offset_s=TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S,
            max_seq_points=TIMELINE_A_SELF_REPEAT_MAX_SEQ_POINTS,
        )
        entry["a_self_repeat"] = {
            "mtime": float(status["mtime"]),
            "params": status["params"],
            "data": _serialize_self_repeat_info(repeat_info),
        }
        self._data[path] = entry
        self._dirty = True
        return repeat_info, False

    def inspect_anchor_cache(
        self,
        path: str,
        *,
        window_s: float = ANCHOR_WINDOW_S,
        step_s: float = ANCHOR_STEP_S,
        force: bool = False,
    ) -> dict:
        """
        Inspect whether anchor hashes can be reused from cache without recomputation.
        Returns: {"cache_hit": bool, "cache_miss_reason": str|None, "params": dict, "mtime": float}
        """
        mtime = os.path.getmtime(path)
        params = {"version": 1, "window_s": float(window_s), "step_s": float(step_s)}
        if force:
            return {
                "cache_hit": False,
                "cache_miss_reason": "forced_rebuild",
                "params": params,
                "mtime": float(mtime),
            }

        entry = self._data.get(path)
        if entry is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_entry",
                "params": params,
                "mtime": float(mtime),
            }

        anc = entry.get("anchors")
        if anc is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_anchor_cache",
                "params": params,
                "mtime": float(mtime),
            }

        try:
            if not isinstance(anc, dict):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "invalid_anchor_cache",
                    "params": params,
                    "mtime": float(mtime),
                }
            if anc.get("params") != params:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "params_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if float(anc.get("mtime", 0.0)) != float(mtime):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "mtime_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "data" not in anc:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_anchor_data",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "trim" not in anc:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_anchor_trim",
                    "params": params,
                    "mtime": float(mtime),
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_anchor_cache",
                "params": params,
                "mtime": float(mtime),
            }

        return {
            "cache_hit": True,
            "cache_miss_reason": None,
            "params": params,
            "mtime": float(mtime),
        }

    def inspect_timeline_cache(
        self,
        path: str,
        *,
        step_s: float = TIMELINE_STEP_S,
        force: bool = False,
    ) -> dict:
        """
        Inspect whether timeline hashes can be reused from cache without recomputation.
        Returns: {"cache_hit": bool, "cache_miss_reason": str|None, "params": dict, "mtime": float}
        """
        mtime = os.path.getmtime(path)
        params = {"version": 1, "step_s": float(step_s)}
        if force:
            return {
                "cache_hit": False,
                "cache_miss_reason": "forced_rebuild",
                "params": params,
                "mtime": float(mtime),
            }

        entry = self._data.get(path)
        if entry is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_entry",
                "params": params,
                "mtime": float(mtime),
            }

        tl = entry.get("timeline")
        if tl is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_timeline_cache",
                "params": params,
                "mtime": float(mtime),
            }

        try:
            if not isinstance(tl, dict):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "invalid_timeline_cache",
                    "params": params,
                    "mtime": float(mtime),
                }
            if tl.get("params") != params:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "params_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if float(tl.get("mtime", 0.0)) != float(mtime):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "mtime_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "data" not in tl:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_timeline_data",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "trim" not in tl:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_timeline_trim",
                    "params": params,
                    "mtime": float(mtime),
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_timeline_cache",
                "params": params,
                "mtime": float(mtime),
            }

        return {
            "cache_hit": True,
            "cache_miss_reason": None,
            "params": params,
            "mtime": float(mtime),
        }

    def inspect_audio_fingerprint_cache(
        self,
        path: str,
        *,
        target_sr: int = AUDIO_TARGET_SR,
        win_s: float = AUDIO_WIN_S,
        hop_s: float = AUDIO_HOP_S,
        n_bands: int = AUDIO_N_BANDS,
        min_freq: float = AUDIO_MIN_FREQ,
        max_freq: float = AUDIO_MAX_FREQ,
        rms_thresh_db: float | None = AUDIO_RMS_THRESH_DB,
        force: bool = False,
    ) -> dict:
        """
        Inspect whether audio fingerprints can be reused from cache without recomputation.
        Returns: {"cache_hit": bool, "cache_miss_reason": str|None, "params": dict, "mtime": float}
        """
        mtime = os.path.getmtime(path)
        params = {
            "version": 1,
            "sr": int(target_sr),
            "win_s": float(win_s),
            "hop_s": float(hop_s),
            "bands": int(n_bands),
            "min_freq": float(min_freq),
            "max_freq": float(max_freq),
            "rms_thresh_db": rms_thresh_db if rms_thresh_db is None else float(rms_thresh_db),
        }
        if force:
            return {
                "cache_hit": False,
                "cache_miss_reason": "forced_rebuild",
                "params": params,
                "mtime": float(mtime),
            }

        entry = self._data.get(path)
        if entry is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_entry",
                "params": params,
                "mtime": float(mtime),
            }

        cached = entry.get("audio")
        if cached is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_audio_cache",
                "params": params,
                "mtime": float(mtime),
            }

        try:
            if not isinstance(cached, dict):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "invalid_audio_cache",
                    "params": params,
                    "mtime": float(mtime),
                }
            if cached.get("params") != params:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "params_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if float(cached.get("mtime", 0.0)) != float(mtime):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "mtime_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "data" not in cached:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_audio_data",
                    "params": params,
                    "mtime": float(mtime),
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_audio_cache",
                "params": params,
                "mtime": float(mtime),
            }

        return {
            "cache_hit": True,
            "cache_miss_reason": None,
            "params": params,
            "mtime": float(mtime),
        }

    def get_audio_fingerprint(
        self,
        path: str,
        *,
        target_sr: int = AUDIO_TARGET_SR,
        win_s: float = AUDIO_WIN_S,
        hop_s: float = AUDIO_HOP_S,
        n_bands: int = AUDIO_N_BANDS,
        min_freq: float = AUDIO_MIN_FREQ,
        max_freq: float = AUDIO_MAX_FREQ,
        rms_thresh_db: float | None = AUDIO_RMS_THRESH_DB,
        force: bool = False,
    ) -> dict:
        cache_status = self.inspect_audio_fingerprint_cache(
            path,
            target_sr=target_sr,
            win_s=win_s,
            hop_s=hop_s,
            n_bands=n_bands,
            min_freq=min_freq,
            max_freq=max_freq,
            rms_thresh_db=rms_thresh_db,
            force=force,
        )
        mtime = float(cache_status["mtime"])
        params = cache_status["params"]
        valid = bool(cache_status["cache_hit"])
        miss_reason = str(cache_status.get("cache_miss_reason") or "unknown")

        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        cached = entry.get("audio")
        if valid and not (isinstance(cached, dict) and ("data" in cached)):
            valid = False
            miss_reason = "invalid_audio_cache"

        if not valid:
            data = compute_audio_fingerprints(
                path,
                target_sr=target_sr,
                win_s=win_s,
                hop_s=hop_s,
                n_bands=n_bands,
                min_freq=min_freq,
                max_freq=max_freq,
                rms_thresh_db=rms_thresh_db,
            )
            cached = {
                "mtime": mtime,
                "params": params,
                "data": [[float(t), int(h)] for t, h in data],
            }
            entry["audio"] = cached
            self._data[path] = entry
            self._dirty = True

        runtime_seq = [(float(t), int(h)) for t, h in cached.get("data", [])]
        return {
            "params": params,
            "data": runtime_seq,
            "cache_hit": bool(valid),
            "cache_miss_reason": None if valid else miss_reason,
        }

    def inspect_timestamp_cache(
        self,
        path: str,
        *,
        step_s: float = TIMESTAMP_SCAN_STEP_S,
        frame_w: int = TIMESTAMP_SCAN_FRAME_W,
        frame_h: int = TIMESTAMP_SCAN_FRAME_H,
        force: bool = False,
    ) -> dict:
        """
        Inspect whether per-video OSD timestamp scan can be reused from cache.
        """
        mtime = os.path.getmtime(path)
        params = {
            "version": int(TIMESTAMP_CACHE_VERSION),
            "step_s": float(step_s),
            "frame_w": int(frame_w),
            "frame_h": int(frame_h),
            "year_min": int(TIMESTAMP_YEAR_MIN),
            "year_max": int(TIMESTAMP_YEAR_MAX),
            "min_repeat": int(TIMESTAMP_MIN_REPEAT_FOR_ROBUST),
            "burst_enable": bool(TIMESTAMP_BURST_ENABLE),
            "burst_window_s": float(TIMESTAMP_BURST_WINDOW_S),
            "burst_step_s": float(TIMESTAMP_BURST_STEP_S),
            "burst_max_windows": int(TIMESTAMP_BURST_MAX_WINDOWS),
        }
        if force:
            return {
                "cache_hit": False,
                "cache_miss_reason": "forced_rebuild",
                "params": params,
                "mtime": float(mtime),
            }

        entry = self._data.get(path)
        if entry is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_entry",
                "params": params,
                "mtime": float(mtime),
            }

        cached = entry.get("osd_dates")
        if cached is None:
            return {
                "cache_hit": False,
                "cache_miss_reason": "missing_timestamp_cache",
                "params": params,
                "mtime": float(mtime),
            }

        try:
            if not isinstance(cached, dict):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "invalid_timestamp_cache",
                    "params": params,
                    "mtime": float(mtime),
                }
            if cached.get("params") != params:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "params_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if float(cached.get("mtime", 0.0)) != float(mtime):
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "mtime_changed",
                    "params": params,
                    "mtime": float(mtime),
                }
            if "data" not in cached:
                return {
                    "cache_hit": False,
                    "cache_miss_reason": "missing_timestamp_data",
                    "params": params,
                    "mtime": float(mtime),
                }
        except Exception:
            return {
                "cache_hit": False,
                "cache_miss_reason": "invalid_timestamp_cache",
                "params": params,
                "mtime": float(mtime),
            }

        return {
            "cache_hit": True,
            "cache_miss_reason": None,
            "params": params,
            "mtime": float(mtime),
        }

    def get_timestamp_scan(
        self,
        path: str,
        *,
        step_s: float = TIMESTAMP_SCAN_STEP_S,
        frame_w: int = TIMESTAMP_SCAN_FRAME_W,
        frame_h: int = TIMESTAMP_SCAN_FRAME_H,
        force: bool = False,
    ) -> dict:
        cache_status = self.inspect_timestamp_cache(
            path,
            step_s=step_s,
            frame_w=frame_w,
            frame_h=frame_h,
            force=force,
        )
        mtime = float(cache_status["mtime"])
        params = cache_status["params"]
        valid = bool(cache_status["cache_hit"])
        miss_reason = str(cache_status.get("cache_miss_reason") or "unknown")

        entry = self._data.get(path)
        if entry is None:
            entry = {}
            self._data[path] = entry

        cached = entry.get("osd_dates")
        if valid and not (isinstance(cached, dict) and ("data" in cached)):
            valid = False
            miss_reason = "invalid_timestamp_cache"

        if not valid:
            data = _compute_video_timestamp_scan_data(
                path,
                step_s=step_s,
                frame_w=frame_w,
                frame_h=frame_h,
            )
            cached = {
                "mtime": mtime,
                "params": params,
                "data": data,
            }
            entry["osd_dates"] = cached
            self._data[path] = entry
            self._dirty = True
        elif isinstance(cached, dict):
            data_existing = dict(cached.get("data", {}) or {})
            candidates_existing = data_existing.get("candidates", []) or []
            if candidates_existing:
                refreshed_summary = _summarize_osd_candidates(candidates_existing)
                if _timestamp_summary_differs(data_existing, refreshed_summary):
                    data_existing.update(refreshed_summary)
                    data_existing["summary_version"] = int(TIMESTAMP_SUMMARY_VERSION)
                    cached["data"] = data_existing
                    entry["osd_dates"] = cached
                    self._data[path] = entry
                    self._dirty = True
            elif int(data_existing.get("summary_version", 0) or 0) != int(TIMESTAMP_SUMMARY_VERSION):
                data_existing["summary_version"] = int(TIMESTAMP_SUMMARY_VERSION)
                cached["data"] = data_existing
                entry["osd_dates"] = cached
                self._data[path] = entry
                self._dirty = True

        runtime_data = dict(cached.get("data", {}))
        return {
            "params": params,
            "data": runtime_data,
            "cache_hit": bool(valid),
            "cache_miss_reason": None if valid else miss_reason,
        }

    def put_timestamp_scan(
        self,
        path: str,
        *,
        data: dict,
        step_s: float = TIMESTAMP_SCAN_STEP_S,
        frame_w: int = TIMESTAMP_SCAN_FRAME_W,
        frame_h: int = TIMESTAMP_SCAN_FRAME_H,
    ) -> None:
        """
        Store externally computed timestamp scan data into unified cache.
        """
        try:
            mtime = float(os.path.getmtime(path))
        except Exception:
            mtime = 0.0
        params = {
            "version": int(TIMESTAMP_CACHE_VERSION),
            "step_s": float(step_s),
            "frame_w": int(frame_w),
            "frame_h": int(frame_h),
            "year_min": int(TIMESTAMP_YEAR_MIN),
            "year_max": int(TIMESTAMP_YEAR_MAX),
            "min_repeat": int(TIMESTAMP_MIN_REPEAT_FOR_ROBUST),
            "burst_enable": bool(TIMESTAMP_BURST_ENABLE),
            "burst_window_s": float(TIMESTAMP_BURST_WINDOW_S),
            "burst_step_s": float(TIMESTAMP_BURST_STEP_S),
            "burst_max_windows": int(TIMESTAMP_BURST_MAX_WINDOWS),
        }
        entry = self._data.get(path)
        if entry is None:
            entry = {}
        entry["osd_dates"] = {
            "mtime": mtime,
            "params": params,
            "data": dict(data or {}),
        }
        self._data[path] = entry
        self._dirty = True

    def save_if_dirty(self):
        """Persist cache to disk if changed."""
        if self._dirty:
            os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
            tmp_path = f"{self.path}.tmp"
            with open(tmp_path, "w") as f:
                json.dump(self._data, f, indent=2)
            os.replace(tmp_path, self.path)
            self._dirty = False
            print(f"[HashStore] cache saved -> {self.path} ({len(self._data)} items)", flush=True)

    def rename_path_key(
        self,
        old_path: str,
        new_path: str,
        *,
        overwrite_existing: bool = False,
    ) -> bool:
        """Move one cache entry to a new absolute-path key."""
        old_key = str(old_path or "").strip()
        new_key = str(new_path or "").strip()
        if not old_key or not new_key:
            raise ValueError("Both old_path and new_path are required.")
        if old_key == new_key:
            return bool(self._data.get(old_key))
        entry = self._data.get(old_key)
        if not isinstance(entry, dict):
            return False
        if (new_key in self._data) and (not overwrite_existing):
            raise KeyError(f"Destination cache key already exists: {new_key}")
        merged = dict(self._data.get(new_key) or {}) if overwrite_existing else {}
        merged.update(entry)
        try:
            file_mtime = float(os.path.getmtime(new_key))
            merged["mtime"] = file_mtime
            for block_name in ("anchors", "timeline", "audio"):
                block = merged.get(block_name)
                if isinstance(block, dict) and ("mtime" in block):
                    block["mtime"] = file_mtime
        except Exception:
            pass
        self._data[new_key] = merged
        del self._data[old_key]
        self._dirty = True
        return True


# 
# Worker for metadata + optional legacy hashes (parallelised)
# 

def _process_video(path_fid, *, need_legacy: bool):
    """
    Threadpool worker:
       Get fps & frame count (for duration).
       If need_legacy=True: compute legacy avg+seq via the cache (slow).
       Else: skip legacy hashes entirely (anchors mode is faster).
       Return metadata tuple.
    """
    path, fid = path_fid
    try:
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            return None
        fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        cnt = cap.get(cv2.CAP_PROP_FRAME_COUNT) or 1
        cap.release()

        file_size = os.path.getsize(path)
        duration = cnt / fps

        if need_legacy:
            avg_hex, seq = VideoHashStore().get(path)
        else:
            avg_hex, seq = None, None

        return (path, fid, fps, cnt, avg_hex, seq, file_size, duration)
    except Exception as e:
        print(f"[find_video_duplicates] ERROR processing {path}: {e}", flush=True)
        return None


# 
# Main dedupe pipeline
# 

def find_video_duplicates(
    directories: List[str],
    faiss_threshold: float      = FAISS_THRESHOLD,
    align_threshold: float      = ALIGN_THRESHOLD,
    align_offset_limit_s: float = ALIGN_OFFSET_LIMIT_S,
    top_k: int                  = TOP_K,
    self_compare: bool          = False,
    use_gpu: bool               = True,
    report_path: str            = DEFAULT_REPORT_PATH,
    refine_mode: str            = "anchors",   # "anchors" | "legacy" | "both" | "audio" | "timeline"
    export_report: bool         = True,
    open_report: bool           = True,
    audio_rms_thresh_db: float | None = AUDIO_RMS_THRESH_DB,
    audio_hamming_thresh: int   = AUDIO_HAMMING_THRESH,
    audio_strict_hamming_thresh: int | None = AUDIO_STRICT_HAMMING_THRESH,
    audio_bin_s: float          = AUDIO_BIN_S,
    audio_min_votes: int        = AUDIO_MIN_VOTES,
    audio_min_vote_fraction: float = AUDIO_MIN_VOTE_FRACTION,
    audio_min_overlap_s: float  = AUDIO_MIN_OVERLAP_S,
    audio_strict_min_overlap_s: float = AUDIO_STRICT_MIN_OVERLAP_S,
    audio_peak_ratio_min: float = AUDIO_PEAK_RATIO_MIN,
    audio_peak_margin: int      = AUDIO_PEAK_MARGIN,
    audio_lsh_chunks: int       = AUDIO_LSH_CHUNKS,
    audio_brute_max: int        = AUDIO_BRUTE_MAX,
    audio_min_bitcount: int = AUDIO_MIN_BITCOUNT,
    audio_max_bitcount: int = AUDIO_MAX_BITCOUNT,
    audio_min_filtered_frames: int = AUDIO_MIN_FILTERED_FRAMES,
    audio_min_unique_ratio: float = AUDIO_MIN_UNIQUE_RATIO,
    audio_duration_ratio_min: float = AUDIO_DURATION_RATIO_MIN,
    audio_min_hashset_intersect_ratio: float = AUDIO_MIN_HASHSET_INTERSECT_RATIO,
    audio_speed_ratio_min: float = AUDIO_SPEED_RATIO_MIN,
    audio_speed_ratio_max: float = AUDIO_SPEED_RATIO_MAX,
    audio_speed_steps: int = AUDIO_SPEED_STEPS,
    audio_max_candidates_per_frame: int = AUDIO_MAX_CANDIDATES_PER_FRAME,
    audio_long_overlap_override_s: float | None = AUDIO_LONG_OVERLAP_OVERRIDE_S,
    audio_long_overlap_vote_mult: float = AUDIO_LONG_OVERLAP_VOTE_MULT,
    audio_run_gap_mult: float = AUDIO_RUN_GAP_MULT,
    audio_mutual_overlap_ratio_min: float = AUDIO_MUTUAL_OVERLAP_RATIO_MIN,
    audio_pair_workers: int | None = None,
    audio_debug_pairs: list[tuple[str, str]] | None = None,
    timeline_pair_workers: int | None = None,
    timeline_hamming_thresh: int = TIMELINE_HAMMING_THRESH,
    timeline_strict_hamming_thresh: int | None = TIMELINE_STRICT_HAMMING_THRESH,
    timeline_bin_s: float = TIMELINE_BIN_S,
    timeline_min_votes: int = TIMELINE_MIN_VOTES,
    timeline_min_vote_fraction: float = TIMELINE_MIN_VOTE_FRACTION,
    timeline_min_overlap_s: float = TIMELINE_MIN_OVERLAP_S,
    timeline_strict_min_overlap_s: float = TIMELINE_STRICT_MIN_OVERLAP_S,
    timeline_peak_ratio_min: float = TIMELINE_PEAK_RATIO_MIN,
    timeline_peak_margin: int = TIMELINE_PEAK_MARGIN,
    timeline_lsh_chunks: int = TIMELINE_LSH_CHUNKS,
    timeline_brute_max: int = TIMELINE_BRUTE_MAX,
    timeline_min_bitcount: int = TIMELINE_MIN_BITCOUNT,
    timeline_max_bitcount: int = TIMELINE_MAX_BITCOUNT,
    timeline_min_filtered_frames: int = TIMELINE_MIN_FILTERED_FRAMES,
    timeline_min_unique_ratio: float = TIMELINE_MIN_UNIQUE_RATIO,
    timeline_duration_ratio_min: float = TIMELINE_DURATION_RATIO_MIN,
    timeline_enable_speed_sweep: bool = TIMELINE_ENABLE_SPEED_SWEEP,
    timeline_speed_ratio_min: float = TIMELINE_SPEED_RATIO_MIN,
    timeline_speed_ratio_max: float = TIMELINE_SPEED_RATIO_MAX,
    timeline_speed_steps: int = TIMELINE_SPEED_STEPS,
    timeline_max_candidates_per_frame: int = TIMELINE_MAX_CANDIDATES_PER_FRAME,
    timeline_long_overlap_override_s: float | None = TIMELINE_LONG_OVERLAP_OVERRIDE_S,
    timeline_long_overlap_vote_mult: float = TIMELINE_LONG_OVERLAP_VOTE_MULT,
    timeline_run_gap_mult: float = TIMELINE_RUN_GAP_MULT,
    focus_file_a_paths: list[str] | None = None,
    focus_file_b_paths: list[str] | None = None,
) -> pd.DataFrame:
    """
    Run dedupe over one or more directories.

    Pipeline:
      1) Walk files (recursive), filter by extensions.
      2) If audio: build audio fingerprints and offset-vote across pairs.
      3) If timeline: build full-timeline hashes and offset-vote across pairs.
      4) Parallel: gather metadata + legacy hashes only if mode needs them.
      5) If anchors are needed: precompute anchors for all files (parallel, cached).
      6) Build candidate pairs:
           - anchors: direct all-pairs (with duration sanity gate).
           - legacy/both: FAISS shortlist on legacy avg hash.
      7) Evaluate pairs with chosen refine(s).
      8) Export a mode-specific report and open it.
    """
    t0 = time.time()
    print(f"[find_video_duplicates] Start {time.strftime('%H:%M:%S')} (mode={refine_mode})", flush=True)
    print(f"[mode] legacy={'yes' if refine_mode in ('legacy','both') else 'no'}, "
          f"anchors={'yes' if refine_mode in ('anchors','both') else 'no'}, "
          f"audio={'yes' if refine_mode == 'audio' else 'no'}, "
          f"timeline={'yes' if refine_mode == 'timeline' else 'no'}, "
          f"faiss={'no' if refine_mode in ('anchors','audio','timeline') else 'yes'}")

    from funcs import get_list_of_files  # local import avoids circularity

    #  1) Gather files (recursive)  build a single grand total
    all_tasks = []
    grand_total = 0
    for fid, folder in enumerate(directories):
        all_files = get_list_of_files(folder)
        print(f"[find_video_duplicates] Folder {fid}: {folder} -> {len(all_files)} files", flush=True)
        video_paths = [f for f in all_files
                       if Path(f).suffix.lower() in EXTS
                       and '_gsdata_' not in f
                       and not Path(f).name.startswith("._")]
        print(f"[find_video_duplicates]   -> {len(video_paths)} video files", flush=True)
        all_tasks += [(p, fid) for p in video_paths]
        grand_total += len(video_paths)

    global TOTAL_VIDEOS
    TOTAL_VIDEOS = grand_total
    if grand_total == 0:
        print("No readable videos found.", flush=True)
        return pd.DataFrame()
    print(f"[totals] {grand_total} videos across {len(directories)} folder(s)", flush=True)

    focus_file_b_set: set[str] | None = None
    if focus_file_b_paths and len(directories) == 2:
        focus_file_b_set = {
            _norm_path(p)
            for p in focus_file_b_paths
            if isinstance(p, str) and p.strip()
        }
        print(f"[focus] limiting pair evaluation to {len(focus_file_b_set)} file_b path(s)", flush=True)
    focus_file_a_set: set[str] | None = None
    if focus_file_a_paths and len(directories) == 2:
        focus_file_a_set = {
            _norm_path(p)
            for p in focus_file_a_paths
            if isinstance(p, str) and p.strip()
        }
        print(f"[focus] limiting pair evaluation to {len(focus_file_a_set)} file_a path(s)", flush=True)

    if refine_mode == "audio":
        prefer_folder = 0 if len(directories) == 2 else None
        vhs = VideoHashStore()
        debug_pairs = set()
        if audio_debug_pairs:
            for a, b in audio_debug_pairs:
                debug_pairs.add((os.path.normpath(a), os.path.normpath(b)))
                debug_pairs.add((os.path.normpath(b), os.path.normpath(a)))
        if audio_pair_workers is None:
            audio_pair_workers = AUDIO_PAIR_WORKERS
        if audio_pair_workers <= 0:
            auto_workers = max(1, (os.cpu_count() or 1) - 1)
            audio_pair_workers = min(auto_workers, max(1, int(AUDIO_PAIR_WORKERS_AUTO_MAX)))
        if audio_debug_pairs:
            audio_pair_workers = 1
        print(f"[audio] pair workers: {audio_pair_workers}", flush=True)

        entries = []
        no_audio_files = []
        low_info_files = []
        total_audio = len(all_tasks)
        print(f"[stage] inspecting audio cache for {total_audio} file(s)", flush=True)
        planned_cache_hits = 0
        planned_cache_miss_reasons = {}
        for path, _fid in all_tasks:
            cache_plan = vhs.inspect_audio_fingerprint_cache(path, rms_thresh_db=audio_rms_thresh_db)
            if cache_plan.get("cache_hit"):
                planned_cache_hits += 1
            else:
                reason = str(cache_plan.get("cache_miss_reason") or "unknown")
                planned_cache_miss_reasons[reason] = planned_cache_miss_reasons.get(reason, 0) + 1
        planned_to_compute = max(0, total_audio - planned_cache_hits)
        print(f"[audio-cache] plan: cached={planned_cache_hits}, to_compute={planned_to_compute}", flush=True)
        if planned_cache_miss_reasons:
            reasons = ", ".join(f"{k}={planned_cache_miss_reasons[k]}" for k in sorted(planned_cache_miss_reasons))
            print(f"[audio-cache] plan reasons: {reasons}", flush=True)
        print(
            f"[stage] loading metadata + audio fingerprints for {total_audio} file(s)"
            f" ({planned_to_compute} to compute)",
            flush=True,
        )
        tick_audio = max(1, total_audio // 10)
        audio_t0 = time.time()
        last_audio_status_t = audio_t0
        next_cache_save_t = audio_t0 + float(CACHE_AUTOSAVE_SECONDS)
        audio_done = 0
        audio_cache_hits = 0
        audio_cache_miss_reasons = {}
        for path, fid in all_tasks:
            try:
                duration = float(ffprobe_duration_seconds(path))
            except Exception:
                duration = 0.0
            try:
                size = os.path.getsize(path)
            except Exception:
                size = 0
            audio = vhs.get_audio_fingerprint(path, rms_thresh_db=audio_rms_thresh_db)
            if audio.get("cache_hit"):
                audio_cache_hits += 1
            else:
                reason = str(audio.get("cache_miss_reason") or "unknown")
                audio_cache_miss_reasons[reason] = audio_cache_miss_reasons.get(reason, 0) + 1
            audio_data_raw = audio["data"]
            if not audio_data_raw:
                no_audio_files.append(path)
                audio_data = _filter_hash_seq(
                audio_data_raw,
                min_bitcount=audio_min_bitcount,
                max_bitcount=audio_max_bitcount,
                collapse_runs=True,
            )
            hash_set = {h for _t, h in audio_data}
            unique_ratio = (
                float(len(hash_set) / max(1, len(audio_data)))
                if audio_data else 0.0
            )
            if audio_data and (len(audio_data) < audio_min_filtered_frames or unique_ratio < audio_min_unique_ratio):
                low_info_files.append({
                    "file_path": path,
                    "usable_fingerprints": len(audio_data),
                    "unique_ratio": round(unique_ratio, 3),
                })
                audio_data = []
            entries.append({
                "path": path,
                "fid": fid,
                "duration": duration,
                "size": size,
                "seq": audio_data,
                "seq_len": len(audio_data),
                "hash_set": hash_set,
                "audio_unique_ratio": unique_ratio,
            })
            audio_done += 1
            now = time.time()
            if (audio_done % tick_audio == 0) or (audio_done == total_audio) or ((now - last_audio_status_t) >= 10.0):
                elapsed = now - audio_t0
                avg = elapsed / audio_done if audio_done else 0.0
                eta = max(0.0, (total_audio - audio_done) * avg)
                pct = int(round(100 * audio_done / total_audio)) if total_audio else 100
                print(f"[progress] audio fingerprints {audio_done}/{total_audio} ({pct}%) - "
                      f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}", flush=True)
                last_audio_status_t = now
            if now >= next_cache_save_t:
                vhs.save_if_dirty()
                next_cache_save_t = now + float(CACHE_AUTOSAVE_SECONDS)
            if audio_debug_pairs:
                norm_path = os.path.normpath(path)
                if any(norm_path in pair for pair in debug_pairs):
                    if not audio_data:
                        print(f"[audio-debug] Empty audio fingerprints for {path}", flush=True)

        audio_cache_misses = max(0, total_audio - audio_cache_hits)
        print(f"[audio-cache] actual: hits={audio_cache_hits}, rebuilt={audio_cache_misses}", flush=True)
        if audio_cache_miss_reasons:
            reasons = ", ".join(f"{k}={audio_cache_miss_reasons[k]}" for k in sorted(audio_cache_miss_reasons))
            print(f"[audio-cache] rebuild reasons: {reasons}", flush=True)
        # Persist rebuilt fingerprints before long pair-evaluation stage.
        vhs.save_if_dirty()

        if low_info_files:
            print(f"[audio] {len(low_info_files)} file(s) skipped due to low-information fingerprints.", flush=True)

        folder_ids = [e["fid"] for e in entries]
        paths = [e["path"] for e in entries]
        raw_pairs = []
        skipped_empty = 0
        skipped_ratio = 0
        skipped_hashset = 0
        skipped_focus_a = 0
        skipped_focus = 0
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                if not self_compare and folder_ids[i] == folder_ids[j]:
                    continue
                if focus_file_a_set is not None:
                    a_path = None
                    if int(entries[i]["fid"]) == 0:
                        a_path = entries[i]["path"]
                    elif int(entries[j]["fid"]) == 0:
                        a_path = entries[j]["path"]
                    if (a_path is None) or (_norm_path(a_path) not in focus_file_a_set):
                        skipped_focus_a += 1
                        continue
                if focus_file_b_set is not None:
                    b_path = None
                    if int(entries[i]["fid"]) == 1:
                        b_path = entries[i]["path"]
                    elif int(entries[j]["fid"]) == 1:
                        b_path = entries[j]["path"]
                    if (b_path is None) or (_norm_path(b_path) not in focus_file_b_set):
                        skipped_focus += 1
                        continue
                if not entries[i]["seq"] or not entries[j]["seq"]:
                    skipped_empty += 1
                    continue
                if max(entries[i]["duration"], entries[j]["duration"]) <= 0:
                    skipped_ratio += 1
                    continue
                shorter_dur = min(entries[i]["duration"], entries[j]["duration"])
                ratio = shorter_dur / max(entries[i]["duration"], entries[j]["duration"])
                eff_ratio_min = _effective_duration_ratio_min(shorter_dur, audio_duration_ratio_min)
                if ratio < eff_ratio_min:
                    skipped_ratio += 1
                    continue
                if audio_min_hashset_intersect_ratio > 0.0:
                    hs_i = entries[i]["hash_set"]
                    hs_j = entries[j]["hash_set"]
                    if hs_i and hs_j:
                        small, big = (hs_i, hs_j) if len(hs_i) <= len(hs_j) else (hs_j, hs_i)
                        common = 0
                        for h in small:
                            if h in big:
                                common += 1
                        shared_ratio = float(common) / float(max(1, len(small)))
                        if shared_ratio < float(audio_min_hashset_intersect_ratio):
                            skipped_hashset += 1
                            continue
                a, b = _order_pair(i, j, folder_ids, paths, prefer_folder)
                raw_pairs.append((a, b))
        print(
            f"[audio] candidate pairs: {len(raw_pairs)} "
            f"(skipped empty={skipped_empty}, ratio={skipped_ratio}, hashset={skipped_hashset}, "
            f"focus_a={skipped_focus_a}, focus_b={skipped_focus})",
            flush=True,
        )

        total_bits = 2 * AUDIO_N_BANDS - 1
        results_rows = []
        eval_total = len(raw_pairs)
        eval_tick = max(1, eval_total // 10)
        eval_t0 = time.time()
        mem_limited_pairs = 0
        use_mp = audio_pair_workers > 1 and eval_total > 0
        if use_mp:
            spawn_ok, launch_desc = _spawn_pool_supported_in_this_launch()
            if not spawn_ok:
                print(
                    f"[audio] multiprocessing disabled for this launch context ({launch_desc}); "
                    "falling back to single worker.",
                    flush=True,
                )
                use_mp = False
        params = {
            "total_bits": total_bits,
            "hamming_thresh": audio_hamming_thresh,
            "bin_s": audio_bin_s,
            "min_votes": audio_min_votes,
            "min_overlap_s": audio_min_overlap_s,
            "hop_s": AUDIO_HOP_S,
            "lsh_chunks": audio_lsh_chunks,
            "brute_limit": audio_brute_max,
            "strict_hamming_thresh": audio_strict_hamming_thresh,
            "strict_min_overlap_s": audio_strict_min_overlap_s,
            "min_vote_fraction": audio_min_vote_fraction,
            "peak_ratio_min": audio_peak_ratio_min,
            "peak_margin": audio_peak_margin,
            "speed_ratio_min": audio_speed_ratio_min,
            "speed_ratio_max": audio_speed_ratio_max,
            "speed_steps": audio_speed_steps,
            "max_candidates_per_frame": audio_max_candidates_per_frame,
            "long_overlap_override_s": audio_long_overlap_override_s,
            "long_overlap_vote_mult": audio_long_overlap_vote_mult,
            "run_gap_mult": audio_run_gap_mult,
            "mutual_overlap_ratio_min": audio_mutual_overlap_ratio_min,
        }
        print(f"[stage] evaluating audio candidates - {eval_total} pair(s)", flush=True)
        if use_mp:
            ctx = mp.get_context("spawn")
            if eval_total <= 5_000:
                chunksize = 8
            elif eval_total <= 20_000:
                chunksize = 16
            else:
                chunksize = 32
            last_status_t = time.time()
            with ctx.Pool(
                processes=audio_pair_workers,
                initializer=_init_pair_worker,
                initargs=(entries, params),
            ) as pool:
                for n, (i, j, ok, offset_s, votes, overlap_s) in enumerate(
                    pool.imap_unordered(_eval_pair_worker, raw_pairs, chunksize=chunksize), 1
                ):
                    if votes < 0:
                        mem_limited_pairs += 1
                        continue
                    if ok:
                        results_rows.append({
                            "file_a": entries[i]["path"],
                            "file_b": entries[j]["path"],
                            "duration_a (s)": round(entries[i]["duration"], 1),
                            "duration_b (s)": round(entries[j]["duration"], 1),
                            "size_a (MB)": round(entries[i]["size"] / (1024 * 1024), 2),
                            "size_b (MB)": round(entries[j]["size"] / (1024 * 1024), 2),
                            "audio_offset_s": round(float(offset_s or 0.0), 1),
                            "audio_votes": int(votes),
                            "audio_overlap_s": round(float(overlap_s), 1),
                            "audio_seq_a": int(entries[i]["seq_len"]),
                            "audio_seq_b": int(entries[j]["seq_len"]),
                            "audio_unique_ratio_a": round(float(entries[i]["audio_unique_ratio"]), 3),
                            "audio_unique_ratio_b": round(float(entries[j]["audio_unique_ratio"]), 3),
                        })
                    now = time.time()
                    if (n % eval_tick == 0) or (n == eval_total) or ((now - last_status_t) >= 10.0):
                        pct = int(round(100 * n / eval_total)) if eval_total else 100
                        elapsed = now - eval_t0
                        avg = elapsed / n if n else 0.0
                        eta = max(0.0, (eval_total - n) * avg)
                        print(f"[progress] audio {n}/{eval_total} ({pct}%) - "
                              f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                              flush=True)
                        last_status_t = now
        else:
            _init_pair_worker(entries, params)
            last_status_t = time.time()
            for n, (i, j) in enumerate(raw_pairs, 1):
                pair_t0 = time.time()
                i, j, ok, offset_s, votes, overlap_s = _eval_pair_worker((i, j))
                pair_elapsed = time.time() - pair_t0
                now = time.time()
                force_tick = (now - last_status_t) >= 10.0
                if votes < 0:
                    mem_limited_pairs += 1
                    continue
                if debug_pairs and (entries[i]["path"], entries[j]["path"]) in debug_pairs:
                    print(
                        "[audio-debug] "
                        f"{entries[i]['path']} | {entries[j]['path']} "
                        f"seqs={len(entries[i]['seq'])}/{len(entries[j]['seq'])} "
                        f"offset={offset_s} votes={votes} overlap_s={overlap_s} ok={ok}",
                        flush=True,
                    )
                if ok:
                    results_rows.append({
                        "file_a": entries[i]["path"],
                        "file_b": entries[j]["path"],
                        "duration_a (s)": round(entries[i]["duration"], 1),
                        "duration_b (s)": round(entries[j]["duration"], 1),
                        "size_a (MB)": round(entries[i]["size"] / (1024 * 1024), 2),
                        "size_b (MB)": round(entries[j]["size"] / (1024 * 1024), 2),
                        "audio_offset_s": round(float(offset_s or 0.0), 1),
                        "audio_votes": int(votes),
                        "audio_overlap_s": round(float(overlap_s), 1),
                        "audio_seq_a": int(entries[i]["seq_len"]),
                        "audio_seq_b": int(entries[j]["seq_len"]),
                        "audio_unique_ratio_a": round(float(entries[i]["audio_unique_ratio"]), 3),
                        "audio_unique_ratio_b": round(float(entries[j]["audio_unique_ratio"]), 3),
                    })

                if (n % eval_tick == 0) or (n == eval_total) or force_tick:
                    pct = int(round(100 * n / eval_total)) if eval_total else 100
                    elapsed = now - eval_t0
                    avg = elapsed / n if n else 0.0
                    eta = max(0.0, (eval_total - n) * avg)
                    print(f"[progress] audio {n}/{eval_total} ({pct}%) - "
                          f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                          flush=True)
                    last_status_t = now
                if pair_elapsed > 8.0 and not ((n % eval_tick == 0) or (n == eval_total) or force_tick):
                    elapsed = now - eval_t0
                    avg = elapsed / n if n else 0.0
                    eta = max(0.0, (eval_total - n) * avg)
                    print(f"[progress] audio {n}/{eval_total} - slow pair {_format_duration_auto(pair_elapsed)}"
                          f" - ETA {_format_duration_auto(eta)}",
                          flush=True)
        if mem_limited_pairs > 0:
            print(f"[audio] memory-limited pairs skipped: {mem_limited_pairs}", flush=True)

        columns = [
            "file_a", "file_b",
            "duration_a (s)", "duration_b (s)",
            "size_a (MB)", "size_b (MB)",
            "audio_offset_s", "audio_votes", "audio_overlap_s",
            "audio_seq_a", "audio_seq_b",
            "audio_unique_ratio_a", "audio_unique_ratio_b",
        ]
        df = pd.DataFrame(results_rows, columns=columns)
        if export_report:
            export_excel(df, report_path)
            if no_audio_files or low_info_files:
                try:
                    wb = load_workbook(report_path)
                    if no_audio_files:
                        ws = wb.create_sheet("No Audio")
                        ws.append(["file_path"])
                        for path in sorted(no_audio_files):
                            ws.append([path])
                        ws.freeze_panes = "A2"
                        max_len = max((len(p) for p in no_audio_files), default=len("file_path"))
                        ws.column_dimensions["A"].width = min(max_len + 2, 120)
                    if low_info_files:
                        ws2 = wb.create_sheet("Low Audio Info")
                        ws2.append(["file_path", "usable_fingerprints", "unique_ratio"])
                        for row in sorted(low_info_files, key=lambda r: r["file_path"]):
                            ws2.append([row["file_path"], row["usable_fingerprints"], row["unique_ratio"]])
                        ws2.freeze_panes = "A2"
                        max_len2 = max((len(r["file_path"]) for r in low_info_files), default=len("file_path"))
                        ws2.column_dimensions["A"].width = min(max_len2 + 2, 120)
                        ws2.column_dimensions["B"].width = 22
                        ws2.column_dimensions["C"].width = 14
                    wb.save(report_path)
                    if no_audio_files:
                        print(f"[audio] {len(no_audio_files)} file(s) have no audio; listed in 'No Audio' sheet.", flush=True)
                    if low_info_files:
                        print(f"[audio] {len(low_info_files)} file(s) listed in 'Low Audio Info' sheet.", flush=True)
                except Exception as e:
                    print(f"[audio] Failed to write audio diagnostics sheets: {e}", flush=True)
            print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
                  f"{len(results_rows)} pairs saved", flush=True)
            if open_report:
                open_excel_file(report_path)
        else:
            print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
                  f"{len(results_rows)} pairs found", flush=True)
        return df

    if refine_mode == "timeline":
        prefer_folder = 0 if len(directories) == 2 else None
        vhs = VideoHashStore()
        if timeline_pair_workers is None:
            timeline_pair_workers = TIMELINE_PAIR_WORKERS
        if timeline_pair_workers <= 0:
            timeline_pair_workers = max(1, (os.cpu_count() or 1) - 1)
        print(f"[timeline] pair workers: {timeline_pair_workers}", flush=True)

        entries = []
        no_timeline_files = []
        low_info_files = []
        total_timeline = len(all_tasks)
        print(f"[stage] inspecting timeline cache for {total_timeline} file(s)", flush=True)
        planned_cache_hits = 0
        planned_cache_miss_reasons = {}
        for path, _fid in all_tasks:
            cache_plan = vhs.inspect_timeline_cache(path, step_s=TIMELINE_STEP_S)
            if cache_plan.get("cache_hit"):
                planned_cache_hits += 1
            else:
                reason = str(cache_plan.get("cache_miss_reason") or "unknown")
                planned_cache_miss_reasons[reason] = planned_cache_miss_reasons.get(reason, 0) + 1
        planned_to_compute = max(0, total_timeline - planned_cache_hits)
        print(f"[timeline-cache] plan: cached={planned_cache_hits}, to_compute={planned_to_compute}", flush=True)
        if planned_cache_miss_reasons:
            reasons = ", ".join(f"{k}={planned_cache_miss_reasons[k]}" for k in sorted(planned_cache_miss_reasons))
            print(f"[timeline-cache] plan reasons: {reasons}", flush=True)
        print(
            f"[stage] loading metadata + timeline fingerprints for {total_timeline} file(s)"
            f" ({planned_to_compute} to compute)",
            flush=True,
        )
        tick_timeline = max(1, total_timeline // 10)
        timeline_t0 = time.time()
        last_timeline_status_t = timeline_t0
        next_cache_save_t = timeline_t0 + float(CACHE_AUTOSAVE_SECONDS)
        timeline_done = 0
        timeline_cache_hits = 0
        timeline_cache_miss_reasons = {}
        for path, fid in all_tasks:
            try:
                duration = float(ffprobe_duration_seconds(path))
            except Exception:
                duration = 0.0
            try:
                size = os.path.getsize(path)
            except Exception:
                size = 0
            timeline_cache_state = vhs.inspect_timeline_cache(path, step_s=TIMELINE_STEP_S)
            if timeline_cache_state.get("cache_hit"):
                timeline_cache_hits += 1
            else:
                reason = str(timeline_cache_state.get("cache_miss_reason") or "unknown")
                timeline_cache_miss_reasons[reason] = timeline_cache_miss_reasons.get(reason, 0) + 1
            timeline_bytes, _trim = vhs.get_timeline(path, step_s=TIMELINE_STEP_S)
            timeline_seq_raw = [(t, _bytes_to_int64(h)) for t, h in timeline_bytes]
            if not timeline_seq_raw:
                no_timeline_files.append(path)
            # Keep full filtered timeline (no run-collapsing) for matching.
            # Collapsing can under-estimate contiguous overlap on repetitive footage.
            timeline_seq = _filter_hash_seq(
                timeline_seq_raw,
                min_bitcount=timeline_min_bitcount,
                max_bitcount=timeline_max_bitcount,
                collapse_runs=False,
            )
            # Use a compact collapsed view only for low-information diagnostics/gating.
            timeline_seq_compact = _filter_hash_seq(
                timeline_seq_raw,
                min_bitcount=timeline_min_bitcount,
                max_bitcount=timeline_max_bitcount,
                collapse_runs=True,
            )
            unique_ratio = (
                float(len({h for _t, h in timeline_seq_compact}) / max(1, len(timeline_seq_compact)))
                if timeline_seq_compact else 0.0
            )
            if timeline_seq and (
                len(timeline_seq_compact) < timeline_min_filtered_frames
                or unique_ratio < timeline_min_unique_ratio
            ):
                low_info_files.append({
                    "file_path": path,
                    "usable_fingerprints": len(timeline_seq_compact),
                    "unique_ratio": round(unique_ratio, 3),
                })
                timeline_seq = []
            entries.append({
                "path": path,
                "fid": fid,
                "duration": duration,
                "size": size,
                "seq": timeline_seq,
                "seq_len": len(timeline_seq),
                "timeline_unique_ratio": unique_ratio,
            })
            timeline_done += 1
            now = time.time()
            if now >= next_cache_save_t:
                vhs.save_if_dirty()
                next_cache_save_t = now + float(CACHE_AUTOSAVE_SECONDS)
            if (timeline_done % tick_timeline == 0) or (timeline_done == total_timeline) or ((now - last_timeline_status_t) >= 10.0):
                elapsed = now - timeline_t0
                avg = elapsed / timeline_done if timeline_done else 0.0
                eta = max(0.0, (total_timeline - timeline_done) * avg)
                pct = int(round(100 * timeline_done / total_timeline)) if total_timeline else 100
                last_timeline_status_t = now
                print(f"[progress] timeline fingerprints {timeline_done}/{total_timeline} ({pct}%) - "
                      f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}", flush=True)

        timeline_cache_misses = max(0, total_timeline - timeline_cache_hits)
        print(f"[timeline-cache] actual: hits={timeline_cache_hits}, rebuilt={timeline_cache_misses}", flush=True)
        if timeline_cache_miss_reasons:
            reasons = ", ".join(f"{k}={timeline_cache_miss_reasons[k]}" for k in sorted(timeline_cache_miss_reasons))
            print(f"[timeline-cache] rebuild reasons: {reasons}", flush=True)
        vhs.save_if_dirty()
        if low_info_files:
            print(f"[timeline] {len(low_info_files)} file(s) skipped due to low-information fingerprints.", flush=True)

        folder_ids = [e["fid"] for e in entries]
        paths = [e["path"] for e in entries]
        raw_pairs = []
        skipped_empty = 0
        skipped_ratio = 0
        skipped_focus = 0
        skipped_focus_a = 0
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                if not self_compare and folder_ids[i] == folder_ids[j]:
                    continue
                if focus_file_a_set is not None:
                    a_path = None
                    if int(entries[i]["fid"]) == 0:
                        a_path = entries[i]["path"]
                    elif int(entries[j]["fid"]) == 0:
                        a_path = entries[j]["path"]
                    if (a_path is None) or (_norm_path(a_path) not in focus_file_a_set):
                        skipped_focus_a += 1
                        continue
                if focus_file_b_set is not None:
                    b_path = None
                    if int(entries[i]["fid"]) == 1:
                        b_path = entries[i]["path"]
                    elif int(entries[j]["fid"]) == 1:
                        b_path = entries[j]["path"]
                    if (b_path is None) or (_norm_path(b_path) not in focus_file_b_set):
                        skipped_focus += 1
                        continue
                if not entries[i]["seq"] or not entries[j]["seq"]:
                    skipped_empty += 1
                    continue
                if max(entries[i]["duration"], entries[j]["duration"]) <= 0:
                    skipped_ratio += 1
                    continue
                shorter_dur = min(entries[i]["duration"], entries[j]["duration"])
                ratio = shorter_dur / max(entries[i]["duration"], entries[j]["duration"])
                eff_ratio_min = _effective_duration_ratio_min(shorter_dur, timeline_duration_ratio_min)
                if ratio < eff_ratio_min:
                    skipped_ratio += 1
                    continue
                a, b = _order_pair(i, j, folder_ids, paths, prefer_folder)
                raw_pairs.append((a, b))
        print(
            f"[timeline] candidate pairs: {len(raw_pairs)} "
            f"(skipped empty={skipped_empty}, ratio={skipped_ratio}, focus={skipped_focus}, focus_a={skipped_focus_a})",
            flush=True,
        )

        total_bits = 64
        results_rows = []
        eval_total = len(raw_pairs)
        eval_tick = max(1, eval_total // 10)
        eval_t0 = time.time()
        use_mp = timeline_pair_workers > 1 and eval_total > 0
        (
            timeline_speed_ratio_min_eff,
            timeline_speed_ratio_max_eff,
            timeline_speed_steps_eff,
        ) = _resolve_timeline_speed_search(
            timeline_enable_speed_sweep,
            timeline_speed_ratio_min,
            timeline_speed_ratio_max,
            timeline_speed_steps,
            fixed_ratio=TIMELINE_SPEED_RATIO_FIXED,
        )
        if timeline_speed_steps_eff > 1:
            print(
                "[timeline] speed sweep: "
                f"on ({timeline_speed_ratio_min_eff:.3f}..{timeline_speed_ratio_max_eff:.3f}, "
                f"steps={timeline_speed_steps_eff})",
                flush=True,
            )
        else:
            print(f"[timeline] speed sweep: off (fixed={timeline_speed_ratio_min_eff:.3f})", flush=True)
        if use_mp:
            spawn_ok, launch_desc = _spawn_pool_supported_in_this_launch()
            if not spawn_ok:
                print(
                    f"[timeline] multiprocessing disabled for this launch context ({launch_desc}); "
                    "falling back to single worker.",
                    flush=True,
                )
                use_mp = False
        print(f"[stage] evaluating timeline candidates - {eval_total} pair(s)", flush=True)
        params = {
            "total_bits": total_bits,
            "hamming_thresh": timeline_hamming_thresh,
            "bin_s": timeline_bin_s,
            "min_votes": timeline_min_votes,
            "min_overlap_s": timeline_min_overlap_s,
            "hop_s": TIMELINE_STEP_S,
            "lsh_chunks": timeline_lsh_chunks,
            "brute_limit": timeline_brute_max,
            "strict_hamming_thresh": timeline_strict_hamming_thresh,
            "strict_min_overlap_s": timeline_strict_min_overlap_s,
            "min_vote_fraction": timeline_min_vote_fraction,
            "peak_ratio_min": timeline_peak_ratio_min,
            "peak_margin": timeline_peak_margin,
            "speed_ratio_min": timeline_speed_ratio_min_eff,
            "speed_ratio_max": timeline_speed_ratio_max_eff,
            "speed_steps": timeline_speed_steps_eff,
            "max_candidates_per_frame": timeline_max_candidates_per_frame,
            "long_overlap_override_s": timeline_long_overlap_override_s,
            "long_overlap_vote_mult": timeline_long_overlap_vote_mult,
            "run_gap_mult": timeline_run_gap_mult,
        }
        # Postfit rescue gates: for strong near-miss pairs rejected by base voting,
        # accept if detailed postfit evidence is clearly strong.
        rescue_votes_min = max(150, int(math.ceil(float(timeline_min_votes) * 3.0)))
        rescue_overlap_min = max(60.0, float(timeline_min_overlap_s) * 0.75)
        rescue_strict_overlap_min = max(float(timeline_strict_min_overlap_s), float(timeline_min_overlap_s) * 0.75)
        rescue_strict_votes_min = max(12, int(math.ceil(float(timeline_min_votes) * 0.35)))
        rescue_attempted = 0
        rescue_accepted = 0

        def _is_timeline_rescue_candidate(ok_flag: bool, votes_n: int, overlap_n: float) -> bool:
            return (
                (not ok_flag)
                and bool(TIMELINE_POSTFIT_ENABLE)
                and (int(votes_n) >= rescue_votes_min)
                and (float(overlap_n) >= rescue_overlap_min)
            )

        def _timeline_row_passes_rescue(row: dict) -> bool:
            strict_overlap = float(row.get("timeline_overlap_strict_s", row.get("timeline_overlap_s", 0.0)) or 0.0)
            strict_votes = int(row.get("timeline_votes", 0) or 0)
            strict_seg_n = int(row.get("timeline_segment_count_strict", 0) or 0)
            return (
                strict_seg_n > 0
                and strict_overlap >= rescue_strict_overlap_min
                and strict_votes >= rescue_strict_votes_min
            )

        def _timeline_match_confidence(i_idx: int, j_idx: int, votes_n: int, overlap_n: float) -> float:
            votes_score = min(1.0, float(votes_n) / max(1.0, timeline_min_votes * 2.0))
            overlap_score = min(1.0, float(overlap_n) / max(1.0, timeline_min_overlap_s * 2.0))
            seq_score = min(
                1.0,
                float(min(entries[i_idx]["seq_len"], entries[j_idx]["seq_len"])) / max(1.0, timeline_min_filtered_frames * 2.0),
            )
            uniq_floor = max(1e-6, timeline_min_unique_ratio * 3.0)
            uniq_score = min(
                1.0,
                float(min(entries[i_idx]["timeline_unique_ratio"], entries[j_idx]["timeline_unique_ratio"])) / uniq_floor,
            )
            return round(0.45 * votes_score + 0.45 * overlap_score + 0.05 * seq_score + 0.05 * uniq_score, 3)

        def _build_timeline_result_row(
            i_idx: int,
            j_idx: int,
            offset_s_raw: float | None,
            votes_raw: int,
            overlap_raw: float,
        ) -> dict:
            details = {}
            if TIMELINE_POSTFIT_ENABLE:
                try:
                    details = _timeline_postfit_details(
                        entries[i_idx]["seq"],
                        entries[j_idx]["seq"],
                        total_bits=total_bits,
                        loose_hamming_thresh=timeline_hamming_thresh,
                        strict_hamming_thresh=timeline_strict_hamming_thresh,
                        bin_s=timeline_bin_s,
                        step_s=TIMELINE_STEP_S,
                        lsh_chunks=timeline_lsh_chunks,
                        brute_limit=timeline_brute_max,
                        speed_ratio_min=timeline_speed_ratio_min_eff,
                        speed_ratio_max=timeline_speed_ratio_max_eff,
                        speed_steps=timeline_speed_steps_eff,
                        max_candidates_per_frame=timeline_max_candidates_per_frame,
                        run_gap_mult=timeline_run_gap_mult,
                        max_matches=TIMELINE_POSTFIT_MAX_MATCHES,
                        seed_gap_s=TIMELINE_POSTFIT_SEED_GAP_S,
                    )
                except Exception as ex:
                    print(
                        f"[timeline-postfit] warning: failed for pair: "
                        f"{entries[i_idx]['path']} | {entries[j_idx]['path']} "
                        f"({ex.__class__.__name__}: {ex})",
                        flush=True,
                    )
                    details = {}

            drift_ratio = float(details.get("drift_ratio", 1.0) or 1.0)
            refined_offset = details.get("offset_s", None)
            offset_final = float(offset_s_raw or 0.0) if refined_offset is None else float(refined_offset)
            overlap_strict = float(details.get("overlap_strict_s", 0.0) or 0.0)
            overlap_loose = float(details.get("overlap_raw_s", 0.0) or 0.0)
            overlap_final = float(overlap_raw)
            if overlap_strict > 0.0:
                overlap_final = overlap_strict
            overlap_loose = max(overlap_loose, overlap_final)
            votes_final = int(votes_raw)
            if details:
                votes_final = max(votes_final, int(details.get("votes_strict", 0) or 0))

            row = {
                "file_a": entries[i_idx]["path"],
                "file_b": entries[j_idx]["path"],
                "duration_a (s)": round(entries[i_idx]["duration"], 1),
                "duration_b (s)": round(entries[j_idx]["duration"], 1),
                "size_a (MB)": round(entries[i_idx]["size"] / (1024 * 1024), 2),
                "size_b (MB)": round(entries[j_idx]["size"] / (1024 * 1024), 2),
                "timeline_offset_s": round(offset_final, 1),
                "timeline_votes": int(votes_final),
                "timeline_overlap_s": round(float(overlap_final), 1),
                "match_confidence": _timeline_match_confidence(i_idx, j_idx, int(votes_final), float(overlap_final)),
                "timeline_seq_a": int(entries[i_idx]["seq_len"]),
                "timeline_seq_b": int(entries[j_idx]["seq_len"]),
                "timeline_unique_ratio_a": round(float(entries[i_idx]["timeline_unique_ratio"]), 3),
                "timeline_unique_ratio_b": round(float(entries[j_idx]["timeline_unique_ratio"]), 3),
                # Timeline-specific diagnostics.
                "timeline_drift_ratio": round(drift_ratio, 4),
                "timeline_overlap_raw_s": round(float(overlap_loose), 1),
                "timeline_overlap_strict_unbridged_s": round(float(details.get("overlap_strict_unbridged_s", overlap_final)), 1),
                "timeline_overlap_strict_s": round(float(overlap_final), 1),
                "timeline_segment_count_raw": int(details.get("segment_count_raw", 0) or 0),
                "timeline_segment_count_strict_unbridged": int(details.get("segment_count_strict_unbridged", 0) or 0),
                "timeline_segment_count_strict": int(details.get("segment_count_strict", 0) or 0),
                "timeline_segments_a_raw": str(details.get("segments_a_raw", "") or ""),
                "timeline_segments_b_raw": str(details.get("segments_b_raw", "") or ""),
                "timeline_segments_a_strict": str(details.get("segments_a_strict", "") or ""),
                "timeline_segments_b_strict": str(details.get("segments_b_strict", "") or ""),
                # Shared normalized fields for cross-mode consolidation.
                "drift_ratio": round(drift_ratio, 4),
                "overlap_raw_s": round(float(overlap_loose), 1),
                "overlap_strict_s": round(float(overlap_final), 1),
                "segment_count_strict": int(details.get("segment_count_strict", 0) or 0),
            }
            return row
        if use_mp:
            ctx = mp.get_context("spawn")
            chunksize = max(1, eval_total // (timeline_pair_workers * 8))
            last_status_t = time.time()
            with ctx.Pool(
                processes=timeline_pair_workers,
                initializer=_init_pair_worker,
                initargs=(entries, params),
            ) as pool:
                for n, (i, j, ok, offset_s, votes, overlap_s) in enumerate(
                    pool.imap_unordered(_eval_pair_worker, raw_pairs, chunksize=chunksize), 1
                ):
                    if ok:
                        results_rows.append(_build_timeline_result_row(i, j, offset_s, int(votes), float(overlap_s)))
                    elif _is_timeline_rescue_candidate(ok, int(votes), float(overlap_s)):
                        rescue_attempted += 1
                        rescue_row = _build_timeline_result_row(i, j, offset_s, int(votes), float(overlap_s))
                        if _timeline_row_passes_rescue(rescue_row):
                            results_rows.append(rescue_row)
                            rescue_accepted += 1
                    now = time.time()
                    if (n % eval_tick == 0) or (n == eval_total) or ((now - last_status_t) >= 10.0):
                        pct = int(round(100 * n / eval_total)) if eval_total else 100
                        elapsed = now - eval_t0
                        avg = elapsed / n if n else 0.0
                        eta = max(0.0, (eval_total - n) * avg)
                        accept_rate = (100.0 * len(results_rows) / n) if n else 0.0
                        print(
                            f"[progress] timeline {n}/{eval_total} ({pct}%) - "
                            f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}"
                            f" - accepted {len(results_rows)} ({accept_rate:.1f}%)",
                            flush=True,
                        )
                        last_status_t = now
        else:
            _init_pair_worker(entries, params)
            last_status_t = time.time()
            for n, (i, j) in enumerate(raw_pairs, 1):
                i, j, ok, offset_s, votes, overlap_s = _eval_pair_worker((i, j))
                if ok:
                    results_rows.append(_build_timeline_result_row(i, j, offset_s, int(votes), float(overlap_s)))
                elif _is_timeline_rescue_candidate(ok, int(votes), float(overlap_s)):
                    rescue_attempted += 1
                    rescue_row = _build_timeline_result_row(i, j, offset_s, int(votes), float(overlap_s))
                    if _timeline_row_passes_rescue(rescue_row):
                        results_rows.append(rescue_row)
                        rescue_accepted += 1

                now = time.time()
                if (n % eval_tick == 0) or (n == eval_total) or ((now - last_status_t) >= 10.0):
                    pct = int(round(100 * n / eval_total)) if eval_total else 100
                    elapsed = now - eval_t0
                    avg = elapsed / n if n else 0.0
                    eta = max(0.0, (eval_total - n) * avg)
                    accept_rate = (100.0 * len(results_rows) / n) if n else 0.0
                    print(
                        f"[progress] timeline {n}/{eval_total} ({pct}%) - "
                        f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}"
                        f" - accepted {len(results_rows)} ({accept_rate:.1f}%)",
                        flush=True,
                    )
                    last_status_t = now
        if rescue_attempted > 0:
            print(
                f"[timeline] postfit rescue: accepted {rescue_accepted}/{rescue_attempted} "
                f"near-miss pair(s)",
                flush=True,
            )

        columns = [
            "file_a", "file_b",
            "duration_a (s)", "duration_b (s)",
            "size_a (MB)", "size_b (MB)",
            "timeline_offset_s", "timeline_votes", "timeline_overlap_s",
            "match_confidence",
            "timeline_seq_a", "timeline_seq_b",
            "timeline_unique_ratio_a", "timeline_unique_ratio_b",
            "timeline_drift_ratio",
            "timeline_overlap_raw_s", "timeline_overlap_strict_s",
            "timeline_overlap_strict_unbridged_s",
            "timeline_segment_count_raw", "timeline_segment_count_strict", "timeline_segment_count_strict_unbridged",
            "timeline_segments_a_raw", "timeline_segments_b_raw",
            "timeline_segments_a_strict", "timeline_segments_b_strict",
            "drift_ratio", "overlap_raw_s", "overlap_strict_s", "segment_count_strict",
        ]
        df = pd.DataFrame(results_rows, columns=columns)
        if export_report:
            export_excel(df, report_path)
            if no_timeline_files or low_info_files:
                try:
                    wb = load_workbook(report_path)
                    if no_timeline_files:
                        ws = wb.create_sheet("No Timeline")
                        ws.append(["file_path"])
                        for path in sorted(no_timeline_files):
                            ws.append([path])
                        ws.freeze_panes = "A2"
                        max_len = max((len(p) for p in no_timeline_files), default=len("file_path"))
                        ws.column_dimensions["A"].width = min(max_len + 2, 120)
                    if low_info_files:
                        ws2 = wb.create_sheet("Low Timeline Info")
                        ws2.append(["file_path", "usable_fingerprints", "unique_ratio"])
                        for row in sorted(low_info_files, key=lambda r: r["file_path"]):
                            ws2.append([row["file_path"], row["usable_fingerprints"], row["unique_ratio"]])
                        ws2.freeze_panes = "A2"
                        max_len2 = max((len(r["file_path"]) for r in low_info_files), default=len("file_path"))
                        ws2.column_dimensions["A"].width = min(max_len2 + 2, 120)
                        ws2.column_dimensions["B"].width = 22
                        ws2.column_dimensions["C"].width = 14
                    wb.save(report_path)
                    if no_timeline_files:
                        print(f"[timeline] {len(no_timeline_files)} file(s) have no timeline hashes; listed in 'No Timeline' sheet.", flush=True)
                    if low_info_files:
                        print(f"[timeline] {len(low_info_files)} file(s) listed in 'Low Timeline Info' sheet.", flush=True)
                except Exception as e:
                    print(f"[timeline] Failed to write timeline diagnostics sheets: {e}", flush=True)
            print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
                  f"{len(results_rows)} pairs saved", flush=True)
            if open_report:
                open_excel_file(report_path)
        else:
            print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
                  f"{len(results_rows)} pairs found", flush=True)
        return df

    # If the user passed exactly two folders, put folder 0 on the A side.
    prefer_folder = 0 if len(directories) == 2 else None

    #  2) Metadata (+ legacy hashes only if needed)
    need_legacy = (refine_mode in ("legacy", "both"))
    results = []
    total_jobs = len(all_tasks)
    tick = max(1, total_jobs // 10)  # ~10 updates
    print(f"[stage] metadata scan - need_legacy={need_legacy} - {total_jobs} file(s)")
    meta_t0 = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as exec:
        futures = {exec.submit(_process_video, t, need_legacy=need_legacy): t for t in all_tasks}
        for i, fut in enumerate(concurrent.futures.as_completed(futures), 1):
            res = fut.result()
            if res:
                results.append(res)
            if (i % tick == 0) or (i == total_jobs):
                pct = int(round(100 * i / total_jobs))
                elapsed = time.time() - meta_t0
                print(f"[progress] metadata {i}/{total_jobs} ({pct}%) - "
                      f"{_format_duration_auto(elapsed)} elapsed", flush=True)

    # Unpack basics
    paths, folder_ids, durations, sizes = [], [], [], []
    for path, fid, fps, cnt, avg_hex, seq, file_size, duration in results:
        paths.append(path)
        folder_ids.append(fid)
        durations.append(duration)
        sizes.append(file_size)

    #  3) Anchors precompute (parallel) if needed
    anchors_cache: dict[int, dict] = {}
    if refine_mode in ("anchors", "both"):
        vhs = VideoHashStore()
        total_jobs = len(paths)
        print(f"[stage] inspecting anchors cache for {total_jobs} file(s)", flush=True)
        planned_anchor_hits = 0
        planned_anchor_miss_reasons = {}
        for p in paths:
            cache_plan = vhs.inspect_anchor_cache(p, window_s=ANCHOR_WINDOW_S, step_s=ANCHOR_STEP_S)
            if cache_plan.get("cache_hit"):
                planned_anchor_hits += 1
            else:
                reason = str(cache_plan.get("cache_miss_reason") or "unknown")
                planned_anchor_miss_reasons[reason] = planned_anchor_miss_reasons.get(reason, 0) + 1
        planned_anchor_build = max(0, total_jobs - planned_anchor_hits)
        print(f"[anchors-cache] plan: cached={planned_anchor_hits}, to_compute={planned_anchor_build}", flush=True)
        if planned_anchor_miss_reasons:
            reasons = ", ".join(f"{k}={planned_anchor_miss_reasons[k]}" for k in sorted(planned_anchor_miss_reasons))
            print(f"[anchors-cache] plan reasons: {reasons}", flush=True)
        print(f"[stage] anchors precompute - {total_jobs} file(s) ({planned_anchor_build} to compute)", flush=True)
        anc_t0 = time.time()
        next_anchor_save_t = anc_t0 + float(CACHE_AUTOSAVE_SECONDS)

        def _build(idx_path):
            idx, p = idx_path
            anc, _ = vhs.get_anchors(
                p,
                window_s=ANCHOR_WINDOW_S,
                step_s=ANCHOR_STEP_S,
            )
            return idx, anc

        with concurrent.futures.ThreadPoolExecutor(max_workers=ANCHOR_WORKERS) as exec:
            futures = {exec.submit(_build, (idx, p)): idx for idx, p in enumerate(paths)}
            done_count = 0
            last_print = 0.0
            pending = set(futures.keys())

            while pending:
                # wait up to 10s to either get completions or emit a heartbeat
                done, pending = concurrent.futures.wait(
                    pending,
                    timeout=10.0,
                    return_when=concurrent.futures.FIRST_COMPLETED,
                )

                # collect finished tasks
                for fut in done:
                    idx, anc = fut.result()
                    anchors_cache[idx] = anc
                    done_count += 1

                # heartbeat every 10s (or on completion)
                now = time.time()
                if now >= next_anchor_save_t:
                    vhs.save_if_dirty()
                    next_anchor_save_t = now + float(CACHE_AUTOSAVE_SECONDS)
                if (now - last_print) >= 10.0 or not pending:
                    pct = int(round(100 * done_count / total_jobs)) if total_jobs else 100
                    elapsed = now - anc_t0
                    avg = (elapsed / done_count) if done_count else 0.0
                    eta = max(0.0, (total_jobs - done_count) * avg) if done_count else 0.0
                    print(f"[progress] anchors {done_count}/{total_jobs} ({pct}%) - "
                          f"{_format_duration_auto(elapsed)} elapsed - ~{_format_duration_auto(avg)}/file"
                          f" - ETA {_format_duration_auto(eta)}",
                          flush=True)
                    last_print = now
        vhs.save_if_dirty()

    #  4) Candidate pairs
    if refine_mode == "anchors":
        # Direct all-pairs (unique), with an ultra-cheap duration sanity gate.
        raw_pairs = set()
        N = len(paths)
        for i in range(N):
            for j in range(i + 1, N):
                if paths[i] == paths[j]:
                    continue  # skip self-pairs even if the same file appears twice
                if not self_compare and folder_ids[i] == folder_ids[j]:
                    continue
                if max(durations[i], durations[j]) <= 0:
                    continue
                shorter_dur = min(durations[i], durations[j])
                ratio = shorter_dur / max(durations[i], durations[j])
                eff_ratio_min = _effective_duration_ratio_min(shorter_dur, DURATION_RATIO_MIN)
                if ratio < eff_ratio_min:
                    continue
                a, b = _order_pair(i, j, folder_ids, paths, prefer_folder)
                raw_pairs.add((a, b))
        print(f"[find_video_duplicates] (anchors) {len(raw_pairs)} candidate pairs (no FAISS)", flush=True)
        faiss_dist = {}  # not used in anchors mode

    else:
        # legacy / both: FAISS shortlist on legacy avg hash
        print(f"[stage] FAISS shortlist (legacy avg hash) - {len(paths)} vector(s)")
        avg_vecs = []
        seqs = []  # legacy sparse sequences for aligner
        for (_, _, _, _, avg_hex, seq, _, _) in results:
            avg_vecs.append(_hex_to_vec(avg_hex))
            seqs.append(seq)

        mat = np.stack(avg_vecs).astype("float32")
        index = faiss.IndexFlatL2(mat.shape[1])
        if use_gpu:
            res = faiss.StandardGpuResources()
            index = faiss.index_cpu_to_gpu(res, 0, index)
        index.add(mat)
        D, I = index.search(mat, top_k + 1)
        print(f"[find_video_duplicates] FAISS search done", flush=True)

        raw_pairs = set()
        for i, (drow, idxrow) in enumerate(zip(D, I)):
            for dist, j in zip(drow, idxrow):
                if paths[i] == paths[j]:
                    continue  # skip self-pairs even if the same file appears twice
                if i == j or dist > faiss_threshold:
                    continue
                if not self_compare and folder_ids[i] == folder_ids[j]:
                    continue
                a, b = _order_pair(i, j, folder_ids, paths, prefer_folder)
                raw_pairs.add((a, b))
        print(f"[find_video_duplicates] {len(raw_pairs)} candidate pairs", flush=True)

        # Distance lookup for report
        faiss_dist = {(i, j): float(D[i][k])
                      for i, (drow, idxrow) in enumerate(zip(D, I))
                      for k, j in enumerate(idxrow) if i != j}

    #  5) Evaluate candidates
    results_rows = []
    eval_total = len(raw_pairs)
    eval_tick = max(1, eval_total // 10)
    print(f"[stage] evaluating candidates - {eval_total} pair(s)")
    eval_t0 = time.time()
    last_eval_status_t = eval_t0
    seen_pairs: set[tuple[str, str]] = set()

    for n, (i, j) in enumerate(raw_pairs, 1):
        row = {
            "file_a": paths[i],
            "file_b": paths[j],
            "size_a (MB)": round(sizes[i] / (1024 * 1024), 2),
            "size_b (MB)": round(sizes[j] / (1024 * 1024), 2),
            "duration_a (s)": round(durations[i], 1),
            "duration_b (s)": round(durations[j], 1),
        }
        # Drop exact duplicates across modes/passes (path-level)
        pair_key = (row["file_a"], row["file_b"])
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)

        if refine_mode != "anchors":
            # Only legacy/both have meaningful FAISS distances to display
            row["avg_frame_diff (064)"] = faiss_dist.get((i, j), faiss_dist.get((j, i)))

        legacy_ok = False
        anchors_ok = False
        legacy_score = 0.0
        anchors_score = 0.0

        # Legacy refine (only if mode needs it)
        if refine_mode in ("legacy", "both"):
            seq_i = results[i][5]
            seq_j = results[j][5]
            sec_per = (durations[i] / len(seq_i)) if seq_i else float('inf')
            max_shift = int(align_offset_limit_s / sec_per) if math.isfinite(sec_per) and sec_per > 0 else 0
            best_h, best_ts = _aligned_distance_and_time_limited(
                seq_i, seq_j, max_shift, align_offset_limit_s
            )
            legacy_ok = (best_h <= align_threshold)
            row.update({
                "legacy_best_aligned_diff (064)": best_h,
                "legacy_time_shift_s": round(best_ts, 1),
            })
            dist_score = max(0.0, 1.0 - (float(best_h) / max(1.0, float(align_threshold) * 1.5)))
            shift_score = max(0.0, 1.0 - (abs(float(best_ts)) / max(1.0, float(align_offset_limit_s) * 1.5)))
            legacy_score = 0.8 * dist_score + 0.2 * shift_score

        # Anchors refine (only if mode needs it)
        if refine_mode in ("anchors", "both"):
            ancA = anchors_cache[i]
            ancB = anchors_cache[j]
            stats = compare_anchors(ancA, ancB, hamming_thresh=ANCHOR_HAMMING_THRESH)
            ok, relation = decide_subset_match(
                stats,
                min_fraction=ANCHOR_MIN_FRACTION,
                max_mad_s=ANCHOR_MAX_MAD_S
            )
            anchors_ok = ok

            # Directional roles by duration (5% tolerance = ambiguous)
            relation_dir = relation
            full_clip = None
            subset_clip = None
            subset_start_in_full_s = None

            if ok and relation in ("first_partfull", "second_partfull"):
                durA, durB = durations[i], durations[j]
                if durA > 1.05 * durB:
                    full_clip, subset_clip = "A", "B"
                elif durB > 1.05 * durA:
                    full_clip, subset_clip = "B", "A"

                if full_clip and subset_clip:
                    if relation == "first_partfull":
                        # STARTSTART offsets, ordered subsetfull
                        if subset_clip == "A":
                            med, _mad, _ = _median_offset_ordered(ancA["start"], ancB["start"],
                                                                  hamming_thresh=ANCHOR_HAMMING_THRESH)
                        else:
                            med, _mad, _ = _median_offset_ordered(ancB["start"], ancA["start"],
                                                                  hamming_thresh=ANCHOR_HAMMING_THRESH)
                        subset_start_in_full_s = round(-med, 1)
                        relation_dir = f"{full_clip}=full, {subset_clip}=first_part"

                    elif relation == "second_partfull":
                        # ENDEND offsets, ordered subsetfull
                        if subset_clip == "A":
                            med, _mad, _ = _median_offset_ordered(ancA["end"], ancB["end"],
                                                                  hamming_thresh=ANCHOR_HAMMING_THRESH)
                        else:
                            med, _mad, _ = _median_offset_ordered(ancB["end"], ancA["end"],
                                                                  hamming_thresh=ANCHOR_HAMMING_THRESH)
                        subset_start_in_full_s = round(-med, 1)
                        relation_dir = f"{full_clip}=full, {subset_clip}=second_part"

            row.update({
                "relation": relation,
                "relation_dir": relation_dir,
                "full_clip": full_clip,
                "subset_clip": subset_clip,
                "subset_start_in_full_s": subset_start_in_full_s,
                "subset_start_mm:ss": _format_mmss(subset_start_in_full_s),
                # symmetric quality stats for transparency
                "start_match_fraction": round(stats["start_fraction"], 3),
                "start_offset_s": round(stats["start_offset_s"], 1),
                "start_offset_mad_s": (round(stats["start_mad_s"], 1)
                                       if math.isfinite(stats["start_mad_s"]) else None),
                "end_match_fraction": round(stats["end_fraction"], 3),
                "end_offset_s": round(stats["end_offset_s"], 1),
                "end_offset_mad_s": (round(stats["end_mad_s"], 1)
                                     if math.isfinite(stats["end_mad_s"]) else None),
            })
            frac_score = max(float(stats["start_fraction"]), float(stats["end_fraction"]))
            mad_candidates = [stats["start_mad_s"], stats["end_mad_s"]]
            finite_mads = [float(m) for m in mad_candidates if math.isfinite(float(m))]
            if finite_mads:
                mad_score = max(0.0, 1.0 - (min(finite_mads) / max(1e-6, float(ANCHOR_MAX_MAD_S) * 2.0)))
            else:
                mad_score = 0.0
            anchors_score = 0.75 * frac_score + 0.25 * mad_score

        keep = ((refine_mode == "legacy"  and legacy_ok) or
                (refine_mode == "anchors" and anchors_ok) or
                (refine_mode == "both"    and (legacy_ok or anchors_ok)))
        if keep:
            if refine_mode == "legacy":
                pair_score = legacy_score
            elif refine_mode == "anchors":
                pair_score = anchors_score
            else:
                pair_score = max(legacy_score, anchors_score)
            row["match_confidence"] = round(float(pair_score), 3)
            results_rows.append(row)

        now = time.time()
        if (n % eval_tick == 0) or (n == eval_total) or ((now - last_eval_status_t) >= 10.0):
            pct = int(round(100 * n / eval_total)) if eval_total else 100
            elapsed = now - eval_t0
            avg = elapsed / n if n else 0.0
            eta = max(0.0, (eval_total - n) * avg)
            accept_rate = (100.0 * len(results_rows) / n) if n else 0.0
            print(
                f"[progress] evaluate {n}/{eval_total} ({pct}%) - "
                f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}"
                f" - accepted {len(results_rows)} ({accept_rate:.1f}%)",
                flush=True,
            )
            last_eval_status_t = now

    #  6) Export
    if refine_mode == "legacy":
        columns = [
            "file_a", "file_b",
            "duration_a (s)", "duration_b (s)",
            "size_a (MB)", "size_b (MB)",
            "match_confidence",
            "avg_frame_diff (064)",
            "legacy_best_aligned_diff (064)",
            "legacy_time_shift_s",
        ]
    elif refine_mode == "anchors":
        columns = [
            "file_a", "file_b",
            "duration_a (s)", "duration_b (s)",
            "size_a (MB)", "size_b (MB)",
            "match_confidence",
            "relation", "relation_dir", "full_clip", "subset_clip",
            "subset_start_in_full_s", "subset_start_mm:ss",
            "start_match_fraction", "start_offset_s", "start_offset_mad_s",
            "end_match_fraction",   "end_offset_s",   "end_offset_mad_s",
        ]
    else:  # both
        columns = [
            "file_a", "file_b",
            "duration_a (s)", "duration_b (s)",
            "size_a (MB)", "size_b (MB)",
            "match_confidence",
            "avg_frame_diff (064)",
            # legacy
            "legacy_best_aligned_diff (064)",
            "legacy_time_shift_s",
            # anchors
            "relation", "relation_dir", "full_clip", "subset_clip",
            "subset_start_in_full_s", "subset_start_mm:ss",
            "start_match_fraction", "start_offset_s", "start_offset_mad_s",
            "end_match_fraction",   "end_offset_s",   "end_offset_mad_s",
        ]

    filtered = [{k: r.get(k, None) for k in columns} for r in results_rows]
    df = pd.DataFrame(filtered, columns=columns)

    if export_report:
        export_excel(df, report_path)
        print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
              f"{len(results_rows)} pairs saved", flush=True)
        if open_report:
            open_excel_file(report_path)
    else:
        print(f"[find_video_duplicates] Done in {_format_duration_auto(time.time()-t0)} - "
              f"{len(results_rows)} pairs found", flush=True)
    return df


def list_videos_missing_duplicates(
    folder_a: str,
    folder_b: str,
    report_path: str | None = None,
    refine_mode: str = "anchors",
    faiss_threshold: float = FAISS_THRESHOLD,
    align_threshold: float = ALIGN_THRESHOLD,
    align_offset_limit_s: float = ALIGN_OFFSET_LIMIT_S,
) -> pd.DataFrame:
    """
    Compare two folders and list videos in folder_a that have no match in folder_b.
    Writes a single-column Excel report to reports/ and returns the DataFrame.
    """
    from funcs import get_list_of_files  # local import avoids circularity

    if report_path is None:
        report_path = os.path.join(REPORTS_DIR, "missing_from_folder_b.xlsx")

    all_a = get_list_of_files(folder_a)
    videos_a = sorted(
        f for f in all_a
        if Path(f).suffix.lower() in EXTS
        and '_gsdata_' not in f
        and not Path(f).name.startswith("._")
    )

    matches_df = find_video_duplicates(
        directories=[folder_a, folder_b],
        faiss_threshold=faiss_threshold,
        align_threshold=align_threshold,
        align_offset_limit_s=align_offset_limit_s,
        self_compare=False,
        report_path=os.path.join(REPORTS_DIR, "temp_duplicates_report.xlsx"),
        refine_mode=refine_mode,
        export_report=False,
        open_report=False,
    )

    matched_a = set(matches_df["file_a"]) if not matches_df.empty else set()
    missing = [path for path in videos_a if path not in matched_a]

    out_df = pd.DataFrame({"missing_in_folder_b": missing})
    export_excel(out_df, report_path)
    return out_df


def scan_video_timestamps(
    directories: list[str],
    *,
    report_path: str = os.path.join(REPORTS_DIR, "dedupe_timestamps.xlsx"),
    export_report: bool = True,
    open_report: bool = False,
    step_s: float = TIMESTAMP_SCAN_STEP_S,
    frame_w: int = TIMESTAMP_SCAN_FRAME_W,
    frame_h: int = TIMESTAMP_SCAN_FRAME_H,
    timestamp_workers: int = TIMESTAMP_SCAN_WORKERS,
    force: bool = False,
) -> pd.DataFrame:
    """
    Optional OSD date scan:
      - Samples frames across each trimmed video.
      - OCRs camcorder timestamp overlay (bottom-right).
      - Produces per-file chronological month-year range for naming workflow.
    """
    t0 = time.time()
    print(f"[scan_timestamps] Start {time.strftime('%H:%M:%S')}", flush=True)
    if _osd_ocr_text is None or _osd_roi_tophat is None:
        print("[scan_timestamps] OCR helpers unavailable; returning empty report.", flush=True)
        return pd.DataFrame()

    from funcs import get_list_of_files  # local import avoids circularity

    all_tasks: list[tuple[str, int]] = []
    for fid, folder in enumerate(directories):
        all_files = get_list_of_files(folder)
        print(f"[scan_timestamps] Folder {fid}: {folder} -> {len(all_files)} files", flush=True)
        video_paths = [f for f in all_files
                       if Path(f).suffix.lower() in EXTS
                       and '_gsdata_' not in f
                       and not Path(f).name.startswith("._")]
        print(f"[scan_timestamps]   -> {len(video_paths)} video files", flush=True)
        all_tasks.extend((p, fid) for p in video_paths)

    total = len(all_tasks)
    if total == 0:
        print("[scan_timestamps] No videos found.", flush=True)
        return pd.DataFrame()

    if timestamp_workers <= 0:
        auto_workers = max(1, (os.cpu_count() or 1) - 1)
        timestamp_workers = min(auto_workers, max(1, int(TIMESTAMP_SCAN_WORKERS_AUTO_MAX)))
    print(f"[timestamps] workers: {int(timestamp_workers)}", flush=True)
    print(
        f"[timestamps] sampling: coarse_step={float(step_s):.1f}s, full-span=on, "
        f"burst={'on' if TIMESTAMP_BURST_ENABLE else 'off'} "
        f"(window={float(TIMESTAMP_BURST_WINDOW_S):.1f}s, step={float(TIMESTAMP_BURST_STEP_S):.1f}s, "
        f"max_windows={int(TIMESTAMP_BURST_MAX_WINDOWS)})",
        flush=True,
    )
    max_inflight_batches = int(TIMESTAMP_OCR_MAX_INFLIGHT_BATCHES)
    if max_inflight_batches <= 0:
        max_inflight_batches = max(2, int(timestamp_workers) * 2)
    if int(TIMESTAMP_OCR_BATCH_FRAMES) <= 0:
        # Keep total queued OCR frames in a moderate range so we amortize task
        # overhead without letting memory/latency balloon.
        target_inflight_frames = max(128, min(256, int(timestamp_workers) * 16))
        batch_frames = max(8, min(24, int(round(target_inflight_frames / max(1, max_inflight_batches)))))
        batch_mode = "auto"
    else:
        batch_frames = max(1, int(TIMESTAMP_OCR_BATCH_FRAMES))
        batch_mode = "manual"
    print(
        f"[timestamps] pipeline: io_worker=1 (sequential), cpu_workers={int(timestamp_workers)}, "
        f"ocr_batch_frames={batch_frames} ({batch_mode}), max_inflight_batches={max_inflight_batches}",
        flush=True,
    )

    vhs = VideoHashStore()
    print(f"[stage] inspecting timestamp cache for {total} file(s)", flush=True)
    planned_hits = 0
    planned_miss_reasons: dict[str, int] = {}
    cache_hit_tasks: list[tuple[str, int]] = []
    compute_tasks: list[tuple[str, int]] = []
    for path, fid in all_tasks:
        plan = vhs.inspect_timestamp_cache(
            path,
            step_s=step_s,
            frame_w=frame_w,
            frame_h=frame_h,
            force=force,
        )
        if plan.get("cache_hit"):
            planned_hits += 1
            cache_hit_tasks.append((path, fid))
        else:
            compute_tasks.append((path, fid))
            reason = str(plan.get("cache_miss_reason") or "unknown")
            planned_miss_reasons[reason] = planned_miss_reasons.get(reason, 0) + 1
    planned_compute = len(compute_tasks)
    print(f"[timestamp-cache] plan: cached={planned_hits}, to_compute={planned_compute}", flush=True)
    if planned_miss_reasons:
        reasons = ", ".join(f"{k}={planned_miss_reasons[k]}" for k in sorted(planned_miss_reasons))
        print(f"[timestamp-cache] plan reasons: {reasons}", flush=True)

    rows: list[dict] = []
    done = 0
    cache_hits = 0
    scan_t0 = time.time()
    tick = max(1, total // 10)
    last_status_t = scan_t0
    next_cache_save_t = scan_t0 + float(CACHE_AUTOSAVE_SECONDS)

    # 1) Load cached entries quickly (single-threaded, safe cache access).
    for path, fid in cache_hit_tasks:
        result = vhs.get_timestamp_scan(
            path,
            step_s=step_s,
            frame_w=frame_w,
            frame_h=frame_h,
            force=False,
        )
        data = result.get("data", {}) if isinstance(result.get("data", {}), dict) else {}
        rows.append(_timestamp_scan_row_from_data(path, fid, data))
        cache_hits += 1
        done += 1
        now = time.time()
        if (done % tick == 0) or (done == total) or ((now - last_status_t) >= 10.0):
            elapsed = now - scan_t0
            avg = elapsed / done if done else 0.0
            eta = max(0.0, (total - done) * avg)
            pct = int(round(100 * done / total))
            print(
                f"[progress] timestamp scan {done}/{total} ({pct}%) - "
                f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                flush=True,
            )
            last_status_t = now

    # 2) Compute cache misses via two-stage pipeline:
    #    stage A: one sequential I/O producer (video frame reads)
    #    stage B: parallel OCR consumers (CPU-heavy parsing)
    if compute_tasks:
        states: dict[str, dict] = {}
        compute_order: list[tuple[str, int]] = []
        prep_total = int(len(compute_tasks))
        prep_t0 = time.time()
        prep_last_status_t = prep_t0
        prep_tick = max(1, prep_total // 10)
        heartbeat_s = max(1.0, float(TIMESTAMP_PROGRESS_HEARTBEAT_S))
        file_tick_stage = max(1, int(TIMESTAMP_PROGRESS_FILE_TICK))
        prep_done = 0
        ready_paths: list[str] = []
        coarse_files_total = 0
        burst_files_budget_total = 0
        estimated_frame_total = 0
        burst_files_with_windows = 0

        def _record_prepared_state(idx: int, path: str, fid: int, st: dict) -> None:
            nonlocal prep_done, prep_last_status_t, coarse_files_total, burst_files_budget_total, estimated_frame_total
            st["fid"] = int(fid)
            states[path] = st
            compute_order.append((path, int(fid)))
            prep_done += 1
            if bool(st.get("ready_for_io", False)):
                ready_paths.append(path)
                coarse_files_total += 1
                estimated_frame_total += int(st.get("estimated_total_frames", 0) or 0)
                if bool(st["data"].get("burst_enabled", False)):
                    burst_files_budget_total += 1
            now = time.time()
            if (idx % prep_tick == 0) or (idx == prep_total) or ((now - prep_last_status_t) >= heartbeat_s):
                elapsed = now - prep_t0
                avg = elapsed / prep_done if prep_done else 0.0
                eta = max(0.0, (prep_total - prep_done) * avg)
                pct = int(round(100 * prep_done / prep_total)) if prep_total else 100
                print(
                    f"[progress] timestamp state prep {prep_done}/{prep_total} ({pct}%) - "
                    f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                    flush=True,
                )
                prep_last_status_t = now

        stage_stats = {
            "coarse": {
                "files_read": 0,
                "batches_submitted": 0,
                "batches_done": 0,
                "frames_enqueued": 0,
                "io_time_s": 0.0,
                "backpressure_wait_s": 0.0,
                "ocr_time_s": 0.0,
                "ocr_frames": 0,
                "ocr_text_hits": 0,
                "ocr_candidate_hits": 0,
            },
            "burst": {
                "files_read": 0,
                "batches_submitted": 0,
                "batches_done": 0,
                "frames_enqueued": 0,
                "io_time_s": 0.0,
                "backpressure_wait_s": 0.0,
                "ocr_time_s": 0.0,
                "ocr_frames": 0,
                "ocr_text_hits": 0,
                "ocr_candidate_hits": 0,
            },
        }
        active_stage = "coarse"
        pending: dict[concurrent.futures.Future, tuple[str, str]] = {}
        last_stage_status_t = time.time()
        pending_highwater = 0
        idle_wait_s = 0.0

        def _print_overall_progress(now_t: float) -> None:
            nonlocal last_status_t
            elapsed = now_t - scan_t0
            avg = elapsed / done if done else 0.0
            eta = max(0.0, (total - done) * avg)
            pct = int(round(100 * done / total)) if total else 100
            print(
                f"[progress] timestamp scan {done}/{total} ({pct}%) - "
                f"{_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                flush=True,
            )
            last_status_t = now_t

        def _print_stage_heartbeat(now_t: float) -> None:
            nonlocal last_stage_status_t
            if (now_t - last_stage_status_t) < heartbeat_s:
                return
            c = stage_stats["coarse"]
            b = stage_stats["burst"]
            processed_frames = int(sum(len(states[p]["processed_keys"]) for p in ready_paths)) if ready_paths else 0
            prep_pct = (100.0 * prep_done / prep_total) if prep_total > 0 else 100.0
            coarse_ready = max(int(coarse_files_total), int(c["files_read"]))
            burst_den = burst_files_with_windows if burst_files_with_windows > 0 else burst_files_budget_total
            burst_pct = (100.0 * b["files_read"] / burst_den) if burst_den > 0 else 100.0
            print(
                f"[pipeline] stage={active_stage} "
                f"| prep={prep_done}/{prep_total} ({prep_pct:.1f}%) "
                f"| frames={processed_frames}"
                f"{f'/{estimated_frame_total}' if estimated_frame_total > 0 else ''} "
                f"| coarse={c['files_read']}/{coarse_ready} "
                f"| burst={b['files_read']}/{burst_den} ({burst_pct:.1f}%) "
                f"| pending={len(pending)} "
                f"| io={_format_duration_auto(c['io_time_s'] + b['io_time_s'])} "
                f"| ocr={_format_duration_auto(c['ocr_time_s'] + b['ocr_time_s'])}",
                flush=True,
            )
            last_stage_status_t = now_t

        def _merge_batch(path_key: str, stage_name: str, batch_rows: list[dict]) -> None:
            st = states[path_key]
            for row in batch_rows:
                key = int(row.get("key", -1))
                if key < 0 or key in st["processed_keys"]:
                    continue
                st["processed_keys"].add(key)
                if bool(row.get("text_hit", False)):
                    st["ocr_text_hits"] += 1
                if stage_name == "coarse" and (bool(row.get("text_hit", False)) or bool(row.get("burst_trigger", False))):
                    st["coarse_text_hit_times"].append(float(row.get("t_s", 0.0)))
                cand = row.get("candidate")
                if isinstance(cand, dict):
                    st["candidates"].append(cand)
                if stage_name == "burst":
                    st["burst_sampled_frames"] += 1

        def _build_timestamp_state_data(st: dict) -> dict:
            data = dict(st["data"])
            if st.get("ready_for_io", False):
                candidates = list(st["candidates"])
                sampled_frames = int(len(st["processed_keys"]))
                ocr_text_hits = int(st["ocr_text_hits"])
                data["sampled_frames"] = sampled_frames
                data["ocr_text_hits"] = ocr_text_hits
                data["osd_candidate_hits"] = int(len(candidates))
                data["candidates"] = candidates
                data["burst_windows_used"] = int(st["burst_windows_used"])
                data["burst_sampled_frames"] = int(st["burst_sampled_frames"])
                data.update(_summarize_osd_candidates(candidates))
                if sampled_frames <= 0:
                    data["status"] = "no_samples"
                elif len(candidates) > 0:
                    data["status"] = "ok"
                elif ocr_text_hits > 0:
                    data["status"] = "ocr_text_no_date"
                else:
                    data["status"] = "no_ocr_text"
            return data

        def _commit_timestamp_state(path_key: str, *, stage_name: str) -> None:
            nonlocal next_cache_save_t
            st = states[path_key]
            fid = int(st.get("fid", -1))
            if fid < 0:
                return
            if stage_name == "coarse":
                if bool(st.get("coarse_committed", False)):
                    return
                st["coarse_committed"] = True
            elif stage_name == "burst":
                if bool(st.get("burst_committed", False)):
                    return
                st["burst_committed"] = True
            else:
                return
            data = _build_timestamp_state_data(st)
            vhs.put_timestamp_scan(
                path_key,
                data=data,
                step_s=step_s,
                frame_w=frame_w,
                frame_h=frame_h,
            )
            vhs.save_if_dirty()
            next_cache_save_t = time.time() + float(CACHE_AUTOSAVE_SECONDS)

        def _maybe_commit_completed_stage(path_key: str, stage_name: str) -> None:
            st = states[path_key]
            if not st.get("ready_for_io", False):
                return
            if stage_name == "coarse":
                if bool(st.get("coarse_submit_finished", False)) and int(st.get("coarse_done_batches", 0)) >= int(st.get("coarse_submitted_batches", 0)):
                    _commit_timestamp_state(path_key, stage_name="coarse")
            elif stage_name == "burst":
                if bool(st.get("burst_submit_finished", False)) and int(st.get("burst_done_batches", 0)) >= int(st.get("burst_submitted_batches", 0)):
                    _commit_timestamp_state(path_key, stage_name="burst")

        def _drain_completed(*, wait_all: bool = False, timeout_s: float = 10.0) -> None:
            nonlocal next_cache_save_t, idle_wait_s
            if not pending:
                return

            if wait_all:
                completed = list(pending.keys())
            else:
                wait_t0 = time.time()
                done_set, _ = concurrent.futures.wait(
                    list(pending.keys()),
                    timeout=float(timeout_s),
                    return_when=concurrent.futures.FIRST_COMPLETED,
                )
                wait_elapsed = max(0.0, time.time() - wait_t0)
                completed = list(done_set)
                now_t = time.time()
                if not completed:
                    idle_wait_s += wait_elapsed
                    _print_stage_heartbeat(now_t)
                    if now_t >= next_cache_save_t:
                        vhs.save_if_dirty()
                        next_cache_save_t = now_t + float(CACHE_AUTOSAVE_SECONDS)
                    return

            for fut in completed:
                path_key, stage_name = pending.pop(fut)
                try:
                    _path_ret, _stage_ret, rows_batch, batch_stats = fut.result()
                except Exception as e:
                    print(f"[warn] timestamp OCR batch failed ({Path(path_key).name}): {e}", flush=True)
                    rows_batch = []
                    batch_stats = {"elapsed_s": 0.0, "frame_count": 0, "text_hits": 0, "candidate_hits": 0}
                _merge_batch(path_key, stage_name, rows_batch)
                st = states[path_key]
                st[f"{stage_name}_done_batches"] = int(st.get(f"{stage_name}_done_batches", 0)) + 1
                stage_stats[stage_name]["batches_done"] += 1
                stage_stats[stage_name]["ocr_time_s"] += float(batch_stats.get("elapsed_s", 0.0) or 0.0)
                stage_stats[stage_name]["ocr_frames"] += int(batch_stats.get("frame_count", 0) or 0)
                stage_stats[stage_name]["ocr_text_hits"] += int(batch_stats.get("text_hits", 0) or 0)
                stage_stats[stage_name]["ocr_candidate_hits"] += int(batch_stats.get("candidate_hits", 0) or 0)
                _maybe_commit_completed_stage(path_key, stage_name)

            now_t = time.time()
            _print_stage_heartbeat(now_t)
            if now_t >= next_cache_save_t:
                vhs.save_if_dirty()
                next_cache_save_t = now_t + float(CACHE_AUTOSAVE_SECONDS)

        def _submit_batch(exec_pool, path_key: str, stage_name: str, frames_batch: list[tuple[float, np.ndarray]]) -> None:
            nonlocal pending_highwater
            if not frames_batch:
                return
            st = states[path_key]
            fut = exec_pool.submit(_timestamp_ocr_batch_worker, (path_key, stage_name, frames_batch))
            pending[fut] = (path_key, stage_name)
            st[f"{stage_name}_submitted_batches"] = int(st.get(f"{stage_name}_submitted_batches", 0)) + 1
            stage_stats[stage_name]["batches_submitted"] += 1
            pending_highwater = max(int(pending_highwater), int(len(pending)))

        def _queue_stage_frames(
            exec_pool,
            path_key: str,
            stage_name: str,
            *,
            start_s_local: float,
            duration_s_local: float,
            fps_local: float,
            limit_new_frames: int | None,
        ) -> int:
            if duration_s_local <= 0.0:
                return 0
            st = states[path_key]
            queued_before = len(st["queued_keys"])
            batch: list[tuple[float, np.ndarray]] = []
            io_t0 = time.time()
            backpressure_wait_s_local = 0.0
            for t_s, frame_bgr in _stream_bgr_frames_ffmpeg(
                path_key,
                start_s_local,
                duration_s_local,
                fps=fps_local,
                w=frame_w,
                h=frame_h,
            ):
                key = int(round(float(t_s) * 10.0))
                if key in st["queued_keys"]:
                    continue
                st["queued_keys"].add(key)
                batch.append((float(t_s), frame_bgr))
                if len(batch) >= batch_frames:
                    _submit_batch(exec_pool, path_key, stage_name, batch)
                    batch = []
                while len(pending) >= max_inflight_batches:
                    bp_t0 = time.time()
                    _drain_completed(wait_all=False)
                    backpressure_wait_s_local += max(0.0, time.time() - bp_t0)
                queued_now = len(st["queued_keys"]) - queued_before
                if (limit_new_frames is not None) and queued_now >= int(limit_new_frames):
                    break
            if batch:
                _submit_batch(exec_pool, path_key, stage_name, batch)
            queued_added = max(0, len(st["queued_keys"]) - queued_before)
            io_elapsed = max(0.0, time.time() - io_t0)
            stage_stats[stage_name]["frames_enqueued"] += int(queued_added)
            stage_stats[stage_name]["backpressure_wait_s"] += float(backpressure_wait_s_local)
            stage_stats[stage_name]["io_time_s"] += max(0.0, float(io_elapsed) - float(backpressure_wait_s_local))
            return queued_added

        cpu_workers = max(1, int(timestamp_workers))
        interrupted = False
        with concurrent.futures.ThreadPoolExecutor(max_workers=cpu_workers) as exec_pool, concurrent.futures.ThreadPoolExecutor(max_workers=1) as prep_pool:
            prep_iter = iter(enumerate(compute_tasks, 1))
            prep_future: concurrent.futures.Future | None = None
            prep_meta: tuple[int, str, int] | None = None

            def _submit_next_prep() -> None:
                nonlocal prep_future, prep_meta
                try:
                    idx, (path, fid) = next(prep_iter)
                except StopIteration:
                    prep_future = None
                    prep_meta = None
                    return
                if idx == 1 or idx == prep_total or (idx % file_tick_stage) == 0:
                    print(f"[pipeline] state prep file {idx}/{prep_total}: {Path(path).name}", flush=True)
                prep_meta = (int(idx), str(path), int(fid))
                cached_trim, cached_duration_s, _trim_source = vhs.get_cached_trim_bounds(path)
                prep_future = prep_pool.submit(
                    _init_timestamp_scan_state,
                    path,
                    step_s=step_s,
                    cached_trim=cached_trim,
                    cached_duration_s=cached_duration_s,
                )

            active_stage = "coarse"
            print("[stage] timestamp pipeline: streaming state prep + coarse read -> OCR queue", flush=True)
            try:
                _submit_next_prep()
                while prep_future is not None:
                    if not prep_future.done():
                        if pending:
                            _drain_completed(wait_all=False)
                        else:
                            wait_t0 = time.time()
                            try:
                                prep_future.result(timeout=heartbeat_s)
                            except concurrent.futures.TimeoutError:
                                idle_wait_s += max(0.0, time.time() - wait_t0)
                                now_t = time.time()
                                _print_stage_heartbeat(now_t)
                                if now_t >= next_cache_save_t:
                                    vhs.save_if_dirty()
                                    next_cache_save_t = now_t + float(CACHE_AUTOSAVE_SECONDS)
                                continue

                    if prep_meta is None:
                        break
                    idx, path, fid = prep_meta
                    st = prep_future.result()
                    _record_prepared_state(idx, path, fid, st)
                    _submit_next_prep()
                    if not st.get("ready_for_io", False):
                        continue
                    stage_stats["coarse"]["files_read"] += 1
                    coarse_read_n = int(stage_stats["coarse"]["files_read"])
                    if coarse_read_n == 1 or (coarse_read_n % file_tick_stage) == 0 or prep_done == prep_total:
                        print(
                            f"[pipeline] coarse read file {coarse_read_n}/{max(coarse_files_total, coarse_read_n)}: {Path(path).name}",
                            flush=True,
                        )
                    _queue_stage_frames(
                        exec_pool,
                        path,
                        "coarse",
                        start_s_local=float(st["scan_start_s"]),
                        duration_s_local=float(st["scan_span_s"]),
                        fps_local=float(st["coarse_fps"]),
                        limit_new_frames=int(st["coarse_target_frames"]),
                    )
                    st["coarse_submit_finished"] = True
                    _maybe_commit_completed_stage(path, "coarse")
                    _drain_completed(wait_all=False)
                _drain_completed(wait_all=True)

                # Build burst windows from coarse text hits.
                for path, _fid in compute_order:
                    st = states[path]
                    if not st.get("ready_for_io", False):
                        continue
                    if not bool(st["data"].get("burst_enabled", False)):
                        continue
                    hits = sorted(set(round(float(t), 1) for t in st["coarse_text_hit_times"]))
                    if not hits:
                        continue
                    windows = _build_timestamp_burst_windows(
                        hits,
                        scan_start_s=float(st["scan_start_s"]),
                        scan_end_s=float(st["scan_end_s"]),
                        coarse_step_s=float(st["data"].get("scan_step_s", step_s) or step_s),
                    )
                    st["burst_windows"] = windows
                    if windows:
                        burst_files_with_windows += 1

                # Phase B: burst scan windows (single sequential reader, queued OCR).
                active_stage = "burst"
                print("[stage] timestamp pipeline: burst read -> OCR queue", flush=True)
                for path, _fid in compute_order:
                    st = states[path]
                    if not st.get("ready_for_io", False):
                        continue
                    windows = st.get("burst_windows", [])
                    if not windows:
                        continue
                    stage_stats["burst"]["files_read"] += 1
                    burst_den = burst_files_with_windows if burst_files_with_windows > 0 else burst_files_budget_total
                    burst_read_n = int(stage_stats["burst"]["files_read"])
                    if (
                        burst_read_n == 1
                        or burst_read_n == burst_den
                        or (burst_read_n % file_tick_stage) == 0
                    ):
                        print(
                            f"[pipeline] burst read file {burst_read_n}/{burst_den}: {Path(path).name}",
                            flush=True,
                        )
                    windows_since_drain = 0
                    for lo, hi in windows:
                        before = len(st["queued_keys"])
                        _queue_stage_frames(
                            exec_pool,
                            path,
                            "burst",
                            start_s_local=float(lo),
                            duration_s_local=max(0.0, float(hi) - float(lo)),
                            fps_local=float(st["burst_fps"]),
                            limit_new_frames=None,
                        )
                        added = max(0, len(st["queued_keys"]) - before)
                        if added > 0:
                            st["burst_windows_used"] += 1
                        windows_since_drain += 1
                        if windows_since_drain >= max(1, int(TIMESTAMP_BURST_DRAIN_EVERY_WINDOWS)):
                            _drain_completed(wait_all=False)
                            windows_since_drain = 0
                    st["burst_submit_finished"] = True
                    _maybe_commit_completed_stage(path, "burst")
                    _drain_completed(wait_all=False)
                _drain_completed(wait_all=True)
            except KeyboardInterrupt:
                interrupted = True
                print("[scan_timestamps] Interrupt received - saving completed timestamp progress.", flush=True)
                _drain_completed(wait_all=False, timeout_s=0.0)
                for path, _fid in compute_order:
                    _maybe_commit_completed_stage(path, "coarse")
                    _maybe_commit_completed_stage(path, "burst")

        # Finalize each rebuilt file and save to cache/report.
        for path, fid in compute_order:
            st = states[path]
            data = _build_timestamp_state_data(st)

            vhs.put_timestamp_scan(
                path,
                data=data,
                step_s=step_s,
                frame_w=frame_w,
                frame_h=frame_h,
            )
            rows.append(_timestamp_scan_row_from_data(path, int(fid), data))
            done += 1
            now = time.time()
            if (done % tick == 0) or (done == total) or ((now - last_status_t) >= 10.0):
                _print_overall_progress(now)
            if now >= next_cache_save_t:
                vhs.save_if_dirty()
                next_cache_save_t = now + float(CACHE_AUTOSAVE_SECONDS)

    vhs.save_if_dirty()
    rebuilt = int(len(compute_tasks))
    print(f"[timestamp-cache] actual: hits={cache_hits}, rebuilt={rebuilt}", flush=True)
    if planned_miss_reasons:
        reasons = ", ".join(f"{k}={planned_miss_reasons[k]}" for k in sorted(planned_miss_reasons))
        print(f"[timestamp-cache] rebuild reasons: {reasons}", flush=True)
    if compute_tasks:
        c = stage_stats["coarse"]
        b = stage_stats["burst"]
        total_io_s = float(c["io_time_s"] + b["io_time_s"])
        total_ocr_s = float(c["ocr_time_s"] + b["ocr_time_s"])
        total_bp_s = float(c["backpressure_wait_s"] + b["backpressure_wait_s"])
        total_ocr_frames = int(c["ocr_frames"] + b["ocr_frames"])
        avg_ocr_ms = (1000.0 * total_ocr_s / total_ocr_frames) if total_ocr_frames > 0 else 0.0
        print(
            "[timestamps] profile: "
            f"io={_format_duration_auto(total_io_s)} "
            f"(coarse={_format_duration_auto(c['io_time_s'])}, burst={_format_duration_auto(b['io_time_s'])}) | "
            f"ocr_batch_wall={_format_duration_auto(total_ocr_s)} "
            f"(frames={total_ocr_frames}, avg={avg_ocr_ms:.1f}ms/frame) | "
            f"backpressure_wait={_format_duration_auto(total_bp_s)} | "
            f"idle_wait={_format_duration_auto(idle_wait_s)} | "
            f"pending_highwater={int(pending_highwater)}",
            flush=True,
        )

    df = pd.DataFrame(rows).sort_values(by=["folder_id", "file_path"], ascending=[True, True]).reset_index(drop=True)
    if export_report:
        export_excel(df, report_path)
        print(
            f"[scan_timestamps] Done in {_format_duration_auto(time.time()-t0)} - "
            f"{len(df)} file(s) written to {report_path}",
            flush=True,
        )
        if open_report:
            open_excel_file(report_path)
    else:
        print(
            f"[scan_timestamps] Done in {_format_duration_auto(time.time()-t0)} - "
            f"{len(df)} file(s) scanned",
            flush=True,
        )
    return df


# 
# Excel export helpers
# 

def _format_excel_sheet(df: pd.DataFrame, ws) -> None:
    """Apply header styles, frozen panes, column widths, and number formats."""
    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Column widths + number formats
    duration_display_cols: set[str] = set()
    for col in df.columns:
        col_norm = str(col).strip().lower()
        if not col_norm.endswith("_s"):
            continue
        series = pd.to_numeric(df[col], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        if series.empty:
            duration_display_cols.add(col_norm)
            continue
        if float(series.min()) >= 0.0:
            duration_display_cols.add(col_norm)
    for idx, col in enumerate(df.columns, start=1):
        col_norm = str(col).strip().lower()
        max_len = df[col].astype(str).map(len).max() if not df.empty else 0
        width = max(len(col), max_len) + 2
        if col_norm in duration_display_cols:
            width = max(width, 14)
        ws.column_dimensions[get_column_letter(idx)].width = width

        if col_norm in duration_display_cols:
            fmt = "[h]:mm:ss"
        elif "file_a" in col.lower() or "file_b" in col.lower():
            fmt = numbers.FORMAT_GENERAL
        elif "%" in col:
            fmt = numbers.FORMAT_PERCENTAGE_00
        else:
            fmt = numbers.FORMAT_NUMBER_00
        if col_norm in duration_display_cols:
            for row_idx, raw_value in enumerate(df[col].tolist(), start=2):
                cell = ws.cell(row=row_idx, column=idx)
                try:
                    if raw_value in ("", None):
                        pass
                    elif isinstance(raw_value, datetime.timedelta):
                        cell.value = float(raw_value.total_seconds()) / 86400.0
                    elif isinstance(raw_value, pd.Timedelta):
                        cell.value = float(raw_value.total_seconds()) / 86400.0
                    else:
                        num = pd.to_numeric(raw_value, errors="coerce")
                        if not pd.isna(num):
                            cell.value = float(num) / 86400.0
                except Exception:
                    pass
                cell.number_format = fmt
        else:
            for cell in ws[get_column_letter(idx)][1:]:
                cell.number_format = fmt


def export_excel(df: pd.DataFrame, path: str):
    """
    Write a single-sheet Excel file with bold headers, frozen top row,
    sane column widths, and simple number formatting.
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    refresh_backup_before_retry = False
    while True:
        try:
            if refresh_backup_before_retry and os.path.exists(path):
                _backup_reports([path], label="retry")
                refresh_backup_before_retry = False
            with ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Duplicates")
                ws = writer.sheets["Duplicates"]
                _format_excel_sheet(df, ws)
            break
        except PermissionError:
            choice = sg.popup_yes_no(
                f"Can't write to file:\n{path}\n\nIt might be open in Excel.\n\nRetry?",
                title="Export Failed",
                keep_on_top=True
            )
            if choice != 'Yes':
                print(f"Export to {path} aborted by user.")
                break
            refresh_backup_before_retry = True
        except Exception as e:
            sg.popup_error(f"Unexpected error while exporting:\n{e}", title="Export Failed", keep_on_top=True)
            break


def _normalize_col_name(col: str) -> str:
    return col.strip().lower().replace(" ", "_")


def _infer_mode_from_path(path: str) -> str:
    stem = Path(path).stem.lower()
    if stem.startswith("dedupe_"):
        stem = stem[len("dedupe_"):]
    # Keep letters/numbers/underscore only for clean column names.
    out = []
    for ch in stem:
        out.append(ch if (ch.isalnum() or ch == "_") else "_")
    return "".join(out) or "report"


def _sanitize_mode_name(mode: str) -> str:
    mode = (mode or "").strip().lower()
    out = []
    for ch in mode:
        out.append(ch if (ch.isalnum() or ch == "_") else "_")
    return "".join(out) or "report"


def _in_folder(path: str, folder: str) -> bool:
    try:
        path_norm = os.path.normcase(os.path.normpath(path))
        folder_norm = os.path.normcase(os.path.normpath(folder))
        if path_norm == folder_norm:
            return True
        return path_norm.startswith(folder_norm + os.path.sep)
    except Exception:
        return False


def _canonicalize_pair(
    file_a: str,
    file_b: str,
    folder_a: str | None,
    folder_b: str | None,
) -> tuple[str, str]:
    """
    Canonicalize ordering of a pair. Prefer A/B folders when provided; fallback
    to stable lexicographic ordering (case-insensitive).
    """
    if folder_a and folder_b:
        if _in_folder(file_a, folder_a) and _in_folder(file_b, folder_b):
            return file_a, file_b
        if _in_folder(file_a, folder_b) and _in_folder(file_b, folder_a):
            return file_b, file_a
    elif folder_a:
        if _in_folder(file_a, folder_a) and not _in_folder(file_b, folder_a):
            return file_a, file_b
        if _in_folder(file_b, folder_a) and not _in_folder(file_a, folder_a):
            return file_b, file_a

    a_key = os.path.normcase(os.path.normpath(file_a))
    b_key = os.path.normcase(os.path.normpath(file_b))
    return (file_a, file_b) if a_key <= b_key else (file_b, file_a)


def _list_videos_in_folder(folder: str) -> list[str]:
    from funcs import get_list_of_files  # local import avoids circularity
    all_files = get_list_of_files(folder)
    out = []
    for f in all_files:
        if Path(f).suffix.lower() in EXTS and '_gsdata_' not in f and not Path(f).name.startswith("._"):
            out.append(f)
    return sorted(out)


def _norm_path(path: str) -> str:
    return os.path.normcase(os.path.normpath(path))


def _pair_key(file_a: str, file_b: str) -> tuple[str, str]:
    a = _norm_path(file_a)
    b = _norm_path(file_b)
    return (a, b) if a <= b else (b, a)


def _backup_reports(report_paths: list[str], *, label: str = "") -> dict[str, str]:
    """
    Move existing reports into reports/archive/backups with a stable naming scheme so reruns
    don't overwrite old outputs.
    Returns {original_path: backup_path}.
    """
    backups: dict[str, str] = {}
    stamp = time.strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(REPORTS_DIR, "archive", "backups")
    os.makedirs(backup_dir, exist_ok=True)
    label_clean = "".join(ch if (ch.isalnum() or ch in ("-", "_")) else "_" for ch in str(label or ""))
    label_clean = label_clean.strip("_")
    label_suffix = f"__{label_clean}" if label_clean else ""
    for path in report_paths:
        if not path:
            continue
        if not os.path.exists(path):
            continue
        base_name = os.path.basename(path)
        stem, ext = os.path.splitext(base_name)
        dst = os.path.join(backup_dir, f"{stem}__{stamp}{label_suffix}{ext}")
        n = 2
        while os.path.exists(dst):
            dst = os.path.join(backup_dir, f"{stem}__{stamp}{label_suffix}__{n}{ext}")
            n += 1
        try:
            os.replace(path, dst)
            backups[path] = dst
            print(f"[reports] backup: {path} -> {dst}", flush=True)
        except Exception as e:
            print(f"[reports] backup failed for {path}: {e}", flush=True)
    return backups


def _latest_archived_report_for(canonical_report_path: str) -> str | None:
    """
    Return the newest archived backup for a canonical report path, or None.
    Example canonical path: reports/dedupe_timeline.xlsx
    Archived layout: reports/archive/backups/dedupe_timeline__YYYYMMDD_HHMMSS*.xlsx
    """
    try:
        stem = Path(canonical_report_path).stem
    except Exception:
        return None
    backup_dir = os.path.join(REPORTS_DIR, "archive", "backups")
    pattern = os.path.join(backup_dir, f"{stem}__*.xlsx")
    candidates = [p for p in glob.glob(pattern) if os.path.isfile(p)]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def _resolve_report_path(canonical_report_path: str) -> tuple[str | None, bool]:
    """
    Resolve a report path with archive fallback.
    Returns (resolved_path_or_none, used_archive_fallback).
    """
    if os.path.exists(canonical_report_path):
        return canonical_report_path, False
    archived = _latest_archived_report_for(canonical_report_path)
    if archived and os.path.exists(archived):
        return archived, True
    return None, False


def _collect_video_tasks(directories: list[str]) -> list[tuple[str, int]]:
    from funcs import get_list_of_files  # local import avoids circularity
    tasks: list[tuple[str, int]] = []
    for fid, folder in enumerate(directories):
        all_files = get_list_of_files(folder)
        video_paths = [
            f for f in all_files
            if Path(f).suffix.lower() in EXTS
            and "_gsdata_" not in f
            and not Path(f).name.startswith("._")
        ]
        tasks.extend((p, fid) for p in video_paths)
    return tasks


def _build_silver_truth_from_consolidated(
    consolidated_path: str,
    *,
    expected_parts_per_b: int = 2,
    min_confidence: float = 70.0,
    min_mode_votes: int = 2,
) -> tuple[dict[str, set[str]], set[tuple[str, str]], set[tuple[str, str]]]:
    """
    Build a pragmatic "silver truth" set from a consolidated report.
    Returns:
      - positives_by_b: {norm(file_b): {norm(file_a), ...}}
      - positive_pairs: {(norm(file_a), norm(file_b)), ...}
      - hard_negative_pairs: competing low-confidence pairs for known file_b groups
    """
    positives_by_b: dict[str, set[str]] = {}
    positive_pairs: set[tuple[str, str]] = set()
    hard_negative_pairs: set[tuple[str, str]] = set()

    if not os.path.exists(consolidated_path):
        print(f"[tune] truth report not found: {consolidated_path}", flush=True)
        return positives_by_b, positive_pairs, hard_negative_pairs

    try:
        df = pd.read_excel(consolidated_path, sheet_name="Consolidated")
    except Exception as e:
        print(f"[tune] failed reading truth report: {e}", flush=True)
        return positives_by_b, positive_pairs, hard_negative_pairs

    if df.empty or ("file_a" not in df.columns) or ("file_b" not in df.columns):
        print("[tune] truth report has no usable pair rows.", flush=True)
        return positives_by_b, positive_pairs, hard_negative_pairs

    if "matched_by_count" not in df.columns:
        mode_cols = [c for c in df.columns if c.startswith("matched_by_") and c != "matched_by_count"]
        if mode_cols:
            df["matched_by_count"] = df[mode_cols].fillna(False).astype(bool).sum(axis=1).astype(int)
        else:
            df["matched_by_count"] = 1

    if "confidence_score" not in df.columns:
        df["confidence_score"] = np.where(df["matched_by_count"] >= 2, 80.0, 60.0)
    df["confidence_score"] = pd.to_numeric(df["confidence_score"], errors="coerce").fillna(0.0)
    df["matched_by_count"] = pd.to_numeric(df["matched_by_count"], errors="coerce").fillna(0).astype(int)

    for file_b, grp in df.groupby("file_b", dropna=True):
        if not isinstance(file_b, str):
            continue
        b_key = _norm_path(file_b)
        g = grp.copy().sort_values(
            by=["confidence_score", "matched_by_count"],
            ascending=[False, False],
        ).reset_index(drop=True)
        if g.empty:
            continue

        strong = g[
            (g["matched_by_count"] >= int(min_mode_votes))
            | (g["confidence_score"] >= float(min_confidence))
        ]
        if strong.empty:
            # Fallback: keep top-ranked pair so every represented file_b contributes.
            strong = g.head(1)

        selected_a: list[str] = []
        top_conf = float(strong.iloc[0]["confidence_score"])
        for _, row in strong.iterrows():
            if len(selected_a) >= max(1, int(expected_parts_per_b)):
                break
            a = row.get("file_a")
            if not isinstance(a, str):
                continue
            a_key = _norm_path(a)
            if a_key in selected_a:
                continue
            conf = float(row.get("confidence_score", 0.0))
            mode_n = int(row.get("matched_by_count", 0))
            # Keep second split part if confidence is not wildly behind, or if multi-mode.
            if selected_a and conf < (top_conf - 18.0) and mode_n < min_mode_votes:
                continue
            selected_a.append(a_key)

        if not selected_a:
            a = g.iloc[0].get("file_a")
            if isinstance(a, str):
                selected_a = [_norm_path(a)]
        if not selected_a:
            continue

        positives_by_b[b_key] = set(selected_a)
        for a_key in selected_a:
            positive_pairs.add((a_key, b_key))

        # Hard negatives: competing rows in same file_b group, clearly weaker.
        for _, row in g.iterrows():
            a = row.get("file_a")
            if not isinstance(a, str):
                continue
            a_key = _norm_path(a)
            if a_key in positives_by_b[b_key]:
                continue
            conf = float(row.get("confidence_score", 0.0))
            mode_n = int(row.get("matched_by_count", 0))
            if mode_n <= 1 and conf <= (top_conf - 10.0):
                hard_negative_pairs.add((a_key, b_key))

    print(
        "[tune] silver truth: "
        f"{len(positive_pairs)} positive pair(s), "
        f"{len(hard_negative_pairs)} hard negative pair(s), "
        f"{len(positives_by_b)} file_b group(s)",
        flush=True,
    )
    return positives_by_b, positive_pairs, hard_negative_pairs


def _extract_timeline_overlap_truth(
    consolidated_path: str,
    *,
    min_timeline_overlap_s: float = 90.0,
) -> dict[tuple[str, str], dict]:
    """
    Build pair-level timeline truth targets from consolidated report.
    Returns {(norm_a, norm_b): {"overlap_s": float, "offset_s": float|None}}.
    """
    targets: dict[tuple[str, str], dict] = {}
    if not os.path.exists(consolidated_path):
        return targets
    try:
        df = pd.read_excel(consolidated_path, sheet_name="Consolidated")
    except Exception:
        return targets
    if df.empty:
        return targets
    if "file_a" not in df.columns or "file_b" not in df.columns:
        return targets

    def _pick_num(*cols: str) -> pd.Series:
        for c in cols:
            if c in df.columns:
                s = pd.to_numeric(df[c], errors="coerce")
                if not s.isna().all():
                    return s
        return pd.Series(np.nan, index=df.index, dtype=float)

    m_tl = df.get("matched_by_timeline", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    if "matched_by_count" in df.columns:
        mode_votes = pd.to_numeric(df["matched_by_count"], errors="coerce").fillna(0).astype(int)
    else:
        mode_cols = [c for c in df.columns if c.startswith("matched_by_") and c != "matched_by_count"]
        mode_votes = (
            df[mode_cols].fillna(False).astype(bool).sum(axis=1).astype(int)
            if mode_cols
            else pd.Series(1, index=df.index, dtype=int)
        )
    if "confidence_score" in df.columns:
        conf = pd.to_numeric(df["confidence_score"], errors="coerce").fillna(0.0)
    else:
        conf = pd.Series(np.where(mode_votes >= 2, 80.0, 60.0), index=df.index, dtype=float)

    trusted = (mode_votes >= 2) | (conf >= 70.0)
    overlap = _pick_num("timeline__timeline_overlap_s", "timeline_overlap_s")
    offset = _pick_num("timeline__timeline_offset_s", "timeline_offset_s")
    keep = m_tl & trusted & overlap.notna() & (overlap >= float(min_timeline_overlap_s))

    for idx, row in df[keep].iterrows():
        a = row.get("file_a")
        b = row.get("file_b")
        if not isinstance(a, str) or not isinstance(b, str):
            continue
        key = (_norm_path(a), _norm_path(b))
        ov = float(overlap.loc[idx])
        off_raw = offset.loc[idx] if idx in offset.index else np.nan
        off_val = None if pd.isna(off_raw) else float(off_raw)
        prev = targets.get(key)
        # Keep the strongest timeline overlap when duplicates exist.
        if (prev is None) or (ov > float(prev.get("overlap_s", 0.0))):
            targets[key] = {"overlap_s": ov, "offset_s": off_val}

    print(f"[tune] timeline overlap truth pairs: {len(targets)}", flush=True)
    return targets


def _load_cached_entries_for_tuning(
    directories: list[str],
    *,
    mode: str,
    audio_rms_thresh_db: float | None = AUDIO_RMS_THRESH_DB,
) -> list[dict]:
    """
    Load per-file sequences from unified cache for tuning replay.
    No report files are touched; cache misses may compute and store entries.
    """
    tasks = _collect_video_tasks(directories)
    vhs = VideoHashStore()
    entries: list[dict] = []
    kept = 0
    empty = 0
    for path, fid in tasks:
        try:
            duration = float(ffprobe_duration_seconds(path))
        except Exception:
            duration = 0.0

        if mode == "audio":
            cache = vhs.get_audio_fingerprint(path, rms_thresh_db=audio_rms_thresh_db)
            seq_raw = cache.get("data", [])
            seq = _filter_hash_seq(
                seq_raw,
                min_bitcount=AUDIO_MIN_BITCOUNT,
                max_bitcount=AUDIO_MAX_BITCOUNT,
                collapse_runs=True,
            )
            unique_ratio = (
                float(len({h for _t, h in seq}) / max(1, len(seq)))
                if seq else 0.0
            )
            if seq and (len(seq) < AUDIO_MIN_FILTERED_FRAMES or unique_ratio < AUDIO_MIN_UNIQUE_RATIO):
                seq = []
            hash_set = {h for _t, h in seq}
        elif mode == "timeline":
            timeline_bytes, _trim = vhs.get_timeline(path, step_s=TIMELINE_STEP_S)
            seq_raw = [(t, _bytes_to_int64(h)) for t, h in timeline_bytes]
            seq = _filter_hash_seq(
                seq_raw,
                min_bitcount=TIMELINE_MIN_BITCOUNT,
                max_bitcount=TIMELINE_MAX_BITCOUNT,
                collapse_runs=False,
            )
            seq_compact = _filter_hash_seq(
                seq_raw,
                min_bitcount=TIMELINE_MIN_BITCOUNT,
                max_bitcount=TIMELINE_MAX_BITCOUNT,
                collapse_runs=True,
            )
            unique_ratio = (
                float(len({h for _t, h in seq_compact}) / max(1, len(seq_compact)))
                if seq_compact else 0.0
            )
            if seq and (len(seq_compact) < TIMELINE_MIN_FILTERED_FRAMES or unique_ratio < TIMELINE_MIN_UNIQUE_RATIO):
                seq = []
            hash_set = {h for _t, h in seq}
        else:
            raise ValueError(f"Unsupported tuning mode: {mode}")

        if seq:
            kept += 1
        else:
            empty += 1
        entries.append({
            "path": path,
            "fid": int(fid),
            "duration": float(duration),
            "seq": seq,
            "hash_set": hash_set,
        })

    vhs.save_if_dirty()
    print(
        f"[tune] loaded {len(entries)} {mode} entries "
        f"(usable={kept}, empty_or_low_info={empty})",
        flush=True,
    )
    return entries


def _build_tuning_dataset(
    entries: list[dict],
    positives_by_b: dict[str, set[str]],
    hard_negative_pairs: set[tuple[str, str]],
    *,
    max_negatives_per_b: int = 8,
) -> list[dict]:
    """
    Build labeled tuning pairs:
      label=1 for silver positives,
      label=0 for hard negatives and near-duration competitors.
    """
    if not entries:
        return []

    fids = sorted({int(e["fid"]) for e in entries})
    if len(fids) < 2:
        return []
    fid_a, fid_b = fids[0], fids[1]

    idx_a = [i for i, e in enumerate(entries) if int(e["fid"]) == fid_a and e["seq"]]
    idx_b = [j for j, e in enumerate(entries) if int(e["fid"]) == fid_b and e["seq"]]

    positives: list[dict] = []
    forced_negatives: list[dict] = []
    negatives_by_b: dict[str, list[dict]] = {}

    for j in idx_b:
        b_key = _norm_path(entries[j]["path"])
        pos_as = positives_by_b.get(b_key)
        if not pos_as:
            continue
        dur_b = float(entries[j]["duration"])
        for i in idx_a:
            dur_a = float(entries[i]["duration"])
            dur_max = max(dur_a, dur_b)
            if dur_max <= 0:
                ratio = 0.0
            else:
                ratio = min(dur_a, dur_b) / dur_max
            a_key = _norm_path(entries[i]["path"])
            rec = {
                "i": int(i),
                "j": int(j),
                "ratio": float(ratio),
                "key": (a_key, b_key),
                "dur_delta": abs(dur_a - dur_b),
            }
            if a_key in pos_as:
                rec["label"] = 1
                positives.append(rec)
            else:
                rec["label"] = 0
                if (a_key, b_key) in hard_negative_pairs:
                    rec["priority"] = 2
                    forced_negatives.append(rec)
                else:
                    negatives_by_b.setdefault(b_key, []).append(rec)

    selected_negatives: list[dict] = []
    seen_keys: set[tuple[str, str]] = set()
    for rec in forced_negatives:
        if rec["key"] in seen_keys:
            continue
        seen_keys.add(rec["key"])
        selected_negatives.append(rec)

    per_b = max(1, int(max_negatives_per_b))
    for b_key, rows in negatives_by_b.items():
        rows_sorted = sorted(rows, key=lambda r: (-float(r["ratio"]), float(r["dur_delta"])))
        added = 0
        for rec in rows_sorted:
            if rec["key"] in seen_keys:
                continue
            seen_keys.add(rec["key"])
            selected_negatives.append(rec)
            added += 1
            if added >= per_b:
                break

    dataset = positives + selected_negatives
    n_pos = sum(1 for r in dataset if int(r["label"]) == 1)
    n_neg = len(dataset) - n_pos
    print(f"[tune] dataset: positives={n_pos}, negatives={n_neg}, total={len(dataset)}", flush=True)
    return dataset


def _evaluate_tuning_combo(
    *,
    entries: list[dict],
    dataset: list[dict],
    matcher_params: dict,
    duration_ratio_min: float,
    min_hashset_intersect_ratio: float = 0.0,
    overlap_targets: dict[tuple[str, str], dict] | None = None,
    overlap_target_weight: float = 0.0,
) -> dict:
    tp = fp = tn = fn = 0
    accepted = 0
    target_overlap_score_sum = 0.0
    target_overlap_score_n = 0
    for rec in dataset:
        label = int(rec["label"])
        ratio = float(rec["ratio"])
        i = int(rec["i"])
        j = int(rec["j"])
        if ratio < float(duration_ratio_min):
            ok = False
            overlap_s = 0.0
        elif min_hashset_intersect_ratio > 0.0:
            hs_i = entries[i].get("hash_set") or set()
            hs_j = entries[j].get("hash_set") or set()
            if hs_i and hs_j:
                small, big = (hs_i, hs_j) if len(hs_i) <= len(hs_j) else (hs_j, hs_i)
                common = 0
                for h in small:
                    if h in big:
                        common += 1
                shared_ratio = float(common) / float(max(1, len(small)))
                if shared_ratio < float(min_hashset_intersect_ratio):
                    ok = False
                    overlap_s = 0.0
                else:
                    seq_a = entries[i]["seq"]
                    seq_b = entries[j]["seq"]
                    ok, _off, _votes, overlap_s = _audio_match_offset(seq_a, seq_b, **matcher_params)
            else:
                ok = False
                overlap_s = 0.0
        else:
            seq_a = entries[i]["seq"]
            seq_b = entries[j]["seq"]
            ok, _off, _votes, overlap_s = _audio_match_offset(seq_a, seq_b, **matcher_params)

        if ok:
            accepted += 1
            if label == 1:
                tp += 1
                if overlap_targets:
                    target = overlap_targets.get(rec["key"])
                    target_ov = None if target is None else target.get("overlap_s")
                    if target_ov is not None:
                        t = float(target_ov)
                        if t > 0.0:
                            # 1.0 is perfect overlap agreement with timeline; decays smoothly with log-ratio error.
                            ratio_ov = max(1e-6, float(overlap_s) / t)
                            err = abs(math.log(ratio_ov))
                            agreement = math.exp(-err)
                            target_overlap_score_sum += float(agreement)
                            target_overlap_score_n += 1
            else:
                fp += 1
        else:
            if label == 1:
                fn += 1
            else:
                tn += 1

    recall = tp / max(1, tp + fn)
    precision = tp / max(1, tp + fp)
    fpr = fp / max(1, fp + tn)
    overlap_agreement = target_overlap_score_sum / max(1, target_overlap_score_n)
    score = (0.68 * recall) + (0.24 * precision) - (0.14 * fpr) + (float(overlap_target_weight) * overlap_agreement)
    return {
        "tp": tp,
        "fp": fp,
        "tn": tn,
        "fn": fn,
        "accepted": accepted,
        "recall": recall,
        "precision": precision,
        "false_positive_rate": fpr,
        "overlap_agreement": overlap_agreement,
        "score": score,
    }


def _choose_best_tuning_combo(rows: list[dict]) -> dict:
    if not rows:
        return {}
    baseline = next((r for r in rows if str(r.get("combo_id")) == "baseline"), rows[0])
    baseline_prec = float(baseline.get("precision", 0.0))
    min_prec = max(0.55, baseline_prec - 0.05)
    eligible = [r for r in rows if float(r.get("precision", 0.0)) >= min_prec]
    if not eligible:
        eligible = rows
    eligible_sorted = sorted(
        eligible,
        key=lambda r: (
            float(r.get("recall", 0.0)),
            float(r.get("score", 0.0)),
            float(r.get("precision", 0.0)),
        ),
        reverse=True,
    )
    return eligible_sorted[0]


def _tune_audio_from_cache(
    directories: list[str],
    positives_by_b: dict[str, set[str]],
    hard_negative_pairs: set[tuple[str, str]],
    *,
    timeline_overlap_targets: dict[tuple[str, str], dict] | None = None,
    audio_rms_thresh_db: float | None = AUDIO_RMS_THRESH_DB,
) -> tuple[dict, pd.DataFrame]:
    entries = _load_cached_entries_for_tuning(
        directories,
        mode="audio",
        audio_rms_thresh_db=audio_rms_thresh_db,
    )
    dataset = _build_tuning_dataset(
        entries,
        positives_by_b,
        hard_negative_pairs,
        max_negatives_per_b=2,
    )
    if dataset:
        # Keep replay tuning practical: deterministic cap for long projects.
        pos_rows = [r for r in dataset if int(r.get("label", 0)) == 1]
        neg_rows = [r for r in dataset if int(r.get("label", 0)) == 0]
        if len(pos_rows) > 70:
            pos_rows = sorted(pos_rows, key=lambda r: (r["key"][1], -float(r.get("ratio", 0.0))))[:70]
        if len(neg_rows) > 120:
            neg_rows = sorted(
                neg_rows,
                key=lambda r: (-float(r.get("ratio", 0.0)), float(r.get("dur_delta", 0.0)), r["key"][1]),
            )[:120]
        dataset = pos_rows + neg_rows
        print(
            f"[tune][audio] replay subset: positives={len(pos_rows)}, negatives={len(neg_rows)}, total={len(dataset)}",
            flush=True,
        )
    if not dataset:
        print("[tune][audio] empty tuning dataset; using defaults.", flush=True)
        return {}, pd.DataFrame()

    total_bits = 2 * AUDIO_N_BANDS - 1
    base_params = {
        "total_bits": total_bits,
        "hamming_thresh": AUDIO_HAMMING_THRESH,
        "bin_s": AUDIO_BIN_S,
        "min_votes": AUDIO_MIN_VOTES,
        "min_overlap_s": AUDIO_MIN_OVERLAP_S,
        "hop_s": AUDIO_HOP_S,
        "lsh_chunks": AUDIO_LSH_CHUNKS,
        "brute_limit": AUDIO_BRUTE_MAX,
        "strict_hamming_thresh": AUDIO_STRICT_HAMMING_THRESH,
        "strict_min_overlap_s": AUDIO_STRICT_MIN_OVERLAP_S,
        "min_vote_fraction": AUDIO_MIN_VOTE_FRACTION,
        "peak_ratio_min": AUDIO_PEAK_RATIO_MIN,
        "peak_margin": AUDIO_PEAK_MARGIN,
        "speed_ratio_min": AUDIO_SPEED_RATIO_MIN,
        "speed_ratio_max": AUDIO_SPEED_RATIO_MAX,
        "speed_steps": min(9, int(AUDIO_SPEED_STEPS)),
        "max_candidates_per_frame": AUDIO_MAX_CANDIDATES_PER_FRAME,
        "long_overlap_override_s": AUDIO_LONG_OVERLAP_OVERRIDE_S,
        "long_overlap_vote_mult": AUDIO_LONG_OVERLAP_VOTE_MULT,
        "run_gap_mult": AUDIO_RUN_GAP_MULT,
        "mutual_overlap_ratio_min": AUDIO_MUTUAL_OVERLAP_RATIO_MIN,
    }

    combos: list[dict] = [{
        "combo_id": "baseline",
        "audio_duration_ratio_min": AUDIO_DURATION_RATIO_MIN,
        "audio_min_hashset_intersect_ratio": AUDIO_MIN_HASHSET_INTERSECT_RATIO,
        "audio_strict_min_overlap_s": AUDIO_STRICT_MIN_OVERLAP_S,
        "audio_peak_ratio_min": AUDIO_PEAK_RATIO_MIN,
        "audio_peak_margin": AUDIO_PEAK_MARGIN,
        "audio_strict_hamming_thresh": AUDIO_STRICT_HAMMING_THRESH,
        "audio_min_vote_fraction": AUDIO_MIN_VOTE_FRACTION,
        "audio_mutual_overlap_ratio_min": AUDIO_MUTUAL_OVERLAP_RATIO_MIN,
    }]
    combos.extend([
        {
            "combo_id": "tl_guided_balanced",
            "audio_duration_ratio_min": 0.10,
            "audio_min_hashset_intersect_ratio": 0.10,
            "audio_strict_min_overlap_s": 55.0,
            "audio_peak_ratio_min": 1.08,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.020,
            "audio_mutual_overlap_ratio_min": 0.60,
        },
        {
            "combo_id": "tl_guided_strict",
            "audio_duration_ratio_min": 0.11,
            "audio_min_hashset_intersect_ratio": 0.12,
            "audio_strict_min_overlap_s": 70.0,
            "audio_peak_ratio_min": 1.10,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.022,
            "audio_mutual_overlap_ratio_min": 0.65,
        },
        {
            "combo_id": "tl_guided_recall",
            "audio_duration_ratio_min": 0.08,
            "audio_min_hashset_intersect_ratio": 0.08,
            "audio_strict_min_overlap_s": 45.0,
            "audio_peak_ratio_min": 1.06,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.018,
            "audio_mutual_overlap_ratio_min": 0.55,
        },
        {
            "combo_id": "tl_guided_recall_plus",
            "audio_duration_ratio_min": 0.12,
            "audio_min_hashset_intersect_ratio": 0.09,
            "audio_strict_min_overlap_s": 50.0,
            "audio_peak_ratio_min": 1.06,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.019,
            "audio_mutual_overlap_ratio_min": 0.58,
        },
        {
            "combo_id": "tl_overlap_chase",
            "audio_duration_ratio_min": 0.10,
            "audio_min_hashset_intersect_ratio": 0.10,
            "audio_strict_min_overlap_s": 60.0,
            "audio_peak_ratio_min": 1.07,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.020,
            "audio_mutual_overlap_ratio_min": 0.62,
        },
        {
            "combo_id": "tl_overlap_chase_wide",
            "audio_duration_ratio_min": 0.10,
            "audio_min_hashset_intersect_ratio": 0.10,
            "audio_strict_min_overlap_s": 55.0,
            "audio_peak_ratio_min": 1.06,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 10,
            "audio_min_vote_fraction": 0.018,
            "audio_mutual_overlap_ratio_min": 0.60,
        },
        {
            "combo_id": "tl_guided_max_precision",
            "audio_duration_ratio_min": 0.12,
            "audio_min_hashset_intersect_ratio": 0.13,
            "audio_strict_min_overlap_s": 85.0,
            "audio_peak_ratio_min": 1.10,
            "audio_peak_margin": 1,
            "audio_strict_hamming_thresh": 9,
            "audio_min_vote_fraction": 0.024,
            "audio_mutual_overlap_ratio_min": 0.68,
        },
    ])

    rows: list[dict] = []
    print(f"[tune][audio] evaluating {len(combos)} combo(s)...", flush=True)
    t0 = time.time()
    for n, combo in enumerate(combos, 1):
        matcher_params = dict(base_params)
        matcher_params["strict_min_overlap_s"] = combo["audio_strict_min_overlap_s"]
        matcher_params["peak_ratio_min"] = combo["audio_peak_ratio_min"]
        matcher_params["peak_margin"] = combo["audio_peak_margin"]
        matcher_params["strict_hamming_thresh"] = combo["audio_strict_hamming_thresh"]
        matcher_params["min_vote_fraction"] = combo["audio_min_vote_fraction"]
        matcher_params["mutual_overlap_ratio_min"] = combo["audio_mutual_overlap_ratio_min"]

        metrics = _evaluate_tuning_combo(
            entries=entries,
            dataset=dataset,
            matcher_params=matcher_params,
            duration_ratio_min=combo["audio_duration_ratio_min"],
            min_hashset_intersect_ratio=combo["audio_min_hashset_intersect_ratio"],
            overlap_targets=timeline_overlap_targets,
            overlap_target_weight=0.12 if timeline_overlap_targets else 0.0,
        )
        row = dict(combo)
        row.update(metrics)
        rows.append(row)
        if (n == len(combos)) or (n % max(1, len(combos) // 5) == 0):
            elapsed = time.time() - t0
            print(
                f"[tune][audio] {n}/{len(combos)} combos - {_format_duration_auto(elapsed)} elapsed",
                flush=True,
            )

    best = _choose_best_tuning_combo(rows)
    if not best:
        return {}, pd.DataFrame(rows)

    tuned_kwargs = {
        "audio_duration_ratio_min": float(best["audio_duration_ratio_min"]),
        "audio_min_hashset_intersect_ratio": float(best["audio_min_hashset_intersect_ratio"]),
        "audio_strict_min_overlap_s": float(best["audio_strict_min_overlap_s"]),
        "audio_peak_ratio_min": float(best["audio_peak_ratio_min"]),
        "audio_peak_margin": int(best["audio_peak_margin"]),
        "audio_strict_hamming_thresh": int(best["audio_strict_hamming_thresh"]),
        "audio_min_vote_fraction": float(best["audio_min_vote_fraction"]),
        "audio_mutual_overlap_ratio_min": float(best["audio_mutual_overlap_ratio_min"]),
    }
    print(
        "[tune][audio] selected "
        f"{best.get('combo_id')} "
        f"(recall={float(best.get('recall', 0.0)):.3f}, "
        f"precision={float(best.get('precision', 0.0)):.3f}, "
        f"overlap_agree={float(best.get('overlap_agreement', 0.0)):.3f}, "
        f"fpr={float(best.get('false_positive_rate', 0.0)):.3f})",
        flush=True,
    )
    return tuned_kwargs, pd.DataFrame(rows)


def _tune_timeline_from_cache(
    directories: list[str],
    positives_by_b: dict[str, set[str]],
    hard_negative_pairs: set[tuple[str, str]],
) -> tuple[dict, pd.DataFrame]:
    entries = _load_cached_entries_for_tuning(directories, mode="timeline")
    dataset = _build_tuning_dataset(
        entries,
        positives_by_b,
        hard_negative_pairs,
        max_negatives_per_b=5,
    )
    if not dataset:
        print("[tune][timeline] empty tuning dataset; using defaults.", flush=True)
        return {}, pd.DataFrame()

    speed_min_eff, speed_max_eff, speed_steps_eff = _resolve_timeline_speed_search(
        TIMELINE_ENABLE_SPEED_SWEEP,
        TIMELINE_SPEED_RATIO_MIN,
        TIMELINE_SPEED_RATIO_MAX,
        TIMELINE_SPEED_STEPS,
        fixed_ratio=TIMELINE_SPEED_RATIO_FIXED,
    )

    base_params = {
        "total_bits": 64,
        "hamming_thresh": TIMELINE_HAMMING_THRESH,
        "bin_s": TIMELINE_BIN_S,
        "min_votes": TIMELINE_MIN_VOTES,
        "min_overlap_s": TIMELINE_MIN_OVERLAP_S,
        "hop_s": TIMELINE_STEP_S,
        "lsh_chunks": TIMELINE_LSH_CHUNKS,
        "brute_limit": TIMELINE_BRUTE_MAX,
        "strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        "strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
        "min_vote_fraction": TIMELINE_MIN_VOTE_FRACTION,
        "peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
        "peak_margin": TIMELINE_PEAK_MARGIN,
        "speed_ratio_min": speed_min_eff,
        "speed_ratio_max": speed_max_eff,
        "speed_steps": speed_steps_eff,
        "max_candidates_per_frame": TIMELINE_MAX_CANDIDATES_PER_FRAME,
        "long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
        "long_overlap_vote_mult": TIMELINE_LONG_OVERLAP_VOTE_MULT,
        "run_gap_mult": TIMELINE_RUN_GAP_MULT,
    }

    combos: list[dict] = [{
        "combo_id": "baseline",
        "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
        "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
        "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
        "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
        "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
        "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
    }]
    combos.extend([
        {
            "combo_id": "dr0.08",
            "timeline_duration_ratio_min": 0.08,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "dr0.12",
            "timeline_duration_ratio_min": 0.12,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "strict90",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": 90.0,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "strict110",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": 110.0,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "long180",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": 180.0,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "long240",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": 240.0,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "peak1.06",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": 1.06,
            "timeline_peak_margin": 1,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "strict_h13",
            "timeline_duration_ratio_min": TIMELINE_DURATION_RATIO_MIN,
            "timeline_strict_min_overlap_s": TIMELINE_STRICT_MIN_OVERLAP_S,
            "timeline_long_overlap_override_s": TIMELINE_LONG_OVERLAP_OVERRIDE_S,
            "timeline_peak_ratio_min": TIMELINE_PEAK_RATIO_MIN,
            "timeline_peak_margin": TIMELINE_PEAK_MARGIN,
            "timeline_strict_hamming_thresh": 13,
        },
        {
            "combo_id": "recall_a",
            "timeline_duration_ratio_min": 0.08,
            "timeline_strict_min_overlap_s": 90.0,
            "timeline_long_overlap_override_s": 180.0,
            "timeline_peak_ratio_min": 1.06,
            "timeline_peak_margin": 1,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "recall_b",
            "timeline_duration_ratio_min": 0.10,
            "timeline_strict_min_overlap_s": 90.0,
            "timeline_long_overlap_override_s": 200.0,
            "timeline_peak_ratio_min": 1.06,
            "timeline_peak_margin": 1,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "balanced_a",
            "timeline_duration_ratio_min": 0.10,
            "timeline_strict_min_overlap_s": 100.0,
            "timeline_long_overlap_override_s": 180.0,
            "timeline_peak_ratio_min": 1.06,
            "timeline_peak_margin": 1,
            "timeline_strict_hamming_thresh": TIMELINE_STRICT_HAMMING_THRESH,
        },
        {
            "combo_id": "strict_combo",
            "timeline_duration_ratio_min": 0.12,
            "timeline_strict_min_overlap_s": 110.0,
            "timeline_long_overlap_override_s": 240.0,
            "timeline_peak_ratio_min": 1.10,
            "timeline_peak_margin": 3,
            "timeline_strict_hamming_thresh": 13,
        },
    ])

    rows: list[dict] = []
    print(f"[tune][timeline] evaluating {len(combos)} combo(s)...", flush=True)
    t0 = time.time()
    for n, combo in enumerate(combos, 1):
        matcher_params = dict(base_params)
        matcher_params["strict_min_overlap_s"] = combo["timeline_strict_min_overlap_s"]
        matcher_params["long_overlap_override_s"] = combo["timeline_long_overlap_override_s"]
        matcher_params["peak_ratio_min"] = combo["timeline_peak_ratio_min"]
        matcher_params["peak_margin"] = combo["timeline_peak_margin"]
        matcher_params["strict_hamming_thresh"] = combo["timeline_strict_hamming_thresh"]

        metrics = _evaluate_tuning_combo(
            entries=entries,
            dataset=dataset,
            matcher_params=matcher_params,
            duration_ratio_min=combo["timeline_duration_ratio_min"],
        )
        row = dict(combo)
        row.update(metrics)
        rows.append(row)
        if (n == len(combos)) or (n % max(1, len(combos) // 5) == 0):
            elapsed = time.time() - t0
            print(
                f"[tune][timeline] {n}/{len(combos)} combos - {_format_duration_auto(elapsed)} elapsed",
                flush=True,
            )

    best = _choose_best_tuning_combo(rows)
    if not best:
        return {}, pd.DataFrame(rows)

    tuned_kwargs = {
        "timeline_duration_ratio_min": float(best["timeline_duration_ratio_min"]),
        "timeline_strict_min_overlap_s": float(best["timeline_strict_min_overlap_s"]),
        "timeline_long_overlap_override_s": float(best["timeline_long_overlap_override_s"]),
        "timeline_peak_ratio_min": float(best["timeline_peak_ratio_min"]),
        "timeline_peak_margin": int(best["timeline_peak_margin"]),
        "timeline_strict_hamming_thresh": int(best["timeline_strict_hamming_thresh"]),
    }
    print(
        "[tune][timeline] selected "
        f"{best.get('combo_id')} "
        f"(recall={float(best.get('recall', 0.0)):.3f}, "
        f"precision={float(best.get('precision', 0.0)):.3f}, "
        f"fpr={float(best.get('false_positive_rate', 0.0)):.3f})",
        flush=True,
    )
    return tuned_kwargs, pd.DataFrame(rows)


def _heuristic_tune_from_truth_report(truth_report_path: str) -> tuple[dict, dict, pd.DataFrame]:
    """
    Fast global tuner from existing consolidated results.
    Uses anchor-confirmed rows as calibration truth to set global audio/timeline knobs.
    """
    if not os.path.exists(truth_report_path):
        return {}, {}, pd.DataFrame()
    try:
        df = pd.read_excel(truth_report_path, sheet_name="Consolidated")
    except Exception:
        return {}, {}, pd.DataFrame()
    if df.empty:
        return {}, {}, pd.DataFrame()

    if "matched_by_count" not in df.columns:
        mode_cols = [c for c in df.columns if c.startswith("matched_by_") and c != "matched_by_count"]
        if mode_cols:
            df["matched_by_count"] = df[mode_cols].fillna(False).astype(bool).sum(axis=1).astype(int)
        else:
            df["matched_by_count"] = 1
    if "confidence_score" not in df.columns:
        df["confidence_score"] = np.where(df["matched_by_count"] >= 2, 80.0, 60.0)

    matched_by_count = pd.to_numeric(df["matched_by_count"], errors="coerce").fillna(0).astype(int)
    confidence_score = pd.to_numeric(df["confidence_score"], errors="coerce").fillna(0.0)
    trusted = (matched_by_count >= 2) | (confidence_score >= 70.0)

    def _pick_num(*candidates: str) -> pd.Series:
        for c in candidates:
            if c in df.columns:
                s = pd.to_numeric(df[c], errors="coerce")
                if not s.isna().all():
                    return s
        return pd.Series(np.nan, index=df.index, dtype=float)

    def _q(series: pd.Series, q: float, default: float) -> float:
        s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        if s.empty:
            return float(default)
        return float(s.quantile(q))

    def _clamp(v: float, lo: float, hi: float) -> float:
        return max(float(lo), min(float(hi), float(v)))

    dur_a = _pick_num("duration_a (s)", "anchors__duration_a (s)", "audio__duration_a (s)", "timeline__duration_a (s)")
    dur_b = _pick_num("duration_b (s)", "anchors__duration_b (s)", "audio__duration_b (s)", "timeline__duration_b (s)")
    dur_max = pd.concat([dur_a, dur_b], axis=1).max(axis=1).replace(0, np.nan)
    dur_ratio = (pd.concat([dur_a, dur_b], axis=1).min(axis=1) / dur_max).replace([np.inf, -np.inf], np.nan)
    expected_overlap = pd.concat([dur_a, dur_b], axis=1).min(axis=1)

    m_anchor = df.get("matched_by_anchors", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    start_frac = _pick_num("anchors__start_match_fraction", "start_match_fraction")
    end_frac = _pick_num("anchors__end_match_fraction", "end_match_fraction")
    start_mad = _pick_num("anchors__start_offset_mad_s", "start_offset_mad_s")
    end_mad = _pick_num("anchors__end_offset_mad_s", "end_offset_mad_s")
    start_off = _pick_num("anchors__start_offset_s", "start_offset_s")
    end_off = _pick_num("anchors__end_offset_s", "end_offset_s")

    start_ok = (start_frac >= float(ANCHOR_MIN_FRACTION)) & (start_mad.isna() | (start_mad <= float(ANCHOR_MAX_MAD_S) * 1.25))
    end_ok = (end_frac >= float(ANCHOR_MIN_FRACTION)) & (end_mad.isna() | (end_mad <= float(ANCHOR_MAX_MAD_S) * 1.25))
    anchor_truth = m_anchor & (start_ok | end_ok) & trusted & (expected_overlap >= 120.0)

    # Estimate global speed drift from anchors when both start+end offsets exist.
    slope_mask = anchor_truth & start_off.notna() & end_off.notna() & (expected_overlap >= 180.0)
    slope_est = (1.0 - ((end_off - start_off) / expected_overlap)).where(slope_mask)
    slope_est = slope_est[(slope_est >= 0.80) & (slope_est <= 1.20)]
    if slope_est.empty:
        slope_lo_a, slope_hi_a = AUDIO_SPEED_RATIO_MIN, AUDIO_SPEED_RATIO_MAX
        if TIMELINE_ENABLE_SPEED_SWEEP:
            slope_lo_t, slope_hi_t = TIMELINE_SPEED_RATIO_MIN, TIMELINE_SPEED_RATIO_MAX
        else:
            slope_lo_t, slope_hi_t = TIMELINE_SPEED_RATIO_FIXED, TIMELINE_SPEED_RATIO_FIXED
    else:
        slope_lo_a = _clamp(_q(slope_est, 0.05, AUDIO_SPEED_RATIO_MIN) - 0.01, 0.85, 1.00)
        slope_hi_a = _clamp(_q(slope_est, 0.95, AUDIO_SPEED_RATIO_MAX) + 0.01, 1.00, 1.15)
        if TIMELINE_ENABLE_SPEED_SWEEP:
            slope_lo_t = _clamp(_q(slope_est, 0.10, TIMELINE_SPEED_RATIO_MIN) - 0.005, 0.90, 1.00)
            slope_hi_t = _clamp(_q(slope_est, 0.90, TIMELINE_SPEED_RATIO_MAX) + 0.005, 1.00, 1.10)
        else:
            slope_lo_t, slope_hi_t = TIMELINE_SPEED_RATIO_FIXED, TIMELINE_SPEED_RATIO_FIXED

    def _odd_steps(width: float, step: float, lo: int, hi: int) -> int:
        n = int(round(float(width) / max(1e-6, float(step)))) + 1
        n = max(int(lo), min(int(hi), n))
        if n % 2 == 0:
            n = min(int(hi), n + 1)
        return n

    # Audio global tune from anchor-truth coverage.
    audio_kwargs: dict = {}
    m_audio = df.get("matched_by_audio", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    audio_overlap = _pick_num("audio__audio_overlap_s", "audio_overlap_s")
    audio_votes = _pick_num("audio__audio_votes", "audio_votes")
    audio_seq_a = _pick_num("audio__audio_seq_a", "audio_seq_a")
    audio_seq_b = _pick_num("audio__audio_seq_b", "audio_seq_b")
    audio_on_truth = anchor_truth & m_audio & audio_overlap.notna()
    if bool(anchor_truth.any()):
        cov_audio = (audio_overlap / expected_overlap).replace([np.inf, -np.inf], np.nan)
        cov50 = _q(cov_audio[audio_on_truth], 0.50, 0.30)
        cov25 = _q(cov_audio[audio_on_truth], 0.25, 0.20)
        overlap20 = _q(audio_overlap[audio_on_truth], 0.20, AUDIO_STRICT_MIN_OVERLAP_S)
        overlap60 = _q(audio_overlap[audio_on_truth], 0.60, AUDIO_LONG_OVERLAP_OVERRIDE_S)
        vote_denom = pd.concat([audio_seq_a, audio_seq_b], axis=1).min(axis=1).replace(0, np.nan)
        vote_frac = (audio_votes / vote_denom).replace([np.inf, -np.inf], np.nan)
        vote_frac20 = _q(vote_frac[audio_on_truth], 0.20, AUDIO_MIN_VOTE_FRACTION)
        match_rate = float((anchor_truth & m_audio).sum()) / max(1.0, float(anchor_truth.sum()))

        run_gap = _clamp(1.9 + (1.0 - cov50) * 2.8 + (0.4 if match_rate < 0.55 else 0.0), 1.8, 5.0)
        strict_overlap = _clamp(overlap20 * 0.75, 35.0, float(AUDIO_STRICT_MIN_OVERLAP_S))
        long_override = _clamp(overlap60 * 0.60, 120.0, 380.0)
        duration_ratio_min = _clamp(_q(dur_ratio[anchor_truth], 0.05, AUDIO_DURATION_RATIO_MIN) * 0.65, 0.06, float(AUDIO_DURATION_RATIO_MIN))
        min_vote_fraction = _clamp(vote_frac20 * 0.35, 0.009, float(AUDIO_MIN_VOTE_FRACTION))
        peak_ratio = 1.03 if match_rate < 0.55 else (1.04 if match_rate < 0.75 else 1.06)
        peak_ratio = min(float(peak_ratio), float(AUDIO_PEAK_RATIO_MIN))
        strict_h = AUDIO_STRICT_HAMMING_THRESH + (1 if (cov25 < 0.20 or match_rate < 0.50) else 0)
        strict_h = int(min(AUDIO_HAMMING_THRESH, max(8, strict_h)))
        speed_steps = _odd_steps(slope_hi_a - slope_lo_a, 0.015, 9, 21)

        audio_kwargs = {
            "audio_duration_ratio_min": round(duration_ratio_min, 3),
            "audio_strict_min_overlap_s": round(strict_overlap, 1),
            "audio_long_overlap_override_s": round(long_override, 1),
            "audio_peak_ratio_min": round(float(peak_ratio), 2),
            "audio_peak_margin": 1,
            "audio_min_vote_fraction": round(min_vote_fraction, 4),
            "audio_strict_hamming_thresh": int(strict_h),
            "audio_speed_ratio_min": round(float(slope_lo_a), 3),
            "audio_speed_ratio_max": round(float(slope_hi_a), 3),
            "audio_speed_steps": int(speed_steps),
            "audio_run_gap_mult": round(float(run_gap), 2),
        }

    # Timeline global tune from anchor-truth coverage (timeline + timeline2 rows).
    timeline_kwargs: dict = {}
    m_tl = df.get("matched_by_timeline", pd.Series(False, index=df.index)).fillna(False).astype(bool)
    m_tl_any = m_tl
    tl_overlap = _pick_num("timeline__timeline_overlap_s", "timeline_overlap_s")
    tl_votes = _pick_num("timeline__timeline_votes", "timeline_votes")
    tl_on_truth = anchor_truth & m_tl_any & tl_overlap.notna()
    if bool(anchor_truth.any()):
        cov_tl = (tl_overlap / expected_overlap).replace([np.inf, -np.inf], np.nan)
        cov50_tl = _q(cov_tl[tl_on_truth], 0.50, 0.35)
        cov25_tl = _q(cov_tl[tl_on_truth], 0.25, 0.25)
        overlap20_tl = _q(tl_overlap[tl_on_truth], 0.20, TIMELINE_STRICT_MIN_OVERLAP_S)
        overlap60_tl = _q(tl_overlap[tl_on_truth], 0.60, TIMELINE_LONG_OVERLAP_OVERRIDE_S)
        match_rate_tl = float((anchor_truth & m_tl_any).sum()) / max(1.0, float(anchor_truth.sum()))

        run_gap_tl = _clamp(1.9 + (1.0 - cov50_tl) * 2.4 + (0.3 if match_rate_tl < 0.60 else 0.0), 1.8, 4.5)
        strict_overlap_tl = _clamp(overlap20_tl * 0.85, 70.0, float(TIMELINE_STRICT_MIN_OVERLAP_S))
        long_override_tl = _clamp(overlap60_tl * 0.85, 130.0, 320.0)
        duration_ratio_min_tl = _clamp(_q(dur_ratio[anchor_truth], 0.05, TIMELINE_DURATION_RATIO_MIN) * 0.65, 0.06, float(TIMELINE_DURATION_RATIO_MIN))
        peak_ratio_tl = 1.05 if match_rate_tl < 0.60 else (1.06 if match_rate_tl < 0.80 else float(TIMELINE_PEAK_RATIO_MIN))
        peak_ratio_tl = min(float(peak_ratio_tl), float(TIMELINE_PEAK_RATIO_MIN))
        peak_margin_tl = 1 if peak_ratio_tl <= 1.06 else int(TIMELINE_PEAK_MARGIN)
        strict_h_tl = TIMELINE_STRICT_HAMMING_THRESH + (1 if (cov25_tl < 0.25 or match_rate_tl < 0.55) else 0)
        strict_h_tl = int(min(TIMELINE_HAMMING_THRESH, max(10, strict_h_tl)))
        speed_steps_tl = _odd_steps(slope_hi_t - slope_lo_t, 0.01, 7, 19) if TIMELINE_ENABLE_SPEED_SWEEP else 1

        timeline_kwargs = {
            "timeline_duration_ratio_min": round(duration_ratio_min_tl, 3),
            "timeline_strict_min_overlap_s": round(strict_overlap_tl, 1),
            "timeline_long_overlap_override_s": round(long_override_tl, 1),
            "timeline_peak_ratio_min": round(float(peak_ratio_tl), 2),
            "timeline_peak_margin": int(peak_margin_tl),
            "timeline_strict_hamming_thresh": int(strict_h_tl),
            "timeline_speed_ratio_min": round(float(slope_lo_t), 3),
            "timeline_speed_ratio_max": round(float(slope_hi_t), 3),
            "timeline_speed_steps": int(speed_steps_tl),
            "timeline_enable_speed_sweep": bool(TIMELINE_ENABLE_SPEED_SWEEP),
            "timeline_run_gap_mult": round(float(run_gap_tl), 2),
        }

    summary = pd.DataFrame([
        {
            "truth_report": truth_report_path,
            "trusted_rows": int(trusted.sum()),
            "anchor_truth_rows": int(anchor_truth.sum()),
            "anchor_with_slope_rows": int(slope_est.notna().sum()),
            "slope_min": round(float(_q(slope_est, 0.05, 1.0)), 4),
            "slope_max": round(float(_q(slope_est, 0.95, 1.0)), 4),
            "audio_rows": int(df.get("matched_by_audio", pd.Series(False)).fillna(False).astype(bool).sum()),
            "timeline_rows": int(df.get("matched_by_timeline", pd.Series(False)).fillna(False).astype(bool).sum()),
        }
    ])
    return audio_kwargs, timeline_kwargs, summary


def auto_tune_timeline_from_reports(
    *,
    directories: list[str],
    truth_report_path: str,
    tuning_report_path: str,
    expected_parts_per_b: int = 2,
    use_cache_replay: bool = False,
    enable_timeline: bool = True,
) -> dict:
    """
    Replay cached timeline fingerprints against a silver-truth set to pick better
    timeline gates. Writes a tuning workbook and returns kwargs for
    `find_video_duplicates`.
    """
    timeline_kwargs: dict = {}
    timeline_table = pd.DataFrame()
    summary_df = pd.DataFrame()

    if use_cache_replay:
        positives_by_b, _positive_pairs, hard_negative_pairs = _build_silver_truth_from_consolidated(
            truth_report_path,
            expected_parts_per_b=expected_parts_per_b,
        )
        timeline_overlap_targets = _extract_timeline_overlap_truth(truth_report_path)
        if not positives_by_b:
            print("[tune] no silver truth available; using default mode parameters.", flush=True)
            return {"timeline": {}}
        if enable_timeline:
            timeline_kwargs, timeline_table = _tune_timeline_from_cache(
                directories,
                positives_by_b,
                hard_negative_pairs,
            )
        summary_df = pd.DataFrame([{
            "truth_report": truth_report_path,
            "positive_file_b_groups": len(positives_by_b),
            "positive_pairs": sum(len(v) for v in positives_by_b.values()),
            "hard_negative_pairs": len(hard_negative_pairs),
            "timeline_overlap_targets": len(timeline_overlap_targets),
            "tuning_mode": "cache_replay",
        }])
    else:
        _audio_kwargs_unused, heuristic_timeline_kwargs, summary_df = _heuristic_tune_from_truth_report(truth_report_path)
        timeline_kwargs = heuristic_timeline_kwargs if enable_timeline else {}
        if timeline_kwargs:
            print(f"[tune][timeline] heuristic params: {timeline_kwargs}", flush=True)

    try:
        os.makedirs(os.path.dirname(tuning_report_path) or ".", exist_ok=True)
        with ExcelWriter(tuning_report_path, engine="openpyxl") as writer:
            chosen_rows = []
            if timeline_kwargs:
                chosen_rows.append({"mode": "timeline", **timeline_kwargs})
            chosen_df = pd.DataFrame(chosen_rows if chosen_rows else [{"mode": "timeline"}])
            chosen_df.to_excel(writer, index=False, sheet_name="Chosen")
            _format_excel_sheet(chosen_df, writer.sheets["Chosen"])

            if summary_df.empty:
                summary_df = pd.DataFrame([{"truth_report": truth_report_path}])
            summary_df.to_excel(writer, index=False, sheet_name="TruthSummary")
            _format_excel_sheet(summary_df, writer.sheets["TruthSummary"])

            if not timeline_table.empty:
                timeline_sorted = timeline_table.sort_values(
                    by=["recall", "score", "precision"],
                    ascending=[False, False, False],
                ).reset_index(drop=True)
                timeline_sorted.to_excel(writer, index=False, sheet_name="TimelineCombos")
                _format_excel_sheet(timeline_sorted, writer.sheets["TimelineCombos"])
        print(f"[tune] wrote tuning report: {tuning_report_path}", flush=True)
    except Exception as e:
        print(f"[tune] failed writing tuning report: {e}", flush=True)

    return {"timeline": timeline_kwargs}


CONSOLIDATED_SHEET_BASE_COLUMNS = [
    "file_a",
    "file_a_osd_month_year_start",
    "file_a_osd_month_year_end",
    "file_a_osd_span_months",
    "file_a_osd_out_of_order",
    "file_a_osd_hits",
    "file_b",
    "file_b_osd_month_year_start",
    "file_b_osd_month_year_end",
    "file_b_osd_span_months",
    "file_b_osd_out_of_order",
    "file_b_osd_hits",
    "file_b_part_order",
    "file_a_candidate_count",
    "group_candidate_count",
    "confidence_score",
    "evidence_score",
    "consensus_score",
    "confidence_tier",
    "matched_by_count",
    "modes",
    "rename_ready",
    "timeline_est_overlap_s",
    "timeline_tuned_offset_s",
    "timeline_est_a_start_s",
    "timeline_est_a_end_s",
    "timeline_est_b_start_s",
    "timeline_est_b_end_s",
    "timeline_a_overlap_multimatch",
    "timeline_b_overlap_multimatch",
    "timeline_a_has_unmatched_parts",
    "timeline_b_has_unmatched_parts",
    "timeline_a_has_internal_unmatched_gaps",
    "timeline_b_has_internal_unmatched_gaps",
    "anchor_full_full_conflict",
    "timeline_gap_verify_status",
    "timeline_gap_verify_notes",
    "match_summary",
    "file_b_group_summary",
    "file_a_group_summary",
    "timeline_a_overlap_clash_count",
    "timeline_a_overlap_clash_s",
    "timeline_a_overlap_clash_pct",
    "timeline_b_overlap_clash_count",
    "timeline_b_overlap_clash_s",
    "timeline_b_overlap_clash_pct",
    "timeline_a_coverage_est_pct",
    "timeline_a_unique_covered_est_pct",
    "timeline_a_coverage_raw_pct",
    "timeline_a_coverage_est_s",
    "timeline_a_unique_covered_est_s",
    "timeline_a_coverage_raw_s",
    "timeline_a_duration_est_s",
    "timeline_a_unique_duration_est_s",
    "timeline_a_self_repeat_redundant_pct",
    "timeline_a_self_repeat_redundant_s",
    "timeline_a_coverage_sections_est",
    "timeline_a_unique_coverage_sections_est",
    "timeline_a_coverage_sections_raw",
    "timeline_b_coverage_est_pct",
    "timeline_b_coverage_raw_pct",
    "timeline_b_coverage_est_s",
    "timeline_b_coverage_raw_s",
    "timeline_b_duration_est_s",
    "timeline_b_coverage_sections_est",
    "timeline_b_coverage_sections_raw",
    "review_priority",
    "needs_review",
    "review_flags",
]

B_COVERAGE_SHEET_COLUMNS = [
    "file_b",
    "osd_month_year_start",
    "osd_month_year_end",
    "osd_span_months",
    "osd_hits",
    "osd_out_of_order",
    "matched_file_a_count",
    "matched_file_a_names",
    "best_pair_confidence_score",
    "avg_pair_confidence_score",
    "timeline_coverage_est_pct",
    "timeline_coverage_raw_pct",
    "timeline_coverage_est_s",
    "timeline_coverage_raw_s",
    "timeline_duration_est_s",
    "timeline_coverage_sections_est",
    "timeline_coverage_sections_raw",
    "coverage_summary",
]

A_COVERAGE_SHEET_COLUMNS = [
    "file_a",
    "osd_month_year_start",
    "osd_month_year_end",
    "osd_span_months",
    "osd_hits",
    "osd_out_of_order",
    "matched_file_b_count",
    "matched_file_b_names",
    "best_pair_confidence_score",
    "avg_pair_confidence_score",
    "timeline_a_unique_covered_est_pct",
    "timeline_a_unique_covered_est_s",
    "timeline_a_unique_duration_est_s",
    "timeline_a_self_repeat_redundant_pct",
    "timeline_a_self_repeat_redundant_s",
    "timeline_a_unique_coverage_sections_est",
    "timeline_coverage_est_pct",
    "timeline_coverage_raw_pct",
    "timeline_coverage_est_s",
    "timeline_coverage_raw_s",
    "timeline_duration_est_s",
    "timeline_coverage_sections_est",
    "timeline_coverage_sections_raw",
    "coverage_summary",
]

B_GROUP_COMPARE_SHEET_COLUMNS = [
    "group_id",
    "file_b_1",
    "file_b_2",
    "timeline_relation",
    "timeline_match_ok",
    "timeline_overlap_b1_pct",
    "timeline_overlap_b2_pct",
    "timeline_overlap_b1_s",
    "timeline_overlap_b2_s",
    "timeline_segment_count",
    "timeline_segments_b1",
    "timeline_segments_b2",
    "anchor_relation",
    "anchor_start_fraction",
    "anchor_end_fraction",
    "summary",
]

RENAME_WORKFLOW_SHEET_COLUMNS = [
    "workflow_status",
    "notes",
    "proposed_new_name",
    "triage_tier",
    "file_side",
    "file_name",
    "osd_month_year_start",
    "osd_month_year_end",
    "matched_file_names",
    "coverage_summary",
    "same_group_file_b_summary",
    "timeline_coverage_est_pct",
    "video_runtime_s",
    "runtime_ex_dead_and_duplicate_s",
    "file_path",
    "group_file_count",
    "group_file_a_count",
    "group_file_b_count",
    "timeline_a_unique_covered_est_pct",
    "group_min_a_unique_covered_est_pct",
    "group_min_b_coverage_est_pct",
    "group_best_confidence_score",
    "group_any_overlap_multimatch",
    "group_any_internal_unmatched_gaps",
    "group_any_middle_unmatched_gap",
    "osd_span_months",
    "osd_hits",
    "timeline_a_self_repeat_redundant_pct",
    "timeline_coverage_sections_est",
    "overlap_multimatch",
    "has_internal_unmatched_gaps",
    "has_middle_unmatched_gap",
    "avg_pair_confidence_score",
]

REMUX_PLAN_SHEET_COLUMNS = [
    "workflow_status",
    "notes",
    "proposed_output_name",
    "unique_segment_status",
    "file_a_name",
    "segment_index",
    "matched_file_b_names",
    "borderline_maybe_file_b_name",
    "borderline_maybe_b_start_s",
    "borderline_maybe_b_end_s",
    "borderline_maybe_summary",
    "unique_segment_duration_s",
    "unique_segment_summary",
    "file_a_unique_summary",
    "unique_segment_osd_month_year_start",
    "unique_segment_osd_month_year_end",
    "file_a_osd_month_year_start",
    "file_a_osd_month_year_end",
    "source_file_path",
    "segment_start_s",
    "segment_end_s",
    "video_runtime_s",
    "runtime_ex_dead_and_duplicate_s",
    "group_id",
]


def _select_existing_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Keep requested columns in order, silently skipping absent ones."""
    keep = [c for c in columns if c in df.columns]
    return df[keep].copy()


def _write_consolidated_workbook(
    *,
    output_path: str,
    row_count: int,
    rename_queue_df: pd.DataFrame,
    rename_done_df: pd.DataFrame,
    remux_plan_df: pd.DataFrame,
    remux_short_df: pd.DataFrame,
    b_group_compare_df: pd.DataFrame,
    consolidated_view: pd.DataFrame,
    a_coverage_df: pd.DataFrame,
    b_coverage_df: pd.DataFrame,
    unmatched_a_df: pd.DataFrame | None,
    unmatched_b_df: pd.DataFrame | None,
) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    prewrite_backup_done = False
    refresh_backup_before_retry = False
    while True:
        try:
            if not prewrite_backup_done and os.path.exists(output_path):
                _backup_reports([output_path])
                prewrite_backup_done = True
            if refresh_backup_before_retry and os.path.exists(output_path):
                _backup_reports([output_path], label="retry")
                refresh_backup_before_retry = False
            with ExcelWriter(output_path, engine="openpyxl") as writer:
                if not rename_queue_df.empty:
                    rename_queue_df.to_excel(writer, index=False, sheet_name="Rename_Queue")
                    _format_excel_sheet(rename_queue_df, writer.sheets["Rename_Queue"])

                if not rename_done_df.empty:
                    rename_done_df.to_excel(writer, index=False, sheet_name="Rename_Done")
                    _format_excel_sheet(rename_done_df, writer.sheets["Rename_Done"])

                if not remux_plan_df.empty:
                    remux_plan_df.to_excel(writer, index=False, sheet_name="Remux_Plan")
                    _format_excel_sheet(remux_plan_df, writer.sheets["Remux_Plan"])

                if not remux_short_df.empty:
                    remux_short_df.to_excel(writer, index=False, sheet_name="Remux_Short")
                    _format_excel_sheet(remux_short_df, writer.sheets["Remux_Short"])

                if not b_group_compare_df.empty:
                    b_group_compare_df.to_excel(writer, index=False, sheet_name="B_Group_Compare")
                    _format_excel_sheet(b_group_compare_df, writer.sheets["B_Group_Compare"])

                consolidated_view.to_excel(writer, index=False, sheet_name="Consolidated")
                _format_excel_sheet(consolidated_view, writer.sheets["Consolidated"])

                a_coverage_df.to_excel(writer, index=False, sheet_name="A_Coverage")
                _format_excel_sheet(a_coverage_df, writer.sheets["A_Coverage"])

                b_coverage_df.to_excel(writer, index=False, sheet_name="B_Coverage")
                _format_excel_sheet(b_coverage_df, writer.sheets["B_Coverage"])

                if unmatched_a_df is not None:
                    unmatched_a_df.to_excel(writer, index=False, sheet_name="Unmatched_A")
                    _format_excel_sheet(unmatched_a_df, writer.sheets["Unmatched_A"])

                if unmatched_b_df is not None:
                    unmatched_b_df.to_excel(writer, index=False, sheet_name="Unmatched_B")
                    _format_excel_sheet(unmatched_b_df, writer.sheets["Unmatched_B"])
            break
        except PermissionError:
            choice = sg.popup_yes_no(
                f"Can't write to file:\n{output_path}\n\nIt might be open in Excel.\n\nRetry?",
                title="Export Failed",
                keep_on_top=True,
            )
            if choice != "Yes":
                print(f"Export to {output_path} aborted by user.")
                break
            refresh_backup_before_retry = True
        except Exception as e:
            sg.popup_error(f"Unexpected error while exporting:\n{e}", title="Export Failed", keep_on_top=True)
            break

    print(f"[consolidate] Wrote {row_count} row(s) to {output_path}")


def _consolidated_sheet_columns(mode_names: list[str], available_columns: list[str]) -> list[str]:
    matched_by_cols = [f"matched_by_{m}" for m in mode_names if f"matched_by_{m}" in available_columns]
    strength_cols = [f"strength_{m}" for m in mode_names if f"strength_{m}" in available_columns]
    cols = list(CONSOLIDATED_SHEET_BASE_COLUMNS)
    cols.extend(matched_by_cols)
    cols.extend(strength_cols)
    return [c for c in cols if c in available_columns]


def _load_timestamp_metadata_from_report(timestamp_report_path: str | None) -> dict[str, dict]:
    timestamp_map: dict[str, dict] = {}
    ts_path = timestamp_report_path
    if ts_path is None:
        default_ts_path = os.path.join(REPORTS_DIR, "dedupe_timestamps.xlsx")
        if os.path.exists(default_ts_path):
            ts_path = default_ts_path
    if not ts_path or (not os.path.exists(ts_path)):
        return timestamp_map
    try:
        _refresh_timestamp_report_from_cache(ts_path)
        ts_df = pd.read_excel(ts_path)
        if ts_df.empty:
            return timestamp_map
        cols_norm = {_normalize_col_name(c): c for c in ts_df.columns}
        c_path = cols_norm.get("file_path") or cols_norm.get("file") or cols_norm.get("path")
        c_start = cols_norm.get("osd_month_year_start")
        c_end = cols_norm.get("osd_month_year_end")
        c_span = cols_norm.get("osd_month_year_span_months")
        c_hits = cols_norm.get("osd_candidate_hits") or cols_norm.get("ym_hits")
        c_status = cols_norm.get("status")
        c_out = cols_norm.get("osd_out_of_order_in_video")

        def _to_int_safe(v) -> int:
            num = pd.to_numeric(v, errors="coerce")
            if pd.isna(num):
                return 0
            try:
                return int(num)
            except Exception:
                return 0

        if not (c_path and c_start and c_end):
            print(
                f"[consolidate] warning: timestamp report missing required columns in {ts_path}",
                flush=True,
            )
            return timestamp_map

        for _, row in ts_df.iterrows():
            p = row.get(c_path)
            if not isinstance(p, str) or not p.strip():
                continue
            norm = _norm_path(p)
            timestamp_map[norm] = {
                "start": str(row.get(c_start, "") or ""),
                "end": str(row.get(c_end, "") or ""),
                "span_months": _to_int_safe(row.get(c_span)) if c_span else 0,
                "hits": _to_int_safe(row.get(c_hits)) if c_hits else 0,
                "status": str(row.get(c_status, "") or "") if c_status else "",
                "out_of_order": bool(row.get(c_out, False)) if c_out else False,
            }
    except Exception as e:
        print(f"[consolidate] warning: failed reading timestamp report {ts_path}: {e}", flush=True)
    return timestamp_map


def _unique_in_order(items: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def _remux_segment_state_key(path: str, start_s: float, end_s: float) -> str:
    return f"{_norm_path(path)}|{float(start_s):.1f}|{float(end_s):.1f}"


def _load_prior_workflow_and_remux_state(
    *,
    output_path: str | None,
    state_source_path: str | None,
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]], dict[str, list[dict[str, object]]], bool]:
    prior_queue_state: dict[str, dict[str, str]] = {}
    prior_remux_state: dict[str, dict[str, str]] = {}
    prior_remux_state_by_path: dict[str, list[dict[str, object]]] = {}
    done_status_tokens = {"done", "complete", "completed", "archived"}

    def _merge_text_state(existing: dict[str, str] | None, incoming: dict[str, str] | None) -> dict[str, str]:
        out = dict(existing or {})
        for k, v in (incoming or {}).items():
            txt = _safe_text_cell(v)
            if txt or (k not in out):
                out[k] = txt
        return out

    def _coerce_seconds_cell(value) -> float | None:
        try:
            if value is None or (pd.isna(value) if not isinstance(value, str) else False):
                return None
        except Exception:
            pass
        try:
            if isinstance(value, pd.Timedelta):
                return float(value.total_seconds())
        except Exception:
            pass
        try:
            if isinstance(value, datetime.timedelta):
                return float(value.total_seconds())
        except Exception:
            pass
        if isinstance(value, str):
            txt = value.strip()
            if not txt:
                return None
            try:
                td = pd.to_timedelta(txt)
                if not pd.isna(td):
                    return float(td.total_seconds())
            except Exception:
                pass
        num = pd.to_numeric(value, errors="coerce")
        if pd.isna(num):
            return None
        return float(num)

    def _read_prior_queue_state_from_workbook(path: str) -> tuple[dict[str, dict[str, str]], bool]:
        state_local: dict[str, dict[str, str]] = {}
        has_done_local = False
        for sheet_name in ["Rename_Queue", "Rename_Done"]:
            try:
                prior_queue = pd.read_excel(path, sheet_name=sheet_name)
            except Exception:
                continue
            if prior_queue.empty or ("file_path" not in prior_queue.columns):
                continue
            for _, prow in prior_queue.iterrows():
                p = prow.get("file_path")
                if not isinstance(p, str) or not p.strip():
                    continue
                workflow_status = _safe_text_cell(prow.get("workflow_status", ""))
                norm_status = re.sub(r"[^a-z]+", "", workflow_status.strip().lower()) if workflow_status else ""
                if norm_status in done_status_tokens:
                    has_done_local = True
                state_local[_norm_path(p)] = _merge_text_state(state_local.get(_norm_path(p)), {
                    "workflow_status": workflow_status,
                    "proposed_new_name": _safe_text_cell(prow.get("proposed_new_name", "")),
                    "decision": _safe_text_cell(prow.get("decision", "")),
                    "notes": _safe_text_cell(prow.get("notes", "")),
                })
        return state_local, has_done_local

    def _read_prior_remux_state_from_workbook(path: str) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, object]]]]:
        state_local: dict[str, dict[str, str]] = {}
        by_path_local: dict[str, list[dict[str, object]]] = {}
        for sheet_name in ["Remux_Plan", "Remux_Short"]:
            try:
                remux_df = pd.read_excel(path, sheet_name=sheet_name)
            except Exception:
                continue
            if remux_df.empty:
                continue
            cols_norm = {_normalize_col_name(c): c for c in remux_df.columns}
            c_path = cols_norm.get("source_file_path") or cols_norm.get("file_path")
            c_start = cols_norm.get("segment_start_s")
            c_end = cols_norm.get("segment_end_s")
            if not c_path or not c_start or not c_end:
                continue
            for _, prow in remux_df.iterrows():
                p = prow.get(c_path)
                if not isinstance(p, str) or not p.strip():
                    continue
                s = _coerce_seconds_cell(prow.get(c_start))
                e = _coerce_seconds_cell(prow.get(c_end))
                if s is None or e is None:
                    continue
                row_state = {
                    "workflow_status": _safe_text_cell(prow.get("workflow_status", "")),
                    "proposed_output_name": _safe_text_cell(prow.get("proposed_output_name", "")),
                    "notes": _safe_text_cell(prow.get("notes", "")),
                }
                key = _remux_segment_state_key(p, float(s), float(e))
                state_local[key] = _merge_text_state(state_local.get(key), row_state)
                by_path_local.setdefault(_norm_path(p), []).append(
                    {"start_s": float(s), "end_s": float(e), "state": row_state}
                )
        return state_local, by_path_local

    state_paths: list[str] = []
    if output_path:
        state_paths.append(str(output_path))
    if state_source_path:
        src = str(state_source_path)
        if src and (src not in state_paths):
            state_paths.append(src)

    has_done_local = False
    for state_path in state_paths:
        if not os.path.exists(state_path):
            continue
        try:
            state_local, has_done_candidate = _read_prior_queue_state_from_workbook(state_path)
            remux_exact_local, remux_by_path_local = _read_prior_remux_state_from_workbook(state_path)
            if not has_done_candidate:
                try:
                    done_sheet = pd.read_excel(state_path, sheet_name="Rename_Done")
                    has_done_candidate = bool((not done_sheet.empty) and ("file_path" in done_sheet.columns))
                except Exception:
                    pass
            for k, v in state_local.items():
                prior_queue_state[k] = _merge_text_state(prior_queue_state.get(k), v)
            for k, v in remux_exact_local.items():
                prior_remux_state[k] = _merge_text_state(prior_remux_state.get(k), v)
            for k, items in remux_by_path_local.items():
                prior_remux_state_by_path.setdefault(k, []).extend(items)
            if has_done_candidate:
                has_done_local = True
                break
        except Exception:
            continue

    if (not prior_queue_state or not has_done_local) and output_path:
        archive_dir = os.path.join(REPORTS_DIR, "archive", "backups")
        if os.path.isdir(archive_dir):
            stem = os.path.splitext(os.path.basename(str(output_path)))[0]
            archive_candidates = sorted(
                glob.glob(os.path.join(archive_dir, f"{stem}*.xlsx")),
                key=lambda p: os.path.getmtime(p),
                reverse=True,
            )
            seen_paths = {os.path.normcase(os.path.normpath(p)) for p in state_paths}
            fallback_state: dict[str, dict[str, str]] = {}
            for candidate in archive_candidates:
                cand_norm = os.path.normcase(os.path.normpath(candidate))
                if cand_norm in seen_paths:
                    continue
                try:
                    state_local, has_done_candidate = _read_prior_queue_state_from_workbook(candidate)
                    remux_exact_local, remux_by_path_local = _read_prior_remux_state_from_workbook(candidate)
                except Exception:
                    continue
                if not state_local:
                    continue
                if not has_done_candidate:
                    try:
                        done_sheet = pd.read_excel(candidate, sheet_name="Rename_Done")
                        has_done_candidate = bool((not done_sheet.empty) and ("file_path" in done_sheet.columns))
                    except Exception:
                        pass
                if has_done_candidate:
                    for k, v in state_local.items():
                        prior_queue_state[k] = _merge_text_state(prior_queue_state.get(k), v)
                    for k, v in remux_exact_local.items():
                        prior_remux_state[k] = _merge_text_state(prior_remux_state.get(k), v)
                    for k, items in remux_by_path_local.items():
                        prior_remux_state_by_path.setdefault(k, []).extend(items)
                    print(f"[consolidate] restored workflow state from archive: {candidate}", flush=True)
                    has_done_local = True
                    break
                if not fallback_state:
                    fallback_state = state_local
            if (not prior_queue_state) and fallback_state:
                for k, v in fallback_state.items():
                    prior_queue_state[k] = _merge_text_state(prior_queue_state.get(k), v)

    return prior_queue_state, prior_remux_state, prior_remux_state_by_path, has_done_local


def _build_coverage_views(
    *,
    df_out: pd.DataFrame,
    folder_a: str | None,
    folder_b: str | None,
    listed_folder_a_files: list[str],
    listed_folder_b_files: list[str],
    timestamp_map: dict[str, dict],
    coverage_b_s_map: dict[str, float],
    coverage_b_raw_s_map: dict[str, float],
    duration_b_map: dict[str, float],
    coverage_b_sections_raw_map: dict[str, int],
    coverage_b_sections_est_map: dict[str, int],
    coverage_b_unmatched_desc_map: dict[str, str],
    coverage_b_uncertain_desc_map: dict[str, str],
    coverage_a_s_map: dict[str, float],
    coverage_a_raw_s_map: dict[str, float],
    duration_a_map: dict[str, float],
    coverage_a_unique_s_map: dict[str, float],
    coverage_a_unique_duration_map: dict[str, float],
    coverage_a_self_repeat_s_map: dict[str, float],
    coverage_a_sections_raw_map: dict[str, int],
    coverage_a_sections_est_map: dict[str, int],
    coverage_a_unique_sections_est_map: dict[str, int],
    coverage_a_unmatched_desc_map: dict[str, str],
    coverage_a_unique_uncertain_desc_map: dict[str, str],
    get_runtime_basis: Callable[[str], tuple[float, float]],
    coverage_phrase_fn: Callable[..., str],
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame | None,
    pd.DataFrame | None,
    dict[str, list[str]],
    dict[str, list[str]],
    dict[str, bool],
    dict[str, bool],
    dict[str, bool],
    dict[str, bool],
    list[str],
    list[str],
    dict[str, float],
    dict[str, float],
    dict[str, float],
    dict[str, float],
]:
    df_work = df_out.copy()
    df_work["_file_a_norm"] = df_work["file_a"].map(lambda x: _norm_path(x) if isinstance(x, str) else "")
    df_work["_file_b_norm"] = df_work["file_b"].map(lambda x: _norm_path(x) if isinstance(x, str) else "")

    coverage_b_s_norm = {_norm_path(k): float(v) for k, v in coverage_b_s_map.items()}
    coverage_b_raw_s_norm = {_norm_path(k): float(v) for k, v in coverage_b_raw_s_map.items()}
    duration_b_norm = {_norm_path(k): float(v) for k, v in duration_b_map.items()}
    coverage_b_sections_raw_norm = {_norm_path(k): int(v) for k, v in coverage_b_sections_raw_map.items()}
    coverage_b_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_b_sections_est_map.items()}
    coverage_b_unmatched_desc_norm = {_norm_path(k): str(v) for k, v in coverage_b_unmatched_desc_map.items()}
    coverage_b_uncertain_desc_norm = {_norm_path(k): str(v) for k, v in coverage_b_uncertain_desc_map.items()}

    coverage_a_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_s_map.items()}
    coverage_a_raw_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_raw_s_map.items()}
    duration_a_norm = {_norm_path(k): float(v) for k, v in duration_a_map.items()}
    coverage_a_unique_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_unique_s_map.items()}
    coverage_a_unique_duration_norm = {_norm_path(k): float(v) for k, v in coverage_a_unique_duration_map.items()}
    coverage_a_self_repeat_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_self_repeat_s_map.items()}
    coverage_a_sections_raw_norm = {_norm_path(k): int(v) for k, v in coverage_a_sections_raw_map.items()}
    coverage_a_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_a_sections_est_map.items()}
    coverage_a_unique_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_a_unique_sections_est_map.items()}
    coverage_a_unmatched_desc_norm = {_norm_path(k): str(v) for k, v in coverage_a_unmatched_desc_map.items()}
    coverage_a_unique_uncertain_desc_norm = {_norm_path(k): str(v) for k, v in coverage_a_unique_uncertain_desc_map.items()}

    ts_start_norm = {k: str(v.get("start", "") or "") for k, v in timestamp_map.items()}
    ts_end_norm = {k: str(v.get("end", "") or "") for k, v in timestamp_map.items()}
    ts_span_norm = {k: int(v.get("span_months", 0) or 0) for k, v in timestamp_map.items()}
    ts_hits_norm = {k: int(v.get("hits", 0) or 0) for k, v in timestamp_map.items()}
    ts_out_order_norm = {k: bool(v.get("out_of_order", False)) for k, v in timestamp_map.items()}

    b_group = df_work.groupby("_file_b_norm", dropna=True)
    b_match_count = b_group["file_a"].nunique().to_dict()
    b_best_conf = b_group["confidence_score"].max().to_dict()
    b_avg_conf = b_group["confidence_score"].mean().to_dict()
    b_display = b_group["file_b"].first().to_dict()

    a_group = df_work.groupby("_file_a_norm", dropna=True)
    a_match_count = a_group["file_b"].nunique().to_dict()
    a_best_conf = a_group["confidence_score"].max().to_dict()
    a_avg_conf = a_group["confidence_score"].mean().to_dict()
    a_display = a_group["file_a"].first().to_dict()

    b_matches_norm: dict[str, list[str]] = {}
    a_matches_norm: dict[str, list[str]] = {}
    b_multimatch_norm: dict[str, bool] = {}
    b_internal_gap_norm: dict[str, bool] = {}
    a_multimatch_norm: dict[str, bool] = {}
    a_internal_gap_norm: dict[str, bool] = {}
    b_sorted_for_names = df_work.sort_values(
        by=["_file_b_norm", "file_b_part_order", "timeline_est_b_start_s", "confidence_score", "_file_a_norm"],
        ascending=[True, True, True, False, True],
    )
    a_sorted_for_names = df_work.sort_values(
        by=["_file_a_norm", "timeline_est_a_start_s", "confidence_score", "_file_b_norm"],
        ascending=[True, True, False, True],
    )
    for norm, grp in b_sorted_for_names.groupby("_file_b_norm", dropna=True):
        if isinstance(norm, str) and norm:
            file_as = [str(v) for v in grp["file_a"].tolist() if isinstance(v, str) and str(v)]
            b_matches_norm[norm] = [os.path.basename(p) for p in _unique_in_order(file_as)]
            b_multimatch_norm[norm] = bool(grp["timeline_b_overlap_multimatch"].fillna(False).astype(bool).any())
            b_internal_gap_norm[norm] = bool(
                grp["timeline_b_has_internal_unmatched_gaps"].fillna(False).astype(bool).any()
            )
    for norm, grp in a_sorted_for_names.groupby("_file_a_norm", dropna=True):
        if isinstance(norm, str) and norm:
            file_bs = [str(v) for v in grp["file_b"].tolist() if isinstance(v, str) and str(v)]
            a_matches_norm[norm] = [os.path.basename(p) for p in _unique_in_order(file_bs)]
            a_multimatch_norm[norm] = bool(grp["timeline_a_overlap_multimatch"].fillna(False).astype(bool).any())
            a_internal_gap_norm[norm] = bool(
                grp["timeline_a_has_internal_unmatched_gaps"].fillna(False).astype(bool).any()
            )

    all_b_files = list(listed_folder_b_files) if folder_b else sorted(str(v) for v in b_display.values() if isinstance(v, str))
    all_a_files = list(listed_folder_a_files) if folder_a else sorted(str(v) for v in a_display.values() if isinstance(v, str))

    b_coverage_rows = []
    for path in all_b_files:
        norm = _norm_path(path)
        _raw_runtime_s, playable_runtime_s = get_runtime_basis(path)
        dur_s = float(duration_b_norm.get(norm, playable_runtime_s))
        cov_est_s = float(coverage_b_s_norm.get(norm, 0.0))
        cov_raw_s = float(coverage_b_raw_s_norm.get(norm, 0.0))
        cov_est_pct = (100.0 * cov_est_s / dur_s) if dur_s > 0 else 0.0
        cov_raw_pct = (100.0 * cov_raw_s / dur_s) if dur_s > 0 else 0.0
        cov_sections_est = int(coverage_b_sections_est_norm.get(norm, 0))
        b_coverage_rows.append(
            {
                "file_b": path,
                "osd_month_year_start": _safe_text_cell(ts_start_norm.get(norm, "")),
                "osd_month_year_end": _safe_text_cell(ts_end_norm.get(norm, "")),
                "osd_span_months": int(ts_span_norm.get(norm, 0)),
                "osd_hits": int(ts_hits_norm.get(norm, 0)),
                "osd_out_of_order": bool(ts_out_order_norm.get(norm, False)),
                "matched_file_a_count": int(b_match_count.get(norm, 0)),
                "matched_file_a_names": " | ".join(b_matches_norm.get(norm, [])),
                "best_pair_confidence_score": float(b_best_conf.get(norm, 0.0)),
                "avg_pair_confidence_score": float(b_avg_conf.get(norm, 0.0)),
                "timeline_coverage_est_pct": round(max(0.0, min(100.0, cov_est_pct)), 1),
                "timeline_coverage_raw_pct": round(max(0.0, min(100.0, cov_raw_pct)), 1),
                "timeline_coverage_est_s": round(cov_est_s, 1),
                "timeline_coverage_raw_s": round(cov_raw_s, 1),
                "timeline_duration_est_s": round(dur_s, 1),
                "timeline_coverage_sections_est": int(cov_sections_est),
                "timeline_coverage_sections_raw": int(coverage_b_sections_raw_norm.get(norm, 0)),
                "coverage_summary": coverage_phrase_fn(
                    "file_b",
                    float(max(0.0, min(100.0, cov_est_pct))),
                    int(b_match_count.get(norm, 0)),
                    int(cov_sections_est),
                    str(coverage_b_unmatched_desc_norm.get(norm, "")),
                    str(coverage_b_uncertain_desc_norm.get(norm, "")),
                    covered_s=float(cov_est_s),
                    duration_s=float(dur_s),
                ),
            }
        )
    b_coverage_df = pd.DataFrame(b_coverage_rows, columns=B_COVERAGE_SHEET_COLUMNS).sort_values(
        by=["timeline_coverage_est_pct", "best_pair_confidence_score", "matched_file_a_count", "file_b"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)

    a_coverage_rows = []
    for path in all_a_files:
        norm = _norm_path(path)
        _raw_runtime_s, playable_runtime_s = get_runtime_basis(path)
        dur_s = float(duration_a_norm.get(norm, playable_runtime_s))
        cov_est_s = float(coverage_a_s_norm.get(norm, 0.0))
        cov_raw_s = float(coverage_a_raw_s_norm.get(norm, 0.0))
        unique_cov_est_s = float(coverage_a_unique_s_norm.get(norm, 0.0))
        unique_dur_s = float(coverage_a_unique_duration_norm.get(norm, max(0.0, playable_runtime_s)))
        repeat_s = float(coverage_a_self_repeat_s_norm.get(norm, 0.0))
        cov_est_pct = (100.0 * cov_est_s / dur_s) if dur_s > 0 else 0.0
        cov_raw_pct = (100.0 * cov_raw_s / dur_s) if dur_s > 0 else 0.0
        unique_cov_est_pct = (100.0 * unique_cov_est_s / unique_dur_s) if unique_dur_s > 0 else 0.0
        repeat_pct = (100.0 * repeat_s / dur_s) if dur_s > 0 else 0.0
        cov_sections_est = int(coverage_a_sections_est_norm.get(norm, 0))
        a_coverage_rows.append(
            {
                "file_a": path,
                "osd_month_year_start": _safe_text_cell(ts_start_norm.get(norm, "")),
                "osd_month_year_end": _safe_text_cell(ts_end_norm.get(norm, "")),
                "osd_span_months": int(ts_span_norm.get(norm, 0)),
                "osd_hits": int(ts_hits_norm.get(norm, 0)),
                "osd_out_of_order": bool(ts_out_order_norm.get(norm, False)),
                "matched_file_b_count": int(a_match_count.get(norm, 0)),
                "matched_file_b_names": " | ".join(a_matches_norm.get(norm, [])),
                "best_pair_confidence_score": float(a_best_conf.get(norm, 0.0)),
                "avg_pair_confidence_score": float(a_avg_conf.get(norm, 0.0)),
                "timeline_a_unique_covered_est_pct": round(max(0.0, min(100.0, unique_cov_est_pct)), 1),
                "timeline_a_unique_covered_est_s": round(unique_cov_est_s, 1),
                "timeline_a_unique_duration_est_s": round(unique_dur_s, 1),
                "timeline_a_self_repeat_redundant_pct": round(max(0.0, min(100.0, repeat_pct)), 1),
                "timeline_a_self_repeat_redundant_s": round(repeat_s, 1),
                "timeline_a_unique_coverage_sections_est": int(coverage_a_unique_sections_est_norm.get(norm, 0)),
                "timeline_coverage_est_pct": round(max(0.0, min(100.0, cov_est_pct)), 1),
                "timeline_coverage_raw_pct": round(max(0.0, min(100.0, cov_raw_pct)), 1),
                "timeline_coverage_est_s": round(cov_est_s, 1),
                "timeline_coverage_raw_s": round(cov_raw_s, 1),
                "timeline_duration_est_s": round(dur_s, 1),
                "timeline_coverage_sections_est": int(cov_sections_est),
                "timeline_coverage_sections_raw": int(coverage_a_sections_raw_norm.get(norm, 0)),
                "coverage_summary": coverage_phrase_fn(
                    "file_a unique content",
                    float(max(0.0, min(100.0, unique_cov_est_pct))),
                    int(a_match_count.get(norm, 0)),
                    int(coverage_a_unique_sections_est_norm.get(norm, 0)),
                    str(coverage_a_unmatched_desc_norm.get(norm, "")),
                    str(coverage_a_unique_uncertain_desc_norm.get(norm, "")),
                    covered_s=float(unique_cov_est_s),
                    duration_s=float(unique_dur_s),
                ),
            }
        )
    a_coverage_df = pd.DataFrame(a_coverage_rows, columns=A_COVERAGE_SHEET_COLUMNS).sort_values(
        by=["timeline_a_unique_covered_est_pct", "best_pair_confidence_score", "matched_file_b_count", "file_a"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)

    unmatched_a_df = None
    if folder_a:
        unmatched_a_df = (
            a_coverage_df[a_coverage_df["matched_file_b_count"] <= 0][["file_a"]]
            .rename(columns={"file_a": "unmatched_in_folder_a"})
            .reset_index(drop=True)
        )

    unmatched_b_df = None
    if folder_b:
        unmatched_b_df = (
            b_coverage_df[b_coverage_df["matched_file_a_count"] <= 0][["file_b"]]
            .rename(columns={"file_b": "unmatched_in_folder_b"})
            .reset_index(drop=True)
        )

    return (
        b_coverage_df,
        a_coverage_df,
        unmatched_a_df,
        unmatched_b_df,
        b_matches_norm,
        a_matches_norm,
        b_multimatch_norm,
        b_internal_gap_norm,
        a_multimatch_norm,
        a_internal_gap_norm,
        all_b_files,
        all_a_files,
        b_best_conf,
        a_best_conf,
        b_avg_conf,
        a_avg_conf,
    )


def consolidate_dedupe_reports(
    report_paths: list[str] | None = None,
    *,
    report_dfs: list[tuple[str, pd.DataFrame]] | None = None,
    folder_a: str | None = None,
    folder_b: str | None = None,
    output_path: str | None = None,
    state_source_path: str | None = None,
    timestamp_report_path: str | None = None,
    metrics_whitelist: dict[str, list[str]] | None = None,
    mode_weights: dict[str, float] | None = None,
) -> pd.DataFrame:
    """
    Consolidate multiple dedupe reports into one file with per-mode flags.

    - report_paths: list of Excel report paths. If None, uses canonical report names
      (anchors/legacy/audio/timeline) when present.
    - report_dfs: optional list of (mode_name, DataFrame) to consolidate without reading Excel.
    - folder_a/folder_b: optional folder paths used to canonicalize A/B ordering
      and to generate unmatched lists.
    - timestamp_report_path: optional per-file OSD timestamp report path
      (from scan_video_timestamps). If present, adds date-range columns for file_a/file_b.
    - metrics_whitelist: optional {mode: [col, ...]} to keep only selected metrics.
    - mode_weights: optional {mode: weight} for confidence scoring (default 1.0 each).

    Output Excel has sheets:
    - Rename_Queue (primary workflow; grouped file rows with separators)
    - Rename_Done (optional; rows marked done/completed/archived on prior run)
    - Remux_Plan (A-side unique segments above threshold, for preservation/remux planning)
    - Consolidated (pair-level detail for deeper inspection)
    - A_Coverage (all folder-A files, counterpart file-B names, and coverage metrics)
    - B_Coverage (all folder-B files, counterpart file-A names, and coverage metrics)
    - Unmatched_A / Unmatched_B (when folders provided)
    Returns the consolidated DataFrame.
    """
    if output_path is None:
        output_path = os.path.join(REPORTS_DIR, "dedupe_consolidated.xlsx")

    # Track mode names and merged rows keyed by canonicalized paths.
    mode_names: list[str] = []
    rows_by_key: dict[tuple[str, str], dict] = {}
    display_by_key: dict[tuple[str, str], tuple[str, str]] = {}

    def _unique_mode_name(base: str) -> str:
        mode = base
        n = 2
        while mode in mode_names:
            mode = f"{base}_{n}"
            n += 1
        mode_names.append(mode)
        return mode

    def _ingest_df(mode: str, df: pd.DataFrame, source_label: str) -> None:
        if df.empty:
            print(f"[consolidate] Empty report: {source_label}")
            return

        cols_norm = {_normalize_col_name(c): c for c in df.columns}
        col_a = cols_norm.get("file_a") or cols_norm.get("file_a_path") or cols_norm.get("file_a_full")
        col_b = cols_norm.get("file_b") or cols_norm.get("file_b_path") or cols_norm.get("file_b_full")
        if not col_a or not col_b:
            raise ValueError(f"Report {source_label} missing file_a/file_b columns.")

        for _, row in df.iterrows():
            file_a = row.get(col_a)
            file_b = row.get(col_b)
            if not isinstance(file_a, str) or not isinstance(file_b, str):
                continue

            a, b = _canonicalize_pair(file_a, file_b, folder_a, folder_b)
            key = (os.path.normcase(os.path.normpath(a)), os.path.normcase(os.path.normpath(b)))

            if key not in rows_by_key:
                rows_by_key[key] = {"file_a": a, "file_b": b}
                display_by_key[key] = (a, b)

            rows_by_key[key][f"matched_by_{mode}"] = True

            # Carry over metrics (mode-prefixed) when present.
            if metrics_whitelist is not None:
                allowed = set(metrics_whitelist.get(mode, []))
            else:
                allowed = None
            for col in df.columns:
                if col in (col_a, col_b):
                    continue
                if allowed is not None and col not in allowed:
                    continue
                val = row.get(col)
                if pd.isna(val):
                    continue
                out_col = f"{mode}__{col}"
                if out_col not in rows_by_key[key]:
                    rows_by_key[key][out_col] = val

    any_sources = False

    if report_dfs:
        for mode_label, df in report_dfs:
            mode = _unique_mode_name(_sanitize_mode_name(mode_label))
            _ingest_df(mode, df, mode_label)
            any_sources = True

    if report_paths is None:
        report_paths = [
            os.path.join(REPORTS_DIR, "dedupe_anchors.xlsx"),
            os.path.join(REPORTS_DIR, "dedupe_legacy.xlsx"),
            os.path.join(REPORTS_DIR, "dedupe_audio.xlsx"),
            os.path.join(REPORTS_DIR, "dedupe_timeline.xlsx"),
        ]
    if report_paths:
        for requested_path in report_paths:
            resolved_path, used_archive = _resolve_report_path(requested_path)
            if not resolved_path:
                print(f"[consolidate] Skipping missing report: {requested_path}")
                continue
            if used_archive:
                print(
                    f"[consolidate] Using archived fallback for {requested_path}: {resolved_path}",
                    flush=True,
                )
            # Infer mode from the requested canonical path so archived timestamped
            # filenames still map to stable mode labels (e.g., timeline/anchors).
            mode = _unique_mode_name(_infer_mode_from_path(requested_path))
            df = pd.read_excel(resolved_path)
            _ingest_df(mode, df, resolved_path)
            any_sources = True

    if not any_sources:
        raise FileNotFoundError(
            "No report paths/DataFrames provided and no canonical dedupe reports found in reports/."
        )

    # Build consolidated DataFrame with missing flags defaulting to False.
    records = []
    for key, rec in rows_by_key.items():
        rec_out = dict(rec)
        for mode in mode_names:
            rec_out.setdefault(f"matched_by_{mode}", False)
        modes_list = [m for m in mode_names if rec_out.get(f"matched_by_{m}")]
        rec_out["modes"] = ", ".join(modes_list)
        records.append(rec_out)

    df_out = pd.DataFrame(records)
    if df_out.empty:
        print("[consolidate] No rows to export.")
        return df_out

    def _norm_metric_key(name: str) -> str:
        return "".join(ch if ch.isalnum() else "_" for ch in name.lower())

    def _metric_map_for_mode(mode: str) -> dict[str, str]:
        prefix = f"{mode}__"
        out = {}
        for col in df_out.columns:
            if col.startswith(prefix):
                key = _norm_metric_key(col[len(prefix):])
                out[key] = col
        return out

    def _series_from(df: pd.DataFrame, col: str) -> pd.Series:
        s = df[col]
        try:
            if pd.api.types.is_timedelta64_dtype(s):
                return pd.Series(s.dt.total_seconds(), index=s.index, dtype=float)
        except Exception:
            pass
        try:
            if pd.api.types.is_object_dtype(s):
                sample = s.dropna().head(5).tolist()
                if sample and all(hasattr(v, "total_seconds") for v in sample):
                    return pd.Series(
                        [float(v.total_seconds()) if hasattr(v, "total_seconds") else np.nan for v in s],
                        index=s.index,
                        dtype=float,
                    )
        except Exception:
            pass
        return pd.to_numeric(s, errors="coerce")

    def _clip01(s: pd.Series) -> pd.Series:
        return s.clip(lower=0.0, upper=1.0)

    # Optional OSD timestamp enrichment (per-file metadata, independent of match mode).
    timestamp_map = _load_timestamp_metadata_from_report(timestamp_report_path)
    if timestamp_map:
        print(f"[consolidate] loaded timestamp metadata for {len(timestamp_map)} file(s)", flush=True)

    timestamp_candidates_cache: dict[str, list[dict]] = {}
    vhs_ts = VideoHashStore()

    def _format_hms_compact(seconds: float) -> str:
        total = max(0, int(round(float(seconds))))
        h = total // 3600
        m = (total % 3600) // 60
        s = total % 60
        return f"{h:d}:{m:02d}:{s:02d}"

    def _get_timestamp_candidates(path: str) -> list[dict]:
        norm = _norm_path(path)
        cached = timestamp_candidates_cache.get(norm)
        if cached is not None:
            return cached
        try:
            entry = vhs_ts._data.get(path) or {}
            osd_entry = entry.get("osd_dates") if isinstance(entry, dict) else {}
            candidates = ((osd_entry or {}).get("data") or {}).get("candidates", []) or []
            out = [c for c in candidates if isinstance(c, dict)]
        except Exception:
            out = []
        timestamp_candidates_cache[norm] = out
        return out

    def _summarize_timestamp_interval(path: str, start_s: float, end_s: float) -> tuple[str, str]:
        if not isinstance(path, str) or not path.strip():
            return "", ""
        lo = max(0.0, float(start_s))
        hi = max(lo, float(end_s))
        candidates = []
        for cand in _get_timestamp_candidates(path):
            t_s = pd.to_numeric(cand.get("t_s"), errors="coerce")
            if pd.isna(t_s):
                continue
            t_f = float(t_s)
            if lo <= t_f <= hi:
                candidates.append(cand)
        if not candidates:
            return "", ""
        summary = _summarize_osd_candidates(candidates)
        return (
            str(summary.get("osd_month_year_start", "") or ""),
            str(summary.get("osd_month_year_end", "") or ""),
        )

    def _map_trimmed_rel_to_abs(
        rel_intervals: list[tuple[float, float]],
        trim_head_s: float,
        dead_regions_abs: list[tuple[float, float]] | None = None,
    ) -> list[tuple[float, float]]:
        """
        Map trim-head-relative intervals back to source-file time.

        If internal dead regions were removed from the playable basis, the relative
        timeline has those holes collapsed out. Reinsert them when mapping back to
        source time so remux/snippet intervals line up with the original file.
        """
        head = max(0.0, float(trim_head_s))
        dead_sorted = _merge_intervals_simple(list(dead_regions_abs or []), gap_tolerance_s=0.0)

        def _map_point(rel_t: float) -> float:
            abs_t = head + float(rel_t)
            for ds, de in dead_sorted:
                ds_f = float(ds)
                de_f = float(de)
                if de_f <= head:
                    continue
                insert_at = max(head, ds_f)
                if abs_t >= insert_at:
                    abs_t += max(0.0, de_f - insert_at)
            return abs_t

        out: list[tuple[float, float]] = []
        for s, e in rel_intervals:
            s_f = float(s)
            e_f = float(e)
            if e_f <= s_f:
                continue
            out.append((_map_point(s_f), _map_point(e_f)))
        return _merge_intervals(out, gap_tolerance_s=0.0)

    # Confidence scoring:
    # - per-mode strengths stay in [0,1]
    # - evidence_score = weighted average over matched modes only (quality of evidence)
    # - consensus_score = fraction of total mode-weight that matched (cross-technique agreement)
    # - confidence_score = blend of quality + agreement
    matched_cols = [f"matched_by_{m}" for m in mode_names]
    if matched_cols:
        weights = {m: 1.0 for m in mode_names}
        if mode_weights:
            for m, w in mode_weights.items():
                if m in weights:
                    try:
                        weights[m] = float(w)
                    except Exception:
                        pass
        total_w = sum(max(0.0, weights[m]) for m in mode_names)
        if total_w <= 0:
            total_w = float(len(mode_names))
            weights = {m: 1.0 for m in mode_names}

        df_out["matched_by_count"] = df_out[matched_cols].sum(axis=1).astype(int)

        strength_cols: dict[str, pd.Series] = {}

        def _ratio_score(series: pd.Series, pivot: float) -> pd.Series:
            p = max(1e-6, float(pivot))
            s = pd.to_numeric(series, errors="coerce").fillna(0.0)
            return _clip01(s / (s + p))

        for mode in mode_names:
            mcols = _metric_map_for_mode(mode)
            strength = None

            # Anchors-style metrics.
            if ("start_match_fraction" in mcols) or ("end_match_fraction" in mcols):
                start = (_series_from(df_out, mcols.get("start_match_fraction", ""))
                         if "start_match_fraction" in mcols else pd.Series(np.nan, index=df_out.index))
                end = (_series_from(df_out, mcols.get("end_match_fraction", ""))
                       if "end_match_fraction" in mcols else pd.Series(np.nan, index=df_out.index))
                frac = pd.concat([start, end], axis=1).max(axis=1)

                mad_start = (_series_from(df_out, mcols.get("start_offset_mad_s", ""))
                             if "start_offset_mad_s" in mcols else pd.Series(np.nan, index=df_out.index))
                mad_end = (_series_from(df_out, mcols.get("end_offset_mad_s", ""))
                           if "end_offset_mad_s" in mcols else pd.Series(np.nan, index=df_out.index))
                mad = pd.concat([mad_start, mad_end], axis=1).min(axis=1)

                mad_score = pd.Series(1.0, index=df_out.index)
                mad_score = mad_score.where(mad.isna(), 1.0 - (mad / max(1.0, ANCHOR_MAX_MAD_S * 2.0)))
                frac_score = _clip01(frac.fillna(0.0))
                strength = _clip01(0.8 * frac_score + 0.2 * _clip01(mad_score.fillna(1.0)))

            # Timeline-style metrics.
            if ("timeline_votes" in mcols) or ("timeline_overlap_s" in mcols) or ("overlap_strict_s" in mcols):
                votes = (_series_from(df_out, mcols.get("timeline_votes", ""))
                         if "timeline_votes" in mcols else pd.Series(np.nan, index=df_out.index))
                overlap_col = (
                    mcols.get("timeline_overlap_strict_s")
                    or mcols.get("overlap_strict_s")
                    or mcols.get("timeline_overlap_s")
                )
                overlap = (_series_from(df_out, overlap_col)
                           if overlap_col else pd.Series(np.nan, index=df_out.index))
                votes_score = _ratio_score(votes, max(600.0, TIMELINE_MIN_VOTES * 20.0))
                overlap_score = _ratio_score(overlap, max(90.0, TIMELINE_MIN_OVERLAP_S * 0.8))
                tl_strength = _clip01(0.55 * votes_score + 0.45 * overlap_score)
                strength = tl_strength if strength is None else _clip01(pd.concat([strength, tl_strength], axis=1).max(axis=1))

            # Audio-style metrics.
            if ("audio_votes" in mcols) or ("audio_overlap_s" in mcols):
                votes = (_series_from(df_out, mcols.get("audio_votes", ""))
                         if "audio_votes" in mcols else pd.Series(np.nan, index=df_out.index))
                overlap = (_series_from(df_out, mcols.get("audio_overlap_s", ""))
                           if "audio_overlap_s" in mcols else pd.Series(np.nan, index=df_out.index))
                votes_score = _ratio_score(votes, max(900.0, AUDIO_MIN_VOTES * 120.0))
                overlap_score = _ratio_score(overlap, max(70.0, AUDIO_MIN_OVERLAP_S * 1.2))
                au_strength = _clip01(0.55 * votes_score + 0.45 * overlap_score)
                strength = au_strength if strength is None else _clip01(pd.concat([strength, au_strength], axis=1).max(axis=1))

            # Legacy-style metrics.
            legacy_key = "legacy_best_aligned_diff_0_64"
            avg_key = "avg_frame_diff_0_64"
            legacy_col = mcols.get(legacy_key) or mcols.get(avg_key)
            if legacy_col:
                diff = _series_from(df_out, legacy_col)
                leg_strength = _clip01(1.0 - (diff / 64.0))
                strength = leg_strength if strength is None else _clip01(pd.concat([strength, leg_strength], axis=1).max(axis=1))

            # Fallback: if matched but no usable metrics, treat as 1.0.
            if strength is None:
                strength = df_out[f"matched_by_{mode}"].astype(float)
            else:
                strength = strength.where(df_out[f"matched_by_{mode}"], 0.0)

            strength_cols[mode] = strength
            df_out[f"strength_{mode}"] = strength.round(3)

        weighted_strength = sum(strength_cols[m] * max(0.0, weights[m]) for m in mode_names)
        weighted_presence = sum(
            df_out[f"matched_by_{m}"].astype(float) * max(0.0, weights[m]) for m in mode_names
        )
        weighted_presence_safe = weighted_presence.replace(0.0, np.nan)
        evidence_score = (weighted_strength / weighted_presence_safe).fillna(0.0)
        consensus_score = _clip01(weighted_presence / total_w)
        confidence_score = _clip01(0.70 * evidence_score + 0.30 * consensus_score)

        df_out["evidence_score"] = (100.0 * evidence_score).round(1)
        df_out["consensus_score"] = (100.0 * consensus_score).round(1)
        df_out["confidence_score"] = (100.0 * confidence_score).round(1)
    else:
        df_out["matched_by_count"] = 0
        df_out["evidence_score"] = 0.0
        df_out["consensus_score"] = 0.0
        df_out["confidence_score"] = 0.0

    # Anchor full<->full support is useful, but not sufficient on its own:
    # start/end can still match while the middle differs. Use it only to decide
    # which timeline-partial rows deserve targeted A-side content recovery.
    anchor_full_full_pre = pd.Series(False, index=df_out.index, dtype=bool)
    for mode in mode_names:
        rel_col = _metric_map_for_mode(mode).get("relation_dir")
        if not rel_col:
            continue
        rel = (
            df_out[rel_col]
            .fillna("")
            .astype(str)
            .str.replace(" ", "", regex=False)
            .str.replace("↔", "<->", regex=False)
            .str.replace("â†”", "<->", regex=False)
            .str.lower()
        )
        matched = df_out[f"matched_by_{mode}"].astype(bool)
        strength = pd.to_numeric(
            df_out.get(f"strength_{mode}", pd.Series(0.0, index=df_out.index)),
            errors="coerce",
        ).fillna(0.0)
        anchor_full_full_pre = anchor_full_full_pre | (matched & rel.eq("full<->full") & (strength >= 0.90))

    def _map_timestamp_field(path_series: pd.Series, field: str, *, default):
        if not timestamp_map:
            return pd.Series([default] * len(path_series), index=path_series.index)
        return path_series.map(
            lambda p: timestamp_map.get(_norm_path(p), {}).get(field, default) if isinstance(p, str) else default
        )

    file_a_series_ts = df_out["file_a"].astype(str)
    file_b_series_ts = df_out["file_b"].astype(str)
    df_out["file_a_osd_month_year_start"] = _map_timestamp_field(file_a_series_ts, "start", default="").astype(str)
    df_out["file_a_osd_month_year_end"] = _map_timestamp_field(file_a_series_ts, "end", default="").astype(str)
    df_out["file_a_osd_span_months"] = pd.to_numeric(
        _map_timestamp_field(file_a_series_ts, "span_months", default=0), errors="coerce"
    ).fillna(0).astype(int)
    df_out["file_a_osd_hits"] = pd.to_numeric(
        _map_timestamp_field(file_a_series_ts, "hits", default=0), errors="coerce"
    ).fillna(0).astype(int)
    df_out["file_a_osd_status"] = _map_timestamp_field(file_a_series_ts, "status", default="").astype(str)
    df_out["file_a_osd_out_of_order"] = _map_timestamp_field(file_a_series_ts, "out_of_order", default=False).astype(bool)

    df_out["file_b_osd_month_year_start"] = _map_timestamp_field(file_b_series_ts, "start", default="").astype(str)
    df_out["file_b_osd_month_year_end"] = _map_timestamp_field(file_b_series_ts, "end", default="").astype(str)
    df_out["file_b_osd_span_months"] = pd.to_numeric(
        _map_timestamp_field(file_b_series_ts, "span_months", default=0), errors="coerce"
    ).fillna(0).astype(int)
    df_out["file_b_osd_hits"] = pd.to_numeric(
        _map_timestamp_field(file_b_series_ts, "hits", default=0), errors="coerce"
    ).fillna(0).astype(int)
    df_out["file_b_osd_status"] = _map_timestamp_field(file_b_series_ts, "status", default="").astype(str)
    df_out["file_b_osd_out_of_order"] = _map_timestamp_field(file_b_series_ts, "out_of_order", default=False).astype(bool)

    # Group-level review helpers.
    df_out["group_candidate_count"] = df_out.groupby("file_b")["file_a"].transform("size").astype(int)
    df_out["group_best_confidence"] = df_out.groupby("file_b")["confidence_score"].transform("max")
    df_out["group_second_confidence"] = df_out.groupby("file_b")["confidence_score"].transform(
        lambda s: float(s.nlargest(2).iloc[-1]) if len(s) >= 2 else np.nan
    )
    df_out["group_confidence_gap"] = (
        df_out["group_best_confidence"] - df_out["group_second_confidence"]
    ).fillna(df_out["group_best_confidence"])
    # Review flags to prioritize manual checks.
    review_flags = pd.Series([""] * len(df_out), index=df_out.index, dtype=object)

    def _add_review_flag(mask: pd.Series, label: str) -> None:
        if not bool(mask.any()):
            return
        current = review_flags.loc[mask]
        review_flags.loc[mask] = np.where(current == "", label, current + ";" + label)

    df_out["confidence_tier"] = pd.cut(
        df_out["confidence_score"],
        bins=[-1.0, 40.0, 60.0, 80.0, 100.0],
        labels=["low", "medium", "high", "very_high"],
    ).astype(str)

    # Timeline-based file_b coverage estimate.
    # This estimates how much of each file_b timeline is covered by matched file_a segments,
    # using timeline offset + overlap to place intervals on file_b and unioning them.
    def _union_interval_length(intervals: list[tuple[float, float]], *, gap_tolerance_s: float = 0.0) -> float:
        if not intervals:
            return 0.0
        segs = sorted((float(a), float(b)) for a, b in intervals if (b > a))
        if not segs:
            return 0.0
        cur_s, cur_e = segs[0]
        total = 0.0
        gap_tol = max(0.0, float(gap_tolerance_s))
        for s, e in segs[1:]:
            if s <= (cur_e + gap_tol):
                if e > cur_e:
                    cur_e = e
            else:
                total += max(0.0, cur_e - cur_s)
                cur_s, cur_e = s, e
        total += max(0.0, cur_e - cur_s)
        return total

    def _merge_intervals(intervals: list[tuple[float, float]], *, gap_tolerance_s: float = 0.0) -> list[tuple[float, float]]:
        if not intervals:
            return []
        segs = sorted((float(a), float(b)) for a, b in intervals if (b > a))
        if not segs:
            return []
        out: list[tuple[float, float]] = []
        cur_s, cur_e = segs[0]
        gap_tol = max(0.0, float(gap_tolerance_s))
        for s, e in segs[1:]:
            if s <= (cur_e + gap_tol):
                if e > cur_e:
                    cur_e = e
            else:
                out.append((cur_s, cur_e))
                cur_s, cur_e = s, e
        out.append((cur_s, cur_e))
        return out

    timeline_mode_infos: list[tuple[str, str, str, str | None, str | None]] = []
    for mode in mode_names:
        mcols = _metric_map_for_mode(mode)
        col_off = mcols.get("timeline_offset_s")
        col_ov = mcols.get("timeline_overlap_s")
        if not col_off or not col_ov:
            continue
        col_dur_a = mcols.get("duration_a__s_")
        if not col_dur_a:
            for k, c in mcols.items():
                if k.startswith("duration_a"):
                    col_dur_a = c
                    break
        col_dur_b = mcols.get("duration_b__s_")
        if not col_dur_b:
            for k, c in mcols.items():
                if k.startswith("duration_b"):
                    col_dur_b = c
                    break
        timeline_mode_infos.append((mode, col_off, col_ov, col_dur_a, col_dur_b))

    # Row-level timeline interval estimate (best available timeline mode per row).
    est_start = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_end = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_overlap = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_overlap_raw = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_dur_b = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_start_a = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_end_a = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_dur_a = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_tl_votes = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_tl_offset = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_tl_unique_min = pd.Series(np.nan, index=df_out.index, dtype=float)
    est_tl_drift = pd.Series(1.0, index=df_out.index, dtype=float)
    est_tl_mode = pd.Series("", index=df_out.index, dtype=object)
    est_tl_segments_a_raw = pd.Series("", index=df_out.index, dtype=object)
    est_tl_segments_b_raw = pd.Series("", index=df_out.index, dtype=object)
    est_tl_segments_a = pd.Series("", index=df_out.index, dtype=object)
    est_tl_segments_b = pd.Series("", index=df_out.index, dtype=object)
    chosen_overlap = pd.Series(-1.0, index=df_out.index, dtype=float)
    any_timeline_match = pd.Series(False, index=df_out.index, dtype=bool)

    for mode, col_off, col_ov, col_dur_a, col_dur_b in timeline_mode_infos:
        mode_mask = df_out[f"matched_by_{mode}"].astype(bool)
        any_timeline_match = any_timeline_match | mode_mask
        mcols = _metric_map_for_mode(mode)
        off = _series_from(df_out, col_off)
        ov = _series_from(df_out, col_ov).clip(lower=0.0)
        overlap_raw_col = (
            mcols.get("timeline_overlap_strict_unbridged_s")
            or mcols.get("overlap_strict_unbridged_s")
            or mcols.get("timeline_overlap_raw_s")
            or mcols.get("overlap_raw_s")
            or col_ov
        )
        ov_raw = _series_from(df_out, overlap_raw_col).clip(lower=0.0)
        # Raw cannot exceed final selected overlap for a row.
        ov_raw = pd.Series(
            np.minimum(ov_raw.fillna(0.0), ov.fillna(0.0)),
            index=df_out.index,
            dtype=float,
        )
        if col_dur_a:
            dur_a_s = _series_from(df_out, col_dur_a)
        else:
            dur_a_s = pd.Series(np.nan, index=df_out.index, dtype=float)
        if col_dur_b:
            dur_b_s = _series_from(df_out, col_dur_b)
        else:
            dur_b_s = pd.Series(np.nan, index=df_out.index, dtype=float)
        if "timeline_votes" in mcols:
            votes_s = _series_from(df_out, mcols["timeline_votes"])
        else:
            votes_s = pd.Series(np.nan, index=df_out.index, dtype=float)
        drift_col = mcols.get("timeline_drift_ratio") or mcols.get("drift_ratio")
        drift_s = _series_from(df_out, drift_col) if drift_col else pd.Series(1.0, index=df_out.index, dtype=float)
        seg_a_col = mcols.get("timeline_segments_a_strict") or mcols.get("segments_a_strict")
        seg_b_col = mcols.get("timeline_segments_b_strict") or mcols.get("segments_b_strict")
        seg_a_raw_col = mcols.get("timeline_segments_a_raw") or mcols.get("segments_a_raw")
        seg_b_raw_col = mcols.get("timeline_segments_b_raw") or mcols.get("segments_b_raw")
        seg_a_s = df_out[seg_a_col].astype(str) if seg_a_col else pd.Series([""] * len(df_out), index=df_out.index, dtype=object)
        seg_b_s = df_out[seg_b_col].astype(str) if seg_b_col else pd.Series([""] * len(df_out), index=df_out.index, dtype=object)
        seg_a_raw_s = df_out[seg_a_raw_col].astype(str) if seg_a_raw_col else pd.Series([""] * len(df_out), index=df_out.index, dtype=object)
        seg_b_raw_s = df_out[seg_b_raw_col].astype(str) if seg_b_raw_col else pd.Series([""] * len(df_out), index=df_out.index, dtype=object)
        uniq_a = _series_from(df_out, mcols.get("timeline_unique_ratio_a", "")) if "timeline_unique_ratio_a" in mcols else pd.Series(np.nan, index=df_out.index, dtype=float)
        uniq_b = _series_from(df_out, mcols.get("timeline_unique_ratio_b", "")) if "timeline_unique_ratio_b" in mcols else pd.Series(np.nan, index=df_out.index, dtype=float)
        uniq_min = pd.concat([uniq_a, uniq_b], axis=1).min(axis=1)

        start_b = off.clip(lower=0.0)
        end_b = start_b + ov
        end_b = np.where(dur_b_s.notna(), np.minimum(end_b, dur_b_s), end_b)
        end_b = pd.Series(end_b, index=df_out.index, dtype=float)
        start_a = (-off).clip(lower=0.0)
        end_a = start_a + ov
        end_a = np.where(dur_a_s.notna(), np.minimum(end_a, dur_a_s), end_a)
        end_a = pd.Series(end_a, index=df_out.index, dtype=float)

        valid = mode_mask & ov.notna() & (ov > 0.0) & start_b.notna() & end_b.notna() & (end_b > start_b)
        valid = valid & (dur_b_s.isna() | (start_b < dur_b_s))
        valid = valid & start_a.notna() & end_a.notna() & (end_a > start_a)
        valid = valid & (dur_a_s.isna() | (start_a < dur_a_s))
        better = valid & (ov > chosen_overlap)
        if bool(better.any()):
            chosen_overlap.loc[better] = ov.loc[better]
            est_start.loc[better] = start_b.loc[better]
            est_end.loc[better] = end_b.loc[better]
            est_overlap.loc[better] = ov.loc[better]
            est_overlap_raw.loc[better] = ov_raw.loc[better]
            est_start_a.loc[better] = start_a.loc[better]
            est_end_a.loc[better] = end_a.loc[better]
            est_dur_a.loc[better] = dur_a_s.loc[better]
            est_dur_b.loc[better] = dur_b_s.loc[better]
            est_tl_votes.loc[better] = votes_s.loc[better]
            est_tl_offset.loc[better] = off.loc[better]
            est_tl_unique_min.loc[better] = uniq_min.loc[better]
            est_tl_drift.loc[better] = drift_s.loc[better]
            est_tl_mode.loc[better] = mode
            est_tl_segments_a.loc[better] = seg_a_s.loc[better]
            est_tl_segments_b.loc[better] = seg_b_s.loc[better]
            est_tl_segments_a_raw.loc[better] = seg_a_raw_s.loc[better]
            est_tl_segments_b_raw.loc[better] = seg_b_raw_s.loc[better]

    df_out["timeline_est_b_start_s"] = est_start.round(1)
    df_out["timeline_est_b_end_s"] = est_end.round(1)
    df_out["timeline_est_overlap_s"] = est_overlap.round(1)
    # Pre-bridge overlap estimate from timeline mode internals.
    # Falls back to final overlap when raw-specific metric is unavailable.
    est_overlap_raw = est_overlap_raw.where(est_overlap_raw.notna(), est_overlap)
    df_out["timeline_est_overlap_raw_s"] = est_overlap_raw.round(1)
    df_out["timeline_est_a_start_s"] = est_start_a.round(1)
    df_out["timeline_est_a_end_s"] = est_end_a.round(1)
    df_out["timeline_est_duration_a_s"] = est_dur_a.round(1)
    df_out["timeline_est_duration_b_s"] = est_dur_b.round(1)
    df_out["timeline_tuned_offset_s"] = est_tl_offset.round(1)
    df_out["timeline_tuned_votes"] = est_tl_votes.round(1)
    df_out["timeline_tuned_unique_ratio_min"] = est_tl_unique_min.round(3)
    df_out["timeline_est_drift_ratio"] = est_tl_drift.round(4)
    df_out["timeline_est_mode"] = est_tl_mode.astype(str)
    df_out["timeline_est_segments_a_raw"] = est_tl_segments_a_raw.astype(str)
    df_out["timeline_est_segments_b_raw"] = est_tl_segments_b_raw.astype(str)
    df_out["timeline_est_segments_a_strict"] = est_tl_segments_a.astype(str)
    df_out["timeline_est_segments_b_strict"] = est_tl_segments_b.astype(str)

    # Coverage semantics:
    # - any timeline match with a valid projected interval contributes.
    # This avoids reporting 0 coverage sections on timeline-matched pairs.
    coverage_any_timeline = est_start.notna() & est_end.notna() & any_timeline_match
    coverage_contributor = coverage_any_timeline.copy()
    df_out["anchor_full_full_conflict"] = False
    df_out["timeline_gap_verify_status"] = ""
    df_out["timeline_gap_verify_notes"] = ""
    df_out["timeline_b_coverage_contributor"] = coverage_contributor

    # Coverage denominator should reflect useful video content, not dead head/tail.
    # Reuse cached trim/dead-gap analysis when available, and only estimate missing
    # trim bounds on demand.

    def _project_intervals_to_playable(
        intervals: list[tuple[float, float]],
        *,
        duration_s: float,
        trim_head_s: float,
        trim_tail_s: float,
        dead_regions_abs: list[tuple[float, float]] | None = None,
    ) -> tuple[list[tuple[float, float]], float]:
        dur = max(0.0, float(duration_s))
        head = max(0.0, min(float(trim_head_s), dur))
        tail = max(0.0, min(float(trim_tail_s), max(0.0, dur - head)))
        play_start = head
        play_end = max(play_start, dur - tail)
        playable_spans: list[tuple[float, float]] = [(play_start, play_end)] if play_end > play_start else []
        if dead_regions_abs:
            dead_clamped = _merge_intervals_simple(
                [
                    (max(play_start, float(s)), min(play_end, float(e)))
                    for s, e in dead_regions_abs
                    if float(e) > float(s)
                ],
                gap_tolerance_s=0.0,
            )
            if dead_clamped:
                kept: list[tuple[float, float]] = []
                for span_s, span_e in playable_spans:
                    work = [(span_s, span_e)]
                    for dead_s, dead_e in dead_clamped:
                        next_work: list[tuple[float, float]] = []
                        for ws, we in work:
                            if dead_e <= ws or dead_s >= we:
                                next_work.append((ws, we))
                                continue
                            if dead_s > ws:
                                next_work.append((ws, dead_s))
                            if dead_e < we:
                                next_work.append((dead_e, we))
                        work = next_work
                        if not work:
                            break
                    kept.extend(work)
                playable_spans = _merge_intervals_simple(kept, gap_tolerance_s=0.0)
        playable_dur = float(sum((e - s) for s, e in playable_spans))
        if playable_dur <= 0.0:
            return [], 0.0
        span_offsets: list[tuple[float, float, float]] = []
        cursor = 0.0
        for span_s, span_e in playable_spans:
            span_offsets.append((float(span_s), float(span_e), float(cursor)))
            cursor += max(0.0, float(span_e) - float(span_s))
        out_intervals: list[tuple[float, float]] = []
        for s, e in intervals:
            s = float(s)
            e = float(e)
            if e <= s:
                continue
            s2 = max(play_start, s)
            e2 = min(play_end, e)
            if e2 <= s2:
                continue
            for span_s, span_e, span_off in span_offsets:
                x0 = max(s2, span_s)
                x1 = min(e2, span_e)
                if x1 > x0:
                    out_intervals.append(
                        (
                            float(span_off) + (float(x0) - float(span_s)),
                            float(span_off) + (float(x1) - float(span_s)),
                        )
                    )
        return out_intervals, playable_dur

    listed_folder_a_files = _list_videos_in_folder(folder_a) if folder_a else []
    listed_folder_b_files = _list_videos_in_folder(folder_b) if folder_b else []
    _all_paths_for_trim = pd.concat(
        [df_out["file_a"].astype(str), df_out["file_b"].astype(str)],
        axis=0,
        ignore_index=True,
    ).tolist()
    if listed_folder_a_files:
        _all_paths_for_trim.extend(str(p) for p in listed_folder_a_files if isinstance(p, str))
    if listed_folder_b_files:
        _all_paths_for_trim.extend(str(p) for p in listed_folder_b_files if isinstance(p, str))
    (
        timeline_trim_bounds_map,
        timeline_hash_active_bounds_map,
        timeline_internal_dead_regions_map,
        timeline_trim_cache_hits,
        timeline_trim_estimated,
        timeline_internal_dead_hits,
    ) = _load_playable_basis_maps(_all_paths_for_trim)
    if timeline_trim_bounds_map:
        print(
            "[consolidate] timeline trim coverage basis available for "
            f"{len(timeline_trim_bounds_map)} file(s) "
            f"(cache={timeline_trim_cache_hits}, estimated={timeline_trim_estimated}, "
            f"hash_bounds={len(timeline_hash_active_bounds_map)}, "
            f"internal_dead={timeline_internal_dead_hits})",
            flush=True,
        )

    print(
        "[consolidate] options: "
        f"a_self_repeat={'on' if TIMELINE_A_SELF_REPEAT_ENABLE else 'off'}",
        flush=True,
    )

    def _resolve_playable_window(
        path: str,
        dur_full_s: float,
    ) -> tuple[float, float, float, float, float, list[tuple[float, float]]]:
        return _resolve_playable_window_from_maps(
            path,
            dur_full_s,
            timeline_trim_bounds_map,
            timeline_hash_active_bounds_map,
            timeline_internal_dead_regions_map,
        )

    gap_verify_cache: dict[int, dict] = {}
    a_candidate_count_pre = df_out.groupby("file_a")["file_b"].transform("size").astype(int)
    vhs_gap = VideoHashStore()
    seq_cache_gap: dict[str, list[tuple[float, int]]] = {}
    self_repeat_cache_hits = 0
    self_repeat_rebuilt = 0

    def _get_filtered_timeline_seq_for_gap(path: str) -> list[tuple[float, int]]:
        return _get_filtered_timeline_seq_cached(vhs_gap, seq_cache_gap, path)

    coverage_b_s_map: dict[str, float] = {}
    coverage_b_raw_s_map: dict[str, float] = {}
    duration_b_map: dict[str, float] = {}
    coverage_b_pairs_map: dict[str, int] = {}
    coverage_b_gap_s_map: dict[str, float] = {}
    coverage_b_sections_raw_map: dict[str, int] = {}
    coverage_b_sections_est_map: dict[str, int] = {}
    coverage_b_missing_start_s_map: dict[str, float] = {}
    coverage_b_missing_end_s_map: dict[str, float] = {}
    coverage_b_unmatched_desc_map: dict[str, str] = {}
    coverage_b_uncertain_desc_map: dict[str, str] = {}

    coverage_a_s_map: dict[str, float] = {}
    coverage_a_raw_s_map: dict[str, float] = {}
    duration_a_map: dict[str, float] = {}
    coverage_a_pairs_map: dict[str, int] = {}
    coverage_a_gap_s_map: dict[str, float] = {}
    coverage_a_sections_raw_map: dict[str, int] = {}
    coverage_a_sections_est_map: dict[str, int] = {}
    coverage_a_unmatched_desc_map: dict[str, str] = {}
    coverage_a_uncertain_desc_map: dict[str, str] = {}
    coverage_a_unique_s_map: dict[str, float] = {}
    coverage_a_unique_duration_map: dict[str, float] = {}
    coverage_a_unique_sections_est_map: dict[str, int] = {}
    coverage_a_unique_uncertain_desc_map: dict[str, str] = {}
    coverage_a_unique_unmatched_segments_map: dict[str, list[tuple[float, float]]] = {}
    coverage_a_unique_uncertain_segments_map: dict[str, list[tuple[float, float]]] = {}
    coverage_a_unique_unmatched_segments_abs_map: dict[str, list[tuple[float, float]]] = {}
    coverage_a_unique_uncertain_segments_abs_map: dict[str, list[tuple[float, float]]] = {}
    coverage_a_self_repeat_s_map: dict[str, float] = {}

    # Fallback duration columns if timeline duration is missing on some rows.
    duration_b_fallback_cols = [c for c in df_out.columns if c.endswith("__duration_b (s)")]

    def _describe_unmatched_regions(
        merged_covered: list[tuple[float, float]],
        duration_s: float,
    ) -> str:
        dur = float(max(0.0, duration_s))
        if dur <= 0.0:
            return "unmatched location unavailable (duration unknown)"
        if not merged_covered:
            return "unmatched: entire file"

        tol = max(5.0, float(TIMELINE_STEP_S) * 2.0)
        covered = _merge_intervals(merged_covered, gap_tolerance_s=0.0)
        gaps: list[tuple[float, float]] = []
        cursor = 0.0
        for s, e in covered:
            s = max(0.0, min(dur, float(s)))
            e = max(0.0, min(dur, float(e)))
            if s > cursor:
                gaps.append((cursor, s))
            cursor = max(cursor, e)
        if cursor < dur:
            gaps.append((cursor, dur))

        gaps = [(s, e) for s, e in gaps if (e - s) > tol]
        if not gaps:
            return "no unmatched region"

        where_tags: list[str] = []
        for s, e in gaps:
            if s <= tol:
                where_tags.append("start")
            elif e >= (dur - tol):
                where_tags.append("end")
            else:
                where_tags.append("middle")
        where_unique = []
        for tag in ("start", "middle", "end"):
            if tag in where_tags:
                where_unique.append(tag)
        where_text = "+".join(where_unique) if where_unique else "unknown"

        # Also report the largest unmatched gap location as a coarse percent span.
        largest = max(gaps, key=lambda x: (x[1] - x[0]))
        g0 = max(0.0, min(100.0, 100.0 * largest[0] / dur))
        g1 = max(0.0, min(100.0, 100.0 * largest[1] / dur))
        return f"unmatched at {where_text}; largest gap ~{g0:.0f}-{g1:.0f}%"

    def _describe_uncertain_regions(
        uncertain_intervals: list[tuple[float, float]],
        duration_s: float,
    ) -> str:
        dur = float(max(0.0, duration_s))
        if dur <= 0.0:
            return ""
        segs = _merge_intervals(uncertain_intervals, gap_tolerance_s=0.0)
        tol = max(5.0, float(TIMELINE_STEP_S) * 2.0)
        segs = [(max(0.0, float(s)), min(dur, float(e))) for s, e in segs if (float(e) - float(s)) > tol]
        if not segs:
            return ""
        where_tags: list[str] = []
        for s, e in segs:
            if s <= tol:
                where_tags.append("start")
            elif e >= (dur - tol):
                where_tags.append("end")
            else:
                where_tags.append("middle")
        where_unique = []
        for tag in ("start", "middle", "end"):
            if tag in where_tags:
                where_unique.append(tag)
        where_text = "+".join(where_unique) if where_unique else "unknown"
        largest = max(segs, key=lambda x: (x[1] - x[0]))
        g0 = max(0.0, min(100.0, 100.0 * largest[0] / dur))
        g1 = max(0.0, min(100.0, 100.0 * largest[1] / dur))
        # Suppress meaningless edge slivers that only show up as 0-0% or 100-100%
        # after percentage rounding. Those are not helpful review hints.
        if (round(g0) == round(g1)) and set(where_unique).issubset({"start", "end"}):
            return ""
        return f"borderline unmatched around {where_text}; largest gap ~{g0:.0f}-{g1:.0f}%"

    def _derive_uncertain_unmatched_regions(
        covered_intervals: list[tuple[float, float]],
        duration_s: float,
        bridge_gap_s: float,
    ) -> list[tuple[float, float]]:
        """
        Mark only borderline unmatched gaps as uncertain.

        This is intentionally gap-side logic: accepted covered regions remain
        covered. The uncertain label is reserved for unmatched regions that are
        small enough, or sufficiently internal, that they may plausibly be missed
        overlap rather than true unique content / true non-match.
        """
        dur = float(max(0.0, duration_s))
        if dur <= 0.0:
            return []
        covered = _merge_intervals(covered_intervals, gap_tolerance_s=0.0)
        gaps = _interval_complement(covered, start_s=0.0, end_s=dur)
        if not gaps:
            return []

        internal_max_s = min(300.0, max(45.0, 0.75 * float(bridge_gap_s)))
        edge_max_s = min(120.0, max(15.0, 0.20 * float(bridge_gap_s)))
        eps = max(1e-6, 0.5 * float(TIMELINE_STEP_S))

        uncertain: list[tuple[float, float]] = []
        for g0, g1 in gaps:
            span_s = max(0.0, float(g1) - float(g0))
            if span_s <= 0.0:
                continue
            at_start = float(g0) <= eps
            at_end = float(g1) >= (dur - eps)
            is_internal = (not at_start) and (not at_end)
            if is_internal and (span_s <= internal_max_s):
                uncertain.append((float(g0), float(g1)))
            elif (at_start or at_end) and (span_s <= edge_max_s):
                uncertain.append((float(g0), float(g1)))
        return _merge_intervals(uncertain, gap_tolerance_s=0.0)

    b_groups = list(df_out.groupby("file_b", dropna=True))
    b_total = len(b_groups)
    b_t0 = time.time()
    b_tick = max(1, b_total // 8) if b_total else 1
    for b_idx, (file_b, grp) in enumerate(b_groups, start=1):
        if not isinstance(file_b, str):
            continue
        # Use a single trusted contributor population for both "raw" and "est":
        # - raw = unbridged union (gap_tolerance=0)
        # - est = bridged union (gap_tolerance=gap_s)
        # This guarantees est >= raw and keeps the comparison intuitive.
        g_cov = grp[coverage_contributor.loc[grp.index]]
        g_any = grp[coverage_any_timeline.loc[grp.index]]

        intervals_est = []
        for _idx, row in g_cov.iterrows():
            seg_b_text = str(row.get("timeline_est_segments_b_strict") or "")
            segs_b_abs = _parse_segments_text(seg_b_text)
            seg_b_truncated = ("..." in seg_b_text)
            if segs_b_abs and (not seg_b_truncated):
                intervals_est.extend([(float(s0), float(s1)) for s0, s1 in segs_b_abs if float(s1) > float(s0)])
                continue
            s = row.get("timeline_est_b_start_s")
            e = row.get("timeline_est_b_end_s")
            if pd.isna(s) or pd.isna(e):
                continue
            s = float(s)
            e = float(e)
            if e > s:
                intervals_est.append((s, e))

        intervals_raw = []
        for _idx, row in g_cov.iterrows():
            seg_b_raw_text = str(row.get("timeline_est_segments_b_raw") or "")
            segs_b_raw = _parse_segments_text(seg_b_raw_text)
            seg_b_raw_truncated = ("..." in seg_b_raw_text)
            if segs_b_raw and (not seg_b_raw_truncated):
                intervals_raw.extend([(float(s0), float(s1)) for s0, s1 in segs_b_raw if float(s1) > float(s0)])
                continue
            seg_b_strict_text = str(row.get("timeline_est_segments_b_strict") or "")
            segs_b_strict = _parse_segments_text(seg_b_strict_text)
            seg_b_strict_truncated = ("..." in seg_b_strict_text)
            if segs_b_strict and (not seg_b_strict_truncated):
                intervals_raw.extend([(float(s0), float(s1)) for s0, s1 in segs_b_strict if float(s1) > float(s0)])
                continue
            s = row.get("timeline_est_b_start_s")
            ov_raw = row.get("timeline_est_overlap_raw_s")
            dur_row = row.get("timeline_est_duration_b_s")
            if pd.isna(s) or pd.isna(ov_raw):
                continue
            s = float(s)
            e = s + float(ov_raw)
            if not pd.isna(dur_row):
                e = min(e, float(dur_row))
            if pd.isna(s) or pd.isna(e):
                continue
            if e > s:
                intervals_raw.append((s, e))

        dur_candidates = pd.to_numeric(g_any.get("timeline_est_duration_b_s", pd.Series([], dtype=float)), errors="coerce")
        dur_candidates = dur_candidates.dropna()
        vals = []
        for col in duration_b_fallback_cols:
            s = pd.to_numeric(grp[col], errors="coerce").dropna()
            if not s.empty:
                vals.extend(s.tolist())
        fallback_dur_s = float(np.median(vals)) if vals else 0.0
        if dur_candidates.empty:
            # Fallback to any available duration_b columns from merged reports.
            dur_full_s = float(fallback_dur_s)
        else:
            # Use the larger basis to avoid clipping absolute timeline segments.
            dur_full_s = max(float(dur_candidates.median()), float(fallback_dur_s))
        trim_head_s, trim_tail_s, _play_start_b, _play_end_b, _play_dur_b, dead_b_abs = _resolve_playable_window(file_b, dur_full_s)
        intervals_est, dur_s = _project_intervals_to_playable(
            intervals_est,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_b_abs,
        )
        intervals_raw, _ = _project_intervals_to_playable(
            intervals_raw,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_b_abs,
        )
        cov_raw_s = _union_interval_length(intervals_raw, gap_tolerance_s=0.0)
        cov_pairs = len(intervals_est)

        gap_s = min(
            float(TIMELINE_COVERAGE_BRIDGE_MAX_S),
            max(float(TIMELINE_COVERAGE_BRIDGE_MIN_S), float(TIMELINE_COVERAGE_BRIDGE_FRAC_OF_B) * float(dur_s)),
        )
        cov_s = _union_interval_length(intervals_est, gap_tolerance_s=gap_s)
        merged_raw = _merge_intervals(intervals_raw, gap_tolerance_s=0.0)
        merged_est = _merge_intervals(intervals_est, gap_tolerance_s=gap_s)
        uncertain_b = _derive_uncertain_unmatched_regions(
            merged_est,
            duration_s=dur_s,
            bridge_gap_s=gap_s,
        )

        coverage_b_s_map[file_b] = float(cov_s)
        coverage_b_raw_s_map[file_b] = float(cov_raw_s)
        duration_b_map[file_b] = float(dur_s)
        coverage_b_pairs_map[file_b] = int(cov_pairs)
        coverage_b_gap_s_map[file_b] = float(gap_s)
        coverage_b_sections_raw_map[file_b] = int(len(merged_raw))
        coverage_b_sections_est_map[file_b] = int(len(merged_est))
        coverage_b_unmatched_desc_map[file_b] = _describe_unmatched_regions(merged_est, dur_s)
        coverage_b_uncertain_desc_map[file_b] = _describe_uncertain_regions(uncertain_b, dur_s)
        if merged_est:
            coverage_b_missing_start_s_map[file_b] = max(0.0, float(merged_est[0][0]))
            coverage_b_missing_end_s_map[file_b] = max(0.0, float(dur_s) - float(merged_est[-1][1]))
        else:
            coverage_b_missing_start_s_map[file_b] = float(max(0.0, dur_s))
            coverage_b_missing_end_s_map[file_b] = 0.0
        if (b_idx % b_tick == 0) or (b_idx == b_total):
            elapsed = time.time() - b_t0
            avg = elapsed / b_idx if b_idx else 0.0
            eta = max(0.0, (b_total - b_idx) * avg)
            print(
                f"[progress] consolidate B coverage {b_idx}/{b_total} "
                f"- {_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                flush=True,
            )

    # A-side coverage from the same contributor rows, projected onto file_a timelines.
    # Fallback duration columns if timeline duration is missing on some rows.
    duration_a_fallback_cols = [c for c in df_out.columns if c.endswith("__duration_a (s)")]
    a_groups = list(df_out.groupby("file_a", dropna=True))
    a_total = len(a_groups)
    a_t0 = time.time()
    a_tick = max(1, a_total // 8) if a_total else 1
    for a_idx, (file_a, grp) in enumerate(a_groups, start=1):
        if not isinstance(file_a, str):
            continue
        # Keep raw/est semantics aligned with B-side coverage.
        g_cov = grp[coverage_contributor.loc[grp.index]]
        g_any = grp[coverage_any_timeline.loc[grp.index]]

        intervals_est = []
        for _idx, row in g_cov.iterrows():
            seg_a_text = str(row.get("timeline_est_segments_a_strict") or "")
            segs_a_abs = _parse_segments_text(seg_a_text)
            seg_a_truncated = ("..." in seg_a_text)
            if segs_a_abs and (not seg_a_truncated):
                intervals_est.extend([(float(s0), float(s1)) for s0, s1 in segs_a_abs if float(s1) > float(s0)])
            else:
                s = row.get("timeline_est_a_start_s")
                e = row.get("timeline_est_a_end_s")
                if not (pd.isna(s) or pd.isna(e)):
                    s = float(s)
                    e = float(e)
                    if e > s:
                        intervals_est.append((s, e))
            verify_result = gap_verify_cache.get(int(_idx))
            if verify_result:
                intervals_est.extend([
                    (float(x0), float(x1))
                    for x0, x1 in verify_result.get("verified_gap_intervals_a", [])
                    if float(x1) > float(x0)
                ])

        intervals_raw = []
        for _idx, row in g_cov.iterrows():
            seg_a_raw_text = str(row.get("timeline_est_segments_a_raw") or "")
            segs_a_raw = _parse_segments_text(seg_a_raw_text)
            seg_a_raw_truncated = ("..." in seg_a_raw_text)
            if segs_a_raw and (not seg_a_raw_truncated):
                intervals_raw.extend([(float(s0), float(s1)) for s0, s1 in segs_a_raw if float(s1) > float(s0)])
                continue
            seg_a_strict_text = str(row.get("timeline_est_segments_a_strict") or "")
            segs_a_strict = _parse_segments_text(seg_a_strict_text)
            seg_a_strict_truncated = ("..." in seg_a_strict_text)
            if segs_a_strict and (not seg_a_strict_truncated):
                intervals_raw.extend([(float(s0), float(s1)) for s0, s1 in segs_a_strict if float(s1) > float(s0)])
                continue
            s = row.get("timeline_est_a_start_s")
            ov_raw = row.get("timeline_est_overlap_raw_s")
            dur_row = row.get("timeline_est_duration_a_s")
            if pd.isna(s) or pd.isna(ov_raw):
                continue
            s = float(s)
            e = s + float(ov_raw)
            if not pd.isna(dur_row):
                e = min(e, float(dur_row))
            if pd.isna(s) or pd.isna(e):
                continue
            if e > s:
                intervals_raw.append((s, e))

        dur_candidates = pd.to_numeric(g_any.get("timeline_est_duration_a_s", pd.Series([], dtype=float)), errors="coerce")
        dur_candidates = dur_candidates.dropna()
        vals = []
        for col in duration_a_fallback_cols:
            s = pd.to_numeric(grp[col], errors="coerce").dropna()
            if not s.empty:
                vals.extend(s.tolist())
        fallback_dur_s = float(np.median(vals)) if vals else 0.0
        if dur_candidates.empty:
            dur_full_s = float(fallback_dur_s)
        else:
            dur_full_s = max(float(dur_candidates.median()), float(fallback_dur_s))
        trim_head_s, trim_tail_s, _play_start_a, _play_end_a, _play_dur_a, dead_a_abs = _resolve_playable_window(file_a, dur_full_s)
        intervals_est_abs = list(intervals_est)
        intervals_raw_abs = list(intervals_raw)
        intervals_est, dur_s = _project_intervals_to_playable(
            intervals_est,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_a_abs,
        )
        intervals_raw, _ = _project_intervals_to_playable(
            intervals_raw,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_a_abs,
        )
        cov_raw_s = _union_interval_length(intervals_raw, gap_tolerance_s=0.0)
        cov_pairs = len(intervals_est)

        gap_s = min(
            float(TIMELINE_COVERAGE_BRIDGE_MAX_S),
            max(float(TIMELINE_COVERAGE_BRIDGE_MIN_S), float(TIMELINE_COVERAGE_BRIDGE_FRAC_OF_B) * float(dur_s)),
        )
        cov_s = _union_interval_length(intervals_est, gap_tolerance_s=gap_s)
        merged_raw = _merge_intervals(intervals_raw, gap_tolerance_s=0.0)
        merged_est = _merge_intervals(intervals_est, gap_tolerance_s=gap_s)
        uncertain_a = _derive_uncertain_unmatched_regions(
            merged_est,
            duration_s=dur_s,
            bridge_gap_s=gap_s,
        )

        coverage_a_s_map[file_a] = float(cov_s)
        coverage_a_raw_s_map[file_a] = float(cov_raw_s)
        duration_a_map[file_a] = float(dur_s)
        coverage_a_pairs_map[file_a] = int(cov_pairs)
        coverage_a_gap_s_map[file_a] = float(gap_s)
        coverage_a_sections_raw_map[file_a] = int(len(merged_raw))
        coverage_a_sections_est_map[file_a] = int(len(merged_est))
        coverage_a_unmatched_desc_map[file_a] = _describe_unmatched_regions(merged_est, dur_s)
        coverage_a_uncertain_desc_map[file_a] = _describe_uncertain_regions(uncertain_a, dur_s)
        unmatched_est_abs = _interval_complement(
            intervals_est_abs,
            start_s=_play_start_a,
            end_s=_play_end_a,
        )
        if dead_a_abs:
            playable_non_dead_abs = _interval_complement(
                dead_a_abs,
                start_s=_play_start_a,
                end_s=_play_end_a,
            )
            unmatched_est_abs = _intersect_intervals_simple(unmatched_est_abs, playable_non_dead_abs)
        edge_guard_s = max(2.0 * float(TIMELINE_STEP_S), 1.0)
        edge_unmatched_abs = [
            (u0, u1)
            for u0, u1 in unmatched_est_abs
            if (float(u0) <= (_play_start_a + edge_guard_s)) or (float(u1) >= (_play_end_a - edge_guard_s))
        ]
        has_edge_unmatched_gap = any(
            (float(u1) - float(u0)) >= float(TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S)
            for u0, u1 in edge_unmatched_abs
        )

        unique_basis_abs = [(_play_start_a, _play_end_a)] if dur_s > 0.0 else []
        unique_covered_est_abs = list(intervals_est_abs)
        cov_pct_a = (100.0 * float(cov_s) / float(dur_s)) if float(dur_s) > 0.0 else 0.0
        if (
            TIMELINE_A_SELF_REPEAT_ENABLE
            and dur_s > 0.0
            and (intervals_est_abs or intervals_raw_abs)
            and (float(cov_s) + max(5.0, 0.005 * float(dur_s)) < float(dur_s))
            and has_edge_unmatched_gap
            and cov_pct_a >= float(TIMELINE_A_SELF_REPEAT_TRIGGER_MIN_COVERAGE_PCT)
        ):
            print(f"[stage] self-repeat check {a_idx}/{a_total}: {Path(file_a).name}", flush=True)
            seq_a_gap = _get_filtered_timeline_seq_for_gap(file_a)
            repeat_info, repeat_cache_hit = vhs_gap.get_a_self_repeat_analysis(
                file_a,
                seq=seq_a_gap,
            )
            if repeat_cache_hit:
                self_repeat_cache_hits += 1
            else:
                self_repeat_rebuilt += 1
            repeated_abs = repeat_info.get("repeated_segments", [])
            if repeated_abs:
                # Only discount repeated playback where A is currently unmatched.
                repeated_abs = _intersect_intervals_simple(repeated_abs, unmatched_est_abs)
            residual_edge_unmatched_abs = _intersect_intervals_simple(
                edge_unmatched_abs,
                _interval_complement(
                    repeated_abs,
                    start_s=_play_start_a,
                    end_s=_play_end_a,
                ),
            )
            edge_repeat_abs = _detect_edge_repeat_segments_by_similarity(
                seq_a_gap,
                edge_intervals=residual_edge_unmatched_abs,
                total_bits=64,
                hamming_thresh=TIMELINE_A_SELF_REPEAT_HAMMING_THRESH,
                step_s=TIMELINE_STEP_S,
                lsh_chunks=TIMELINE_LSH_CHUNKS,
                max_candidates_per_frame=TIMELINE_MAX_CANDIDATES_PER_FRAME,
                run_gap_mult=TIMELINE_RUN_GAP_MULT,
                min_segment_s=TIMELINE_A_SELF_REPEAT_MIN_SEGMENT_S,
                min_offset_s=TIMELINE_A_SELF_REPEAT_MIN_OFFSET_S,
                min_match_fraction=0.90,
            )
            edge_repeat_abs = _intersect_intervals_simple(edge_repeat_abs, unmatched_est_abs)
            if edge_repeat_abs:
                repeated_abs = _merge_intervals_simple(
                    list(repeated_abs) + list(edge_repeat_abs),
                    gap_tolerance_s=0.0,
                )
            if repeated_abs:
                unique_basis_abs = _interval_complement(
                    repeated_abs,
                    start_s=_play_start_a,
                    end_s=_play_end_a,
                )
            for pair_info in repeat_info.get("repeat_pairs", []):
                try:
                    offset_s = float(pair_info.get("offset_s", 0.0))
                except Exception:
                    offset_s = 0.0
                if offset_s <= 0.0:
                    continue
                for later_s, later_e in pair_info.get("later_segments", []):
                    later_s = float(later_s)
                    later_e = float(later_e)
                    if later_e <= later_s:
                        continue
                    for cov_s_abs, cov_e_abs in intervals_est_abs:
                        x0 = max(float(cov_s_abs), later_s)
                        x1 = min(float(cov_e_abs), later_e)
                        if x1 > x0:
                            unique_covered_est_abs.append((x0 - offset_s, x1 - offset_s))

        unique_basis_rel, _playable_dur_a = _project_intervals_to_playable(
            unique_basis_abs,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_a_abs,
        )
        unique_covered_est_rel, _ = _project_intervals_to_playable(
            unique_covered_est_abs,
            duration_s=dur_full_s,
            trim_head_s=trim_head_s,
            trim_tail_s=trim_tail_s,
            dead_regions_abs=dead_a_abs,
        )
        unique_covered_est_rel = _merge_intervals(unique_covered_est_rel, gap_tolerance_s=gap_s)
        unique_basis_rel = _merge_intervals(unique_basis_rel, gap_tolerance_s=0.0)
        unique_dur_s = _union_interval_length(unique_basis_rel, gap_tolerance_s=0.0)
        unique_cov_segments = _intersect_intervals_simple(unique_covered_est_rel, unique_basis_rel)
        unique_unmatched_segments = _intersect_intervals_simple(
            unique_basis_rel,
            _interval_complement(unique_cov_segments, start_s=0.0, end_s=dur_s),
        )
        unique_uncertain_segments = _intersect_intervals_simple(unique_unmatched_segments, uncertain_a)
        unique_cov_s = _union_interval_length(unique_cov_segments, gap_tolerance_s=0.0)
        if merged_est:
            a_missing_start_s = max(0.0, float(merged_est[0][0]))
            a_missing_end_s = max(0.0, float(dur_s) - float(merged_est[-1][1]))
        else:
            a_missing_start_s = float(max(0.0, dur_s))
            a_missing_end_s = 0.0
        if (
            TIMELINE_A_COMMON_EDGE_RECOVERY_ENABLE
            and unique_dur_s > 0.0
            and int(len(merged_est)) == 1
        ):
            b_path = ""
            conf_val = 0.0
            try:
                grp_conf = pd.to_numeric(grp.get("confidence_score", pd.Series([], dtype=float)), errors="coerce").fillna(0.0)
                if len(grp_conf) == len(grp):
                    best_idx = grp_conf.idxmax()
                    b_path = str(grp.at[best_idx, "file_b"])
                    conf_val = float(grp_conf.loc[best_idx])
                else:
                    conf_val = float(grp_conf.max()) if (len(grp_conf) > 0) else 0.0
                    b_path = str(grp["file_b"].dropna().astype(str).iloc[0]) if int(grp["file_b"].dropna().shape[0]) > 0 else ""
            except Exception:
                b_path = ""
                conf_val = 0.0
            b_sections_est = int(coverage_b_sections_est_map.get(b_path, 0))
            if (
                b_path
                and (b_sections_est == 1)
                and (conf_val >= float(TIMELINE_A_COMMON_EDGE_RECOVERY_MIN_CONF))
            ):
                a_missing_s = max(0.0, float(unique_dur_s) - float(unique_cov_s))
                b_missing_start_s = float(coverage_b_missing_start_s_map.get(b_path, 0.0))
                b_missing_end_s = float(coverage_b_missing_end_s_map.get(b_path, 0.0))
                shared_start_s = min(float(a_missing_start_s), float(b_missing_start_s))
                shared_end_s = min(float(a_missing_end_s), float(b_missing_end_s))
                shared_edge_s = max(0.0, float(shared_start_s + shared_end_s))
                if shared_edge_s > 0.0:
                    unique_cov_s = min(float(unique_dur_s), float(unique_cov_s) + min(float(a_missing_s), float(shared_edge_s)))
                residual_s = max(0.0, float(unique_dur_s) - float(unique_cov_s))
                if residual_s <= float(TIMELINE_A_COMMON_EDGE_RECOVERY_FULL_IF_RESIDUAL_LE_S):
                    unique_cov_s = float(unique_dur_s)
        unique_missing_s = max(0.0, float(unique_dur_s) - float(unique_cov_s))
        unique_full = False
        if unique_dur_s > 0.0:
            unique_cov_pct_local = 100.0 * float(unique_cov_s) / float(unique_dur_s)
            if unique_cov_pct_local >= float(TIMELINE_COVERAGE_FULL_MIN_PCT):
                unique_full = True
        if unique_missing_s <= float(TIMELINE_COVERAGE_FULL_MAX_MISSING_S):
            unique_full = True
        if unique_full:
            unique_unmatched_segments = []
            unique_uncertain_segments = []
        unique_unmatched_segments_abs = _map_trimmed_rel_to_abs(
            unique_unmatched_segments,
            _play_start_a,
            dead_regions_abs=dead_a_abs,
        )
        unique_uncertain_segments_abs = _map_trimmed_rel_to_abs(
            unique_uncertain_segments,
            _play_start_a,
            dead_regions_abs=dead_a_abs,
        )
        redundant_s = max(0.0, float(dur_s) - float(unique_dur_s))
        coverage_a_unique_s_map[file_a] = float(unique_cov_s)
        coverage_a_unique_duration_map[file_a] = float(unique_dur_s)
        coverage_a_unique_sections_est_map[file_a] = int(len(_merge_intervals(unique_cov_segments, gap_tolerance_s=0.0)))
        coverage_a_unique_uncertain_desc_map[file_a] = _describe_uncertain_regions(unique_uncertain_segments, unique_dur_s)
        coverage_a_unique_unmatched_segments_map[file_a] = list(unique_unmatched_segments)
        coverage_a_unique_uncertain_segments_map[file_a] = list(unique_uncertain_segments)
        coverage_a_unique_unmatched_segments_abs_map[file_a] = list(unique_unmatched_segments_abs)
        coverage_a_unique_uncertain_segments_abs_map[file_a] = list(unique_uncertain_segments_abs)
        coverage_a_self_repeat_s_map[file_a] = float(redundant_s)
        if (a_idx % a_tick == 0) or (a_idx == a_total):
            elapsed = time.time() - a_t0
            avg = elapsed / a_idx if a_idx else 0.0
            eta = max(0.0, (a_total - a_idx) * avg)
            print(
                f"[progress] consolidate A coverage {a_idx}/{a_total} "
                f"- {_format_duration_auto(elapsed)} elapsed - ETA {_format_duration_auto(eta)}",
                flush=True,
            )
    if vhs_gap._dirty:
        vhs_gap.save_if_dirty()
    if TIMELINE_A_SELF_REPEAT_ENABLE:
        print(
            f"[consolidate] a_self_repeat cache: hits={self_repeat_cache_hits}, rebuilt={self_repeat_rebuilt}",
            flush=True,
        )

    file_b_series = df_out["file_b"].astype(str)
    cov_s_series = file_b_series.map(lambda b: coverage_b_s_map.get(b, 0.0)).astype(float)
    cov_raw_s_series = file_b_series.map(lambda b: coverage_b_raw_s_map.get(b, 0.0)).astype(float)
    dur_s_series = file_b_series.map(lambda b: duration_b_map.get(b, 0.0)).astype(float)
    cov_pairs_series = file_b_series.map(lambda b: coverage_b_pairs_map.get(b, 0)).astype(int)
    cov_gap_s_series = file_b_series.map(lambda b: coverage_b_gap_s_map.get(b, 0.0)).astype(float)
    cov_sections_raw_series = file_b_series.map(lambda b: coverage_b_sections_raw_map.get(b, 0)).astype(int)
    cov_sections_est_series = file_b_series.map(lambda b: coverage_b_sections_est_map.get(b, 0)).astype(int)

    cov_raw_pct_series = pd.Series(
        np.where(dur_s_series > 0.0, (100.0 * cov_raw_s_series / dur_s_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)
    cov_pct_series = pd.Series(
        np.where(dur_s_series > 0.0, (100.0 * cov_s_series / dur_s_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)

    df_out["timeline_b_coverage_est_s"] = cov_s_series.round(1)
    df_out["timeline_b_coverage_raw_s"] = cov_raw_s_series.round(1)
    df_out["timeline_b_duration_est_s"] = dur_s_series.round(1)
    df_out["timeline_b_coverage_est_pct"] = cov_pct_series.round(1)
    df_out["timeline_b_coverage_raw_pct"] = cov_raw_pct_series.round(1)
    df_out["timeline_b_coverage_bridge_gap_s"] = cov_gap_s_series.round(1)
    df_out["timeline_b_coverage_pair_count"] = cov_pairs_series
    df_out["timeline_b_coverage_sections_raw"] = cov_sections_raw_series
    df_out["timeline_b_coverage_sections_est"] = cov_sections_est_series

    file_a_series = df_out["file_a"].astype(str)
    cov_a_s_series = file_a_series.map(lambda a: coverage_a_s_map.get(a, 0.0)).astype(float)
    cov_a_raw_s_series = file_a_series.map(lambda a: coverage_a_raw_s_map.get(a, 0.0)).astype(float)
    dur_a_series = file_a_series.map(lambda a: duration_a_map.get(a, 0.0)).astype(float)
    cov_a_unique_s_series = file_a_series.map(lambda a: coverage_a_unique_s_map.get(a, 0.0)).astype(float)
    dur_a_unique_series = file_a_series.map(lambda a: coverage_a_unique_duration_map.get(a, 0.0)).astype(float)
    cov_a_self_repeat_s_series = file_a_series.map(lambda a: coverage_a_self_repeat_s_map.get(a, 0.0)).astype(float)
    cov_a_pairs_series = file_a_series.map(lambda a: coverage_a_pairs_map.get(a, 0)).astype(int)
    cov_a_gap_s_series = file_a_series.map(lambda a: coverage_a_gap_s_map.get(a, 0.0)).astype(float)
    cov_a_sections_raw_series = file_a_series.map(lambda a: coverage_a_sections_raw_map.get(a, 0)).astype(int)
    cov_a_sections_est_series = file_a_series.map(lambda a: coverage_a_sections_est_map.get(a, 0)).astype(int)
    cov_a_unique_sections_est_series = file_a_series.map(lambda a: coverage_a_unique_sections_est_map.get(a, 0)).astype(int)
    cov_a_raw_pct_series = pd.Series(
        np.where(dur_a_series > 0.0, (100.0 * cov_a_raw_s_series / dur_a_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)
    cov_a_pct_series = pd.Series(
        np.where(dur_a_series > 0.0, (100.0 * cov_a_s_series / dur_a_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)
    cov_a_unique_pct_series = pd.Series(
        np.where(dur_a_unique_series > 0.0, (100.0 * cov_a_unique_s_series / dur_a_unique_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)
    cov_a_self_repeat_pct_series = pd.Series(
        np.where(dur_a_series > 0.0, (100.0 * cov_a_self_repeat_s_series / dur_a_series), 0.0),
        index=df_out.index,
        dtype=float,
    ).clip(lower=0.0, upper=100.0)
    df_out["timeline_a_coverage_est_s"] = cov_a_s_series.round(1)
    df_out["timeline_a_coverage_raw_s"] = cov_a_raw_s_series.round(1)
    df_out["timeline_a_duration_est_s"] = dur_a_series.round(1)
    df_out["timeline_a_unique_covered_est_s"] = cov_a_unique_s_series.round(1)
    df_out["timeline_a_unique_duration_est_s"] = dur_a_unique_series.round(1)
    df_out["timeline_a_self_repeat_redundant_s"] = cov_a_self_repeat_s_series.round(1)
    df_out["timeline_a_coverage_est_pct"] = cov_a_pct_series.round(1)
    df_out["timeline_a_coverage_raw_pct"] = cov_a_raw_pct_series.round(1)
    df_out["timeline_a_unique_covered_est_pct"] = cov_a_unique_pct_series.round(1)
    df_out["timeline_a_self_repeat_redundant_pct"] = cov_a_self_repeat_pct_series.round(1)
    df_out["timeline_a_coverage_pair_count"] = cov_a_pairs_series
    df_out["timeline_a_coverage_bridge_gap_s"] = cov_a_gap_s_series.round(1)
    df_out["timeline_a_coverage_sections_raw"] = cov_a_sections_raw_series
    df_out["timeline_a_coverage_sections_est"] = cov_a_sections_est_series
    df_out["timeline_a_unique_coverage_sections_est"] = cov_a_unique_sections_est_series

    # Pair-level overlap-clash metrics:
    # "clash" means this row's matched interval overlaps another matched interval
    # on the same timeline (B-side or A-side). Useful to spot multi-match ambiguity.
    def _overlap_clash_metrics(group_key: str, start_col: str, end_col: str) -> tuple[pd.Series, pd.Series, pd.Series]:
        clash_count = pd.Series(0, index=df_out.index, dtype=int)
        clash_seconds = pd.Series(0.0, index=df_out.index, dtype=float)
        for _group_value, grp in df_out.groupby(group_key, dropna=True):
            items: list[tuple[int, float, float]] = []
            for idx, row in grp.iterrows():
                if not bool(coverage_any_timeline.loc[idx]):
                    continue
                s = row.get(start_col)
                e = row.get(end_col)
                if pd.isna(s) or pd.isna(e):
                    continue
                s = float(s)
                e = float(e)
                if e > s:
                    items.append((int(idx), s, e))
            if len(items) < 2:
                continue
            for idx, s, e in items:
                others = [(s2, e2) for oid, s2, e2 in items if oid != idx and (e2 > s) and (s2 < e)]
                if not others:
                    continue
                clash_count.loc[idx] = len(others)
                merged_others = _merge_intervals(others, gap_tolerance_s=0.0)
                overlap_s = 0.0
                for s2, e2 in merged_others:
                    x0 = max(s, s2)
                    x1 = min(e, e2)
                    if x1 > x0:
                        overlap_s += (x1 - x0)
                clash_seconds.loc[idx] = overlap_s
        clash_pct = pd.Series(
            np.where(est_overlap > 0.0, 100.0 * clash_seconds / est_overlap, 0.0),
            index=df_out.index,
            dtype=float,
        ).clip(lower=0.0, upper=100.0)
        return clash_count, clash_seconds, clash_pct

    b_clash_count, b_clash_seconds, b_clash_pct = _overlap_clash_metrics(
        "file_b", "timeline_est_b_start_s", "timeline_est_b_end_s"
    )
    a_clash_count, a_clash_seconds, a_clash_pct = _overlap_clash_metrics(
        "file_a", "timeline_est_a_start_s", "timeline_est_a_end_s"
    )
    df_out["timeline_b_overlap_clash_count"] = b_clash_count
    df_out["timeline_b_overlap_clash_s"] = b_clash_seconds.round(1)
    df_out["timeline_b_overlap_clash_pct"] = b_clash_pct.round(1)
    df_out["timeline_a_overlap_clash_count"] = a_clash_count
    df_out["timeline_a_overlap_clash_s"] = a_clash_seconds.round(1)
    df_out["timeline_a_overlap_clash_pct"] = a_clash_pct.round(1)
    df_out["timeline_b_overlap_multimatch"] = b_clash_count > 0
    df_out["timeline_a_overlap_multimatch"] = a_clash_count > 0
    b_missing_s_series = (dur_s_series - cov_s_series).clip(lower=0.0)
    a_missing_s_series = (dur_a_series - cov_a_s_series).clip(lower=0.0)
    b_is_full_series = (
        (pd.to_numeric(df_out["timeline_b_coverage_est_pct"], errors="coerce").fillna(0.0) >= float(TIMELINE_COVERAGE_FULL_MIN_PCT))
        | (b_missing_s_series <= float(TIMELINE_COVERAGE_FULL_MAX_MISSING_S))
    )
    a_is_full_series = (
        (pd.to_numeric(df_out["timeline_a_coverage_est_pct"], errors="coerce").fillna(0.0) >= float(TIMELINE_COVERAGE_FULL_MIN_PCT))
        | (a_missing_s_series <= float(TIMELINE_COVERAGE_FULL_MAX_MISSING_S))
    )
    df_out["timeline_b_has_unmatched_parts"] = ~b_is_full_series
    df_out["timeline_a_has_unmatched_parts"] = ~a_is_full_series
    df_out["timeline_b_has_internal_unmatched_gaps"] = (
        (pd.to_numeric(df_out["timeline_b_coverage_sections_est"], errors="coerce").fillna(0).astype(int) >= 2)
        & (~b_is_full_series)
    )
    df_out["timeline_a_has_internal_unmatched_gaps"] = (
        (pd.to_numeric(df_out["timeline_a_coverage_sections_est"], errors="coerce").fillna(0).astype(int) >= 2)
        & (~a_is_full_series)
    )

    def _format_clash_hint(overlap_s: float, overlap_pct: float) -> str:
        s = max(0.0, float(overlap_s))
        p = max(0.0, min(100.0, float(overlap_pct)))
        # Short overlap is easier to reason about in seconds; larger overlap in percent.
        if s < 180.0:
            return f"~{s:.0f}s overlap"
        return f"~{p:.1f}% overlap"

    def _coverage_phrase(
        subject: str,
        pct: float,
        count: int,
        sections: int,
        unmatched_desc: str = "",
        uncertain_desc: str = "",
        *,
        covered_s: float | None = None,
        duration_s: float | None = None,
    ) -> str:
        if count <= 0:
            return f"{subject}: no matches"
        is_full = False
        if pct >= float(TIMELINE_COVERAGE_FULL_MIN_PCT):
            is_full = True
        if (covered_s is not None) and (duration_s is not None):
            try:
                missing_s = max(0.0, float(duration_s) - float(covered_s))
                if missing_s <= float(TIMELINE_COVERAGE_FULL_MAX_MISSING_S):
                    is_full = True
            except Exception:
                pass
        if is_full:
            base = f"{subject}: fully covered by {count} match(es)"
        elif pct >= float(TIMELINE_COVERAGE_MOSTLY_MIN_PCT):
            base = f"{subject}: mostly covered ({pct:.1f}%) by {count} match(es)"
        else:
            base = f"{subject}: partially covered ({pct:.1f}%) by {count} match(es)"
        if sections <= 1:
            detail = "contiguous coverage"
        else:
            detail = f"split into {sections} matched sections"
        include_uncertain = bool(uncertain_desc) and (not is_full)
        if include_uncertain and unmatched_desc and ("no unmatched region" not in unmatched_desc.lower()):
            include_uncertain = False
        if include_uncertain:
            detail = f"{detail}; {uncertain_desc}"
        if (not is_full) and unmatched_desc and ("no unmatched region" not in unmatched_desc.lower()):
            detail = f"{detail}; {unmatched_desc}"
        return f"{base}; {detail}"

    a_cov_num = pd.to_numeric(df_out["timeline_a_coverage_est_pct"], errors="coerce").fillna(0.0)
    a_unique_cov_num = pd.to_numeric(df_out["timeline_a_unique_covered_est_pct"], errors="coerce").fillna(0.0)
    b_cov_num = pd.to_numeric(df_out["timeline_b_coverage_est_pct"], errors="coerce").fillna(0.0)
    a_sec_num = pd.to_numeric(df_out["timeline_a_coverage_sections_est"], errors="coerce").fillna(0).astype(int)
    a_unique_sec_num = pd.to_numeric(df_out["timeline_a_unique_coverage_sections_est"], errors="coerce").fillna(0).astype(int)
    b_sec_num = pd.to_numeric(df_out["timeline_b_coverage_sections_est"], errors="coerce").fillna(0).astype(int)
    a_cnt_num = df_out.groupby("file_a")["file_b"].transform("size").astype(int)
    b_cnt_num = pd.to_numeric(df_out["group_candidate_count"], errors="coerce").fillna(0).astype(int)
    cov_b_desc_norm_local = {_norm_path(k): str(v) for k, v in coverage_b_unmatched_desc_map.items()}
    cov_b_uncertain_desc_norm_local = {_norm_path(k): str(v) for k, v in coverage_b_uncertain_desc_map.items()}
    cov_a_desc_norm_local = {_norm_path(k): str(v) for k, v in coverage_a_unmatched_desc_map.items()}
    cov_a_unique_uncertain_desc_norm_local = {_norm_path(k): str(v) for k, v in coverage_a_unique_uncertain_desc_map.items()}

    df_out["file_b_group_summary"] = [
        _coverage_phrase(
            "file_b",
            float(p),
            int(c),
            int(s),
            str(cov_b_desc_norm_local.get(_norm_path(str(fb)), "")),
            str(cov_b_uncertain_desc_norm_local.get(_norm_path(str(fb)), "")),
            covered_s=float(cs),
            duration_s=float(ds),
        )
        for p, c, s, fb, cs, ds in zip(
            b_cov_num,
            b_cnt_num,
            b_sec_num,
            df_out["file_b"].astype(str),
            cov_s_series,
            dur_s_series,
        )
    ]
    df_out["file_a_group_summary"] = [
        _coverage_phrase(
            "file_a unique content",
            float(pu),
            int(c),
            int(su),
            str(cov_a_desc_norm_local.get(_norm_path(str(fa)), "")),
            str(cov_a_unique_uncertain_desc_norm_local.get(_norm_path(str(fa)), "")),
            covered_s=float(cus),
            duration_s=float(dus),
        )
        for pu, c, su, fa, cus, dus in zip(
            a_unique_cov_num,
            a_cnt_num,
            a_unique_sec_num,
            df_out["file_a"].astype(str),
            cov_a_unique_s_series,
            dur_a_unique_series,
        )
    ]

    pair_summaries: list[str] = []
    for idx in df_out.index:
        parts: list[str] = []
        if bool(df_out.at[idx, "anchor_full_full_conflict"]):
            verify_status = str(df_out.at[idx, "timeline_gap_verify_status"] or "")
            verify_notes = str(df_out.at[idx, "timeline_gap_verify_notes"] or "")
            if verify_status == "verified":
                parts.append("anchors full<->full conflict resolved by A-side recovery")
            elif verify_status == "partial":
                parts.append("anchors full<->full conflict partially resolved by A-side recovery")
            elif verify_status:
                parts.append(f"anchors full<->full conflict remains ({verify_status})")
            else:
                parts.append("anchors full<->full conflict remains")
            if verify_notes:
                parts.append(verify_notes)
        if bool(df_out.at[idx, "timeline_b_overlap_multimatch"]):
            b_s = pd.to_numeric(df_out.at[idx, "timeline_b_overlap_clash_s"], errors="coerce")
            b_p = pd.to_numeric(df_out.at[idx, "timeline_b_overlap_clash_pct"], errors="coerce")
            b_s = 0.0 if pd.isna(b_s) else float(b_s)
            b_p = 0.0 if pd.isna(b_p) else float(b_p)
            parts.append(
                f"B-interval overlaps another file_a match ({_format_clash_hint(b_s, b_p)})"
            )
        if bool(df_out.at[idx, "timeline_a_overlap_multimatch"]):
            a_s = pd.to_numeric(df_out.at[idx, "timeline_a_overlap_clash_s"], errors="coerce")
            a_p = pd.to_numeric(df_out.at[idx, "timeline_a_overlap_clash_pct"], errors="coerce")
            a_s = 0.0 if pd.isna(a_s) else float(a_s)
            a_p = 0.0 if pd.isna(a_p) else float(a_p)
            parts.append(
                f"A-interval overlaps another file_b match ({_format_clash_hint(a_s, a_p)})"
            )
        if bool(df_out.at[idx, "timeline_b_has_internal_unmatched_gaps"]):
            parts.append("internal unmatched gap(s) remain in file_b")
        if bool(df_out.at[idx, "timeline_a_has_internal_unmatched_gaps"]):
            parts.append("internal unmatched gap(s) remain in file_a")
        b_uncertain_desc = str(cov_b_uncertain_desc_norm_local.get(_norm_path(str(df_out.at[idx, "file_b"])), ""))
        a_uncertain_desc = str(cov_a_unique_uncertain_desc_norm_local.get(_norm_path(str(df_out.at[idx, "file_a"])), ""))
        b_cov_pct = float(pd.to_numeric(df_out.at[idx, "timeline_b_coverage_est_pct"], errors="coerce"))
        a_unique_cov_pct = float(pd.to_numeric(df_out.at[idx, "timeline_a_unique_covered_est_pct"], errors="coerce"))
        if b_uncertain_desc and (float(TIMELINE_COVERAGE_MOSTLY_MIN_PCT) <= b_cov_pct < float(TIMELINE_COVERAGE_FULL_MIN_PCT)):
            parts.append(f"file_b has {b_uncertain_desc}")
        if a_uncertain_desc and (float(TIMELINE_COVERAGE_MOSTLY_MIN_PCT) <= a_unique_cov_pct < float(TIMELINE_COVERAGE_FULL_MIN_PCT)):
            parts.append(f"file_a unique-content has {a_uncertain_desc}")
        a_repeat_pct = pd.to_numeric(df_out.at[idx, "timeline_a_self_repeat_redundant_pct"], errors="coerce")
        a_repeat_pct = 0.0 if pd.isna(a_repeat_pct) else float(a_repeat_pct)
        if a_repeat_pct >= 5.0:
            parts.append(f"file_a contains ~{a_repeat_pct:.1f}% repeated playback")
        if not parts:
            parts.append("clean timeline placement (no overlap clash / no internal gap flags)")
        pair_summaries.append("; ".join(parts))
    df_out["match_summary"] = pair_summaries

    # Final review flags: keep this focused on true manual-risk cases.
    # Partial coverage is expected for split tapes and is shown in coverage sheets,
    # but does not by itself require row-level review.
    SAFE_A_COVERAGE_MIN_PCT = 95.0
    REVIEW_LOW_CONFIDENCE_MIN = 70.0
    b_clash_pct_num = pd.to_numeric(df_out["timeline_b_overlap_clash_pct"], errors="coerce").fillna(0.0)
    b_clash_count_num = pd.to_numeric(df_out["timeline_b_overlap_clash_count"], errors="coerce").fillna(0).astype(int)
    a_clash_pct_num = pd.to_numeric(df_out["timeline_a_overlap_clash_pct"], errors="coerce").fillna(0.0)
    a_clash_count_num = pd.to_numeric(df_out["timeline_a_overlap_clash_count"], errors="coerce").fillna(0).astype(int)
    a_cov_pct_num = pd.to_numeric(df_out["timeline_a_unique_covered_est_pct"], errors="coerce").fillna(0.0)
    conf_num = pd.to_numeric(df_out["confidence_score"], errors="coerce").fillna(0.0)
    review_has_clash = (
        (b_clash_pct_num > 5.0)
        | (a_clash_pct_num > 5.0)
        | (((b_clash_count_num + a_clash_count_num) > 0) & (conf_num < 70.0))
    )
    review_low_conf = conf_num < REVIEW_LOW_CONFIDENCE_MIN
    review_a_coverage_gap = a_cov_pct_num < SAFE_A_COVERAGE_MIN_PCT
    verify_status_s = df_out["timeline_gap_verify_status"].fillna("").astype(str)
    review_anchor_timeline_conflict = (
        df_out["anchor_full_full_conflict"].fillna(False).astype(bool)
        & verify_status_s.isin(["failed", "partial", "open_failed", "missing_file"])
    )

    _add_review_flag(review_has_clash, "overlap_multimatch")
    _add_review_flag(review_a_coverage_gap, "a_unique_coverage_lt_95pct")
    _add_review_flag(review_low_conf, "low_confidence")
    _add_review_flag(review_anchor_timeline_conflict, "anchor_timeline_conflict")
    df_out["review_flags"] = review_flags
    df_out["needs_review"] = (df_out["review_flags"] != "")
    df_out["review_priority"] = np.where(
        review_has_clash,
        "high",
        np.where(review_a_coverage_gap | review_low_conf, "medium", "none"),
    )

    # Rename workflow helpers.
    df_out["file_a_candidate_count"] = df_out.groupby("file_a")["file_b"].transform("size").astype(int)
    df_out["file_b_candidate_count"] = df_out["group_candidate_count"].astype(int)
    b_start_num = pd.to_numeric(df_out["timeline_est_b_start_s"], errors="coerce")
    df_out["file_b_part_order"] = (
        b_start_num.groupby(df_out["file_b"]).rank(method="dense", na_option="bottom").fillna(0).astype(int)
    )
    min_safe_overlap_s = 120.0
    rename_ready = (
        (~df_out["needs_review"])
        & (pd.to_numeric(df_out["confidence_score"], errors="coerce").fillna(0.0) >= 85.0)
        & (pd.to_numeric(df_out["timeline_est_overlap_s"], errors="coerce").fillna(0.0) >= min_safe_overlap_s)
        & (pd.to_numeric(df_out["timeline_a_unique_covered_est_pct"], errors="coerce").fillna(0.0) >= SAFE_A_COVERAGE_MIN_PCT)
    )
    if "matched_by_timeline" in df_out.columns:
        rename_ready = rename_ready & df_out["matched_by_timeline"].astype(bool)
    df_out["rename_ready"] = rename_ready

    # Sorting: strong groups first, keep file_b grouped, strong rows first within each group.
    df_out["_file_b_sort"] = df_out["file_b"].map(
        lambda x: os.path.normcase(os.path.normpath(x)) if isinstance(x, str) else ""
    )
    df_out["_file_a_sort"] = df_out["file_a"].map(
        lambda x: os.path.normcase(os.path.normpath(x)) if isinstance(x, str) else ""
    )
    df_out = df_out.sort_values(
        by=[
            "group_best_confidence",
            "group_confidence_gap",
            "_file_b_sort",
            "confidence_score",
            "evidence_score",
            "matched_by_count",
            "_file_a_sort",
        ],
        ascending=[False, False, True, False, False, False, True],
    ).drop(columns=["_file_b_sort", "_file_a_sort"]).reset_index(drop=True)

    # Export a compact Consolidated sheet (reduce noise from raw mode metrics).
    consolidated_view = _select_existing_columns(
        df_out,
        _consolidated_sheet_columns(mode_names, list(df_out.columns)),
    )

    # Build full-file coverage sheets for A and B (include matched and unmatched files).

    runtime_basis_cache: dict[str, tuple[float, float]] = {}

    def _get_runtime_basis(path: str) -> tuple[float, float]:
        """
        Return (raw_duration_s, playable_duration_s) for report rows.

        playable_duration_s excludes dead head/tail and any inferred internal dead sections,
        but not A-side self-repeat removal. That stays in the A-unique coverage metrics.
        """
        norm = _norm_path(path)
        cached = runtime_basis_cache.get(norm)
        if cached is not None:
            return cached

        raw_duration_s = 0.0
        for source_map in (duration_a_map, duration_b_map):
            if path in source_map:
                try:
                    raw_duration_s = max(raw_duration_s, float(source_map[path]))
                except Exception:
                    pass
        try:
            probe_duration_s = float(ffprobe_duration_seconds(path))
        except Exception:
            probe_duration_s = 0.0
        raw_duration_s = max(raw_duration_s, max(0.0, probe_duration_s))

        if raw_duration_s > 0.0:
            _trim_head, _trim_tail, _play_start, _play_end, playable_duration_s, _dead = _resolve_playable_window(
                path,
                raw_duration_s,
            )
        else:
            playable_duration_s = 0.0

        out = (float(raw_duration_s), float(playable_duration_s))
        runtime_basis_cache[norm] = out
        return out

    timeline_speed_ratio_min_eff, timeline_speed_ratio_max_eff, timeline_speed_steps_eff = _resolve_timeline_speed_search(
        TIMELINE_ENABLE_SPEED_SWEEP,
        TIMELINE_SPEED_RATIO_MIN,
        TIMELINE_SPEED_RATIO_MAX,
        TIMELINE_SPEED_STEPS,
        fixed_ratio=TIMELINE_SPEED_RATIO_FIXED,
    )
    group_timeline_seq_cache: dict[str, list[tuple[float, int]]] = {}
    group_anchor_cache: dict[str, dict[str, list[tuple[float, bytes]]]] = {}

    def _get_group_timeline_seq(path: str) -> list[tuple[float, int]]:
        return _get_filtered_timeline_seq_cached(vhs_gap, group_timeline_seq_cache, path)

    def _get_group_anchors(path: str) -> dict[str, list[tuple[float, bytes]]]:
        return _get_cached_anchors(vhs_gap, group_anchor_cache, path)

    def _interval_total_s(intervals: list[tuple[float, float]]) -> float:
        return float(sum(max(0.0, float(e) - float(s)) for s, e in intervals if float(e) > float(s)))

    def _classify_group_b_relation(cov1_pct: float, cov2_pct: float, ok_timeline: bool) -> str:
        if cov1_pct >= float(TIMELINE_COVERAGE_FULL_MIN_PCT) and cov2_pct >= float(TIMELINE_COVERAGE_FULL_MIN_PCT):
            return "full_duplicate"
        if cov1_pct >= float(TIMELINE_COVERAGE_FULL_MIN_PCT):
            return "file_b_1_within_file_b_2"
        if cov2_pct >= float(TIMELINE_COVERAGE_FULL_MIN_PCT):
            return "file_b_2_within_file_b_1"
        if ok_timeline and max(cov1_pct, cov2_pct) >= 30.0:
            return "partial_overlap"
        return "no_clear_overlap"

    def _compare_group_b_pair(group_id: str, file_b_1: str, file_b_2: str) -> dict[str, object]:
        anc_1 = _get_group_anchors(file_b_1)
        anc_2 = _get_group_anchors(file_b_2)
        anchor_stats = compare_anchors(anc_1, anc_2, hamming_thresh=ANCHOR_HAMMING_THRESH)
        anchor_ok, anchor_relation = decide_subset_match(
            anchor_stats,
            min_fraction=ANCHOR_MIN_FRACTION,
            max_mad_s=ANCHOR_MAX_MAD_S,
        )
        if not anchor_ok:
            anchor_relation = "ambiguous"

        seq_1 = _get_group_timeline_seq(file_b_1)
        seq_2 = _get_group_timeline_seq(file_b_2)
        ok_timeline = False
        details = {
            "overlap_strict_s": 0.0,
            "segment_count_strict": 0,
            "segments_a_strict": "",
            "segments_b_strict": "",
        }
        if seq_1 and seq_2:
            ok_timeline, _offset_s, _votes_n, _overlap_n = _audio_match_offset(
                seq_1,
                seq_2,
                total_bits=64,
                hamming_thresh=TIMELINE_HAMMING_THRESH,
                bin_s=TIMELINE_BIN_S,
                min_votes=TIMELINE_MIN_VOTES,
                min_overlap_s=TIMELINE_MIN_OVERLAP_S,
                hop_s=TIMELINE_STEP_S,
                lsh_chunks=TIMELINE_LSH_CHUNKS,
                brute_limit=TIMELINE_BRUTE_MAX,
                strict_hamming_thresh=TIMELINE_STRICT_HAMMING_THRESH,
                strict_min_overlap_s=TIMELINE_STRICT_MIN_OVERLAP_S,
                min_vote_fraction=TIMELINE_MIN_VOTE_FRACTION,
                peak_ratio_min=TIMELINE_PEAK_RATIO_MIN,
                peak_margin=TIMELINE_PEAK_MARGIN,
                speed_ratio_min=timeline_speed_ratio_min_eff,
                speed_ratio_max=timeline_speed_ratio_max_eff,
                speed_steps=timeline_speed_steps_eff,
                max_candidates_per_frame=TIMELINE_MAX_CANDIDATES_PER_FRAME,
                long_overlap_override_s=TIMELINE_LONG_OVERLAP_OVERRIDE_S,
                long_overlap_vote_mult=TIMELINE_LONG_OVERLAP_VOTE_MULT,
                run_gap_mult=TIMELINE_RUN_GAP_MULT,
            )
            details = _timeline_postfit_details(
                seq_1,
                seq_2,
                total_bits=64,
                loose_hamming_thresh=TIMELINE_HAMMING_THRESH,
                strict_hamming_thresh=TIMELINE_STRICT_HAMMING_THRESH,
                bin_s=TIMELINE_BIN_S,
                step_s=TIMELINE_STEP_S,
                lsh_chunks=TIMELINE_LSH_CHUNKS,
                brute_limit=TIMELINE_BRUTE_MAX,
                speed_ratio_min=timeline_speed_ratio_min_eff,
                speed_ratio_max=timeline_speed_ratio_max_eff,
                speed_steps=timeline_speed_steps_eff,
                max_candidates_per_frame=TIMELINE_MAX_CANDIDATES_PER_FRAME,
                run_gap_mult=TIMELINE_RUN_GAP_MULT,
                max_matches=TIMELINE_POSTFIT_MAX_MATCHES,
                seed_gap_s=TIMELINE_POSTFIT_SEED_GAP_S,
                return_debug=False,
            )

        overlap_b1_s = float(details.get("overlap_strict_s") or 0.0)
        segments_b2 = _parse_segments_text(str(details.get("segments_b_strict") or ""))
        overlap_b2_s = _interval_total_s(segments_b2)
        _raw_1, play_1 = _get_runtime_basis(file_b_1)
        _raw_2, play_2 = _get_runtime_basis(file_b_2)
        overlap_b1_pct = (100.0 * overlap_b1_s / play_1) if play_1 > 0 else 0.0
        overlap_b2_pct = (100.0 * overlap_b2_s / play_2) if play_2 > 0 else 0.0
        timeline_relation = _classify_group_b_relation(overlap_b1_pct, overlap_b2_pct, bool(ok_timeline))
        summary = (
            f"{timeline_relation}; "
            f"{os.path.basename(file_b_1)}={overlap_b1_pct:.1f}%, "
            f"{os.path.basename(file_b_2)}={overlap_b2_pct:.1f}%"
        )
        return {
            "group_id": group_id,
            "file_b_1": file_b_1,
            "file_b_2": file_b_2,
            "timeline_relation": timeline_relation,
            "timeline_match_ok": bool(ok_timeline),
            "timeline_overlap_b1_pct": round(max(0.0, min(100.0, overlap_b1_pct)), 1),
            "timeline_overlap_b2_pct": round(max(0.0, min(100.0, overlap_b2_pct)), 1),
            "timeline_overlap_b1_s": round(max(0.0, overlap_b1_s), 1),
            "timeline_overlap_b2_s": round(max(0.0, overlap_b2_s), 1),
            "timeline_segment_count": int(details.get("segment_count_strict") or 0),
            "timeline_segments_b1": str(details.get("segments_a_strict") or ""),
            "timeline_segments_b2": str(details.get("segments_b_strict") or ""),
            "anchor_relation": str(anchor_relation),
            "anchor_start_fraction": round(float(anchor_stats.get("start_fraction", 0.0) or 0.0), 3),
            "anchor_end_fraction": round(float(anchor_stats.get("end_fraction", 0.0) or 0.0), 3),
            "summary": summary,
        }

    (
        b_coverage_df,
        a_coverage_df,
        unmatched_a_df,
        unmatched_b_df,
        b_matches_norm,
        a_matches_norm,
        b_multimatch_norm,
        b_internal_gap_norm,
        a_multimatch_norm,
        a_internal_gap_norm,
        all_b_files,
        all_a_files,
        b_best_conf,
        a_best_conf,
        b_avg_conf,
        a_avg_conf,
    ) = _build_coverage_views(
        df_out=df_out,
        folder_a=folder_a,
        folder_b=folder_b,
        listed_folder_a_files=listed_folder_a_files,
        listed_folder_b_files=listed_folder_b_files,
        timestamp_map=timestamp_map,
        coverage_b_s_map=coverage_b_s_map,
        coverage_b_raw_s_map=coverage_b_raw_s_map,
        duration_b_map=duration_b_map,
        coverage_b_sections_raw_map=coverage_b_sections_raw_map,
        coverage_b_sections_est_map=coverage_b_sections_est_map,
        coverage_b_unmatched_desc_map=coverage_b_unmatched_desc_map,
        coverage_b_uncertain_desc_map=coverage_b_uncertain_desc_map,
        coverage_a_s_map=coverage_a_s_map,
        coverage_a_raw_s_map=coverage_a_raw_s_map,
        duration_a_map=duration_a_map,
        coverage_a_unique_s_map=coverage_a_unique_s_map,
        coverage_a_unique_duration_map=coverage_a_unique_duration_map,
        coverage_a_self_repeat_s_map=coverage_a_self_repeat_s_map,
        coverage_a_sections_raw_map=coverage_a_sections_raw_map,
        coverage_a_sections_est_map=coverage_a_sections_est_map,
        coverage_a_unique_sections_est_map=coverage_a_unique_sections_est_map,
        coverage_a_unmatched_desc_map=coverage_a_unmatched_desc_map,
        coverage_a_unique_uncertain_desc_map=coverage_a_unique_uncertain_desc_map,
        get_runtime_basis=_get_runtime_basis,
        coverage_phrase_fn=_coverage_phrase,
    )
    df_work = df_out.copy()
    df_work["_file_a_norm"] = df_work["file_a"].map(lambda x: _norm_path(x) if isinstance(x, str) else "")
    df_work["_file_b_norm"] = df_work["file_b"].map(lambda x: _norm_path(x) if isinstance(x, str) else "")
    ts_start_norm = {k: str(v.get("start", "") or "") for k, v in timestamp_map.items()}
    ts_end_norm = {k: str(v.get("end", "") or "") for k, v in timestamp_map.items()}
    ts_span_norm = {k: int(v.get("span_months", 0) or 0) for k, v in timestamp_map.items()}
    ts_hits_norm = {k: int(v.get("hits", 0) or 0) for k, v in timestamp_map.items()}
    ts_out_order_norm = {k: bool(v.get("out_of_order", False)) for k, v in timestamp_map.items()}
    coverage_b_s_norm = {_norm_path(k): float(v) for k, v in coverage_b_s_map.items()}
    duration_b_norm = {_norm_path(k): float(v) for k, v in duration_b_map.items()}
    coverage_b_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_b_sections_est_map.items()}
    coverage_b_unmatched_desc_norm = {_norm_path(k): str(v) for k, v in coverage_b_unmatched_desc_map.items()}
    coverage_b_uncertain_desc_norm = {_norm_path(k): str(v) for k, v in coverage_b_uncertain_desc_map.items()}
    coverage_a_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_s_map.items()}
    duration_a_norm = {_norm_path(k): float(v) for k, v in duration_a_map.items()}
    coverage_a_unique_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_unique_s_map.items()}
    coverage_a_unique_duration_norm = {_norm_path(k): float(v) for k, v in coverage_a_unique_duration_map.items()}
    coverage_a_self_repeat_s_norm = {_norm_path(k): float(v) for k, v in coverage_a_self_repeat_s_map.items()}
    coverage_a_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_a_sections_est_map.items()}
    coverage_a_unique_sections_est_norm = {_norm_path(k): int(v) for k, v in coverage_a_unique_sections_est_map.items()}
    coverage_a_unmatched_desc_norm = {_norm_path(k): str(v) for k, v in coverage_a_unmatched_desc_map.items()}
    coverage_a_unique_uncertain_desc_norm = {_norm_path(k): str(v) for k, v in coverage_a_unique_uncertain_desc_map.items()}
    b_match_count = {
        _norm_path(str(row["file_b"])): int(row["matched_file_a_count"])
        for _, row in b_coverage_df.iterrows()
        if isinstance(row.get("file_b"), str)
    }
    a_match_count = {
        _norm_path(str(row["file_a"])): int(row["matched_file_b_count"])
        for _, row in a_coverage_df.iterrows()
        if isinstance(row.get("file_a"), str)
    }

    # Preserve manual workflow status/notes across report rebuilds.
    prior_queue_state, prior_remux_state, prior_remux_state_by_path, has_done_local = _load_prior_workflow_and_remux_state(
        output_path=output_path,
        state_source_path=state_source_path,
    )
    done_status_tokens = {"done", "complete", "completed", "archived"}
    if prior_queue_state:
        loaded_done_count = int(
            sum(
                1
                for v in prior_queue_state.values()
                if re.sub(r"[^a-z]+", "", str(v.get("workflow_status", "")).strip().lower()) in done_status_tokens
            )
        )
        print(
            f"[consolidate] prior workflow states loaded: {len(prior_queue_state)} (done={loaded_done_count})",
            flush=True,
        )

    # Build connected match groups (A/B bipartite components) for workflow.
    def _path_sort_key(path: str) -> str:
        return os.path.normcase(os.path.normpath(path)) if isinstance(path, str) else ""

    node_to_path: dict[str, str] = {}
    node_to_side: dict[str, str] = {}
    adjacency: dict[str, set[str]] = {}
    for path in all_b_files:
        norm = _norm_path(path)
        node = f"B|{norm}"
        node_to_path[node] = path
        node_to_side[node] = "B"
        adjacency.setdefault(node, set())
    for path in all_a_files:
        norm = _norm_path(path)
        node = f"A|{norm}"
        node_to_path[node] = path
        node_to_side[node] = "A"
        adjacency.setdefault(node, set())

    for _, row in df_work.iterrows():
        fa = row.get("file_a")
        fb = row.get("file_b")
        if not isinstance(fa, str) or not isinstance(fb, str):
            continue
        an = _norm_path(fa)
        bn = _norm_path(fb)
        a_node = f"A|{an}"
        b_node = f"B|{bn}"
        if a_node not in adjacency or b_node not in adjacency:
            continue
        adjacency[a_node].add(b_node)
        adjacency[b_node].add(a_node)

    def _node_sort_key(node: str) -> tuple[int, str]:
        side = node_to_side.get(node, "Z")
        return (0 if side == "B" else 1, _path_sort_key(node_to_path.get(node, "")))

    components: list[list[str]] = []
    seen_nodes: set[str] = set()
    for node in sorted(adjacency.keys(), key=_node_sort_key):
        if node in seen_nodes:
            continue
        stack = [node]
        comp: list[str] = []
        while stack:
            cur = stack.pop()
            if cur in seen_nodes:
                continue
            seen_nodes.add(cur)
            comp.append(cur)
            for nb in adjacency.get(cur, set()):
                if nb not in seen_nodes:
                    stack.append(nb)
        components.append(comp)

    def _component_sort_key(comp: list[str]) -> tuple[int, str]:
        b_paths = sorted(_path_sort_key(node_to_path[n]) for n in comp if node_to_side.get(n) == "B")
        a_paths = sorted(_path_sort_key(node_to_path[n]) for n in comp if node_to_side.get(n) == "A")
        if b_paths:
            return (0, b_paths[0])
        if a_paths:
            return (1, a_paths[0])
        return (2, "")

    components = sorted(components, key=_component_sort_key)

    rename_queue_rows: list[dict[str, object]] = []
    b_group_compare_rows: list[dict[str, object]] = []
    b_group_summary_map: dict[str, list[str]] = {}
    for g_idx, comp in enumerate(components, start=1):
        group_id = f"G{g_idx:03d}"
        members = sorted(comp, key=_node_sort_key)
        group_a_count = int(sum(1 for n in members if node_to_side.get(n) == "A"))
        group_b_count = int(sum(1 for n in members if node_to_side.get(n) == "B"))
        if group_a_count > 0 and group_b_count > 0:
            group_kind = "matched_group"
        elif group_a_count > 0:
            group_kind = "unmatched_a"
        else:
            group_kind = "unmatched_b"

        if group_b_count >= 2:
            b_paths_in_group = sorted(
                [node_to_path[n] for n in members if node_to_side.get(n) == "B" and isinstance(node_to_path.get(n), str)],
                key=_path_sort_key,
            )
            for i in range(len(b_paths_in_group)):
                for j in range(i + 1, len(b_paths_in_group)):
                    pair_row = _compare_group_b_pair(group_id, b_paths_in_group[i], b_paths_in_group[j])
                    b_group_compare_rows.append(pair_row)
                    rel = str(pair_row.get("timeline_relation", "") or "")
                    p1 = str(pair_row.get("file_b_1", "") or "")
                    p2 = str(pair_row.get("file_b_2", "") or "")
                    pct1 = float(pair_row.get("timeline_overlap_b1_pct", 0.0) or 0.0)
                    pct2 = float(pair_row.get("timeline_overlap_b2_pct", 0.0) or 0.0)
                    if p1:
                        b_group_summary_map.setdefault(_norm_path(p1), []).append(
                            f"{os.path.basename(p2)}: {rel} ({pct1:.1f}% of this, {pct2:.1f}% of other)"
                        )
                    if p2:
                        b_group_summary_map.setdefault(_norm_path(p2), []).append(
                            f"{os.path.basename(p1)}: {rel} ({pct2:.1f}% of this, {pct1:.1f}% of other)"
                        )

        for m_idx, node in enumerate(members, start=1):
            side = node_to_side.get(node, "")
            path = node_to_path.get(node, "")
            norm = _norm_path(path)
            state = prior_queue_state.get(norm, {})

            base_row = {
                "row_type": "file",
                "group_id": group_id,
                "group_kind": group_kind,
                "group_file_count": int(len(members)),
                "group_file_a_count": group_a_count,
                "group_file_b_count": group_b_count,
                "file_side": side,
                "file_path": path,
                "file_name": os.path.basename(path) if isinstance(path, str) else "",
                "osd_month_year_start": _safe_text_cell(ts_start_norm.get(norm, "")),
                "osd_month_year_end": _safe_text_cell(ts_end_norm.get(norm, "")),
                "osd_span_months": int(ts_span_norm.get(norm, 0)),
                "osd_hits": int(ts_hits_norm.get(norm, 0)),
                "same_group_file_b_summary": "",
                "workflow_status": _safe_text_cell(state.get("workflow_status", "")),
                "proposed_new_name": _safe_text_cell(state.get("proposed_new_name", "")),
                "decision": _safe_text_cell(state.get("decision", "")),
                "notes": _safe_text_cell(state.get("notes", "")),
                "_group_order": int(g_idx),
                "_member_order": int(m_idx),
            }

            if side == "B":
                raw_runtime_s, playable_runtime_s = _get_runtime_basis(path)
                dur_s = float(duration_b_norm.get(norm, playable_runtime_s))
                cov_est_s = float(coverage_b_s_norm.get(norm, 0.0))
                cov_est_pct = (100.0 * cov_est_s / dur_s) if dur_s > 0 else 0.0
                sections = int(coverage_b_sections_est_norm.get(norm, 0))
                unmatched_desc = str(coverage_b_unmatched_desc_norm.get(norm, ""))
                uncertain_desc = str(coverage_b_uncertain_desc_norm.get(norm, ""))
                has_middle_gap = ("middle" in unmatched_desc.lower())
                coverage_summary = _coverage_phrase(
                    "file_b",
                    float(max(0.0, min(100.0, cov_est_pct))),
                    int(b_match_count.get(norm, 0)),
                    sections,
                    unmatched_desc,
                    uncertain_desc,
                    covered_s=float(cov_est_s),
                    duration_s=float(dur_s),
                )
                row = {
                    **base_row,
                    "matched_file_count": int(b_match_count.get(norm, 0)),
                    "matched_file_names": " | ".join(b_matches_norm.get(norm, [])),
                    "same_group_file_b_summary": " | ".join(b_group_summary_map.get(norm, [])),
                    "best_pair_confidence_score": float(b_best_conf.get(norm, 0.0)),
                    "avg_pair_confidence_score": float(b_avg_conf.get(norm, 0.0)),
                    "video_runtime_s": round(raw_runtime_s, 1),
                    # B-side currently has dead-section removal only; no B self-repeat model yet.
                    "runtime_ex_dead_and_duplicate_s": round(playable_runtime_s, 1),
                    "timeline_a_unique_covered_est_pct": "",
                    "timeline_a_self_repeat_redundant_pct": "",
                    "timeline_coverage_est_pct": round(max(0.0, min(100.0, cov_est_pct)), 1),
                    "timeline_coverage_sections_est": sections,
                    "overlap_multimatch": bool(b_multimatch_norm.get(norm, False)),
                    "has_internal_unmatched_gaps": bool(b_internal_gap_norm.get(norm, False)),
                    "has_middle_unmatched_gap": bool(has_middle_gap),
                    "coverage_summary": coverage_summary,
                }
            else:
                raw_runtime_s, playable_runtime_s = _get_runtime_basis(path)
                dur_s = float(duration_a_norm.get(norm, playable_runtime_s))
                cov_est_s = float(coverage_a_s_norm.get(norm, 0.0))
                cov_est_pct = (100.0 * cov_est_s / dur_s) if dur_s > 0 else 0.0
                unique_cov_est_s = float(coverage_a_unique_s_norm.get(norm, 0.0))
                unique_dur_s = float(coverage_a_unique_duration_norm.get(norm, max(0.0, playable_runtime_s)))
                unique_cov_est_pct = (100.0 * unique_cov_est_s / unique_dur_s) if unique_dur_s > 0 else 0.0
                repeat_s = float(coverage_a_self_repeat_s_norm.get(norm, 0.0))
                repeat_pct = (100.0 * repeat_s / dur_s) if dur_s > 0 else 0.0
                sections = int(coverage_a_sections_est_norm.get(norm, 0))
                unmatched_desc = str(coverage_a_unmatched_desc_norm.get(norm, ""))
                uncertain_desc = str(coverage_a_unique_uncertain_desc_norm.get(norm, ""))
                has_middle_gap = ("middle" in unmatched_desc.lower())
                coverage_summary = _coverage_phrase(
                    "file_a unique content",
                    float(max(0.0, min(100.0, unique_cov_est_pct))),
                    int(a_match_count.get(norm, 0)),
                    int(coverage_a_unique_sections_est_norm.get(norm, 0)),
                    unmatched_desc,
                    uncertain_desc,
                    covered_s=float(unique_cov_est_s),
                    duration_s=float(unique_dur_s),
                )
                row = {
                    **base_row,
                    "matched_file_count": int(a_match_count.get(norm, 0)),
                    "matched_file_names": " | ".join(a_matches_norm.get(norm, [])),
                    "best_pair_confidence_score": float(a_best_conf.get(norm, 0.0)),
                    "avg_pair_confidence_score": float(a_avg_conf.get(norm, 0.0)),
                    "video_runtime_s": round(raw_runtime_s, 1),
                    "runtime_ex_dead_and_duplicate_s": round(unique_dur_s if unique_dur_s > 0 else playable_runtime_s, 1),
                    "timeline_a_unique_covered_est_pct": round(max(0.0, min(100.0, unique_cov_est_pct)), 1),
                    "timeline_a_self_repeat_redundant_pct": round(max(0.0, min(100.0, repeat_pct)), 1),
                    "timeline_coverage_est_pct": round(max(0.0, min(100.0, cov_est_pct)), 1),
                    "timeline_coverage_sections_est": sections,
                    "overlap_multimatch": bool(a_multimatch_norm.get(norm, False)),
                    "has_internal_unmatched_gaps": bool(a_internal_gap_norm.get(norm, False)),
                    "has_middle_unmatched_gap": bool(has_middle_gap),
                    "coverage_summary": coverage_summary,
                }
            rename_queue_rows.append(row)

    rename_queue_df = pd.DataFrame(rename_queue_rows)
    b_group_compare_df = pd.DataFrame(b_group_compare_rows, columns=B_GROUP_COMPARE_SHEET_COLUMNS)
    if not rename_queue_df.empty:
        group_min_a_cov_map: dict[str, float] = {}
        group_min_b_cov_map: dict[str, float] = {}
        group_best_conf_map: dict[str, float] = {}
        group_any_multi_map: dict[str, bool] = {}
        group_any_gap_map: dict[str, bool] = {}
        group_any_middle_gap_map: dict[str, bool] = {}
        group_reco_map: dict[str, str] = {}
        group_tier_map: dict[str, str] = {}

        for gid, grp in rename_queue_df.groupby("group_id", dropna=True):
            gid_s = str(gid)
            a_grp = grp[grp["file_side"].fillna("").astype(str) == "A"]
            b_grp = grp[grp["file_side"].fillna("").astype(str) == "B"]
            a_cov = pd.to_numeric(a_grp.get("timeline_a_unique_covered_est_pct", pd.Series([], dtype=float)), errors="coerce").dropna()
            b_cov = pd.to_numeric(b_grp.get("timeline_coverage_est_pct", pd.Series([], dtype=float)), errors="coerce").dropna()
            conf = pd.to_numeric(grp.get("best_pair_confidence_score", pd.Series([], dtype=float)), errors="coerce").dropna()
            min_a_cov = float(a_cov.min()) if not a_cov.empty else float("nan")
            med_a_cov = float(a_cov.median()) if not a_cov.empty else float("nan")
            min_b_cov = float(b_cov.min()) if not b_cov.empty else float("nan")
            med_b_cov = float(b_cov.median()) if not b_cov.empty else float("nan")
            best_conf = float(conf.max()) if not conf.empty else 0.0
            any_multi = bool(grp["overlap_multimatch"].fillna(False).astype(bool).any())
            any_gap = bool(grp["has_internal_unmatched_gaps"].fillna(False).astype(bool).any())
            any_middle_gap = bool(grp["has_middle_unmatched_gap"].fillna(False).astype(bool).any())
            kind_vals = grp["group_kind"].dropna().astype(str)
            group_kind = str(kind_vals.iloc[0]) if not kind_vals.empty else ""

            reco = "manual_review"
            tier = "C"
            if group_kind == "matched_group":
                if (
                    (not np.isnan(min_a_cov))
                    and (not np.isnan(min_b_cov))
                    and (min_a_cov >= 95.0)
                    and (min_b_cov >= 90.0)
                    and (best_conf >= 82.0)
                    and (not any_middle_gap)
                ):
                    reco = "likely_safe_retire_a"
                    tier = "A"
                elif (
                    (((not np.isnan(med_a_cov)) and (med_a_cov >= 85.0)) or ((not np.isnan(med_b_cov)) and (med_b_cov >= 85.0)))
                    and (best_conf >= 72.0)
                    and (not any_middle_gap)
                ):
                    reco = "spot_check"
                    tier = "B"
                elif (
                    (best_conf >= 78.0)
                    and (not any_middle_gap)
                    and (not any_gap)
                ):
                    reco = "spot_check"
                    tier = "B"
                if any_multi and reco == "likely_safe_retire_a":
                    # Multiple overlaps are expected for split/overlapping captures.
                    # Only downgrade when the multimatch still leaves material ambiguity.
                    multimatch_still_ambiguous = (
                        any_gap
                        or any_middle_gap
                        or np.isnan(min_a_cov)
                        or np.isnan(min_b_cov)
                        or (min_a_cov < 99.5)
                        or (min_b_cov < 99.5)
                    )
                    if multimatch_still_ambiguous:
                        reco = "spot_check"
                        tier = "B"
            elif group_kind == "unmatched_a":
                reco = "unmatched_keep_a"
                tier = "U"
            elif group_kind == "unmatched_b":
                reco = "unmatched_keep_b"
                tier = "U"

            group_min_a_cov_map[gid_s] = min_a_cov
            group_min_b_cov_map[gid_s] = min_b_cov
            group_best_conf_map[gid_s] = best_conf
            group_any_multi_map[gid_s] = any_multi
            group_any_gap_map[gid_s] = any_gap
            group_any_middle_gap_map[gid_s] = any_middle_gap
            group_reco_map[gid_s] = reco
            group_tier_map[gid_s] = tier

        rename_queue_df["group_min_a_unique_covered_est_pct"] = rename_queue_df["group_id"].map(
            lambda g: group_min_a_cov_map.get(str(g), np.nan)
        )
        rename_queue_df["group_min_b_coverage_est_pct"] = rename_queue_df["group_id"].map(
            lambda g: group_min_b_cov_map.get(str(g), np.nan)
        )
        rename_queue_df["group_best_confidence_score"] = rename_queue_df["group_id"].map(
            lambda g: group_best_conf_map.get(str(g), np.nan)
        )
        rename_queue_df["group_any_overlap_multimatch"] = rename_queue_df["group_id"].map(
            lambda g: group_any_multi_map.get(str(g), False)
        )
        rename_queue_df["group_any_internal_unmatched_gaps"] = rename_queue_df["group_id"].map(
            lambda g: group_any_gap_map.get(str(g), False)
        )
        rename_queue_df["group_any_middle_unmatched_gap"] = rename_queue_df["group_id"].map(
            lambda g: group_any_middle_gap_map.get(str(g), False)
        )
        rename_queue_df["group_recommendation"] = rename_queue_df["group_id"].map(
            lambda g: group_reco_map.get(str(g), "")
        )
        rename_queue_df["triage_tier"] = rename_queue_df["group_id"].map(
            lambda g: group_tier_map.get(str(g), "U")
        )
        tier_order = {"A": 0, "B": 1, "C": 2, "U": 3}
        kind_order = {"matched_group": 0, "unmatched_a": 1, "unmatched_b": 2}
        rename_queue_df["_tier_order"] = rename_queue_df["triage_tier"].map(lambda t: tier_order.get(str(t), 9)).astype(int)
        rename_queue_df["_group_kind_order"] = rename_queue_df["group_kind"].map(lambda k: kind_order.get(str(k), 9)).astype(int)
        rename_queue_df = rename_queue_df.sort_values(
            by=["_tier_order", "_group_kind_order", "_group_order", "_member_order"],
            ascending=[True, True, True, True],
        ).reset_index(drop=True)
        rename_queue_df = rename_queue_df.drop(columns=["_tier_order", "_group_kind_order"], errors="ignore")

    remux_plan_rows: list[dict[str, object]] = []
    remux_short_rows: list[dict[str, object]] = []
    a_pair_rows_for_remux: dict[str, pd.DataFrame] = {
        str(k): v.copy()
        for k, v in df_out.groupby("file_a", dropna=True)
        if isinstance(k, str)
    }
    remux_borderline_b_grid_cache: dict[tuple[str, float], dict[float, tuple[np.ndarray, np.ndarray]]] = {}

    def _estimate_borderline_b_match_for_segment(
        file_a: str,
        seg_start_s: float,
        seg_end_s: float,
    ) -> dict[str, object]:
        grp = a_pair_rows_for_remux.get(str(file_a))
        if grp is None or grp.empty:
            return {}
        seg_len_s = max(0.0, float(seg_end_s) - float(seg_start_s))
        if seg_len_s <= 0.0 or seg_len_s > float(REMUX_BORDERLINE_MAX_SEGMENT_S):
            return {}
        probe_descs = _sample_segment_probe_descriptors(
            file_a,
            float(seg_start_s),
            float(seg_end_s),
            probe_count=int(REMUX_BORDERLINE_PROBE_COUNT),
        )
        if len(probe_descs) < 2:
            return {}
        best: dict[str, object] | None = None
        best_score: tuple[float, float, float] | None = None
        for _, row in grp.iterrows():
            file_b = str(row.get("file_b", "") or "")
            if not file_b:
                continue
            dur_b = pd.to_numeric(pd.Series([row.get("timeline_est_duration_b_s")]), errors="coerce").iloc[0]
            if pd.isna(dur_b) or float(dur_b) <= float(seg_len_s):
                continue
            grid_key = (str(file_b), float(REMUX_BORDERLINE_SEARCH_STEP_S))
            b_grid = remux_borderline_b_grid_cache.get(grid_key)
            if b_grid is None:
                b_grid = _build_timeline_descriptor_grid(
                    file_b,
                    duration_s=float(dur_b),
                    step_s=float(REMUX_BORDERLINE_SEARCH_STEP_S),
                )
                remux_borderline_b_grid_cache[grid_key] = b_grid
            if not b_grid:
                continue
            best_coarse_start = None
            best_coarse_score = -1.0
            second_coarse_score = -1.0
            max_start_s = max(0.0, float(dur_b) - float(seg_len_s))
            start_s = 0.0
            while start_s <= max_start_s + 1e-6:
                avg_score, used_n = _score_segment_against_descriptor_grid(
                    probe_descs,
                    b_grid,
                    float(start_s),
                )
                if used_n >= 2:
                    if avg_score > best_coarse_score:
                        second_coarse_score = best_coarse_score
                        best_coarse_score = avg_score
                        best_coarse_start = float(start_s)
                    elif avg_score > second_coarse_score:
                        second_coarse_score = avg_score
                start_s += float(REMUX_BORDERLINE_SEARCH_STEP_S)
            if best_coarse_start is None:
                continue
            refine_start = max(0.0, float(best_coarse_start) - float(REMUX_BORDERLINE_REFINE_RADIUS_S))
            refine_end = min(max_start_s, float(best_coarse_start) + float(REMUX_BORDERLINE_REFINE_RADIUS_S))
            refine_grid = _build_timeline_descriptor_grid(
                file_b,
                start_s=float(refine_start),
                duration_s=max(0.0, (float(refine_end) - float(refine_start)) + float(seg_len_s) + float(REMUX_BORDERLINE_REFINE_STEP_S)),
                step_s=float(REMUX_BORDERLINE_REFINE_STEP_S),
            )
            best_refined_start = best_coarse_start
            best_refined_score = best_coarse_score
            refine_pos = refine_start
            while refine_pos <= refine_end + 1e-6:
                avg_score, used_n = _score_segment_against_descriptor_grid(
                    probe_descs,
                    refine_grid,
                    float(refine_pos),
                )
                if used_n >= 2 and avg_score > best_refined_score:
                    best_refined_score = avg_score
                    best_refined_start = float(refine_pos)
                refine_pos += float(REMUX_BORDERLINE_REFINE_STEP_S)
            if best_refined_score < float(REMUX_BORDERLINE_SCORE_MIN):
                continue
            if second_coarse_score > -0.5 and (best_refined_score - float(second_coarse_score)) < float(REMUX_BORDERLINE_SCORE_MARGIN):
                continue
            conf = float(pd.to_numeric(pd.Series([row.get("confidence_score")]), errors="coerce").fillna(0.0).iloc[0])
            score = (-float(best_refined_score), -float(conf), float(best_refined_start))
            if best_score is None or score < best_score:
                b_start = max(0.0, min(float(dur_b), float(best_refined_start)))
                b_end = max(0.0, min(float(dur_b), float(best_refined_start) + float(seg_len_s)))
                best_score = score
                best = {
                    "borderline_maybe_file_b_name": Path(file_b).name,
                    "borderline_maybe_b_start_s": round(float(b_start), 1),
                    "borderline_maybe_b_end_s": round(float(b_end), 1),
                    "borderline_maybe_summary": (
                        f"possible visual match to {Path(file_b).name} around "
                        f"{_format_hms_compact(max(0.0, float(b_start)))}-"
                        f"{_format_hms_compact(max(0.0, float(b_end)))} "
                        f"(descriptor search score {best_refined_score:.2f})"
                    ),
                }
        return best or {}

    if not rename_queue_df.empty:
        a_rows_for_remux = rename_queue_df[
            (rename_queue_df["row_type"].fillna("").astype(str) == "file")
            & (rename_queue_df["file_side"].fillna("").astype(str) == "A")
        ].copy()
        for _, arow in a_rows_for_remux.iterrows():
            path = str(arow.get("file_path", "") or "")
            if not path:
                continue
            norm = _norm_path(path)
            unmatched_abs = list(coverage_a_unique_unmatched_segments_abs_map.get(path, []))
            uncertain_abs = list(coverage_a_unique_uncertain_segments_abs_map.get(path, []))
            if not unmatched_abs:
                continue
            seg_idx = 0
            for seg_start_s, seg_end_s in unmatched_abs:
                seg_len_s = max(0.0, float(seg_end_s) - float(seg_start_s))
                if seg_len_s < float(REMUX_PLAN_SHORT_MIN_UNIQUE_SEGMENT_S):
                    continue
                seg_idx += 1
                uncertain_overlap_s = _union_interval_length(
                    _intersect_intervals_simple([(float(seg_start_s), float(seg_end_s))], uncertain_abs),
                    gap_tolerance_s=0.0,
                )
                if uncertain_overlap_s <= 0.0:
                    seg_status = "definite_unique"
                elif uncertain_overlap_s >= (seg_len_s - 1e-6):
                    seg_status = "borderline_maybe_match"
                else:
                    seg_status = "mixed_unique_and_borderline"
                seg_osd_start, seg_osd_end = _summarize_timestamp_interval(path, seg_start_s, seg_end_s)
                borderline_match_info = {}
                if seg_status != "definite_unique":
                    borderline_match_info = _estimate_borderline_b_match_for_segment(path, seg_start_s, seg_end_s)
                    if seg_status == "borderline_maybe_match" and not borderline_match_info:
                        seg_status = "uncertain_unresolved"
                seg_key = _remux_segment_state_key(path, seg_start_s, seg_end_s)
                seg_state = prior_remux_state.get(seg_key, {})
                if not seg_state:
                    candidates = prior_remux_state_by_path.get(_norm_path(path), [])
                    best_state = None
                    best_score = None
                    for item in candidates:
                        try:
                            old_s = float(item.get("start_s", 0.0))
                            old_e = float(item.get("end_s", 0.0))
                        except Exception:
                            continue
                        inter = max(0.0, min(float(seg_end_s), old_e) - max(float(seg_start_s), old_s))
                        union = max(float(seg_end_s), old_e) - min(float(seg_start_s), old_s)
                        overlap_frac = (inter / union) if union > 0.0 else 0.0
                        edge_delta = abs(float(seg_start_s) - old_s) + abs(float(seg_end_s) - old_e)
                        if overlap_frac <= 0.0:
                            continue
                        score = (-overlap_frac, edge_delta)
                        if best_score is None or score < best_score:
                            best_score = score
                            best_state = item.get("state", {})
                    if best_state and (best_score is not None):
                        overlap_frac = -float(best_score[0])
                        edge_delta = float(best_score[1])
                        if overlap_frac >= 0.60 or edge_delta <= 30.0:
                            seg_state = dict(best_state)
                row_payload = {
                    "workflow_status": _safe_text_cell(seg_state.get("workflow_status", "")),
                    "proposed_output_name": _safe_text_cell(seg_state.get("proposed_output_name", "")),
                    "unique_segment_status": seg_status,
                    "file_a_name": _safe_text_cell(arow.get("file_name", "")),
                    "segment_index": int(seg_idx),
                    "matched_file_b_names": _safe_text_cell(arow.get("matched_file_names", "")),
                    "borderline_maybe_file_b_name": _safe_text_cell(borderline_match_info.get("borderline_maybe_file_b_name", "")),
                    "borderline_maybe_b_start_s": borderline_match_info.get("borderline_maybe_b_start_s", ""),
                    "borderline_maybe_b_end_s": borderline_match_info.get("borderline_maybe_b_end_s", ""),
                    "borderline_maybe_summary": _safe_text_cell(borderline_match_info.get("borderline_maybe_summary", "")),
                    "unique_segment_duration_s": round(seg_len_s, 1),
                    "unique_segment_summary": _remux_segment_summary(seg_status, seg_len_s, uncertain_overlap_s),
                    "file_a_unique_summary": _safe_text_cell(arow.get("coverage_summary", "")),
                    "unique_segment_osd_month_year_start": _safe_text_cell(seg_osd_start),
                    "unique_segment_osd_month_year_end": _safe_text_cell(seg_osd_end),
                    "file_a_osd_month_year_start": _safe_text_cell(arow.get("osd_month_year_start", "")),
                    "file_a_osd_month_year_end": _safe_text_cell(arow.get("osd_month_year_end", "")),
                    "source_file_path": path,
                    "segment_start_s": round(float(seg_start_s), 1),
                    "segment_end_s": round(float(seg_end_s), 1),
                    "video_runtime_s": arow.get("video_runtime_s", ""),
                    "runtime_ex_dead_and_duplicate_s": arow.get("runtime_ex_dead_and_duplicate_s", ""),
                    "group_id": _safe_text_cell(arow.get("group_id", "")),
                    "notes": _safe_text_cell(seg_state.get("notes", "")),
                }
                if seg_len_s >= float(REMUX_PLAN_MIN_UNIQUE_SEGMENT_S):
                    remux_plan_rows.append(row_payload)
                else:
                    remux_short_rows.append(row_payload)
    remux_plan_df = pd.DataFrame(remux_plan_rows)
    remux_short_df = pd.DataFrame(remux_short_rows)
    remux_source_candidates = int(
        sum(
            1
            for segs in coverage_a_unique_unmatched_segments_abs_map.values()
            if any((float(e) - float(s)) >= float(REMUX_PLAN_MIN_UNIQUE_SEGMENT_S) for s, e in segs)
        )
    )
    print(
        (
            f"[consolidate] remux plan candidates: files={remux_source_candidates}, "
            f"long_rows={len(remux_plan_rows)}, short_rows={len(remux_short_rows)}"
        ),
        flush=True,
    )
    if not remux_plan_df.empty:
        remux_plan_df = remux_plan_df.sort_values(
            by=["group_id", "file_a_name", "segment_start_s"],
            ascending=[True, True, True],
        ).reset_index(drop=True)
        remux_plan_df = _select_existing_columns(remux_plan_df, REMUX_PLAN_SHEET_COLUMNS)
        for col in [
            "workflow_status",
            "proposed_output_name",
            "unique_segment_status",
            "borderline_maybe_file_b_name",
            "borderline_maybe_summary",
            "unique_segment_osd_month_year_start",
            "unique_segment_osd_month_year_end",
            "file_a_osd_month_year_start",
            "file_a_osd_month_year_end",
            "matched_file_b_names",
            "file_a_unique_summary",
            "unique_segment_summary",
            "file_a_name",
            "source_file_path",
            "group_id",
            "notes",
        ]:
            if col in remux_plan_df.columns:
                remux_plan_df[col] = remux_plan_df[col].map(_safe_text_cell)

    if not remux_short_df.empty:
        remux_short_df = remux_short_df.sort_values(
            by=["group_id", "file_a_name", "segment_start_s"],
            ascending=[True, True, True],
        ).reset_index(drop=True)
        remux_short_df = _select_existing_columns(remux_short_df, REMUX_PLAN_SHEET_COLUMNS)
        for col in [
            "workflow_status",
            "proposed_output_name",
            "unique_segment_status",
            "borderline_maybe_file_b_name",
            "borderline_maybe_summary",
            "unique_segment_osd_month_year_start",
            "unique_segment_osd_month_year_end",
            "file_a_osd_month_year_start",
            "file_a_osd_month_year_end",
            "matched_file_b_names",
            "file_a_unique_summary",
            "unique_segment_summary",
            "file_a_name",
            "source_file_path",
            "group_id",
            "notes",
        ]:
            if col in remux_short_df.columns:
                remux_short_df[col] = remux_short_df[col].map(_safe_text_cell)

    def _done_status_mask(series: pd.Series) -> pd.Series:
        done_values = {"done", "complete", "completed", "archived"}
        norm = (
            series.fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .str.replace(r"[^a-z]", "", regex=True)
        )
        return norm.isin(done_values)

    rename_done_df = pd.DataFrame(columns=rename_queue_df.columns)
    if not rename_queue_df.empty:
        is_done = _done_status_mask(rename_queue_df["workflow_status"])
        rename_done_df = rename_queue_df[is_done].copy().reset_index(drop=True)
        rename_queue_df = rename_queue_df[~is_done].copy().reset_index(drop=True)
        # Hard guard: prevent a file from appearing in both queue and done tabs.
        if (not rename_done_df.empty) and ("file_path" in rename_done_df.columns) and ("file_path" in rename_queue_df.columns):
            done_keys = set(
                rename_done_df["file_path"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )
            done_keys.discard("")
            if done_keys:
                queue_keys = (
                    rename_queue_df["file_path"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .str.lower()
                )
                rename_queue_df = rename_queue_df[~queue_keys.isin(done_keys)].copy().reset_index(drop=True)

    def _insert_group_separators(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df.copy()
        cols = list(df.columns)
        out_rows: list[dict[str, object]] = []
        prev_gid = None
        for _, row in df.iterrows():
            gid = str(row.get("group_id", ""))
            if prev_gid is not None and gid != prev_gid:
                sep = {c: "" for c in cols}
                sep["row_type"] = "separator"
                out_rows.append(sep)
            out_rows.append(row.to_dict())
            prev_gid = gid
        return pd.DataFrame(out_rows, columns=cols)

    rename_queue_df = _insert_group_separators(rename_queue_df)
    rename_done_df = _insert_group_separators(rename_done_df)

    rename_queue_df = _select_existing_columns(rename_queue_df, RENAME_WORKFLOW_SHEET_COLUMNS)
    if not rename_done_df.empty:
        rename_done_df = _select_existing_columns(rename_done_df, RENAME_WORKFLOW_SHEET_COLUMNS)

    def _validate_report_invariants(named_frames: dict[str, pd.DataFrame]) -> None:
        issues: list[str] = []

        def _check_est_vs_raw(df: pd.DataFrame, sheet: str, est_col: str, raw_col: str, id_cols: list[str]) -> None:
            if est_col not in df.columns or raw_col not in df.columns:
                return
            est = pd.to_numeric(df[est_col], errors="coerce")
            raw = pd.to_numeric(df[raw_col], errors="coerce")
            bad = df[(est + 1e-9) < raw]
            if bad.empty:
                return
            show_cols = [c for c in id_cols + [est_col, raw_col] if c in bad.columns]
            issues.append(f"{sheet}: {est_col} < {raw_col}: {bad[show_cols].head(3).to_dict('records')}")

        def _check_sections(df: pd.DataFrame, sheet: str, est_col: str, raw_col: str, id_cols: list[str]) -> None:
            if est_col not in df.columns or raw_col not in df.columns:
                return
            est = pd.to_numeric(df[est_col], errors="coerce")
            raw = pd.to_numeric(df[raw_col], errors="coerce")
            bad = df[(est - 1e-9) > raw]
            if bad.empty:
                return
            show_cols = [c for c in id_cols + [est_col, raw_col] if c in bad.columns]
            issues.append(f"{sheet}: {est_col} > {raw_col}: {bad[show_cols].head(3).to_dict('records')}")

        def _check_absurd_seconds(df: pd.DataFrame, sheet: str, cols: list[str], id_cols: list[str]) -> None:
            for col in cols:
                if col not in df.columns:
                    continue
                vals = pd.to_numeric(df[col], errors="coerce")
                bad = df[vals.abs() > 1_000_000.0]
                if bad.empty:
                    continue
                show_cols = [c for c in id_cols + [col] if c in bad.columns]
                issues.append(f"{sheet}: absurd seconds in {col}: {bad[show_cols].head(3).to_dict('records')}")

        for sheet, df in named_frames.items():
            if df is None or df.empty:
                continue
            ids = [c for c in ["file_a", "file_b", "file_name"] if c in df.columns]
            _check_est_vs_raw(df, sheet, "timeline_b_coverage_est_pct", "timeline_b_coverage_raw_pct", ids)
            _check_est_vs_raw(df, sheet, "timeline_coverage_est_pct", "timeline_coverage_raw_pct", ids)
            _check_sections(df, sheet, "timeline_b_coverage_sections_est", "timeline_b_coverage_sections_raw", ids)
            _check_sections(df, sheet, "timeline_coverage_sections_est", "timeline_coverage_sections_raw", ids)
            _check_absurd_seconds(
                df,
                sheet,
                [
                    "timeline_est_overlap_s",
                    "timeline_est_overlap_raw_s",
                    "timeline_est_a_start_s",
                    "timeline_est_a_end_s",
                    "timeline_est_b_start_s",
                    "timeline_est_b_end_s",
                    "timeline_coverage_est_s",
                    "timeline_coverage_raw_s",
                    "timeline_b_coverage_est_s",
                    "timeline_b_coverage_raw_s",
                    "segment_start_s",
                    "segment_end_s",
                    "video_runtime_s",
                ],
                ids,
            )

        if issues:
            raise RuntimeError("Report invariant check failed:\n" + "\n".join(issues))

    _validate_report_invariants(
        {
            "Consolidated": consolidated_view,
            "A_Coverage": a_coverage_df,
            "B_Coverage": b_coverage_df,
            "Rename_Queue": rename_queue_df,
            "Rename_Done": rename_done_df,
            "Remux_Plan": remux_plan_df,
            "Remux_Short": remux_short_df,
        }
    )

    def _blankify_string_nans(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        out = df.copy()
        for col in out.columns:
            if pd.api.types.is_object_dtype(out[col]) or pd.api.types.is_string_dtype(out[col]):
                out[col] = out[col].map(_safe_text_cell)
        return out

    rename_queue_df = _blankify_string_nans(rename_queue_df)
    rename_done_df = _blankify_string_nans(rename_done_df)
    remux_plan_df = _blankify_string_nans(remux_plan_df)
    remux_short_df = _blankify_string_nans(remux_short_df)
    b_group_compare_df = _blankify_string_nans(b_group_compare_df)
    consolidated_view = _blankify_string_nans(consolidated_view)
    a_coverage_df = _blankify_string_nans(a_coverage_df)
    b_coverage_df = _blankify_string_nans(b_coverage_df)
    if unmatched_a_df is not None:
        unmatched_a_df = _blankify_string_nans(unmatched_a_df)
    if unmatched_b_df is not None:
        unmatched_b_df = _blankify_string_nans(unmatched_b_df)

    _write_consolidated_workbook(
        output_path=output_path,
        row_count=len(df_out),
        rename_queue_df=rename_queue_df,
        rename_done_df=rename_done_df,
        remux_plan_df=remux_plan_df,
        remux_short_df=remux_short_df,
        b_group_compare_df=b_group_compare_df,
        consolidated_view=consolidated_view,
        a_coverage_df=a_coverage_df,
        b_coverage_df=b_coverage_df,
        unmatched_a_df=unmatched_a_df,
        unmatched_b_df=unmatched_b_df,
    )
    return df_out


def open_excel_file(path):
    """Open the Excel file with the OS default application (best-effort)."""
    try:
        if platform.system() == "Windows":
            os.startfile(path)  # type: ignore
        elif platform.system() == "Darwin":
            subprocess.call(["open", path])
        elif platform.system() == "Linux":
            subprocess.call(["xdg-open", path])
    except Exception as e:
        print(f"Could not open Excel file: {e}")


# 
# PyCharm-friendly harness (edit paths + mode, then Run)
# 

def run_dedupe_pipeline(
    *,
    directories: list[str],
    run_anchors_scan: bool = False,
    run_legacy_scan: bool = False,
    run_audio_scan: bool = False,
    run_timeline_scan: bool = True,
    run_timestamp_scan: bool = False,
    consolidate_include_anchors: bool = True,
    consolidate_include_legacy: bool = True,
    consolidate_include_audio: bool = False,
    consolidate_include_timeline: bool = True,
    consolidate_include_timestamps: bool = True,
    auto_tune_timeline: bool = False,
    timeline_enable_speed_sweep: bool = TIMELINE_ENABLE_SPEED_SWEEP,
    timestamp_step_s: float = TIMESTAMP_SCAN_STEP_S,
    timestamp_frame_w: int = TIMESTAMP_SCAN_FRAME_W,
    timestamp_frame_h: int = TIMESTAMP_SCAN_FRAME_H,
    timestamp_workers: int = TIMESTAMP_SCAN_WORKERS,
    expected_parts_per_b: int = 2,
    backup_label: str = "",
    open_consolidated_report: bool = False,
) -> dict:
    """
    End-to-end project pipeline:
      1) optional backups
      2) optional auto-tuning per method
      3) optional rescans (anchors/legacy/audio/timeline/timestamps)
      4) consolidated report rebuild
    """
    if len(directories) < 2:
        raise ValueError("run_dedupe_pipeline expects at least two directories (folder A and folder B).")
    folder_a = directories[0]
    folder_b = directories[1]

    # Explicit report names only (no glob), so old suffixed files are ignored.
    anchors_report = os.path.join(REPORTS_DIR, "dedupe_anchors.xlsx")
    legacy_report = os.path.join(REPORTS_DIR, "dedupe_legacy.xlsx")
    audio_report = os.path.join(REPORTS_DIR, "dedupe_audio.xlsx")
    timeline_report = os.path.join(REPORTS_DIR, "dedupe_timeline.xlsx")
    timestamp_report = os.path.join(REPORTS_DIR, "dedupe_timestamps.xlsx")
    consolidated_report = os.path.join(REPORTS_DIR, "dedupe_consolidated.xlsx")
    tuning_report = os.path.join(REPORTS_DIR, "dedupe_tuning.xlsx")
    truth_seed_report = os.path.join(REPORTS_DIR, "dedupe_truth_seed.xlsx")

    overwrite_reports = [consolidated_report, tuning_report]
    if run_anchors_scan:
        overwrite_reports.append(anchors_report)
    if run_legacy_scan:
        overwrite_reports.append(legacy_report)
    if run_audio_scan:
        overwrite_reports.append(audio_report)
    if run_timeline_scan:
        overwrite_reports.append(timeline_report)
    if run_timestamp_scan:
        overwrite_reports.append(timestamp_report)
    backups = _backup_reports(overwrite_reports, label=backup_label)

    # Tune using the previous consolidated report when available.
    truth_report = backups.get(consolidated_report)
    if not truth_report:
        truth_report = consolidated_report if os.path.exists(consolidated_report) else ""

    # If there is no previous consolidated report, synthesize one from available mode reports.
    if not truth_report:
        truth_inputs = []
        seed_reports = []
        if consolidate_include_anchors:
            seed_reports.append(anchors_report)
        if consolidate_include_legacy:
            seed_reports.append(legacy_report)
        if consolidate_include_audio:
            seed_reports.append(audio_report)
        if consolidate_include_timeline:
            seed_reports.append(timeline_report)
        for p in seed_reports:
            src = backups.get(p, p)
            if os.path.exists(src):
                truth_inputs.append(src)
        if truth_inputs:
            consolidate_dedupe_reports(
                report_paths=truth_inputs,
                folder_a=folder_a,
                folder_b=folder_b,
                output_path=truth_seed_report,
            )
            truth_report = truth_seed_report

    tuned = {"timeline": {}}
    want_any_tune = bool(auto_tune_timeline)
    if want_any_tune and truth_report and os.path.exists(truth_report):
        tuned = auto_tune_timeline_from_reports(
            directories=directories,
            truth_report_path=truth_report,
            tuning_report_path=tuning_report,
            expected_parts_per_b=expected_parts_per_b,
            use_cache_replay=True,
            enable_timeline=auto_tune_timeline,
        )
        tuned["timeline"] = tuned.get("timeline", {}) if auto_tune_timeline else {}

        print(
            f"[tune] enabled modes: timeline={'yes' if auto_tune_timeline else 'no'}",
            flush=True,
        )
    elif want_any_tune:
        print("[tune] no truth report available; running with defaults for untuned mode(s).", flush=True)
    else:
        print("[tune] auto-tuning disabled; using defaults.", flush=True)

    # Re-scan selected modes.
    if run_anchors_scan:
        find_video_duplicates(
            directories=directories,
            refine_mode="anchors",
            report_path=anchors_report,
            open_report=False,
            export_report=True,
        )
    if run_legacy_scan:
        find_video_duplicates(
            directories=directories,
            refine_mode="legacy",
            report_path=legacy_report,
            open_report=False,
            export_report=True,
        )
    if run_audio_scan:
        find_video_duplicates(
            directories=directories,
            refine_mode="audio",
            report_path=audio_report,
            open_report=False,
            export_report=True,
        )
    if run_timeline_scan:
        timeline_kwargs = dict(tuned.get("timeline", {}))
        timeline_kwargs.setdefault("timeline_enable_speed_sweep", bool(timeline_enable_speed_sweep))
        find_video_duplicates(
            directories=directories,
            refine_mode="timeline",
            report_path=timeline_report,
            open_report=False,
            export_report=True,
            **timeline_kwargs,
        )
    if run_timestamp_scan:
        scan_video_timestamps(
            directories=directories,
            report_path=timestamp_report,
            export_report=True,
            open_report=False,
            step_s=timestamp_step_s,
            frame_w=timestamp_frame_w,
            frame_h=timestamp_frame_h,
            timestamp_workers=timestamp_workers,
        )

    # Rebuild consolidated report from selected fresh outputs and stable reports.
    fresh_report_candidates = []
    if consolidate_include_anchors:
        fresh_report_candidates.append(anchors_report)
    if consolidate_include_legacy:
        fresh_report_candidates.append(legacy_report)
    if consolidate_include_audio:
        fresh_report_candidates.append(audio_report)
    if consolidate_include_timeline:
        fresh_report_candidates.append(timeline_report)
    resolvable_reports = [p for p in fresh_report_candidates if _resolve_report_path(p)[0]]
    if not resolvable_reports:
        raise FileNotFoundError(
            "No reports selected/found for consolidation. "
            "Set consolidate_include_* flags or generate matching reports first."
        )
    prior_state_report = backups.get(consolidated_report)
    if not prior_state_report and os.path.exists(consolidated_report):
        prior_state_report = consolidated_report
    consolidate_dedupe_reports(
        report_paths=fresh_report_candidates,
        folder_a=folder_a,
        folder_b=folder_b,
        output_path=consolidated_report,
        state_source_path=prior_state_report,
        timestamp_report_path=(
            timestamp_report
            if (consolidate_include_timestamps and os.path.exists(timestamp_report))
            else None
        ),
    )
    if open_consolidated_report:
        open_excel_file(consolidated_report)
    print(f"[pipeline] completed. consolidated report: {consolidated_report}", flush=True)

    return {
        "audio_report": audio_report,
        "timeline_report": timeline_report,
        "timestamp_report": timestamp_report,
        "consolidated_report": consolidated_report,
        "tuning_report": tuning_report,
        "truth_report": truth_report,
        "tuned": tuned,
    }


if __name__ == "__main__":
    print()
    # Example: compare two folders. Edit these before running directly.
    folder_a = r""
    folder_b = r""
    directories = [folder_a, folder_b]

    # Run knobs (PyCharm-friendly; edit and Run)
    PIPELINE_CONFIG = {
        # Default: rerun timeline, then rebuild consolidated workbook.
        # Set run_timestamp_scan=True to build/update dedupe_timestamps.xlsx.
        "run_anchors_scan": False,
        "run_legacy_scan": False,
        "run_audio_scan": False,
        "run_timeline_scan": False,
        "run_timestamp_scan": False,
        "consolidate_include_anchors": True,
        "consolidate_include_legacy": False,
        "consolidate_include_audio": False,
        "consolidate_include_timeline": True,
        "consolidate_include_timestamps": True,
        "auto_tune_timeline": False,
        "timeline_enable_speed_sweep": TIMELINE_ENABLE_SPEED_SWEEP,
        "timestamp_step_s": TIMESTAMP_SCAN_STEP_S,
        "timestamp_frame_w": TIMESTAMP_SCAN_FRAME_W,
        "timestamp_frame_h": TIMESTAMP_SCAN_FRAME_H,
        "timestamp_workers": TIMESTAMP_SCAN_WORKERS,
        "expected_parts_per_b": 2,
        "backup_label": "",
        "open_consolidated_report": True,
    }

    if not folder_a or not folder_b:
        print("Set folder_a and folder_b in __main__ before running this file directly.")
    else:
        run_dedupe_pipeline(
            directories=directories,
            **PIPELINE_CONFIG,
        )
