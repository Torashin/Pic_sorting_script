from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl

import video_dedupe as vd


DONE_STATUS_VALUES = {"done", "complete", "completed", "archived"}
EXECUTE_STATUS_VALUES = {"run", "ready", "queued", "queue", "selected", "go", "yes", "approved"}
WORKBOOK_USER_EDIT_HEADERS = {"notes", "proposed_new_name", "proposed_output_name"}
WORKBOOK_PATH_HEADERS = {"file_path", "source_file_path"}
WORKBOOK_DIRECT_NAME_HEADERS = {"file_name", "file_a", "file_b", "file_a_name", "borderline_maybe_file_b_name"}
WORKBOOK_NAME_LIST_HEADERS = {"matched_file_names", "matched_file_a_names", "matched_file_b_names", "matched_file_b_names"}


@dataclass
class RenameAction:
    sheet_name: str
    row_index: int
    source_path: str
    target_path: str
    proposed_new_name: str
    workflow_status: str
    execution_status: str


@dataclass
class RemuxAction:
    sheet_name: str
    row_index: int
    source_path: str
    output_path: str
    start_s: float
    end_s: float
    start_keyframe_mode: str
    end_keyframe_mode: str
    proposed_output_name: str
    workflow_status: str
    execution_status: str


@dataclass
class EditAction:
    sheet_name: str
    row_index: int
    edit_group: str
    part_index: int
    source_path: str
    output_path: str
    start_s: float
    end_s: float
    start_keyframe_mode: str
    end_keyframe_mode: str
    output_name: str
    workflow_status: str
    execution_status: str


@dataclass(frozen=True)
class KeyframeAlignedClip:
    requested_start_s: float
    requested_end_s: float
    remux_start_s: float
    remux_end_s: float
    padded_head_s: float
    padded_tail_s: float


def _run_subprocess(cmd: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        cmd,
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )


def ffprobe_duration_seconds(path: str) -> float:
    proc = _run_subprocess(
        [
            "ffprobe",
            "-hide_banner",
            "-v",
            "error",
            "-show_entries",
            "format=duration",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            path,
        ]
    )
    if proc.returncode != 0:
        raise RuntimeError(f"ffprobe duration failed for {path}:\n{proc.stderr}")
    try:
        return float(proc.stdout.strip())
    except ValueError as exc:
        raise RuntimeError(f"Could not parse duration for {path}") from exc


def ffprobe_stream_signatures(path: str) -> list[dict[str, object]]:
    proc = _run_subprocess(
        [
            "ffprobe",
            "-hide_banner",
            "-v",
            "error",
            "-show_streams",
            "-of",
            "json",
            path,
        ]
    )
    if proc.returncode != 0:
        raise RuntimeError(f"ffprobe streams failed for {path}:\n{proc.stderr}")
    try:
        data = json.loads(proc.stdout or "{}")
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Could not decode ffprobe streams for {path}") from exc
    out: list[dict[str, object]] = []
    for stream in data.get("streams", []):
        if not isinstance(stream, dict):
            continue
        codec_type = str(stream.get("codec_type") or "")
        codec_name = str(stream.get("codec_name") or "")
        if not codec_type or not codec_name:
            continue
        sig: dict[str, object] = {
            "codec_type": codec_type,
            "codec_name": codec_name,
            "codec_tag_string": str(stream.get("codec_tag_string") or ""),
            "pix_fmt": str(stream.get("pix_fmt") or ""),
            "width": int(stream.get("width") or 0) if codec_type == "video" else 0,
            "height": int(stream.get("height") or 0) if codec_type == "video" else 0,
            "sample_rate": str(stream.get("sample_rate") or "") if codec_type == "audio" else "",
            "channels": int(stream.get("channels") or 0) if codec_type == "audio" else 0,
            "channel_layout": str(stream.get("channel_layout") or "") if codec_type == "audio" else "",
        }
        out.append(sig)
    return out


def _stream_signature_key(signature: dict[str, object]) -> tuple[object, ...]:
    return (
        signature.get("codec_type"),
        signature.get("codec_name"),
        signature.get("codec_tag_string"),
        signature.get("pix_fmt"),
        signature.get("width"),
        signature.get("height"),
        signature.get("sample_rate"),
        signature.get("channels"),
        signature.get("channel_layout"),
    )


def _stream_signature_summary(signatures: list[dict[str, object]]) -> str:
    parts: list[str] = []
    for sig in signatures:
        kind = str(sig.get("codec_type") or "?")
        codec = str(sig.get("codec_name") or "?")
        if kind == "video":
            dims = ""
            if int(sig.get("width") or 0) and int(sig.get("height") or 0):
                dims = f" {int(sig.get('width'))}x{int(sig.get('height'))}"
            parts.append(f"video {codec}{dims}")
        elif kind == "audio":
            rate = str(sig.get("sample_rate") or "")
            chans = int(sig.get("channels") or 0)
            suffix = ""
            if rate:
                suffix += f" {rate}Hz"
            if chans:
                suffix += f" {chans}ch"
            parts.append(f"audio {codec}{suffix}")
        else:
            parts.append(f"{kind} {codec}")
    return "; ".join(parts)


def get_video_keyframes(path: str) -> list[float]:
    proc = _run_subprocess(
        [
            "ffprobe",
            "-hide_banner",
            "-v",
            "error",
            "-select_streams",
            "v:0",
            "-skip_frame",
            "nokey",
            "-show_frames",
            "-show_entries",
            "frame=best_effort_timestamp_time,pkt_dts_time,pkt_pts_time",
            "-of",
            "json",
            path,
        ]
    )
    if proc.returncode != 0:
        raise RuntimeError(f"ffprobe keyframes failed for {path}:\n{proc.stderr}")
    try:
        data = json.loads(proc.stdout or "{}")
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Could not decode ffprobe keyframes for {path}") from exc
    out: list[float] = []
    for frame in data.get("frames", []):
        if not isinstance(frame, dict):
            continue
        val = frame.get("best_effort_timestamp_time")
        if val in (None, ""):
            val = frame.get("pkt_pts_time")
        if val in (None, ""):
            val = frame.get("pkt_dts_time")
        if val in (None, ""):
            continue
        try:
            t_s = float(val)
        except Exception:
            continue
        if math.isfinite(t_s):
            out.append(t_s)
    return sorted(set(round(x, 6) for x in out if x >= 0.0))


def align_clip_to_keyframes(
    path: str,
    *,
    start_s: float,
    end_s: float,
    start_keyframe_mode: str = "cover",
    end_keyframe_mode: str = "cover",
    keyframes: list[float] | None = None,
) -> KeyframeAlignedClip:
    req_start = max(0.0, float(start_s))
    req_end = max(req_start, float(end_s))
    duration_s = ffprobe_duration_seconds(path)
    req_end = min(req_end, duration_s)
    start_mode = _normalize_keyframe_mode(start_keyframe_mode)
    end_mode = _normalize_keyframe_mode(end_keyframe_mode)
    if keyframes is None:
        keyframes = get_video_keyframes(path)
    if not keyframes:
        return KeyframeAlignedClip(
            requested_start_s=req_start,
            requested_end_s=req_end,
            remux_start_s=req_start,
            remux_end_s=req_end,
            padded_head_s=0.0,
            padded_tail_s=0.0,
        )
    prev_start_keys = [t for t in keyframes if t <= req_start]
    next_start_keys = [t for t in keyframes if t >= req_start]
    prev_end_keys = [t for t in keyframes if t <= req_end]
    next_end_keys = [t for t in keyframes if t >= req_end]

    if req_start <= 0.0:
        remux_start = 0.0
    elif start_mode == "inside":
        remux_start = next_start_keys[0] if next_start_keys else duration_s
    else:
        remux_start = prev_start_keys[-1] if prev_start_keys else 0.0

    if req_end >= duration_s:
        remux_end = duration_s
    elif end_mode == "inside":
        remux_end = prev_end_keys[-1] if prev_end_keys else 0.0
    else:
        remux_end = next_end_keys[0] if next_end_keys else duration_s

    remux_start = max(0.0, min(remux_start, duration_s))
    remux_end = max(remux_start, min(remux_end, duration_s))
    if remux_end <= remux_start and req_end > req_start:
        raise RuntimeError(
            "No keyframe-safe content remains after applying "
            f"start_keyframe_mode={start_mode!r}, end_keyframe_mode={end_mode!r} "
            f"to requested clip {req_start:.3f}-{req_end:.3f}s"
        )
    return KeyframeAlignedClip(
        requested_start_s=req_start,
        requested_end_s=req_end,
        remux_start_s=remux_start,
        remux_end_s=remux_end,
        padded_head_s=max(0.0, req_start - remux_start),
        padded_tail_s=max(0.0, remux_end - req_end),
    )


def remux_clip_lossless(
    input_path: str,
    output_path: str,
    *,
    start_s: float,
    end_s: float,
    align_to_keyframes: bool = True,
    start_keyframe_mode: str = "cover",
    end_keyframe_mode: str = "cover",
    overwrite: bool = True,
) -> KeyframeAlignedClip:
    input_path = str(input_path)
    output_path = str(output_path)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    aligned = align_clip_to_keyframes(
        input_path,
        start_s=start_s,
        end_s=end_s,
        start_keyframe_mode=start_keyframe_mode,
        end_keyframe_mode=end_keyframe_mode,
    ) if align_to_keyframes else KeyframeAlignedClip(
        requested_start_s=max(0.0, float(start_s)),
        requested_end_s=max(0.0, float(end_s)),
        remux_start_s=max(0.0, float(start_s)),
        remux_end_s=max(0.0, float(end_s)),
        padded_head_s=0.0,
        padded_tail_s=0.0,
    )
    cmd = [
        "ffmpeg",
        "-hide_banner",
        "-nostdin",
        "-y" if overwrite else "-n",
        "-ss",
        f"{aligned.remux_start_s:.3f}",
        "-to",
        f"{aligned.remux_end_s:.3f}",
        "-i",
        input_path,
        "-map",
        "0",
        "-c",
        "copy",
        "-avoid_negative_ts",
        "make_zero",
        output_path,
    ]
    proc = _run_subprocess(cmd)
    if proc.returncode != 0:
        raise RuntimeError(f"ffmpeg remux failed for {input_path}:\n{proc.stderr}")
    return aligned


def concat_clips_lossless(
    input_paths: list[str],
    output_path: str,
    *,
    overwrite: bool = True,
) -> None:
    if not input_paths:
        raise ValueError("concat_clips_lossless requires at least one input path")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="video_edit_concat_") as tmp_dir:
        list_path = Path(tmp_dir) / "concat_list.txt"
        with list_path.open("w", encoding="utf-8") as fh:
            for path in input_paths:
                safe_path = str(Path(path).resolve()).replace("'", "'\\''")
                fh.write(f"file '{safe_path}'\n")
        cmd = [
            "ffmpeg",
            "-hide_banner",
            "-nostdin",
            "-y" if overwrite else "-n",
            "-f",
            "concat",
            "-safe",
            "0",
            "-i",
            str(list_path),
            "-c",
            "copy",
            output_path,
        ]
        proc = _run_subprocess(cmd)
        if proc.returncode != 0:
            raise RuntimeError(f"ffmpeg concat failed for {output_path}:\n{proc.stderr}")


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        return "" if text.lower() in {"nan", "nat", "none"} else text
    return str(value).strip()


def _coerce_seconds(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if hasattr(value, "total_seconds"):
        try:
            return float(value.total_seconds())
        except Exception:
            return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = _safe_text(value)
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        pass
    parts = text.split(":")
    if len(parts) == 3:
        try:
            return (float(parts[0]) * 3600.0) + (float(parts[1]) * 60.0) + float(parts[2])
        except Exception:
            return None
    return None


def _coerce_index(value: object) -> int:
    seconds = _coerce_seconds(value)
    if seconds is None:
        return 0
    try:
        return int(seconds)
    except Exception:
        return 0


def _coerce_segment_bound(
    value: object,
    *,
    bound: str,
    source_path: str,
    duration_cache: dict[str, float],
) -> float | None:
    text = _safe_text(value).lower()
    if text == "start":
        return 0.0
    if text == "end":
        if source_path not in duration_cache:
            duration_cache[source_path] = ffprobe_duration_seconds(source_path)
        return duration_cache[source_path]
    return _coerce_seconds(value)


def _normalize_keyframe_mode(value: object, *, default: str = "cover") -> str:
    text = _safe_text(value).lower()
    if text in {"inside", "inward", "exclude", "trim"}:
        return "inside"
    if text in {"cover", "outward", "expand", "include", ""}:
        return default
    return default


def _normalize_done_status(value: str) -> bool:
    norm = "".join(ch for ch in _safe_text(value).lower() if ch.isalpha())
    return norm in DONE_STATUS_VALUES


def _normalize_execution_status(value: str) -> str:
    return "".join(ch for ch in _safe_text(value).lower() if ch.isalpha())


def _should_execute_status(value: str) -> bool:
    return _normalize_execution_status(value) in EXECUTE_STATUS_VALUES


def _sanitize_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = "".join("_" if ch in bad else ch for ch in _safe_text(name))
    return out.strip().rstrip(". ")


def _build_header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = _safe_text(ws.cell(row=1, column=col_idx).value)
        if header:
            out[header] = col_idx
    return out


def _unique_preserve_order(items: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def _replace_list_member(text: str, old_name: str, new_name: str) -> str:
    if not text or not old_name or old_name == new_name:
        return text
    parts = [part.strip() for part in text.split(";")]
    changed = False
    for i, part in enumerate(parts):
        if part == old_name:
            parts[i] = new_name
            changed = True
    return "; ".join(parts) if changed else text


def _replace_name_token(text: str, old_name: str, new_name: str) -> str:
    if not text or not old_name or old_name == new_name:
        return text
    pattern = re.compile(rf"(?<!\S){re.escape(old_name)}(?!\S)")
    return pattern.sub(new_name, text)


def _save_workbook(path: Path, wb: openpyxl.Workbook, *, backup_label: str) -> None:
    prewrite_backup_done = False
    refresh_backup_before_retry = False
    while True:
        try:
            if not prewrite_backup_done and path.exists():
                vd._backup_reports([str(path)], label=backup_label)
                prewrite_backup_done = True
            if refresh_backup_before_retry and path.exists():
                vd._backup_reports([str(path)], label="retry")
                refresh_backup_before_retry = False
            wb.save(path)
            return
        except PermissionError:
            choice = vd.sg.popup_yes_no(
                f"Can't write to file:\n{path}\n\nIt might be open in Excel.\n\nRetry?",
                title="Workflow Save Failed",
                keep_on_top=True,
            )
            if choice != "Yes":
                raise PermissionError(f"Can't write workbook {path}. It may be open in Excel.")
            refresh_backup_before_retry = True


def _update_workbook_after_renames(workbook_path: str, results: list[dict[str, object]]) -> None:
    renamed = [row for row in results if row.get("status") == "renamed"]
    if not renamed:
        return
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    rename_specs: list[tuple[str, str, str, str]] = []
    for row in renamed:
        source_path = _safe_text(row.get("source_path"))
        target_path = _safe_text(row.get("target_path"))
        if not source_path or not target_path:
            continue
        rename_specs.append((source_path, target_path, Path(source_path).name, Path(target_path).name))

        sheet_name = _safe_text(row.get("sheet"))
        row_idx = int(row.get("row") or 0)
        if sheet_name not in wb.sheetnames or row_idx < 2:
            continue
        ws = wb[sheet_name]
        headers = _build_header_map(ws)
        if "execution_status" in headers:
            ws.cell(row=row_idx, column=headers["execution_status"]).value = "done"
        if "file_path" in headers:
            ws.cell(row=row_idx, column=headers["file_path"]).value = target_path
        if "file_name" in headers:
            ws.cell(row=row_idx, column=headers["file_name"]).value = Path(target_path).name

    for ws in wb.worksheets:
        headers = _build_header_map(ws)
        for row_idx in range(2, ws.max_row + 1):
            for header, col_idx in headers.items():
                if header in WORKBOOK_USER_EDIT_HEADERS:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if not isinstance(value, str) or not value:
                    continue
                new_value = value
                for old_path, new_path, old_name, new_name in rename_specs:
                    if header in WORKBOOK_PATH_HEADERS:
                        if new_value == old_path:
                            new_value = new_path
                    elif header in WORKBOOK_DIRECT_NAME_HEADERS:
                        if new_value == old_name:
                            new_value = new_name
                    elif header in WORKBOOK_NAME_LIST_HEADERS:
                        new_value = _replace_list_member(new_value, old_name, new_name)
                    elif header in {"workflow_status", "execution_status"}:
                        continue
                if new_value != value:
                    cell.value = new_value

    _save_workbook(Path(workbook_path), wb, backup_label="workflow")


def _update_workbook_after_remux(workbook_path: str, results: list[dict[str, object]]) -> None:
    remuxed = [row for row in results if row.get("status") == "remuxed"]
    if not remuxed:
        return
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    for row in remuxed:
        sheet_name = _safe_text(row.get("sheet"))
        row_idx = int(row.get("row") or 0)
        if sheet_name not in wb.sheetnames or row_idx < 2:
            continue
        ws = wb[sheet_name]
        headers = _build_header_map(ws)
        if "execution_status" in headers:
            ws.cell(row=row_idx, column=headers["execution_status"]).value = "done"
    _save_workbook(Path(workbook_path), wb, backup_label="workflow")


def _update_workbook_after_edit(workbook_path: str, results: list[dict[str, object]]) -> None:
    edited = [row for row in results if row.get("status") == "edited"]
    if not edited:
        return
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    for row in edited:
        sheet_name = _safe_text(row.get("sheet"))
        row_idx = int(row.get("row") or 0)
        if sheet_name not in wb.sheetnames or row_idx < 2:
            continue
        ws = wb[sheet_name]
        headers = _build_header_map(ws)
        if "execution_status" in headers:
            ws.cell(row=row_idx, column=headers["execution_status"]).value = "done"
    _save_workbook(Path(workbook_path), wb, backup_label="workflow")


def _split_ext(name: str, default_ext: str) -> tuple[str, str]:
    raw = _sanitize_filename(name)
    base, ext = os.path.splitext(raw)
    if ext:
        return base or raw, ext
    return raw, default_ext


def _format_hms(seconds: float) -> str:
    total = max(0, int(round(float(seconds))))
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h:02d}-{m:02d}-{s:02d}"


def _apply_path_maps(path: str, mappings: list[tuple[str, str]]) -> str:
    raw = str(path)
    norm_raw = os.path.normcase(os.path.normpath(raw))
    best: tuple[int, str] | None = None
    for src, dst in mappings:
        norm_src = os.path.normcase(os.path.normpath(src))
        if norm_raw == norm_src or norm_raw.startswith(norm_src + os.sep):
            suffix = raw[len(src):].lstrip("\\/")
            candidate = os.path.join(dst, suffix) if suffix else dst
            rank = len(norm_src)
            if best is None or rank > best[0]:
                best = (rank, candidate)
    return best[1] if best else raw


def _open_hash_store(cache_path: str | None) -> vd.VideoHashStore:
    desired = os.path.abspath(cache_path or vd.VIDEO_HASH_STORE_PATH)
    current = getattr(vd.VideoHashStore, "_inst", None)
    if current is not None and os.path.abspath(getattr(current, "path", "")) != desired:
        vd.VideoHashStore._inst = None
    return vd.VideoHashStore(desired)


def _iter_sheet_rows(ws) -> Iterable[tuple[int, dict[str, object]]]:
    headers = [_safe_text(c.value) for c in ws[1]]
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
        row = dict(zip(headers, values))
        yield row_idx, row


def _split_excel_args(text: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    i = 0
    while i < len(text):
        ch = text[i]
        if ch == '"':
            in_string = not in_string
            current.append(ch)
        elif not in_string and ch == '(':
            depth += 1
            current.append(ch)
        elif not in_string and ch == ')':
            depth = max(0, depth - 1)
            current.append(ch)
        elif not in_string and depth == 0 and ch == ',':
            parts.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
        i += 1
    parts.append("".join(current).strip())
    return parts


def _excel_col_letters_to_index(letters: str) -> int | None:
    letters = "".join(ch for ch in letters.upper() if "A" <= ch <= "Z")
    if not letters:
        return None
    out = 0
    for ch in letters:
        out = (out * 26) + (ord(ch) - ord("A") + 1)
    return out - 1


def _eval_excel_condition(expr: str, row_values: list[object]) -> bool | None:
    expr = expr.strip()
    m_eq = re.fullmatch(r'([A-Z]+)\d+\s*=\s*"([^"]*)"', expr, flags=re.IGNORECASE)
    if m_eq:
        idx = _excel_col_letters_to_index(m_eq.group(1))
        if idx is None or idx >= len(row_values):
            return None
        return _safe_text(row_values[idx]) == m_eq.group(2)
    m_neq = re.fullmatch(r'([A-Z]+)\d+\s*<>\s*"([^"]*)"', expr, flags=re.IGNORECASE)
    if m_neq:
        idx = _excel_col_letters_to_index(m_neq.group(1))
        if idx is None or idx >= len(row_values):
            return None
        return _safe_text(row_values[idx]) != m_neq.group(2)
    return None


def _eval_excel_text_formula(formula: str, row_values: list[object]) -> str | None:
    text = _safe_text(formula)
    if not text.startswith("="):
        return None
    body = text[1:].strip()
    m_if = re.fullmatch(r'IF\((.*)\)', body, flags=re.IGNORECASE)
    if not m_if:
        return None
    args = _split_excel_args(m_if.group(1))
    if len(args) != 3:
        return None
    cond_expr, true_expr, false_expr = args
    cond_expr = cond_expr.strip()
    cond_result: bool | None = None
    m_and = re.fullmatch(r'AND\((.*)\)', cond_expr, flags=re.IGNORECASE)
    m_or = re.fullmatch(r'OR\((.*)\)', cond_expr, flags=re.IGNORECASE)
    if m_and:
        conds = _split_excel_args(m_and.group(1))
        vals = [_eval_excel_condition(c, row_values) for c in conds]
        if all(v is not None for v in vals):
            cond_result = all(bool(v) for v in vals)
    elif m_or:
        conds = _split_excel_args(m_or.group(1))
        vals = [_eval_excel_condition(c, row_values) for c in conds]
        if all(v is not None for v in vals):
            cond_result = any(bool(v) for v in vals)
    else:
        cond_result = _eval_excel_condition(cond_expr, row_values)
    if cond_result is None:
        return None

    def _parse_text_literal(expr: str) -> str | None:
        expr = expr.strip()
        m = re.fullmatch(r'"([^"]*)"', expr)
        if m:
            return m.group(1)
        return None

    return _parse_text_literal(true_expr) if cond_result else _parse_text_literal(false_expr)


def _resolve_execution_status(raw_value: object, value_value: object, row_values: list[object]) -> str:
    resolved = _safe_text(value_value)
    if resolved:
        return resolved
    raw_text = _safe_text(raw_value)
    if raw_text.startswith("="):
        evaluated = _eval_excel_text_formula(raw_text, row_values)
        if evaluated is not None:
            return evaluated
    return raw_text


def _collect_rename_actions(
    workbook_path: str,
    *,
    path_maps: list[tuple[str, str]],
) -> list[RenameAction]:
    wb_raw = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    actions: list[RenameAction] = []
    for sheet_name in ("Rename_Queue", "Rename_Done"):
        if sheet_name not in wb_raw.sheetnames or sheet_name not in wb_values.sheetnames:
            continue
        ws_raw = wb_raw[sheet_name]
        ws_values = wb_values[sheet_name]
        headers = [_safe_text(c.value) for c in ws_raw[1]]
        for row_idx in range(2, ws_raw.max_row + 1):
            raw_values = [ws_raw.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
            value_values = [ws_values.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
            row = dict(zip(headers, raw_values))
            value_row = dict(zip(headers, value_values))
            source_path = _safe_text(row.get("file_path"))
            proposed = _safe_text(row.get("proposed_new_name"))
            if not source_path or not proposed:
                continue
            workflow_status = _safe_text(row.get("workflow_status"))
            execution_status = _resolve_execution_status(
                row.get("execution_status"),
                value_row.get("execution_status"),
                raw_values,
            )
            if not _should_execute_status(execution_status):
                continue
            mapped_source = _apply_path_maps(source_path, path_maps)
            base_name, ext = _split_ext(proposed, Path(source_path).suffix or ".mkv")
            target_name = base_name + ext
            target_path = str(Path(mapped_source).with_name(target_name))
            actions.append(
                RenameAction(
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    source_path=mapped_source,
                    target_path=target_path,
                    proposed_new_name=proposed,
                    workflow_status=workflow_status,
                    execution_status=execution_status,
                )
            )
    return actions


def _derive_remux_output_name(source_path: str, row: dict[str, object], start_s: float, end_s: float) -> str:
    proposed = _safe_text(row.get("proposed_output_name"))
    if proposed:
        return proposed
    stem = Path(source_path).stem
    seg_idx = _safe_text(row.get("segment_index")) or "seg"
    return f"{stem}__seg{seg_idx}__{_format_hms(start_s)}__{_format_hms(end_s)}"


def _derive_edit_output_name(row: dict[str, object], source_path: str) -> str:
    proposed = _safe_text(row.get("output_name"))
    if proposed:
        return proposed
    stem = Path(source_path).stem
    group = _safe_text(row.get("edit_group")) or "edit"
    return f"{stem}__{group}"


def _collect_remux_actions(
    workbook_path: str,
    *,
    path_maps: list[tuple[str, str]],
    output_root: str,
    include_short: bool,
) -> list[RemuxAction]:
    wb_raw = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    actions: list[RemuxAction] = []
    duration_cache: dict[str, float] = {}
    sheet_names = ["Remux_Plan"] + (["Remux_Short"] if include_short else [])
    for sheet_name in sheet_names:
        if sheet_name not in wb_raw.sheetnames or sheet_name not in wb_values.sheetnames:
            continue
        ws_raw = wb_raw[sheet_name]
        ws_values = wb_values[sheet_name]
        headers = [_safe_text(c.value) for c in ws_raw[1]]
        for row_idx in range(2, ws_raw.max_row + 1):
            raw_values = [ws_raw.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
            value_values = [ws_values.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
            row = dict(zip(headers, raw_values))
            value_row = dict(zip(headers, value_values))
            source_path = _safe_text(row.get("source_file_path"))
            mapped_source = _apply_path_maps(source_path, path_maps) if source_path else ""
            start_s = _coerce_segment_bound(
                row.get("segment_start_s"),
                bound="start",
                source_path=mapped_source,
                duration_cache=duration_cache,
            )
            end_s = _coerce_segment_bound(
                row.get("segment_end_s"),
                bound="end",
                source_path=mapped_source,
                duration_cache=duration_cache,
            )
            if not source_path or start_s is None or end_s is None or end_s <= start_s:
                continue
            workflow_status = _safe_text(row.get("workflow_status"))
            execution_status = _resolve_execution_status(
                row.get("execution_status"),
                value_row.get("execution_status"),
                raw_values,
            )
            if not _should_execute_status(execution_status):
                continue
            start_keyframe_mode = _normalize_keyframe_mode(row.get("start_keyframe_mode"))
            end_keyframe_mode = _normalize_keyframe_mode(row.get("end_keyframe_mode"))
            output_name = _derive_remux_output_name(source_path, row, start_s, end_s)
            base_name, ext = _split_ext(output_name, ".mkv")
            output_path = str(Path(output_root) / f"{base_name}{ext or '.mkv'}")
            actions.append(
                RemuxAction(
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    source_path=mapped_source,
                    output_path=output_path,
                    start_s=float(start_s),
                    end_s=float(end_s),
                    start_keyframe_mode=start_keyframe_mode,
                    end_keyframe_mode=end_keyframe_mode,
                    proposed_output_name=_safe_text(row.get("proposed_output_name")),
                    workflow_status=workflow_status,
                    execution_status=execution_status,
                )
            )
    return actions


def _collect_edit_actions(
    workbook_path: str,
    *,
    path_maps: list[tuple[str, str]],
    output_root: str,
) -> list[EditAction]:
    wb_raw = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Edit_Plan" not in wb_raw.sheetnames or "Edit_Plan" not in wb_values.sheetnames:
        return []
    ws_raw = wb_raw["Edit_Plan"]
    ws_values = wb_values["Edit_Plan"]
    headers = [_safe_text(c.value) for c in ws_raw[1]]
    actions: list[EditAction] = []
    duration_cache: dict[str, float] = {}
    for row_idx in range(2, ws_raw.max_row + 1):
        raw_values = [ws_raw.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
        value_values = [ws_values.cell(row=row_idx, column=i + 1).value for i in range(len(headers))]
        row = dict(zip(headers, raw_values))
        value_row = dict(zip(headers, value_values))
        source_path = _safe_text(row.get("source_file_path"))
        mapped_source = _apply_path_maps(source_path, path_maps) if source_path else ""
        start_s = _coerce_segment_bound(
            row.get("segment_start_s"),
            bound="start",
            source_path=mapped_source,
            duration_cache=duration_cache,
        )
        end_s = _coerce_segment_bound(
            row.get("segment_end_s"),
            bound="end",
            source_path=mapped_source,
            duration_cache=duration_cache,
        )
        if not source_path or start_s is None or end_s is None or end_s <= start_s:
            continue
        execution_status = _resolve_execution_status(
            row.get("execution_status"),
            value_row.get("execution_status"),
            raw_values,
        )
        if not _should_execute_status(execution_status):
            continue
        start_keyframe_mode = _normalize_keyframe_mode(row.get("start_keyframe_mode"))
        end_keyframe_mode = _normalize_keyframe_mode(row.get("end_keyframe_mode"))
        edit_group = _safe_text(row.get("edit_group")) or _safe_text(row.get("output_name")) or Path(source_path).stem
        output_name = _derive_edit_output_name(row, source_path)
        base_name, ext = _split_ext(output_name, ".mkv")
        output_path = str(Path(output_root) / f"{base_name}{ext or '.mkv'}")
        part_index = _coerce_index(row.get("part_index"))
        actions.append(
            EditAction(
                sheet_name="Edit_Plan",
                row_index=row_idx,
                edit_group=edit_group,
                part_index=part_index,
                source_path=mapped_source,
                output_path=output_path,
                start_s=float(start_s),
                end_s=float(end_s),
                start_keyframe_mode=start_keyframe_mode,
                end_keyframe_mode=end_keyframe_mode,
                output_name=output_name,
                workflow_status=_safe_text(row.get("workflow_status")),
                execution_status=execution_status,
            )
        )
    actions.sort(key=lambda a: (a.output_path.lower(), a.edit_group.lower(), a.part_index, a.start_s, a.row_index))
    return actions


def execute_rename_actions(
    actions: list[RenameAction],
    *,
    cache_path: str | None,
    apply_changes: bool,
    overwrite: bool,
) -> list[dict[str, object]]:
    vhs = _open_hash_store(cache_path)
    results: list[dict[str, object]] = []
    for action in actions:
        status = "planned"
        message = ""
        source = Path(action.source_path)
        target = Path(action.target_path)
        if not source.exists():
            status = "missing_source"
            message = "source file does not exist"
        elif source.resolve() == target.resolve():
            status = "noop"
            message = "source already has target name"
        elif target.exists() and not overwrite:
            status = "target_exists"
            message = "target already exists"
        elif apply_changes:
            target.parent.mkdir(parents=True, exist_ok=True)
            if target.exists() and overwrite:
                raise FileExistsError(f"Refusing to overwrite existing rename target: {target}")
            source.rename(target)
            moved = vhs.rename_path_key(str(source), str(target), overwrite_existing=False)
            status = "renamed"
            message = "cache entry moved" if moved else "no cache entry to move"
        results.append(
            {
                "sheet": action.sheet_name,
                "row": action.row_index,
                "source_path": str(source),
                "target_path": str(target),
                "status": status,
                "message": message,
            }
        )
    if apply_changes:
        vhs.save_if_dirty()
    return results


def execute_remux_actions(
    actions: list[RemuxAction],
    *,
    apply_changes: bool,
    overwrite: bool,
) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    for action in actions:
        source = Path(action.source_path)
        target = Path(action.output_path)
        status = "planned"
        message = ""
        aligned: KeyframeAlignedClip | None = None
        if not source.exists():
            status = "missing_source"
            message = "source file does not exist"
        elif target.exists() and not overwrite:
            status = "target_exists"
            message = "target already exists"
        elif apply_changes:
            aligned = remux_clip_lossless(
                str(source),
                str(target),
                start_s=action.start_s,
                end_s=action.end_s,
                align_to_keyframes=True,
                start_keyframe_mode=action.start_keyframe_mode,
                end_keyframe_mode=action.end_keyframe_mode,
                overwrite=overwrite,
            )
            status = "remuxed"
        results.append(
            {
                "sheet": action.sheet_name,
                "row": action.row_index,
                "source_path": str(source),
                "output_path": str(target),
                "start_s": action.start_s,
                "end_s": action.end_s,
                "status": status,
                "message": message,
                "aligned": {
                    "requested_start_s": aligned.requested_start_s,
                    "requested_end_s": aligned.requested_end_s,
                    "remux_start_s": aligned.remux_start_s,
                    "remux_end_s": aligned.remux_end_s,
                    "padded_head_s": aligned.padded_head_s,
                    "padded_tail_s": aligned.padded_tail_s,
                } if aligned else None,
            }
        )
    return results


def execute_edit_actions(
    actions: list[EditAction],
    *,
    apply_changes: bool,
    overwrite: bool,
) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    grouped: dict[tuple[str, str], list[EditAction]] = {}
    for action in actions:
        grouped.setdefault((action.output_path, action.edit_group), []).append(action)

    for (_, _), group_actions in grouped.items():
        group_actions = sorted(group_actions, key=lambda a: (a.part_index, a.start_s, a.row_index))
        target = Path(group_actions[0].output_path)
        source_paths = [str(Path(a.source_path)) for a in group_actions]
        status = "planned"
        message = ""
        concat_inputs: list[str] = []
        aligned_payloads: list[dict[str, float]] = []

        missing = [p for p in source_paths if not Path(p).exists()]
        if missing:
            status = "missing_source"
            message = f"missing source file(s): {len(missing)}"
        elif target.exists() and not overwrite:
            status = "target_exists"
            message = "target already exists"
        else:
            try:
                signatures_by_source = {p: ffprobe_stream_signatures(p) for p in _unique_preserve_order(source_paths)}
                signature_keys = {
                    tuple(_stream_signature_key(sig) for sig in sigs)
                    for sigs in signatures_by_source.values()
                }
                if len(signature_keys) != 1:
                    status = "codec_mismatch"
                    parts = [
                        f"{Path(src).name}: {_stream_signature_summary(sigs)}"
                        for src, sigs in signatures_by_source.items()
                    ]
                    message = "stream signatures differ: " + " | ".join(parts)
                elif apply_changes:
                    target.parent.mkdir(parents=True, exist_ok=True)
                    with tempfile.TemporaryDirectory(prefix="video_edit_parts_") as tmp_dir:
                        for idx, action in enumerate(group_actions, start=1):
                            part_path = str(Path(tmp_dir) / f"part_{idx:03d}.mkv")
                            aligned = remux_clip_lossless(
                                action.source_path,
                                part_path,
                                start_s=action.start_s,
                                end_s=action.end_s,
                                align_to_keyframes=True,
                                start_keyframe_mode=action.start_keyframe_mode,
                                end_keyframe_mode=action.end_keyframe_mode,
                                overwrite=True,
                            )
                            concat_inputs.append(part_path)
                            aligned_payloads.append(
                                {
                                    "requested_start_s": aligned.requested_start_s,
                                    "requested_end_s": aligned.requested_end_s,
                                    "remux_start_s": aligned.remux_start_s,
                                    "remux_end_s": aligned.remux_end_s,
                                    "padded_head_s": aligned.padded_head_s,
                                    "padded_tail_s": aligned.padded_tail_s,
                                }
                            )
                        if len(concat_inputs) == 1:
                            os.makedirs(os.path.dirname(str(target)) or ".", exist_ok=True)
                            if target.exists() and overwrite:
                                raise FileExistsError(f"Refusing to overwrite existing edit target: {target}")
                            Path(concat_inputs[0]).replace(target)
                        else:
                            concat_clips_lossless(concat_inputs, str(target), overwrite=overwrite)
                    status = "edited"
                    message = f"built from {len(group_actions)} part(s)"
            except Exception as exc:
                status = "error"
                message = str(exc)

        for idx, action in enumerate(group_actions):
            results.append(
                {
                    "sheet": action.sheet_name,
                    "row": action.row_index,
                    "edit_group": action.edit_group,
                    "source_path": action.source_path,
                    "output_path": action.output_path,
                    "start_s": action.start_s,
                    "end_s": action.end_s,
                    "status": status,
                    "message": message,
                    "aligned": aligned_payloads[idx] if idx < len(aligned_payloads) else None,
                }
            )
    return results


def _parse_path_map(values: list[str]) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for raw in values:
        if "=" not in raw:
            raise ValueError(f"Invalid --path-map value: {raw}")
        src, dst = raw.split("=", 1)
        src = src.strip().strip('"')
        dst = dst.strip().strip('"')
        if not src or not dst:
            raise ValueError(f"Invalid --path-map value: {raw}")
        out.append((src, dst))
    return out


def _write_manifest(path: str | None, payload: dict[str, object]) -> None:
    if not path:
        return
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)


def _ensure_edit_plan_sheet(workbook_path: str) -> None:
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    if "Edit_Plan" in wb.sheetnames:
        ws = wb["Edit_Plan"]
        existing_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        normalized = [_safe_text(h) for h in existing_headers]
        if normalized == list(vd.EDIT_PLAN_SHEET_COLUMNS):
            return
        # Preserve existing cell values by rebuilding the sheet with the canonical header order.
        rows: list[dict[str, object]] = []
        header_map = {str(h): idx + 1 for idx, h in enumerate(existing_headers) if h is not None}
        for row_idx in range(2, ws.max_row + 1):
            payload = {col: "" for col in vd.EDIT_PLAN_SHEET_COLUMNS}
            has_any = False
            for col in vd.EDIT_PLAN_SHEET_COLUMNS:
                cidx = header_map.get(col)
                if cidx is None:
                    continue
                val = ws.cell(row=row_idx, column=cidx).value
                payload[col] = val
                has_any = has_any or (val not in (None, ""))
            if has_any:
                rows.append(payload)
        del wb["Edit_Plan"]
        ws = wb.create_sheet("Edit_Plan")
        for col_idx, header in enumerate(vd.EDIT_PLAN_SHEET_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx).value = header
        for row_idx, payload in enumerate(rows, start=2):
            for col_idx, header in enumerate(vd.EDIT_PLAN_SHEET_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx).value = payload.get(header, "")
    else:
        ws = wb.create_sheet("Edit_Plan")
        for col_idx, header in enumerate(vd.EDIT_PLAN_SHEET_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx).value = header
    _save_workbook(Path(workbook_path), wb, backup_label="workflow")


def run_workflow(
    *,
    workbook: str = str(Path("reports") / "dedupe_consolidated.xlsx"),
    mode: str = "all",
    apply: bool = False,
    overwrite: bool = False,
    include_short: bool = False,
    output_root: str = str(Path("Other") / "remux_output"),
    cache_path: str = vd.VIDEO_HASH_STORE_PATH,
    manifest: str | None = str(Path("reports") / "workflow_run_manifest.json"),
    path_map: list[str] | None = None,
) -> int:
    workbook = Path(workbook)
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")
    _ensure_edit_plan_sheet(str(workbook))

    if mode not in {"rename", "remux", "edit", "all"}:
        raise ValueError(f"Unsupported mode: {mode}")

    path_maps = _parse_path_map(list(path_map or []))
    payload: dict[str, object] = {
        "workbook": str(workbook),
        "mode": mode,
        "apply": bool(apply),
        "path_maps": path_maps,
        "rename_results": [],
        "remux_results": [],
        "edit_results": [],
    }

    if mode in {"rename", "all"}:
        rename_actions = _collect_rename_actions(
            str(workbook),
            path_maps=path_maps,
        )
        payload["rename_results"] = execute_rename_actions(
            rename_actions,
            cache_path=cache_path,
            apply_changes=bool(apply),
            overwrite=bool(overwrite),
        )
        if apply:
            _update_workbook_after_renames(str(workbook), list(payload["rename_results"]))

    if mode in {"remux", "all"}:
        remux_actions = _collect_remux_actions(
            str(workbook),
            path_maps=path_maps,
            output_root=output_root,
            include_short=bool(include_short),
        )
        payload["remux_results"] = execute_remux_actions(
            remux_actions,
            apply_changes=bool(apply),
            overwrite=bool(overwrite),
        )
        if apply:
            _update_workbook_after_remux(str(workbook), list(payload["remux_results"]))

    if mode in {"edit", "all"}:
        edit_actions = _collect_edit_actions(
            str(workbook),
            path_maps=path_maps,
            output_root=output_root,
        )
        payload["edit_results"] = execute_edit_actions(
            edit_actions,
            apply_changes=bool(apply),
            overwrite=bool(overwrite),
        )
        if apply:
            _update_workbook_after_edit(str(workbook), list(payload["edit_results"]))

    _write_manifest(manifest, payload)

    rename_done = sum(1 for row in payload["rename_results"] if row.get("status") == "renamed")
    remux_done = sum(1 for row in payload["remux_results"] if row.get("status") == "remuxed")
    edit_done = sum(1 for row in payload["edit_results"] if row.get("status") == "edited")
    print(
        f"[workflow] mode={mode} apply={'yes' if apply else 'no'} "
        f"rename_actions={len(payload['rename_results'])} renamed={rename_done} "
        f"remux_actions={len(payload['remux_results'])} remuxed={remux_done} "
        f"edit_parts={len(payload['edit_results'])} edited={edit_done}",
        flush=True,
    )
    if manifest:
        print(f"[workflow] manifest: {manifest}", flush=True)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Execute workbook-driven rename and remux actions.")
    parser.add_argument("--workbook", default=str(Path("reports") / "dedupe_consolidated.xlsx"))
    parser.add_argument("--mode", choices=["rename", "remux", "edit", "all"], default="all")
    parser.add_argument("--apply", action="store_true", help="Perform filesystem changes. Default is dry-run.")
    parser.add_argument("--overwrite", action="store_true", help="Allow remux outputs to overwrite existing files.")
    parser.add_argument("--include-short", action="store_true", help="Include Remux_Short rows as well as Remux_Plan.")
    parser.add_argument("--output-root", default=str(Path("Other") / "remux_output"))
    parser.add_argument("--cache-path", default=vd.VIDEO_HASH_STORE_PATH)
    parser.add_argument("--manifest", default=str(Path("reports") / "workflow_run_manifest.json"))
    parser.add_argument("--path-map", action="append", default=[], help="Prefix remap OLD=NEW for safe test runs.")
    args = parser.parse_args()
    return run_workflow(
        workbook=args.workbook,
        mode=args.mode,
        apply=bool(args.apply),
        overwrite=bool(args.overwrite),
        include_short=bool(args.include_short),
        output_root=args.output_root,
        cache_path=args.cache_path,
        manifest=args.manifest,
        path_map=list(args.path_map),
    )


if __name__ == "__main__":
    print()
    # PyCharm-friendly run block. Edit values below, then Run this file directly.
    WORKFLOW_CONFIG = {
        "workbook": str(Path("reports") / "dedupe_consolidated.xlsx"),
        "mode": "edit",  # "rename", "remux", "edit", or "all"
        "apply": True,  # set True to perform real filesystem changes
        "overwrite": False,
        "include_short": False,
        "output_root": str(Path("Other") / "remux_output"),
        "cache_path": vd.VIDEO_HASH_STORE_PATH,
        "manifest": str(Path("reports") / "workflow_run_manifest.json"),
        # Example safe test remap:
        # "path_map": [
        #     r"D:\Example Videos\Archive A=C:\path\to\test\archive_a",
        #     r"E:\Example Videos\Archive B=C:\path\to\test\archive_b",
        # ],
        "path_map": [],
    }
    raise SystemExit(run_workflow(**WORKFLOW_CONFIG))
